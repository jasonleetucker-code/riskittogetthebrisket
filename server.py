"""
Dynasty Trade Calculator — Unified Server
==========================================
Single command to run everything:
    python server.py

Serves the dashboard at http://localhost:8000
Scrapes all sites every 2 hours automatically.
Manual scrape: POST http://localhost:8000/api/scrape

Requirements:
    pip install fastapi uvicorn --break-system-packages
    (Playwright + other scraper deps assumed already installed)
"""

import asyncio
import json
import math
import os
import sys
import signal
import threading
import time
import logging
import traceback
import smtplib
import gzip
import hashlib
import hmac
import shutil
import uuid
import urllib.request
import urllib.error
import urllib.parse
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from datetime import datetime, timezone
from pathlib import Path
from contextlib import asynccontextmanager
from typing import Any

from fastapi import FastAPI, BackgroundTasks, Request
from fastapi.concurrency import run_in_threadpool
from fastapi.middleware.gzip import GZipMiddleware

# RedirectResponse went with the page routes (#555). Its last four users
# were _auth_redirect_response, serve_index_alias, serve_draft_capital's
# 308, and GET /logout — all deleted. Nothing this backend serves redirects
# any more: /api/* either answers or 401s, and every redirect a user sees
# now comes from Next (frontend/middleware.js, next.config.mjs).
from fastapi.responses import (
    FileResponse,
    JSONResponse,
    Response,
)
from fastapi.staticfiles import StaticFiles

try:
    import anthropic
except ImportError:  # pragma: no cover — optional dep; chat endpoint degrades gracefully
    anthropic = None

from src.api.data_contract import (
    CONTRACT_VERSION as API_DATA_CONTRACT_VERSION,
    build_api_data_contract,
    build_api_startup_payload,
    build_rankings_delta_payload,
    get_ranking_source_registry,
    normalize_source_overrides,
    normalize_tep_multiplier,
    normalize_tep_native_multiplier,
    stamp_optimal_lineups as _stamp_optimal_lineups_owner,
    validate_api_data_contract,
)
from src.api import gameplan as _gameplan
from src.api import roster_intelligence as _roster_intelligence
from src.api import guest_passes as _guest_passes
from src.api import rank_history as _rank_history
from src.api import source_history as _source_history
from src.history import record as _history_record
from src.api import push_delivery as _push_delivery
from src.api import signal_alerts as _signal_alerts
from src.api import terminal as _terminal
from src.api import trade_simulator as _trade_simulator
from src.api import user_kv as _user_kv
from src.api import league_registry as _league_registry
from src.api import sleeper_overlay as _sleeper_overlay
from src.news import NewsService, build_default_service
from src.news import custom_alerts as _custom_alerts
from src.news.providers.espn_player import DEFAULT_MAX_TARGETS as _ESPN_NEWS_TARGET_LIMIT

# ── CONFIG ──────────────────────────────────────────────────────────────
SCRAPE_INTERVAL_HOURS = 2
PORT = 8000
HOST = "0.0.0.0"  # accessible from local network; use "127.0.0.1" for local only
SCRAPE_STALL_SECONDS = int(os.getenv("SCRAPE_STALL_SECONDS", "900"))
SCRAPE_RUN_TIMEOUT_SECONDS = int(os.getenv("SCRAPE_RUN_TIMEOUT_SECONDS", "7200"))
# The async scraper launches a Playwright Chromium without a try/finally,
# so a run-timeout cancellation skips browser.close() and orphans the
# Chromium process tree → RAM leak across repeated 2h timeouts → OOM.
# A single scrape_run_lock guarantees that once the scrape coroutine has
# exited, any surviving Chromium WE spawned is orphaned, so the finalize
# path SIGKILLs Chromium descendants of this process.  Set "0" to disable.
SCRAPE_REAP_ORPHAN_BROWSERS = os.getenv("SCRAPE_REAP_ORPHAN_BROWSERS", "1") != "0"
# /api/trade/simulate-mc is a pure-Python Monte Carlo (no numpy) run
# twice for A→B/B→A symmetrization.  On this box: 50k≈0.9s, 200k≈3.9s.
# It must never run on the event loop unbounded — that freezes every
# other request + health checks.  50k sims is statistically ample for a
# win-probability point estimate; the timeout is a backstop for a
# loaded box, not the primary control.
SIMULATE_MC_MAX_SIMS = int(os.getenv("SIMULATE_MC_MAX_SIMS", "50000"))
SIMULATE_MC_TIMEOUT_SECONDS = int(os.getenv("SIMULATE_MC_TIMEOUT_SECONDS", "10"))


# ── EMAIL ALERTS ────────────────────────────────────────────────────────
# Configure alerts via environment variables (no hardcoded secrets):
#   ALERT_ENABLED=true|false
#   ALERT_TO=you@example.com
#   ALERT_FROM=sender@gmail.com
#   ALERT_PASSWORD=<gmail app password>
# Optional alias:
#   GMAIL_APP_PASSWORD=<gmail app password>
def _env_bool(name: str, default: bool = False) -> bool:
    raw = os.getenv(name)
    if raw is None:
        return default
    return raw.strip().lower() in {"1", "true", "yes", "on"}


# FRONTEND_URL and FRONTEND_RUNTIME lived here until the page proxy was
# deleted (#555).  Both described a relationship this process no longer
# has: FRONTEND_URL was the proxy target, and FRONTEND_RUNTIME recorded
# which frontend the backend served pages from.  This backend serves no
# pages, so neither has a reader — FRONTEND_RUNTIME already had none, and
# docs/OWNER_ACTION_AUDIT_2026-07-29.md:1133 called it "obsolete as a
# variable" before this change.
#
# nginx routes pages straight to Next in production and always did
# (deploy/nginx/chaseupside-proxy.conf: only `location /api/` reaches this
# process), so Next's location is nginx's business, not ours.

ALERT_ENABLED = _env_bool("ALERT_ENABLED", False)
ALERT_TO = os.getenv("ALERT_TO", "")
ALERT_FROM = os.getenv("ALERT_FROM", "")
ALERT_PASSWORD = os.getenv("ALERT_PASSWORD") or os.getenv("GMAIL_APP_PASSWORD", "")

# Shared secret for the systemd signal-alert timer.  When set, any
# request to ``POST /api/signal-alerts/run`` with a matching
# ``Authorization: Bearer <token>`` header is accepted in place of
# the usual password-session gate.  Empty = cron auth disabled
# (only browser-sessioned admins can trigger the sweep).  Generate
# with e.g. ``openssl rand -hex 32`` and store in .env alongside the
# ALERT_* creds.  Treat it like a password: leaking it lets anyone
# force a full alert send.
SIGNAL_ALERT_CRON_TOKEN = os.getenv("SIGNAL_ALERT_CRON_TOKEN", "").strip()

# ── UPTIME WATCHDOG ────────────────────────────────────────────────────
UPTIME_CHECK_ENABLED = _env_bool("UPTIME_CHECK_ENABLED", True)
UPTIME_CHECK_URL = os.getenv(
    "UPTIME_CHECK_URL",
    "https://chaseupside.com/api/health",
).strip()
UPTIME_CHECK_INTERVAL_SEC = int(os.getenv("UPTIME_CHECK_INTERVAL_SEC", "300"))
UPTIME_CHECK_TIMEOUT_SEC = float(os.getenv("UPTIME_CHECK_TIMEOUT_SEC", "5"))
UPTIME_ALERT_FAIL_THRESHOLD = int(os.getenv("UPTIME_ALERT_FAIL_THRESHOLD", "2"))

# ── LIGHTWEIGHT AUTH GATE (PRIVATE-USE) ────────────────────────────────
# App UI is intentionally gated behind Jason login.
#
# ``JASON_LOGIN_PASSWORD`` MUST come from the environment.  Earlier
# revisions defaulted to a real string baked into source — the
# pattern is dangerous (the default IS the password until rotated)
# and shows up in any clone of the repo.  Now: read the env var,
# accept ``ALLOW_DEFAULT_LOGIN_DEV=1`` as an explicit local-dev
# escape hatch (uses ``"changeme"``, matching ``.env.example``), or
# fail fast at import time.  Production sets ``JASON_LOGIN_PASSWORD``
# in ``.env`` (loaded by the systemd unit's ``EnvironmentFile=``);
# this guard prevents a misconfigured restart from silently shipping
# a known password.
JASON_LOGIN_USERNAME = (os.getenv("JASON_LOGIN_USERNAME") or "jasonleetucker").strip()
_JASON_LOGIN_PASSWORD_RAW = (os.getenv("JASON_LOGIN_PASSWORD") or "").strip()
if not _JASON_LOGIN_PASSWORD_RAW:
    if _env_bool("ALLOW_DEFAULT_LOGIN_DEV", False):
        _JASON_LOGIN_PASSWORD_RAW = "changeme"
        logging.getLogger(__name__).warning(
            "JASON_LOGIN_PASSWORD unset; ALLOW_DEFAULT_LOGIN_DEV=1 — "
            "using local-dev placeholder.  DO NOT USE IN PRODUCTION."
        )
    else:
        raise RuntimeError(
            "JASON_LOGIN_PASSWORD env var is required.  Set it on "
            "the production .env (loaded via the systemd unit), or "
            "set ALLOW_DEFAULT_LOGIN_DEV=1 for a placeholder password "
            "during local development."
        )
JASON_LOGIN_PASSWORD = _JASON_LOGIN_PASSWORD_RAW
JASON_AUTH_COOKIE_NAME = "jason_session"
JASON_AUTH_COOKIE_SECURE = _env_bool("JASON_AUTH_COOKIE_SECURE", True)


# Match the SQLite session TTL (SESSION_TTL_DAYS, default 30) so the
# browser cookie outlives an iOS Safari tab eviction instead of dying
# as a session cookie the first time the OS reclaims memory.  Parse
# defensively so a malformed env var (e.g. ``30d``) doesn't take the
# whole server down at import time — fall back to the 30-day default.
def _session_ttl_days_seconds(default_days: float = 30.0) -> int:
    raw = os.getenv("SESSION_TTL_DAYS", "")
    try:
        days = float(raw) if raw else default_days
        if not math.isfinite(days) or days <= 0:
            days = default_days
        return int(days * 86400)
    except (TypeError, ValueError, OverflowError):
        return int(default_days * 86400)


JASON_AUTH_COOKIE_MAX_AGE = _session_ttl_days_seconds()

# How often an active session's ``last_seen_at`` is refreshed on disk.
# The sliding TTL only needs a coarse heartbeat, so we throttle the
# write to at most once per interval per session — keeping ``/api/data``
# reads off the SQLite write path while still keeping active users
# signed in indefinitely.
SESSION_TOUCH_INTERVAL_SECONDS = 6 * 3600

# Private-app allowlist.  Gates admin-only endpoints to operator
# usernames.  Env var is comma-separated, lowercase-normalised at
# load time.
PRIVATE_APP_ALLOWED_USERNAMES = frozenset(
    u.strip().lower()
    for u in (os.getenv("PRIVATE_APP_ALLOWED_USERNAMES") or "jasonleetucker").split(",")
    if u.strip()
)

# Rate limit: max 1 email per hour to avoid spam on repeated failures
_last_alert_time = 0
ALERT_COOLDOWN_SEC = 3600


def send_alert(subject: str, body: str):
    """Send an email alert. Fails silently if not configured."""
    global _last_alert_time

    if not ALERT_ENABLED or not ALERT_FROM or not ALERT_PASSWORD:
        return

    now = time.time()
    if now - _last_alert_time < ALERT_COOLDOWN_SEC:
        log.info(f"Alert suppressed (cooldown): {subject}")
        return

    try:
        msg = MIMEMultipart()
        msg["From"] = ALERT_FROM
        msg["To"] = ALERT_TO
        msg["Subject"] = f"[Dynasty Server] {subject}"

        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        html = f"""
        <div style="font-family:monospace;font-size:14px;padding:16px;">
            <h2 style="color:#ff4060;">⚠ Dynasty Server Alert</h2>
            <p><strong>Time:</strong> {timestamp}</p>
            <p><strong>Issue:</strong> {subject}</p>
            <hr>
            <pre style="background:#1a1a2e;color:#e2e8f8;padding:12px;border-radius:8px;overflow-x:auto;">{body}</pre>
        </div>
        """
        msg.attach(MIMEText(html, "html"))

        with smtplib.SMTP_SSL("smtp.gmail.com", 465) as server:
            server.login(ALERT_FROM, ALERT_PASSWORD)
            server.send_message(msg)

        _last_alert_time = now
        log.info(f"Alert sent: {subject}")
    except Exception as e:
        log.error(f"Failed to send alert email: {e}")


# ── PATHS ───────────────────────────────────────────────────────────────
BASE_DIR = Path(__file__).parent.resolve()
DATA_DIR = BASE_DIR / "data"
STATIC_DIR = BASE_DIR / "static"
SCRAPER_PATH = BASE_DIR / "Dynasty Scraper.py"

DATA_DIR.mkdir(exist_ok=True)
STATIC_DIR.mkdir(exist_ok=True)

# ── LOGGING ─────────────────────────────────────────────────────────────
# R-8: Structured JSON logging when LOG_FORMAT=json (for log aggregation).
# Default is human-readable for local dev and journalctl.
LOG_FORMAT = os.getenv("LOG_FORMAT", "text").strip().lower()

if LOG_FORMAT == "json":

    class _JsonFormatter(logging.Formatter):
        """Minimal JSON log formatter for structured log aggregation."""

        def format(self, record):
            entry = {
                "ts": datetime.now(timezone.utc).isoformat(),
                "level": record.levelname,
                "logger": record.name,
                "msg": record.getMessage(),
            }
            if record.exc_info and record.exc_info[0]:
                entry["exception"] = self.formatException(record.exc_info)
            return json.dumps(entry, ensure_ascii=False)

    _handler = logging.StreamHandler()
    _handler.setFormatter(_JsonFormatter())
    logging.basicConfig(level=logging.INFO, handlers=[_handler])
else:
    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s [%(levelname)s] %(message)s",
        datefmt="%Y-%m-%d %H:%M:%S",
    )
log = logging.getLogger("dynasty-server")

# ── STATE ───────────────────────────────────────────────────────────────
# In-memory cache of latest scrape data
latest_data: dict | None = None
latest_contract_data: dict | None = None
latest_data_bytes: bytes | None = None
latest_data_gzip_bytes: bytes | None = None
latest_data_etag: str | None = None
# Lean runtime payload (drops heavy contract-only arrays not needed by frontend startup).
latest_runtime_data: dict | None = None
latest_runtime_data_bytes: bytes | None = None
latest_runtime_data_gzip_bytes: bytes | None = None
latest_runtime_data_etag: str | None = None
# Startup-slim payload for first paint and early interaction.
latest_startup_data: dict | None = None
latest_startup_data_bytes: bytes | None = None
latest_startup_data_gzip_bytes: bytes | None = None
latest_startup_data_etag: str | None = None
# Array payload: full contract minus the LEGACY ``players`` dict.  The
# dict and ``playersArray`` are parallel encodings of the same players
# (~5.8MB + ~6.6MB of a ~12MB payload); ``playersArray`` is the richer
# one and the only one the frontend materializer reads when present, so
# desktop clients request this view and cut the wire/parse cost roughly
# in half with zero field loss.
latest_array_data: dict | None = None
latest_array_data_bytes: bytes | None = None
latest_array_data_gzip_bytes: bytes | None = None
latest_array_data_etag: str | None = None
# Compact payload for mobile / slow networks (~90% smaller).  Precomputed
# (bytes + gzip + etag) at refresh time so the ``?view=compact`` fast path
# never re-runs ``compact_contract`` + ``json.dumps`` + gzip on the event
# loop per request.
latest_compact_data: dict | None = None
latest_compact_data_bytes: bytes | None = None
latest_compact_data_gzip_bytes: bytes | None = None
latest_compact_data_etag: str | None = None
# Serialized-bytes cache for the live-overlay /api/data responses.  The
# overlay path splices per-league Sleeper data onto the rankings payload
# and must re-serialize, but the overlay itself is cached ~15 min per
# league, so the serialized result is stable within that window.
#
# Keyed by a STABLE slot — (kind, leagueKey, loadedLeague, view,
# sleeper_matches) — with the freshness stamp carried INSIDE the value:
# (etag, raw_bytes, gzip_bytes, version).  Refreshes replace their slot
# rather than minting a new key beside it, so the cache holds at most one
# multi-MB generation per slot no matter how many refresh cycles pass.
# The key space is bounded by the registry (leagues × views), so the cap
# below is a safety net, not the primary bound.
_OVERLAY_RESPONSE_CACHE: dict = {}
# Per-key single-flight locks so concurrent misses coalesce onto one
# encode instead of each launching the multi-MB serialization.  Touched
# only from the async handler on the single event-loop thread, so the
# dict needs no extra synchronization.
_OVERLAY_ENCODE_LOCKS: dict = {}
_OVERLAY_RESPONSE_CACHE_MAX = 32
# ``POST /api/rankings/overrides`` response memo.  Nearly every client
# posts the identical stock body ({"tep_multiplier": 1.15} — the
# /settings default), so one cached entry serves the whole user base
# between scrapes.  Same shape as the overlay cache above:
# ``key -> (raw_bytes, gzip_bytes, version)`` where ``version`` is the
# ``latest_data_etag`` generation the entry was built from.  Key =
# (hash of the NORMALIZED build inputs — override map, tep knobs,
# valuation mode, warnings — plus delta_view, league key,
# sleeper_matches): everything that can alter the response body, and
# nothing that cannot.  It used to hash the raw posted body, which with
# custom source weighting withdrawn (#875) meant every distinct
# source-toggle body paid a full pipeline rebuild for a byte-identical
# response; ``normalize_source_overrides`` maps them all to the same
# empty map.  Entries whose version no longer matches are rebuilt in
# place (one generation per slot); ``_prime_latest_payload`` clears the
# dict outright on scrape promotion.  The leagueAdjusted path is cached
# like any other since B9a withdrew the lens (no gameplan factors left
# to go stale).
_OVERRIDES_RESPONSE_CACHE: dict = {}
_OVERRIDES_ENCODE_LOCKS: dict = {}
_OVERRIDES_RESPONSE_CACHE_MAX = 16
latest_data_source: dict = {
    "type": "",
    "path": "",
    # When THIS PROCESS loaded a payload.  A real fact about the process,
    # and NOT a statement about the board — see ``producedAt`` below.
    "loadedAt": "",
    # When the BOARD was produced, taken from the payload's own
    # ``scrapeTimestamp``.  Empty means the payload did not say, which is
    # UNKNOWN — never a licence to substitute ``loadedAt`` (audit F-19).
    "producedAt": "",
}
contract_health: dict = {
    "ok": False,
    "status": "unknown",
    "errors": ["contract not initialized"],
    "warnings": [],
    "errorCount": 1,
    "warningCount": 0,
    "checkedAt": None,
    "contractVersion": API_DATA_CONTRACT_VERSION,
    "playerCount": 0,
}
# Per-source coverage of the CURRENTLY-SERVED contract:
# ``{sourceKey: playerCount}`` counted from
# ``playersArray[*].sourceRankMeta`` — how many players each source
# actually contributed to the live board.  Recomputed once per prime
# (not per request).  This is the ONLY reliable "is the served board
# fully enriched?" signal: ``sites`` / ``siteStats`` only ever list
# the 3 legacy-scraper sources even when the contract is fully
# CSV-enriched, so they cannot tell a healthy ~11-source board apart
# from a degraded 3-source one (the exact silent failure that left
# OTCFFB — and ~8 other sources — off the live board for days).  The
# deploy-gate (deploy/verify-deploy.sh) and the scheduled health
# check assert on this so a non-re-primed/degraded board fails
# loudly instead of silently serving wrong values.
served_source_coverage: dict = {}


def _compute_served_source_coverage(contract: dict | None) -> dict:
    """Count, per source, how many served players carry it in
    ``sourceRankMeta`` (== "contributed to the blend").  Defensive:
    any shape surprise yields ``{}`` rather than raising inside the
    prime path."""
    cov: dict[str, int] = {}
    try:
        for row in (contract or {}).get("playersArray") or []:
            srm = row.get("sourceRankMeta") if isinstance(row, dict) else None
            if isinstance(srm, dict):
                for k in srm:
                    cov[str(k)] = cov.get(str(k), 0) + 1
    except Exception:  # noqa: BLE001
        return {}
    return cov


# ── NEWS SERVICE ───────────────────────────────────────────────────────
# Lazy-built singleton.  Built on first request rather than at import
# so unit tests can monkey-patch the factory and the server can boot
# even if a transient DNS failure would block provider construction.
_news_service: NewsService | None = None
_news_service_lock = threading.Lock()


def _get_news_service() -> NewsService:
    global _news_service
    if _news_service is not None:
        return _news_service
    with _news_service_lock:
        if _news_service is None:
            _news_service = build_default_service(
                provider_config={
                    # Per-player ESPN news: the provider trickle-
                    # refreshes espn_id targets from the live board
                    # through its own per-player TTL cache — the
                    # supplier is read lazily on each provider fetch,
                    # so it always sees the current contract.
                    "espn_player": {"targets_supplier": _live_espn_news_targets},
                }
            )
    return _news_service


def _reset_news_service_for_tests(svc: NewsService | None = None) -> None:
    """Test hook — inject a stubbed service or clear the singleton."""
    global _news_service
    with _news_service_lock:
        _news_service = svc


# Memo for the full-contract scans below (~2,000-row walks that were
# re-run on every /api/terminal and /api/news request).  Key: the
# contract generation (``latest_data_etag``) — a scrape promotion mints
# a new etag, which invalidates every entry implicitly.  Values are
# shared references: CALLERS MUST TREAT THEM AS READ-ONLY (they already
# do — the news service only reads names/meta).  No caching while the
# etag is None (mid-prime / startup).
_LIVE_CONTRACT_SCAN_MEMO: dict[str, tuple[str, Any]] = {}


def _live_contract_scan(kind: str, builder):
    etag = latest_data_etag
    if not etag:
        return builder()
    hit = _LIVE_CONTRACT_SCAN_MEMO.get(kind)
    if hit is not None and hit[0] == etag:
        return hit[1]
    val = builder()
    _LIVE_CONTRACT_SCAN_MEMO[kind] = (etag, val)
    return val


def _live_player_names() -> list[str]:
    """Return every player name visible in the live contract.

    ESPN's RSS provider uses this set to tag headlines with matched
    players; returning an empty list when the contract hasn't
    loaded yet degrades gracefully — headlines still surface,
    they just arrive with empty ``players[]``.

    Memoized per contract generation — treat the result as read-only.
    """

    def _build() -> list[str]:
        contract = latest_contract_data or {}
        rows = contract.get("playersArray") or []
        names: list[str] = []
        for row in rows:
            if not isinstance(row, dict):
                continue
            for key in ("displayName", "name", "canonicalName", "fullName"):
                v = row.get(key)
                if isinstance(v, str) and v.strip():
                    names.append(v.strip())
                    break
        return names

    return _live_contract_scan("player_names", _build)


# How many top-board players get per-player ESPN news coverage.
# Rostered players cluster at the top of the board, and the provider
# trickle-refreshes ~8 ids per 3-minute aggregate cycle, so 150
# targets fully refresh roughly hourly at a polite request rate.
#
# Single source of truth: ``_ESPN_NEWS_TARGET_LIMIT`` is imported at
# the top of this file as an alias of the provider's own
# ``DEFAULT_MAX_TARGETS`` — the supplier below and the provider's
# ``_valid_targets`` truncation can never drift apart (Codex P2:
# a local 150 here vs the provider's default 100 silently discarded
# targets 101-150).


def _live_espn_news_targets() -> list[dict[str, str | None]]:
    """Top-board players joined to their ESPN athlete ids.

    Walks the live contract's ``playersArray`` in consensus-rank
    order, resolves each row's Sleeper ``playerId`` through the
    contract's Sleeper player directory (``sleeper.players`` — the
    ``/v1/players/nfl`` shape, which carries ``espn_id``), and emits
    ``{name, espnId, position, team}`` targets for the
    ``espn_player`` news provider.  Rows without an espn_id mapping
    are skipped; an unloaded contract yields [] and the provider
    stays quiet.
    """
    contract = latest_contract_data or {}
    rows = contract.get("playersArray") or []
    players_dir = (contract.get("sleeper") or {}).get("players") or {}
    if not isinstance(players_dir, dict) or not rows:
        return []

    def _rank(row: dict) -> float:
        try:
            r = float(row.get("canonicalConsensusRank") or row.get("rank") or 0)
            return r if r > 0 else float("inf")
        except (TypeError, ValueError):
            return float("inf")

    targets: list[dict[str, str | None]] = []
    for row in sorted((r for r in rows if isinstance(r, dict)), key=_rank):
        sid = str(row.get("playerId") or "").strip()
        if not sid:
            continue
        p = players_dir.get(sid)
        if not isinstance(p, dict):
            continue
        espn_id = str(p.get("espn_id") or "").strip()
        if not espn_id:
            continue
        name = ""
        for key in ("displayName", "name", "canonicalName", "fullName"):
            v = row.get(key)
            if isinstance(v, str) and v.strip():
                name = v.strip()
                break
        if not name:
            continue
        position = row.get("position")
        team = row.get("team")
        targets.append(
            {
                "name": name,
                "espnId": espn_id,
                "position": position.strip()
                if isinstance(position, str) and position.strip()
                else None,
                "team": team.strip().upper() if isinstance(team, str) and team.strip() else None,
            }
        )
        if len(targets) >= _ESPN_NEWS_TARGET_LIMIT:
            break
    return targets


def _live_player_meta() -> dict[str, dict[str, str | None]]:
    """Map exact contract display names → {position, team}.

    The news service stamps these identity discriminators onto
    tagged mentions (``NewsService._enrich_player_mentions``) so
    name-collision players (CJ Allen the LB vs C.J. Allen the WR)
    can be told apart on per-player surfaces.  ``team`` is sparsely
    populated until the next scrape cycle — null when absent;
    position is always available on contract rows.

    Memoized per contract generation — treat the result as read-only.
    """

    def _build() -> dict[str, dict[str, str | None]]:
        contract = latest_contract_data or {}
        rows = contract.get("playersArray") or []
        meta: dict[str, dict[str, str | None]] = {}
        for row in rows:
            if not isinstance(row, dict):
                continue
            name = ""
            for key in ("displayName", "name", "canonicalName", "fullName"):
                v = row.get(key)
                if isinstance(v, str) and v.strip():
                    name = v.strip()
                    break
            if not name:
                continue
            position = row.get("position")
            team = row.get("team")
            entry = {
                "position": position.strip()
                if isinstance(position, str) and position.strip()
                else None,
                "team": team.strip().upper() if isinstance(team, str) and team.strip() else None,
            }
            prior = meta.get(name)
            if prior is None:
                meta[name] = entry
            elif prior != entry:
                # Two contract rows share the EXACT display string with
                # different identities — an exact-name lookup can't
                # attribute a mention safely, so stamp nothing rather
                # than the wrong player's identity.
                meta[name] = {"position": None, "team": None}
        return meta

    return _live_contract_scan("player_meta", _build)


# R-9: Lightweight metrics counters
_metrics: dict = {
    "server_start_time": None,
    "request_count": 0,
    "scrape_total": 0,
    "scrape_failures": 0,
    "scrape_duration_seconds_last": 0.0,
    "data_age_seconds": 0.0,
}
# Canonical scrape lifecycle state.
# Compatibility aliases are maintained:
#   running -> is_running
#   error   -> last_error
scrape_status = {
    "running": False,
    "is_running": False,  # legacy alias for UI compatibility
    "hung": False,
    "stalled": False,
    # Verdict on the LAST run, not on the current moment. Set by
    # _reconcile_orphaned_running_state, cleared by _start_scrape_run.
    "interrupted": False,
    "interrupted_at": None,
    "started_at": None,
    "finished_at": None,
    "last_heartbeat": None,
    "last_scrape": None,  # last successful scrape ISO timestamp
    "last_success_at": None,
    "last_failure_at": None,
    "last_duration_sec": None,
    "next_scrape": None,  # ISO timestamp
    "error": None,
    "last_error": None,  # legacy alias for UI compatibility
    "current_step": None,
    "current_source": None,
    "progress_step_index": 0,
    "progress_step_total": 0,
    "worker_id": None,
    "scrape_count": 0,
    "run_events": [],
}
# R-4: Rolling scrape history for success rate tracking.
SCRAPE_HISTORY_MAX = 50
scrape_history: list[dict] = []

# Single-owner run lock: only one scrape run can own mutable active state.
scrape_run_lock = asyncio.Lock()
uptime_status = {
    "enabled": UPTIME_CHECK_ENABLED,
    "target_url": UPTIME_CHECK_URL,
    "last_check": None,
    "last_ok": None,
    "last_error": None,
    "last_http_status": None,
    "consecutive_failures": 0,
}
# ``frontend_runtime_status`` used to sit here and was stamped onto
# /api/status and /api/health as {"configured": "next", "active": "next",
# "reason": "next_only", "fallbackFrom": None}.
#
# It is gone with the page proxy (#555). This process has no "active
# frontend runtime" and no fallback to report — it serves no pages. The
# field survived the FRONTEND_RUNTIME deletion because it is a separate
# hardcoded dict, not a reader of that constant, which is exactly how a
# stale claim outlives the thing it described.
#
# Removed rather than left as harmless: it had zero consumers anywhere
# (frontend, deploy, scripts, tests and workflows all checked), and it
# sat on /api/health — the endpoint the production uptime watchdog
# probes — telling anyone debugging that the backend's active frontend
# runtime is "next". That is the precise misconception this change
# exists to remove.
# In-memory auth sessions for private-use gate.
auth_sessions: dict[str, dict] = {}

# Startup validation summary — populated by lifespan; surfaced via
# /api/health.  Default to an empty summary so the endpoint never
# references an unbound name before lifespan runs.
_startup_checks_summary: dict = {"total": 0, "ok": 0, "failed": 0, "fatal": 0, "checks": []}


def _utc_now_iso() -> str:
    return datetime.now(timezone.utc).isoformat()


# R-10: Disk space guard — minimum free space before writing data files (in MB)
DISK_SPACE_MIN_MB = int(os.getenv("DISK_SPACE_MIN_MB", "500"))


def _check_disk_space(path: Path | None = None) -> tuple[bool, int]:
    """Check if there's enough disk space. Returns (ok, free_mb)."""
    target = path or DATA_DIR
    try:
        usage = shutil.disk_usage(str(target))
        free_mb = usage.free // (1024 * 1024)
        return free_mb >= DISK_SPACE_MIN_MB, free_mb
    except OSError:
        # If we can't check, allow the write (fail-open)
        return True, -1


def _sanitize_next_path(raw: str | None, default: str = "/") -> str:
    """Reduce a ``?next=`` value to a same-origin path, or refuse it.

    PARSE AND COMPARE, not blocklist.  The previous implementation was a
    list of string tests — reject ``http://``, ``https://``, anything not
    starting with ``/``, anything starting with ``//``, CR and LF — and it
    was bypassed by a character it did not name: **the backslash**.

    ``/\\attacker.tld`` starts with ``/``, does not start with ``//``, has
    no scheme prefix and contains no newline, so every guard passed and the
    value was returned verbatim.  Browsers then normalise ``\\`` to ``/``
    when resolving a URL, turning it into the protocol-relative
    ``//attacker.tld`` — a working post-login open redirect to an arbitrary
    host, on the real domain.  The same trick works with a tab or a raw
    control character, which browsers strip *before* resolving
    (``/\\tattacker.tld``).

    So: normalise the input the way a browser would, then judge the result
    structurally.  Anything that parses with a scheme or a netloc is not a
    path on this origin, whatever it looked like as a string.
    """
    value = str(raw or "").strip()
    if not value:
        return default

    # Refuse rather than repair.  Browsers strip TAB/LF/CR from a URL and
    # treat a backslash as a path separator, so both are load-bearing to the
    # attack — but no path this application mints contains either, so the
    # honest answer to one appearing is "no", not a normalised guess at what
    # was meant.  Rejecting also keeps this function's output a subset of its
    # input, which is what makes it auditable.
    if any(ord(ch) < 0x20 or ord(ch) == 0x7F for ch in value):
        return default
    if "\\" in value:
        return default

    try:
        parts = urllib.parse.urlsplit(value)
    except ValueError:
        return default

    # A scheme or a netloc means it addresses some other origin.  This is
    # the check the string tests were approximating; it does not care
    # whether the author remembered a particular character.
    if parts.scheme or parts.netloc:
        return default
    if not parts.path.startswith("/") or parts.path.startswith("//"):
        return default

    # Re-emit from the parsed parts so the caller receives the normalised
    # form that was actually judged, never the raw string.
    return urllib.parse.urlunsplit(("", "", parts.path, parts.query, parts.fragment))


def _get_auth_session(request: Request) -> dict | None:
    session_id = str(request.cookies.get(JASON_AUTH_COOKIE_NAME, "")).strip()
    if not session_id:
        return None
    session = auth_sessions.get(session_id)
    if not isinstance(session, dict):
        return None
    # Per-session expiry — set when the session originated from a
    # time-bounded guest pass (see ``/api/auth/login`` fall-through).
    # Cookie max-age is the client-side ceiling but a tampered or
    # stretched cookie wouldn't extend the pass's authority.  We
    # double-check on every request and silently evict expired
    # sessions so guests can't outlive their pass.
    expires_at = session.get("expires_at_epoch")
    if isinstance(expires_at, (int, float)) and expires_at > 0:
        if time.time() >= float(expires_at):
            auth_sessions.pop(session_id, None)
            try:
                from src.api import session_store as _ss

                _ss.evict(session_id)
            except Exception as exc:  # noqa: BLE001
                log.warning(
                    "session_store evict on expiry failed: %s",
                    exc,
                )
            return None
    # Sliding TTL heartbeat.  Bump ``last_seen_at`` at most once per
    # SESSION_TOUCH_INTERVAL_SECONDS so an actively-used session never
    # expires, without writing to SQLite on every request.  Best-effort:
    # a failed touch just means the session ages toward its idle TTL.
    now = time.time()
    last_touch = session.get("_last_touch_epoch")
    if not isinstance(last_touch, (int, float)):
        last_touch = session.get("last_seen_epoch")
    # A session with no recorded heartbeat is touched immediately;
    # otherwise the write is throttled to once per interval.  Missing
    # data drives an explicit branch rather than a fabricated epoch.
    if (
        not isinstance(last_touch, (int, float))
        or now - float(last_touch) >= SESSION_TOUCH_INTERVAL_SECONDS
    ):
        session["_last_touch_epoch"] = now
        try:
            from src.api import session_store as _ss

            _ss.touch(session_id)
        except Exception as exc:  # noqa: BLE001
            log.warning("session_store touch failed: %s", exc)
    return session


def _is_authenticated(request: Request) -> bool:
    return _get_auth_session(request) is not None


def _create_auth_session(
    username: str,
    *,
    sleeper_user_id: str | None = None,
    display_name: str | None = None,
    avatar: str | None = None,
    auth_method: str = "password",
    expires_at_epoch: float | None = None,
    guest_pass_id: int | None = None,
) -> str:
    """Create an in-memory session for ``username``.

    ``sleeper_user_id`` / ``display_name`` / ``avatar`` are optional
    and consumed by downstream handlers that resolve the user's
    Sleeper team by ownerId (populated today only by the E2E test
    login path).  ``auth_method`` tags the session for logs and
    audit tooling.

    ``expires_at_epoch`` (optional) bounds the session's authority.
    Guest-pass logins set this to the pass's expiry so the session
    silently dies when the pass does — see ``_get_auth_session``.
    ``guest_pass_id`` records which pass minted the session, so a
    revoked pass can be cross-referenced if needed (audit + future
    bulk-revoke-on-pass-revoke if we want it).
    """
    session_id = uuid.uuid4().hex
    payload: dict[str, Any] = {
        "username": str(username or ""),
        "sleeper_user_id": str(sleeper_user_id or ""),
        "display_name": str(display_name or username or ""),
        "avatar": str(avatar or ""),
        "auth_method": str(auth_method or "password"),
        "created_at": _utc_now_iso(),
    }
    if isinstance(expires_at_epoch, (int, float)) and expires_at_epoch > 0:
        payload["expires_at_epoch"] = float(expires_at_epoch)
    if isinstance(guest_pass_id, int) and guest_pass_id > 0:
        payload["guest_pass_id"] = int(guest_pass_id)
    auth_sessions[session_id] = payload
    # Persist the session so a deploy/restart doesn't force a re-login.
    # Best-effort: any SQLite failure falls through to in-memory-only
    # behavior (the existing behavior — no regression).
    try:
        from src.api import session_store as _ss

        _ss.persist(
            session_id,
            payload,
            allowlist=PRIVATE_APP_ALLOWED_USERNAMES,
        )
    except Exception as exc:  # noqa: BLE001
        log.warning("session_store persist in _create_auth_session failed: %s", exc)
    if len(auth_sessions) > 5000:
        oldest = sorted(
            auth_sessions.items(),
            key=lambda kv: str((kv[1] or {}).get("created_at") or ""),
        )[:500]
        for sid, _ in oldest:
            auth_sessions.pop(sid, None)
    return session_id


def _clear_auth_session(request: Request) -> None:
    session_id = str(request.cookies.get(JASON_AUTH_COOKIE_NAME, "")).strip()
    if session_id:
        auth_sessions.pop(session_id, None)
        try:
            from src.api import session_store as _ss

            _ss.evict(session_id)
        except Exception as exc:  # noqa: BLE001
            log.warning("session_store evict failed: %s", exc)


def _parse_iso(ts: str | None) -> datetime | None:
    if not ts:
        return None
    try:
        return datetime.fromisoformat(ts.replace("Z", "+00:00"))
    except Exception:
        return None


def _seconds_since_iso(ts: str | None) -> float | None:
    dt = _parse_iso(ts)
    if dt is None:
        return None
    return max(0.0, (datetime.now(timezone.utc) - dt).total_seconds())


def _trim_run_events(limit: int = 50) -> None:
    events = scrape_status.get("run_events") or []
    if len(events) > limit:
        scrape_status["run_events"] = events[-limit:]


def _record_scrape_event(event: str, level: str = "info", message: str = "", **meta) -> None:
    payload = {
        "ts": _utc_now_iso(),
        "event": event,
        "message": message,
    }
    if meta:
        payload["meta"] = meta
    scrape_status.setdefault("run_events", []).append(payload)
    _trim_run_events()

    log_line = f"[Scrape] {event}"
    if message:
        log_line += f" — {message}"
    if meta:
        log_line += f" | {meta}"
    if level == "error":
        log.error(log_line)
    elif level == "warning":
        log.warning(log_line)
    else:
        log.info(log_line)


def _touch_scrape_heartbeat() -> None:
    scrape_status["last_heartbeat"] = _utc_now_iso()


def _is_scrape_stalled() -> bool:
    if not scrape_status.get("running"):
        return False
    age = _seconds_since_iso(scrape_status.get("last_heartbeat"))
    if age is None:
        return False
    return age > SCRAPE_STALL_SECONDS


def _sync_scrape_alias_fields() -> None:
    scrape_status["is_running"] = bool(scrape_status.get("running"))
    scrape_status["last_error"] = scrape_status.get("error")


def _reconcile_orphaned_running_state() -> None:
    # Safety net: if status says running but lock is free, a prior worker exited
    # unexpectedly before state cleanup. Reset running state explicitly.
    if scrape_status.get("running") and not scrape_run_lock.locked():
        _record_scrape_event(
            "orphaned_running_reset",
            level="warning",
            message="Detected running=True without active lock; resetting state",
            worker_id=scrape_status.get("worker_id"),
        )
        scrape_status["running"] = False
        # NOT ``stalled``/``hung``.  Both were set here and both were
        # dead assignments: ``_scrape_status_payload`` calls this
        # reconciler and then immediately evaluates ``_is_scrape_stalled()``,
        # which returns False once ``running`` is False, and the else
        # branch resets ``stalled`` and ``hung`` to False three lines
        # later.  An orphaned worker therefore reported
        # ``status_summary: "idle"`` with no error — indistinguishable
        # from a healthy server that simply had not scraped recently.
        # /api/health's ``scrape_stalled`` stayed false too, so
        # StaleDataBanner never fired.
        #
        # ``interrupted`` is a property of the LAST RUN rather than of
        # the current moment, so nothing downstream recomputes it away.
        # ``_start_scrape_run`` clears it, because a new run supersedes
        # the verdict on the old one.
        scrape_status["interrupted"] = True
        scrape_status["interrupted_at"] = _utc_now_iso()
        scrape_status["finished_at"] = _utc_now_iso()
        scrape_status["current_step"] = "stale_state_reset"
        scrape_status["current_source"] = None
        _touch_scrape_heartbeat()
        _sync_scrape_alias_fields()


def _start_scrape_run(trigger: str) -> str:
    run_id = f"run-{uuid.uuid4().hex[:12]}"
    now_iso = _utc_now_iso()
    scrape_status.update(
        {
            "running": True,
            "hung": False,
            "stalled": False,
            # A fresh run supersedes any verdict on the previous one.
            "interrupted": False,
            "interrupted_at": None,
            "started_at": now_iso,
            "finished_at": None,
            "last_heartbeat": now_iso,
            "current_step": "bootstrap",
            "current_source": "server",
            "progress_step_index": 0,
            "progress_step_total": 0,
            "worker_id": run_id,
        }
    )
    _sync_scrape_alias_fields()
    _record_scrape_event(
        "scrape_started", message=f"trigger={trigger}", trigger=trigger, worker_id=run_id
    )
    return run_id


def _update_scrape_progress(
    *,
    step: str | None = None,
    source: str | None = None,
    step_index: int | None = None,
    step_total: int | None = None,
    event: str | None = None,
    message: str | None = None,
    level: str = "info",
    meta: dict | None = None,
) -> None:
    if step is not None:
        scrape_status["current_step"] = step
    if source is not None:
        scrape_status["current_source"] = source
    if step_index is not None:
        scrape_status["progress_step_index"] = int(step_index)
    if step_total is not None:
        scrape_status["progress_step_total"] = int(step_total)
    scrape_status["hung"] = False
    scrape_status["stalled"] = False
    _touch_scrape_heartbeat()
    _sync_scrape_alias_fields()
    if event:
        _record_scrape_event(event, level=level, message=message or "", **(meta or {}))


def _mark_scrape_success(
    elapsed: float, player_count: int, site_count: int, total_sites: int
) -> None:
    now_iso = _utc_now_iso()
    scrape_status.update(
        {
            "running": False,
            "hung": False,
            "stalled": False,
            "finished_at": now_iso,
            "last_scrape": now_iso,
            "last_success_at": now_iso,
            "last_duration_sec": round(elapsed, 1),
            "error": None,
            "current_step": "complete",
            "current_source": None,
            "scrape_count": int(scrape_status.get("scrape_count", 0)) + 1,
        }
    )
    _touch_scrape_heartbeat()
    _sync_scrape_alias_fields()
    _record_scrape_event(
        "scrape_succeeded",
        message=f"{player_count} players, {site_count}/{total_sites} sites, {elapsed:.1f}s",
        player_count=player_count,
        site_count=site_count,
        total_sites=total_sites,
        duration_sec=round(elapsed, 1),
    )
    # R-4: Record to rolling history
    _record_scrape_history(
        "success",
        elapsed,
        player_count=player_count,
        site_count=site_count,
        total_sites=total_sites,
    )
    # R-9: Update metrics counters
    _metrics["scrape_total"] = _metrics.get("scrape_total", 0) + 1
    _metrics["scrape_duration_seconds_last"] = round(elapsed, 1)


def _missing_expected_sites(result: dict | None) -> list[str]:
    """Anchor sources the scrape was expected to produce and did not.

    Audit O-3.  The payload declares its own load-bearing inputs in
    ``coverageAudit.expectedSites`` — ``{"offense": ["ktc"], "idp":
    ["idpTradeCalc"]}`` on live data — so "did we lose an anchor?" is
    answerable without inventing a threshold or hardcoding a source
    name here.  A source counts as produced only if it actually carried
    players; present-but-empty is exactly the degraded case the guard
    exists to catch.

    Returns ``[]`` on any shape surprise.  This runs on the scrape path
    and a diagnostic that can crash the scrape is worse than the defect
    it reports — but note that an empty list from a MALFORMED payload
    means "no anchors known to be missing", not "all anchors present",
    which is why the ratio test is kept alongside it rather than
    replaced by it.
    """
    try:
        audit = (result or {}).get("coverageAudit") or {}
        expected_block = audit.get("expectedSites") or {}
        expected: set[str] = set()
        for names in expected_block.values():
            if isinstance(names, (list, tuple)):
                expected.update(str(n) for n in names if n)
        if not expected:
            return []

        def _reported_rows(block: object, field: str) -> bool:
            """True only when the block states a positive row count.

            Absent, null or non-numeric means the source did not report
            producing anything — which is treated the same as zero HERE
            because this guard's question is "can we prove the anchor
            arrived?", and unproven must not read as arrived.  Written
            out rather than as ``or 0`` so the reasoning is visible:
            the coercion gate flags that shape precisely because it
            usually hides this decision instead of stating it.
            """
            if not isinstance(block, dict):
                return False
            count = block.get(field)
            return isinstance(count, (int, float)) and count > 0

        produced: set[str] = set()
        for site in (result or {}).get("sites") or []:
            if _reported_rows(site, "playerCount"):
                produced.add(str(site.get("key") or ""))
        for key, stats in ((result or {}).get("siteStats") or {}).items():
            if _reported_rows(stats, "count"):
                produced.add(str(key))

        return sorted(expected - produced)
    except Exception:  # noqa: BLE001 — never break the scrape over a diagnostic
        return []


def _mark_scrape_blocked(
    reason: str, elapsed: float, player_count: int, site_count: int, total_sites: int
) -> None:
    """Record a scrape that ran but whose output was REFUSED.

    Audit O-2.  The partial-scrape guard called ``_mark_scrape_success``,
    so a scrape that was rejected for being too degraded to publish was
    filed as a success — in ``scrape_status.last_success_at``, in the
    event log, and in the rolling history that ``_scrape_success_rate_24h``
    reads.  The rate therefore read 100% while every scrape was being
    thrown away, which is the precise inversion that made the "success
    rate < 50%" alert useless even once O-1 let it fire at all.

    A blocked run is neither a success nor a crash: the scraper worked,
    the server is healthy, and it is still serving last-known-good data
    — but nothing new was published, and that must be visible.  So
    ``last_success_at`` is deliberately NOT advanced, and the history
    outcome is ``blocked``, which counts against the success rate
    because ``_scrape_success_rate_24h`` counts only ``success``.
    """
    now_iso = _utc_now_iso()
    scrape_status.update(
        {
            "running": False,
            "hung": False,
            "stalled": False,
            "finished_at": now_iso,
            "last_scrape": now_iso,
            # NOT last_success_at — nothing was promoted.
            "last_blocked_at": now_iso,
            "last_duration_sec": round(elapsed, 1),
            "error": None,
            "current_step": "blocked",
            "current_source": None,
        }
    )
    _touch_scrape_heartbeat()
    _sync_scrape_alias_fields()
    _record_scrape_event(
        "scrape_blocked",
        level="warning",
        message=reason,
        player_count=player_count,
        site_count=site_count,
        total_sites=total_sites,
        duration_sec=round(elapsed, 1),
    )
    _record_scrape_history(
        "blocked",
        elapsed,
        player_count=player_count,
        site_count=site_count,
        total_sites=total_sites,
        reason=reason,
    )
    _metrics["scrape_total"] = _metrics.get("scrape_total", 0) + 1
    _metrics["scrape_blocked"] = _metrics.get("scrape_blocked", 0) + 1
    _metrics["scrape_duration_seconds_last"] = round(elapsed, 1)


def _mark_scrape_failure(exc: Exception, elapsed: float) -> None:
    now_iso = _utc_now_iso()
    error_text = f"{type(exc).__name__}: {str(exc)[:400]}"
    failed_step = scrape_status.get("current_step")
    failed_source = scrape_status.get("current_source")
    scrape_status.update(
        {
            "running": False,
            "hung": False,
            "stalled": False,
            "finished_at": now_iso,
            "last_failure_at": now_iso,
            "last_duration_sec": round(elapsed, 1),
            "error": error_text,
            "current_step": "failed",
        }
    )
    _touch_scrape_heartbeat()
    _sync_scrape_alias_fields()
    _record_scrape_event(
        "scrape_failed",
        level="error",
        message=error_text,
        failed_step=failed_step,
        failed_source=failed_source,
        duration_sec=round(elapsed, 1),
    )
    # R-4: Record to rolling history
    _record_scrape_history("failure", elapsed, error=error_text)
    # R-9: Update metrics counters
    _metrics["scrape_total"] = _metrics.get("scrape_total", 0) + 1
    _metrics["scrape_failures"] = _metrics.get("scrape_failures", 0) + 1
    _metrics["scrape_duration_seconds_last"] = round(elapsed, 1)


def _record_scrape_history(outcome: str, duration: float, **meta) -> None:
    """R-4: Append to rolling scrape history for success rate tracking."""
    entry = {
        "ts": _utc_now_iso(),
        "outcome": outcome,
        "duration_sec": round(duration, 1),
    }
    entry.update(meta)
    scrape_history.append(entry)
    # Trim to max size
    while len(scrape_history) > SCRAPE_HISTORY_MAX:
        scrape_history.pop(0)


def _scrape_success_rate_24h() -> dict:
    """R-4: Calculate scrape success rate over the last 24 hours."""
    now = datetime.now(timezone.utc)
    recent = []
    for entry in scrape_history:
        try:
            ts = datetime.fromisoformat(entry["ts"])
            if (now - ts).total_seconds() <= 86400:
                recent.append(entry)
        except (ValueError, TypeError, KeyError):
            continue
    total = len(recent)
    if total == 0:
        return {"total": 0, "success": 0, "failure": 0, "rate": None}
    successes = sum(1 for e in recent if e.get("outcome") == "success")
    return {
        "total": total,
        "success": successes,
        "failure": total - successes,
        "rate": round(successes / total, 2),
    }


def _looks_like_playwright_chromium(cmdline: str) -> bool:
    """True for a Playwright-spawned headless Chromium command line.

    Deliberately conservative: requires BOTH a chromium binary token
    AND an automation/headless marker, so it can't match an unrelated
    process even before the descendant-scoping guard.
    """
    c = cmdline.lower()
    if "chrome" not in c and "chromium" not in c:
        return False
    return (
        "--headless" in c
        or "--remote-debugging-" in c
        or "--remote-debugging-pipe" in c
        or "/ms-playwright/" in c
        or "playwright" in c
    )


def _collect_descendant_pids(root_pid: int) -> set[int]:
    """All transitive child PIDs of ``root_pid`` via /proc (Linux only).

    Uses the ppid field of /proc/<pid>/stat (field after the comm
    parenthesis) — no kernel CONFIG_PROC_CHILDREN dependency.
    """
    children: dict[int, list[int]] = {}
    try:
        entries = os.listdir("/proc")
    except OSError:
        return set()
    for entry in entries:
        if not entry.isdigit():
            continue
        try:
            with open(f"/proc/{entry}/stat", "rb") as fh:
                data = fh.read().decode("latin-1")
        except OSError:
            continue
        rparen = data.rfind(")")
        if rparen == -1:
            continue
        fields = data[rparen + 2 :].split()
        try:
            ppid = int(fields[1])  # fields[0]=state, fields[1]=ppid
        except (IndexError, ValueError):
            continue
        children.setdefault(ppid, []).append(int(entry))
    out: set[int] = set()
    stack = [root_pid]
    while stack:
        p = stack.pop()
        for child in children.get(p, ()):
            if child not in out:
                out.add(child)
                stack.append(child)
    return out


def _reap_orphan_browsers(root_pid: int | None = None, match=None) -> int:
    """SIGKILL orphaned Playwright Chromium descendants of this process.

    Best-effort; never raises.  Scoped to descendants of ``root_pid``
    (our own PID) so it can never touch unrelated system processes.
    The scrape that spawned them already exited (single scrape_run_lock),
    so a graceful shutdown buys nothing — SIGKILL directly, with no
    sleep, so this stays safe to call from the async finally path.
    """
    try:
        root = os.getpid() if root_pid is None else root_pid
        predicate = _looks_like_playwright_chromium if match is None else match
        killed = 0
        for pid in _collect_descendant_pids(root):
            try:
                with open(f"/proc/{pid}/cmdline", "rb") as fh:
                    cmd = fh.read().replace(b"\x00", b" ").decode("utf-8", "replace").strip()
            except OSError:
                continue
            if not cmd or not predicate(cmd):
                continue
            try:
                os.kill(pid, signal.SIGKILL)
                killed += 1
            except OSError:
                pass
        return killed
    except Exception:  # noqa: BLE001 — finalize must never raise
        return 0


def _finalize_scrape_run(worker_id: str) -> None:
    # Guaranteed cleanup path (always called in run_scraper finally).
    if scrape_status.get("worker_id") != worker_id:
        return
    if SCRAPE_REAP_ORPHAN_BROWSERS:
        reaped = _reap_orphan_browsers()
        if reaped:
            log.warning(
                "reaped %d orphaned Chromium process(es) after scrape "
                "(timeout/cancel left them behind)",
                reaped,
            )
    if scrape_status.get("running"):
        scrape_status["running"] = False
        if not scrape_status.get("finished_at"):
            scrape_status["finished_at"] = _utc_now_iso()
        if scrape_status.get("current_step") not in {"complete", "failed"}:
            scrape_status["current_step"] = "finalized"
            _record_scrape_event(
                "scrape_finalized_with_running_true",
                level="warning",
                message="Forced running=False during finally cleanup",
                worker_id=worker_id,
            )
    if scrape_status.get("current_step") == "complete":
        scrape_status["current_source"] = None
    _touch_scrape_heartbeat()
    _sync_scrape_alias_fields()


def _build_scrape_progress_callback(worker_id: str):
    async def _on_progress(payload: dict):
        if scrape_status.get("worker_id") != worker_id:
            return
        if not isinstance(payload, dict):
            return
        _update_scrape_progress(
            step=payload.get("step"),
            source=payload.get("source"),
            step_index=payload.get("step_index"),
            step_total=payload.get("step_total"),
            event=payload.get("event"),
            message=payload.get("message"),
            level=payload.get("level", "info"),
            meta=payload.get("meta"),
        )

    return _on_progress


def _scrape_status_payload() -> dict:
    _reconcile_orphaned_running_state()
    stalled = _is_scrape_stalled()
    was_stalled = bool(scrape_status.get("stalled"))
    if stalled:
        scrape_status["hung"] = True
        scrape_status["stalled"] = True
        if not was_stalled:
            _record_scrape_event(
                "scrape_stalled_detected",
                level="warning",
                message=(
                    f"No heartbeat update for >{SCRAPE_STALL_SECONDS}s "
                    f"(step={scrape_status.get('current_step')}, "
                    f"source={scrape_status.get('current_source')})"
                ),
                stall_threshold_sec=SCRAPE_STALL_SECONDS,
                current_step=scrape_status.get("current_step"),
                current_source=scrape_status.get("current_source"),
            )
    else:
        scrape_status["hung"] = False
        scrape_status["stalled"] = False
    _sync_scrape_alias_fields()

    payload = dict(scrape_status)
    payload["stall_threshold_sec"] = SCRAPE_STALL_SECONDS
    payload["run_timeout_sec"] = SCRAPE_RUN_TIMEOUT_SECONDS
    # Audit O-1: ops_alerts._check_scrape_rate reads
    # ``scrape_success_rate_24h``, and this payload — the one
    # ``check_and_alert`` is actually called with — never carried it.
    # The key existed only inside the /api/status route handler, so the
    # "scrape success rate < 50%" alert returned None on every sweep and
    # could not fire once, ever. Attaching it here is the whole fix on
    # this side; the reader also had to learn this dict shape (see
    # ops_alerts).
    payload["scrape_success_rate_24h"] = _scrape_success_rate_24h()
    payload["status_summary"] = (
        "stalled"
        if payload.get("stalled")
        else "running"
        if payload.get("running")
        else "failed"
        if payload.get("error")
        else "interrupted"
        if payload.get("interrupted")
        else "idle"
    )
    return payload


def _set_latest_data_source(
    source_type: str, path: str | None = None, produced_at: str | None = None
) -> None:
    """Record what we loaded, when we loaded it, and when it was PRODUCED.

    ``produced_at`` is the payload's own ``scrapeTimestamp``.  Audit F-19:
    four surfaces used to compute "how old is the data" from ``loadedAt``,
    which measures how long THIS PROCESS has been holding it — so every
    restart, including every production deploy, reset the answer to zero.
    """
    latest_data_source.update(
        {
            "type": str(source_type or ""),
            "path": str(path or ""),
            "loadedAt": _utc_now_iso(),
            "producedAt": str(produced_at or ""),
        }
    )


#: How far into the future a board's production stamp may sit before it is
#: refused as UNKNOWN rather than published as an age (audit F-28).
#:
#: Bounded from both sides rather than picked:
#:
#: * it must be LARGER than real clock skew between two NTP-synced hosts,
#:   which is seconds — five minutes is three orders of magnitude of headroom;
#: * it must be SMALLER than the smallest timezone quantum in use anywhere,
#:   which is 15 minutes (UTC+05:45, UTC+12:45).  Any tolerance at or above
#:   that could swallow a genuine timezone misreading, which is the entire
#:   condition this guard exists to catch.
_BOARD_AGE_CLOCK_SKEW_TOLERANCE_HOURS = 5.0 / 60.0


def _board_age_hours() -> float | None:
    """Hours since the loaded BOARD was produced, or ``None`` if unknown.

    THE ONE OWNER of that question (audit F-19).  ``None`` means the payload
    did not state when it was produced; it is never approximated from
    ``loadedAt``, because doing so is the defect this replaced — MISSING IS
    NEVER ZERO, and here zero is the most reassuring number the field can
    carry.

    ``scrapeTimestamp`` is tz-aware UTC as of audit F-28.  Payloads produced
    before that — every committed archive, and any board still on disk from an
    older scraper — are NAIVE (``"2026-08-18T11:04:55.664246"``).  Subtracting
    a naive datetime from a tz-aware ``now`` raises ``TypeError``, which every
    caller here swallows, so without the explicit attach below this function
    would return ``None`` for those payloads and the F-19 repair would be
    silently inert for them.  Reading a LEGACY naive stamp as UTC remains an
    assumption; it is stated here rather than left implicit, and it is now
    confined to legacy payloads instead of describing the live path.

    **A board produced in the FUTURE is UNKNOWN, never a number** (F-28).  A
    negative age is not a small error to publish and move on from: it is below
    every staleness threshold by construction, so it reports maximum freshness
    with maximum confidence at precisely the moment the stamp cannot be
    trusted.  Production served ``data_age_hours: -1.0`` for a board scraped an
    hour earlier, which made ``data_stale`` unreachable rather than merely
    inaccurate.  ``ops_alerts._check_data_freshness`` already treats ``None``
    as "no opinion" and raises no alert, so UNKNOWN degrades quietly instead of
    inventing an alarm.

    A small negative tolerance absorbs ordinary clock skew between the
    producing host and this one; beyond it the stamp is refused.
    """
    produced_at = latest_data_source.get("producedAt")
    if not produced_at:
        return None
    try:
        produced_dt = datetime.fromisoformat(str(produced_at))
    except (ValueError, TypeError):
        return None
    if produced_dt.tzinfo is None:
        produced_dt = produced_dt.replace(tzinfo=timezone.utc)
    try:
        age_h = (datetime.now(timezone.utc) - produced_dt).total_seconds() / 3600.0
    except (ValueError, TypeError, OverflowError):
        return None
    if age_h < -_BOARD_AGE_CLOCK_SKEW_TOLERANCE_HOURS:
        return None
    return max(age_h, 0.0)


# Identity-keyed cache for live rankDerivedValue lookups used by the
# trade-suggestions overlay.  ``latest_contract_data`` is replaced
# (never mutated) each time ``_prime_latest_payload`` runs, so the
# cache invalidates automatically when a fresh payload arrives.
_LIVE_BY_NAME_CACHE: dict = {"contract_id": None, "value": {}}


def _live_by_name_from_contract(contract: dict | None) -> dict[str, int]:
    """Return ``{displayName: rankDerivedValue}`` for the live contract.

    Cached by ``id(contract)`` so repeat trade-suggestion requests
    between scrapes skip the N-row walk.  Returns the cached dict by
    reference; callers must not mutate it.
    """
    cid = id(contract) if contract is not None else None
    if _LIVE_BY_NAME_CACHE["contract_id"] == cid and cid is not None:
        return _LIVE_BY_NAME_CACHE["value"]
    built: dict[str, int] = {}
    try:
        live_rows = (contract or {}).get("playersArray") or []
    except Exception:  # noqa: BLE001
        live_rows = []
    for row in live_rows:
        name = str(row.get("canonicalName") or row.get("displayName") or "").strip()
        if not name:
            continue
        rdv = row.get("rankDerivedValue")
        try:
            v = int(rdv) if rdv is not None else None
        except (TypeError, ValueError):
            continue
        if v is not None and v > 0:
            built[name] = v
    _LIVE_BY_NAME_CACHE["contract_id"] = cid
    _LIVE_BY_NAME_CACHE["value"] = built
    return built


def _build_source_health_snapshot(
    data: dict | None, coverage: dict[str, int] | None = None
) -> dict:
    """Source health over the population we are ENTITLED TO EXPECT.

    F-7 / census S-1.  The denominator used to be ``payload["sites"]``,
    which the scraper emits carrying only the two ANCHOR markets —
    measured on the live 2026-08-18 export it is exactly
    ``[ktc, idpTradeCalc]``.  So this block reported **2 of 2 healthy,
    0 missing** for a board carried by 21 registered production voters,
    and a source that went silent could not appear in
    ``missing_sources`` because it had never been in the population.

    That is ``MISSING IS NEVER ZERO`` at the health layer, and the
    repair is the rule ``src/api/confidence.py`` already applies to
    coverage: the denominator is what COULD have been observed, so
    silence is permanent missing evidence rather than a smaller
    population.

    ``coverage`` is the per-source contribution to the SERVED board
    (``_compute_served_source_coverage``, counted off ``sourceRankMeta``
    = "voted in the blend").  It is deliberately not the count of
    positive ``canonicalSiteValues``: DraftSharks IDP publishes real
    NEGATIVE values, so a ``> 0`` predicate reports 143 of its 269
    actual votes.  Sign-bearing sources make ">0" the wrong question.

    Without ``coverage`` the per-source contribution is UNKNOWN, and
    unknown is not zero — the counts come back ``None`` and the keys
    land in ``unmeasured_sources`` rather than being accused of going
    silent.  The scraper's own anchor row counts are kept, under a name
    that says what they are (``anchor_row_counts``), because they are a
    different quantity from board contribution and merging the two into
    one map is the defect being repaired.

    NOT ``coverageAudit.expectedSites``: that block is an anchor-loss
    detector and 2 is correct for it.  Different question, different
    owner.
    """
    payload = data or {}
    sites = payload.get("sites")
    if not isinstance(sites, list):
        sites = []
    anchor_row_counts: dict[str, int | None] = {}
    for row in sites:
        if not isinstance(row, dict):
            continue
        key = str(row.get("key") or "").strip()
        if not key:
            continue
        # A row with no ``playerCount`` has not told us its size.  That
        # is unknown, not zero — and this map is a diagnostic, so it can
        # carry the distinction rather than flatten it.
        raw_player_count = row.get("playerCount")
        anchor_row_counts[key] = (
            int(raw_player_count) if isinstance(raw_player_count, (int, float)) else None
        )

    # The population owner is the ranking-source registry.  Falling back
    # to the anchor list would silently restore the defect, so a
    # registry that cannot be read is reported as such instead.
    try:
        from src.api.data_contract import get_ranking_source_keys

        registered = sorted({str(k) for k in get_ranking_source_keys() if k})
    except Exception:  # pragma: no cover - registry import failure
        registered = []

    # An EMPTY coverage map is "not measured yet", not "all 21 sources
    # went silent".  ``served_source_coverage`` is a module global that
    # starts ``{}`` and is filled at contract promotion, so treating
    # empty as measured-zero would accuse every source of silence on
    # every cold boot — the same missing-vs-zero mistake this function
    # exists to fix, reintroduced one level down.  A genuinely empty
    # served board cannot coexist with a loaded contract: the contract
    # is built FROM the sources.
    measured = bool(coverage)

    source_counts: dict[str, int | None] = {}
    missing: list[str] = []
    unmeasured: list[str] = []
    available = 0
    for key in registered:
        if not measured:
            source_counts[key] = None
            unmeasured.append(key)
            continue
        # Absent from the coverage map means the scan of the served
        # board found this source on ZERO rows.  That is a real
        # measurement over a complete pass — the map is built by walking
        # every row — not a stand-in for evidence we do not have.  The
        # "we were not given a board at all" case is the ``measured``
        # branch above, and it reports ``None``.
        raw_count = coverage.get(key)
        count = int(raw_count) if isinstance(raw_count, (int, float)) else 0
        source_counts[key] = count
        if count > 0:
            available += 1
        else:
            missing.append(key)

    failures: list[dict] = []
    seen_failures: set[tuple[str, str, str]] = set()

    def _push_failure(source: str, reason: str, details: dict | None = None) -> None:
        src = str(source or "").strip()
        rsn = str(reason or "").strip() or "unknown"
        d = details if isinstance(details, dict) else {}
        detail_sig = str(d.get("error") or d.get("message") or "")
        key = (src, rsn, detail_sig)
        if key in seen_failures:
            return
        seen_failures.add(key)
        failures.append(
            {
                "source": src,
                "reason": rsn,
                "details": d,
            }
        )

    settings = payload.get("settings") if isinstance(payload.get("settings"), dict) else {}
    dlf_import = settings.get("dlfImport") if isinstance(settings.get("dlfImport"), dict) else {}
    for src_key, meta in dlf_import.items():
        if not isinstance(meta, dict):
            continue
        if not meta.get("loaded", False):
            _push_failure(
                str(src_key),
                "not_loaded",
                {
                    "file": meta.get("file"),
                    "parseMode": meta.get("parseMode"),
                    "badRows": meta.get("badRows"),
                },
            )
        elif meta.get("stale", False):
            _push_failure(
                str(src_key),
                "stale_csv",
                {
                    "file": meta.get("file"),
                    "ageDays": meta.get("ageDays"),
                },
            )

    source_run_summary = settings.get("sourceRunSummary")
    source_runtime = {}
    partial_run = False
    if isinstance(source_run_summary, dict):
        enabled_sources = source_run_summary.get("enabledSources")
        complete_sources = source_run_summary.get("completeSources")
        partial_sources = source_run_summary.get("partialSources")
        timed_out_sources = source_run_summary.get("timedOutSources")
        failed_sources = source_run_summary.get("failedSources")
        source_rows = source_run_summary.get("sources")
        if not isinstance(enabled_sources, list):
            enabled_sources = []
        if not isinstance(complete_sources, list):
            complete_sources = []
        if not isinstance(partial_sources, list):
            partial_sources = []
        if not isinstance(timed_out_sources, list):
            timed_out_sources = []
        if not isinstance(failed_sources, list):
            failed_sources = []
        if not isinstance(source_rows, dict):
            source_rows = {}

        for src in timed_out_sources:
            row = source_rows.get(src) if isinstance(source_rows.get(src), dict) else {}
            _push_failure(
                str(src),
                "timeout",
                {
                    "error": row.get("error"),
                    "message": row.get("message"),
                    "timeoutSec": row.get("timeoutSec"),
                    "valueCount": row.get("valueCount"),
                },
            )
        for src in failed_sources:
            row = source_rows.get(src) if isinstance(source_rows.get(src), dict) else {}
            _push_failure(
                str(src),
                "failed",
                {
                    "error": row.get("error"),
                    "message": row.get("message"),
                    "valueCount": row.get("valueCount"),
                },
            )
        for src in partial_sources:
            row = source_rows.get(src) if isinstance(source_rows.get(src), dict) else {}
            _push_failure(
                str(src),
                "partial",
                {
                    "message": row.get("message"),
                    "valueCount": row.get("valueCount"),
                },
            )

        partial_run = bool(
            source_run_summary.get("partialRun")
            or partial_sources
            or timed_out_sources
            or failed_sources
        )
        source_runtime = {
            "overall_status": source_run_summary.get("overallStatus"),
            "partial_run": partial_run,
            "started_at": source_run_summary.get("startedAt"),
            "finished_at": source_run_summary.get("finishedAt"),
            "duration_sec": source_run_summary.get("durationSec"),
            "enabled_sources": sorted([str(s) for s in enabled_sources]),
            "complete_sources": sorted([str(s) for s in complete_sources]),
            "partial_sources": sorted([str(s) for s in partial_sources]),
            "timed_out_sources": sorted([str(s) for s in timed_out_sources]),
            "failed_sources": sorted([str(s) for s in failed_sources]),
        }

    if not partial_run:
        partial_run = len(failures) > 0

    # Per-source freshness — the CSV mtime tells us when each source
    # was last successfully refreshed, independent of the most recent
    # *full-pipeline* scrape time.  Stamp this here so the staleness
    # alert system in src.api.source_health_alerts has the
    # ``sources[src].lastFetched`` field it expects, and the source-
    # health page can render per-source ages.
    sources_meta = _per_source_freshness()

    return {
        # The population we are entitled to expect, not what one run
        # happened to emit.
        "total_sources": len(registered),
        "registered_sources": registered,
        "sources_with_data": available,
        "source_counts": source_counts,
        # Known to have contributed nothing.  Distinct from
        # ``unmeasured_sources``, which is "we were not given the
        # served board and therefore cannot say".
        "missing_sources": sorted(missing),
        "unmeasured_sources": sorted(unmeasured),
        # The scraper's own anchor row counts, kept under a name that
        # says what they are.  A different quantity from board
        # contribution above.
        "anchor_row_counts": anchor_row_counts,
        "partial_run": bool(partial_run),
        "source_runtime": source_runtime,
        "source_failures": failures,
        "sources": sources_meta,
    }


# 30s memo for the disk-probing observability helpers below.  They
# glob + stat dozens of files per call, run on the event loop, and are
# hit by /api/status (polled by /settings + tools pages) and
# /api/health (polled every 60s by StaleDataBanner on EVERY page).
# Pure observability data: key = helper name, TTL-only invalidation,
# results treated as read-only, ≤30s staleness is invisible for
# freshness ages measured in hours.
_DISK_PROBE_MEMO: dict[str, tuple[float, Any]] = {}
_DISK_PROBE_TTL_SEC = 30.0


def _disk_probe_memo(kind: str, builder):
    now = time.time()
    hit = _DISK_PROBE_MEMO.get(kind)
    if hit is not None and (now - hit[0]) < _DISK_PROBE_TTL_SEC:
        return hit[1]
    val = builder()
    _DISK_PROBE_MEMO[kind] = (now, val)
    return val


def _per_source_freshness() -> dict[str, dict]:
    """Per-source ``{lastFetched, ageHours}`` for every source registered
    in ``_SOURCE_CSV_PATHS``.

    Memoized 30s (see ``_DISK_PROBE_MEMO``) — treat as read-only.

    Freshness signal preference, in order:

    1. ``data/scrape_state/{key}_last_success`` — a unix-epoch stamp
       written by the source's fetcher on a successful run, regardless
       of whether the CSV content changed.  This is the load-bearing
       signal for monthly-cadence vendors (Fitzmaurice, Yahoo Boone,
       etc.) where the fetcher succeeds on every CI cycle but writes
       byte-identical content most of the month.  ``git checkout
       --force`` skips rewriting unchanged files, so the CSV mtime on
       prod gets stuck on the last *content* change instead of the
       last fetcher *success*.  The stamp tracks success directly, so
       a silent fetcher failure trips the 24h alert within one cycle
       even when the vendor hasn't published new content.

    2. CSV mtime — fallback for sources that don't write a stamp yet.
       Reliable signal for sources whose fetchers produce content that
       varies on every run (rank jitter, timestamps, etc.) since each
       run produces a new blob and ``git checkout --force`` does
       rewrite differing files.

    Sources missing both the stamp and the CSV file are omitted from
    the output entirely.  Returns ``{}`` if the source registry can't
    be loaded.
    """

    def _build() -> dict[str, dict]:
        try:
            from src.api.data_contract import _SOURCE_CSV_PATHS
        except Exception:  # pragma: no cover
            return {}
        repo_root = Path(__file__).resolve().parent
        state_dir = repo_root / "data" / "scrape_state"
        out: dict[str, dict] = {}
        now_epoch = time.time()
        for src_key, entry in _SOURCE_CSV_PATHS.items():
            csv_rel = entry if isinstance(entry, str) else (entry or {}).get("path")
            if not csv_rel:
                continue
            csv_path = repo_root / csv_rel

            # Prefer the stamp content over CSV mtime; fall back when the
            # stamp is missing or unparseable.
            stamp_path = state_dir / f"{src_key}_last_success"
            last_epoch: float | None = None
            try:
                stamp_text = stamp_path.read_text().strip()
            except OSError:
                stamp_text = ""
            if stamp_text:
                try:
                    last_epoch = float(stamp_text)
                except ValueError:
                    last_epoch = None
            if last_epoch is None:
                try:
                    last_epoch = csv_path.stat().st_mtime
                except OSError:
                    continue

            age_hours = max(0.0, (now_epoch - last_epoch) / 3600.0)
            out[src_key] = {
                "lastFetched": datetime.fromtimestamp(last_epoch, tz=timezone.utc).isoformat(
                    timespec="seconds"
                ),
                "ageHours": round(age_hours, 2),
            }
        return out

    return _disk_probe_memo("per_source_freshness", _build)


def _backup_freshness() -> dict:
    """Age + location of the newest SQLite backup.

    Mirrors deploy/backup_user_kv.sh: the primary dir ``/var/backups/
    riskit`` needs root; the self-healing script falls back to
    ``/home/dynasty/backups/riskit`` (owned by the service user).  This
    surfaces ``newestBackupAgeHours`` so monitoring trips an alert
    within a day of a silent backup failure — the exact gap that let
    backups die unnoticed for ~2 weeks before the P0 fix.  Never raises;
    returns null age when no backup is found.

    Memoized 30s (see ``_DISK_PROBE_MEMO``) — treat as read-only.
    """
    return _disk_probe_memo("backup_freshness", _backup_freshness_uncached)


def _backup_freshness_uncached() -> dict:
    result = {
        "newestBackupAgeHours": None,
        "newestBackupPath": None,
        "backupDirUsed": "none",
        "dbCount": 0,
    }
    try:
        now_epoch = time.time()
        newest_epoch: float | None = None
        for label, d in (
            ("primary", Path("/var/backups/riskit/daily")),
            ("fallback", Path("/home/dynasty/backups/riskit/daily")),
        ):
            try:
                files = list(d.glob("*.sqlite.gz"))
            except OSError:
                continue
            for f in files:
                try:
                    mt = f.stat().st_mtime
                except OSError:
                    continue
                if newest_epoch is None or mt > newest_epoch:
                    newest_epoch = mt
                    result["newestBackupPath"] = str(f)
                    result["backupDirUsed"] = label
        if newest_epoch is not None:
            result["newestBackupAgeHours"] = round(max(0.0, (now_epoch - newest_epoch) / 3600.0), 2)
            try:
                newest = Path(result["newestBackupPath"])
                parts = newest.name.split(".")
                stamp = parts[1] if len(parts) >= 3 else ""
                if stamp:
                    result["dbCount"] = len(
                        {f.name.split(".")[0] for f in newest.parent.glob(f"*.{stamp}.sqlite.gz")}
                    )
            except OSError:
                pass
    except Exception:  # noqa: BLE001 — monitoring helper must never raise
        pass
    return result


# ── SCRAPER INTEGRATION ────────────────────────────────────────────────
def _warm_overlays_in_background(contract_payload: dict) -> None:
    """Force-refresh the Sleeper overlay cache for every active league
    on a daemon thread.

    Called by ``_prime_latest_payload`` after a scrape / startup prime.
    Must never run inline on the event loop — see the call site for the
    deploy-failure history.  Thread-safe: ``fetch_sleeper_overlay`` is
    the same sync callable the request path already invokes via
    ``run_in_threadpool``, and it guards its own per-league cache.
    """

    def _worker() -> None:
        try:
            loaded_sleeper = contract_payload.get("sleeper") or {}
            id_map = loaded_sleeper.get("idToPlayer") if isinstance(loaded_sleeper, dict) else {}
            warmed: list[str] = []
            warm_failed: list[str] = []
            scoring_refreshed: list[str] = []
            for cfg in _league_registry.active_leagues():
                # Factual scoring card (W18-F001).  The compatibility
                # gate reads a snapshot and never fetches inside a
                # request, so this pass is what keeps it current.  A
                # failure leaves the previous snapshot alone.
                if _league_registry.refresh_scoring_snapshot(cfg):
                    scoring_refreshed.append(cfg.key)
                try:
                    overlay = _sleeper_overlay.fetch_sleeper_overlay(
                        sleeper_league_id=cfg.sleeper_league_id,
                        id_to_player=id_map if isinstance(id_map, dict) else {},
                        force_refresh=True,
                    )
                    if overlay and overlay.get("teams"):
                        warmed.append(cfg.key)
                    else:
                        warm_failed.append(cfg.key)
                except Exception as inner:  # noqa: BLE001
                    log.warning(
                        "post-scrape overlay warm failed for %s: %s",
                        cfg.key,
                        inner,
                    )
                    warm_failed.append(cfg.key)
            if warmed or warm_failed:
                log.info(
                    "post-scrape overlay warm: %d warmed, %d failed "
                    "(warmed=%s failed=%s scoringSnapshots=%s)",
                    len(warmed),
                    len(warm_failed),
                    warmed,
                    warm_failed,
                    scoring_refreshed,
                )
            # Sleeper trending-adds warm (FAAB v2) — global (not
            # per-league), tiny payload, 15-min TTL.  Failure is
            # non-fatal: the recommend endpoint fetches on demand
            # and degrades to the contract's sleeperTrending
            # fallback when the adapter has nothing.
            try:
                from src.adapters import sleeper_trending as _sleeper_trending  # noqa: PLC0415

                if not _sleeper_trending.warm():
                    log.warning("post-scrape trending warm: no snapshot available")
                else:
                    # C1-RET-05.  The adapter holds a 15-minute TTL and
                    # persists nothing, so the waiver-heat series has
                    # never existed.  This is a CACHE READ of the
                    # snapshot warm() just fetched — no second
                    # round-trip — recorded off the request path on the
                    # scrape cadence.  Keyed on the snapshot's own
                    # fetchedAt, so re-recording a cached snapshot is a
                    # structural no-op rather than a duplicate.
                    try:
                        from src.retention import evidence_store as _evidence  # noqa: PLC0415

                        snap = _sleeper_trending.get_trending_adds()
                        result = _evidence.observe_trending_snapshot(snap)
                        if result.get("action") == "recorded":
                            log.info(
                                "trending retention: recorded %d observation(s) @ %s",
                                result.get("inserted", 0),
                                result.get("observedAt"),
                            )
                    except Exception as ret_exc:  # noqa: BLE001
                        log.warning("trending retention: record failed: %s", ret_exc)
            except Exception as trend_exc:  # noqa: BLE001
                log.warning("post-scrape trending warm failed: %s", trend_exc)
        except Exception as exc:  # noqa: BLE001
            log.warning("post-scrape overlay warm pass failed: %s", exc)

    try:
        threading.Thread(
            target=_worker,
            name="sleeper-overlay-warm",
            daemon=True,
        ).start()
    except Exception as exc:  # noqa: BLE001
        log.warning("post-scrape overlay warm thread failed to start: %s", exc)


# ── Scoring-identity gate (W18-F001) ─────────────────────────────────
#
# ONE place decides "may the loaded contract's rankings be served for
# this league?".  Before this existed the question was answered by six
# scattered comparisons of the hand-typed ``scoringProfile`` registry
# label — on /api/data, /api/rankings/overrides, /api/terminal,
# /api/trade/simulate, /api/draft-capital's rookie pool and the
# signal-alerts sweep.  Four of them also short-circuited to
# "compatible" whenever the loaded contract carried no label at all,
# which is why an unidentified contract could be served for any league.
#
# The prose deliberately does not quote that condition verbatim:
# ``tests/api/test_scoring_compatibility.py`` greps this file for the
# fail-open shape, and a comment reproducing it would blind the check.

#: One-entry memo for ``_contract_scoring_fingerprint`` — see its body.
_CONTRACT_FINGERPRINT_MEMO: dict[str, tuple[tuple, str | None]] = {}


def _contract_scoring_fingerprint(contract: Any) -> str | None:
    """The factual scoring identity of a loaded contract, or ``None``.

    **The contract's own scoring card is the evidence; the stamp is a
    cache of it.**  That ordering is the whole justification for putting
    an identity on the contract at all — it can be recomputed from the
    artifact it describes, unlike a value copied out of config.  So:

    * card present, stamp agrees → that fingerprint;
    * card present, no stamp → recompute (the migration path: the live
      board carries 141 scoring keys and identifies immediately);
    * card present, stamp DISAGREES → ``None``.  A stale, corrupted or
      hand-edited stamp contradicting the contract's own card is exactly
      the kind of unverified claim W18-F001 exists to refuse, and a
      version prefix from older normalization rules cannot be compared
      at all;
    * **no card, stamp only → ``None``.**  Decided explicitly rather than
      left to fall out of the lookup order: a stamp with nothing to check
      it against is unverifiable, and unverifiable fails closed.  The
      documented migration policy never depended on this branch — it
      depended on the card, which every real contract carries.
    """
    if not isinstance(contract, dict):
        return None
    meta = contract.get("meta") if isinstance(contract.get("meta"), dict) else {}
    stamped = str(meta.get("scoringFingerprint") or "").strip()

    sleeper = contract.get("sleeper")
    if not isinstance(sleeper, dict):
        return None
    card = sleeper.get("scoringSettings")
    # Hashing the card is what makes the stamp checkable, and the live
    # board's card is 141 keys — ~85 us, on a gate that runs per request.
    # One memo entry, keyed on the identity and size of the card dict plus
    # the stamp being checked, so a replaced contract (every scrape swaps
    # the object) recomputes.  Purely an accelerator: a miss recomputes,
    # and the comparison it accelerates is unchanged.
    memo_key = (id(sleeper), id(card), len(card) if isinstance(card, dict) else -1, stamped)
    cached = _CONTRACT_FINGERPRINT_MEMO.get("k")
    if cached is not None and cached[0] == memo_key:
        return cached[1]
    try:
        from src.league_comparison.sleeper_scoring import scoring_fingerprint  # noqa: PLC0415

        derived = scoring_fingerprint(card)
    except Exception:  # noqa: BLE001 — a gate must not raise
        return None
    if not derived:
        _CONTRACT_FINGERPRINT_MEMO["k"] = (memo_key, None)
        return None
    if stamped and stamped != derived:
        log.warning(
            "contract scoring identity is self-contradictory: meta stamp %s vs "
            "its own scoring card %s — refusing to prove compatibility",
            stamped,
            derived,
        )
        _CONTRACT_FINGERPRINT_MEMO["k"] = (memo_key, None)
        return None
    _CONTRACT_FINGERPRINT_MEMO["k"] = (memo_key, derived)
    return derived


def _scoring_identity_error(contract: Any, league_cfg: Any) -> JSONResponse | None:
    """``None`` when ``contract``'s rankings may be served for ``league_cfg``.

    Otherwise the ``503 data_not_ready`` this repo already returns for
    incompatible rankings.

    Fails CLOSED on an unproven identity — either side missing — because
    the alternative is the live defect: one league's board published
    verbatim under another league's name with nothing on the response
    saying so.  The cost is bounded and self-healing: the loaded league's
    own requests never reach here (same key short-circuits), and the
    scoring snapshot + contract stamp are both refreshed every scrape.
    """
    if not isinstance(contract, dict) or league_cfg is None:
        return None
    meta = contract.get("meta") if isinstance(contract.get("meta"), dict) else {}
    loaded_key = str(meta.get("leagueKey") or "")
    if loaded_key and loaded_key == league_cfg.key:
        return None

    loaded_fp = _contract_scoring_fingerprint(contract)
    try:
        requested_fp = _league_registry.scoring_fingerprint_for_league(league_cfg)
    except Exception:  # noqa: BLE001
        requested_fp = None
    if loaded_fp and requested_fp and loaded_fp == requested_fp:
        return None

    if not loaded_fp:
        reason = (
            "the loaded contract carries no scoring identity, so its rankings "
            f"cannot be shown to apply to league {league_cfg.key!r}"
        )
    elif not requested_fp:
        reason = (
            f"league {league_cfg.key!r} has no verified scoring snapshot, so its "
            "scoring cannot be compared with the loaded contract's"
        )
    else:
        reason = (
            f"league {league_cfg.key!r} scores players differently from the "
            "scoring the loaded rankings were built under"
        )
    return JSONResponse(
        status_code=503,
        content={
            "error": "data_not_ready",
            "message": f"Rankings are not compatible: {reason}.",
            "leagueKey": league_cfg.key,
            "scoringProfile": league_cfg.scoring_profile,
            "scoringFingerprint": requested_fp,
            "loadedScoringFingerprint": loaded_fp,
        },
    )


def _prime_latest_payload(data: dict | None, *, is_fresh_scrape: bool = False) -> None:
    """Pre-serialize latest payload once so /api/data returns instantly.

    ``is_fresh_scrape`` gates rank-history appends: startup priming
    from cached disk data must NOT append a new "today" entry (which
    would fabricate a history point after every server restart).
    Scrape-promotion callers pass ``is_fresh_scrape=True``; startup
    lifespan priming leaves it False so the history log stays read-
    only until a real scrape lands.
    """
    global latest_contract_data, latest_data_bytes, latest_data_gzip_bytes, latest_data_etag
    global \
        latest_runtime_data, \
        latest_runtime_data_bytes, \
        latest_runtime_data_gzip_bytes, \
        latest_runtime_data_etag
    global \
        latest_startup_data, \
        latest_startup_data_bytes, \
        latest_startup_data_gzip_bytes, \
        latest_startup_data_etag
    global \
        latest_array_data, \
        latest_array_data_bytes, \
        latest_array_data_gzip_bytes, \
        latest_array_data_etag
    global \
        latest_compact_data, \
        latest_compact_data_bytes, \
        latest_compact_data_gzip_bytes, \
        latest_compact_data_etag
    global contract_health
    global served_source_coverage

    def _swap_to_empty() -> None:
        """Publish the 'no payload' generation (falsy data / failed
        build).  Plain reference assignments, no I/O — near-atomic."""
        global latest_contract_data, latest_data_bytes, latest_data_gzip_bytes, latest_data_etag
        global \
            latest_runtime_data, \
            latest_runtime_data_bytes, \
            latest_runtime_data_gzip_bytes, \
            latest_runtime_data_etag
        global \
            latest_startup_data, \
            latest_startup_data_bytes, \
            latest_startup_data_gzip_bytes, \
            latest_startup_data_etag
        global \
            latest_array_data, \
            latest_array_data_bytes, \
            latest_array_data_gzip_bytes, \
            latest_array_data_etag
        global \
            latest_compact_data, \
            latest_compact_data_bytes, \
            latest_compact_data_gzip_bytes, \
            latest_compact_data_etag
        global served_source_coverage
        served_source_coverage = {}
        latest_data_bytes = None
        latest_data_gzip_bytes = None
        latest_data_etag = None
        latest_contract_data = None
        latest_runtime_data = None
        latest_runtime_data_bytes = None
        latest_runtime_data_gzip_bytes = None
        latest_runtime_data_etag = None
        latest_startup_data = None
        latest_startup_data_bytes = None
        latest_startup_data_gzip_bytes = None
        latest_startup_data_etag = None
        latest_array_data = None
        latest_array_data_bytes = None
        latest_array_data_gzip_bytes = None
        latest_array_data_etag = None
        latest_compact_data = None
        latest_compact_data_bytes = None
        latest_compact_data_gzip_bytes = None
        latest_compact_data_etag = None
        # A new contract generation invalidates every memoized overrides
        # response (they are versioned on ``latest_data_etag``, but clearing
        # eagerly also frees the multi-MB byte payloads immediately).
        _OVERRIDES_RESPONSE_CACHE.clear()
        # Draft-capital results price picks off the live contract on the
        # non-default-league path — drop them with the old generation.
        _DRAFT_CAPITAL_CACHE.clear()

    if not data:
        _swap_to_empty()
        return

    # ── Stage 1: compute the ENTIRE new generation into locals.
    # This function is offloaded to a worker thread at the scrape-
    # completion call site, so requests can interleave with the
    # multi-second build.  No global is touched until every variant is
    # encoded; the publish happens in the tight swap at the end.  (The
    # previous reset-globals-first shape was only safe because the
    # whole build used to block the event loop.)
    try:
        contract_payload = build_api_data_contract(data, data_source=latest_data_source)
        contract_report = validate_api_data_contract(contract_payload)
        contract_payload["contractHealth"] = {
            "ok": bool(contract_report.get("ok")),
            "status": contract_report.get("status"),
            "errorCount": int(contract_report.get("errorCount", 0)),
            "warningCount": int(contract_report.get("warningCount", 0)),
            "checkedAt": contract_report.get("checkedAt"),
        }
        # Rank-history integration:
        # - Append a new "today" snapshot ONLY on fresh scrape
        #   promotions.  Startup priming from cached disk data must
        #   stay read-only or every restart fabricates a redundant
        #   history entry (and /api/data/rank-history misleads
        #   consumers into thinking a scrape ran).
        # - Stamp ``rankHistory`` onto every row regardless of
        #   source — it's a pure read of the existing log and the
        #   frontend glyph needs it on startup-primed payloads too.
        try:
            if is_fresh_scrape:
                # Each of the three recorders is isolated in its own
                # try so one failing append can never silently skip the
                # others (the ledger record used to sit downstream of
                # the rank-history append inside one block — an
                # asymmetric coupling the C1-U4 final review flagged).
                try:
                    _rank_history.append_snapshot(contract_payload)
                except Exception as inner_exc:  # noqa: BLE001
                    log.warning("rank_history: append failed: %s", inner_exc)
                # Sister snapshot: per-source value history.  Stored in
                # a separate JSONL so the rank-history log stays small
                # and readable while the popup chart can stream a
                # richer per-source series on demand.  Failures are
                # isolated so a source-history write error doesn't
                # nuke the rank-history append we just did.
                try:
                    _source_history.append_snapshot(contract_payload)
                except Exception as inner_exc:  # noqa: BLE001
                    log.warning(
                        "source_history: append failed: %s",
                        inner_exc,
                    )
                # Canonical temporal ledger (src/history, C1-U4): the
                # one as-of owner records this build — including the
                # tethered slot-pick rows the rank-gated log above
                # structurally drops (C1-HIST-02).  Fresh scrapes only,
                # same discipline as the appends above; isolated so a
                # ledger failure cannot nuke either sibling append.
                try:
                    ledger_result = _history_record.record_contract(contract_payload)
                    log.info(
                        "temporal_ledger: recorded %d observations for %s "
                        "(%d duplicate, %d unresolved rows)",
                        ledger_result.get("written", 0),
                        ledger_result.get("boardDate"),
                        ledger_result.get("duplicates", 0),
                        ledger_result.get("unresolved", 0),
                    )
                except Exception as inner_exc:  # noqa: BLE001
                    log.warning("temporal_ledger: record failed: %s", inner_exc)
            stamped = _rank_history.stamp_contract_with_history(contract_payload)
            if stamped:
                log.info("rank_history: stamped %d rows with history series", stamped)
        except Exception as exc:  # noqa: BLE001
            # Non-fatal: a history log failure must NOT break the
            # contract response.  The glyph degrades gracefully when
            # rankHistory is absent.
            log.warning("rank_history: append/stamp failed: %s", exc)
        # Tag the contract with the league + scoring profile it was
        # built for.  Two different roles:
        #
        #   * ``meta.leagueKey`` — which specific league's Sleeper
        #     block (teams, rosters, ownerIds) is stamped here.
        #     Team-requiring endpoints (/api/terminal, /api/trade/*)
        #     reject requests for other leagues with 503.
        #   * ``meta.scoringProfile`` — which scoring rules produced
        #     these rankings.  Rankings endpoints (/api/data,
        #     /api/rankings/overrides) serve the same rankings to
        #     any league that shares the profile, and only 503 when
        #     profiles actually differ.
        #
        # This split is the core of the "scoring drives rankings,
        # league drives context" architecture — see CLAUDE.md.
        #   * ``meta.scoringFingerprint`` — the FACTUAL identity of the
        #     scoring that produced them (W18-F001).  Derived from the
        #     contract's OWN ``sleeper.scoringSettings``, i.e. the card
        #     the scrape actually fetched from the host, and NOT copied
        #     from the registry: a stamp taken from a second file proves
        #     only that the second file said so, while this one can be
        #     recomputed from the artifact it describes.  Absent — never
        #     a hash of ``{}`` — when the scrape carried no card.
        try:
            _default_cfg = _league_registry.get_default_league()
            if _default_cfg and isinstance(contract_payload, dict):
                meta_block = contract_payload.setdefault("meta", {})
                meta_block["leagueKey"] = _default_cfg.key
                meta_block["scoringProfile"] = _default_cfg.scoring_profile
                _fp = _contract_scoring_fingerprint(contract_payload)
                if _fp:
                    meta_block["scoringFingerprint"] = _fp
                else:
                    meta_block.pop("scoringFingerprint", None)
        except Exception:  # noqa: BLE001
            pass
        new_coverage = _compute_served_source_coverage(contract_payload)

        # Post-scrape overlay warm — for every ACTIVE league
        # (including the default league the scraper just built for),
        # force-refresh the Sleeper overlay so the first user request
        # after a scrape hits a warm 15-min cache instead of round-
        # tripping to Sleeper.  The default league used to be skipped
        # because /api/data served its baked sleeper block directly;
        # since /api/data now splices the overlay onto every response
        # (default + cross-league), warming the default league makes
        # the very first /api/data after a scrape return overlay-fresh
        # rosters too.  Non-fatal: any failure is logged + skipped.
        #
        # Runs on a background daemon thread (same pattern as
        # ``_kick_background_refresh``), NEVER inline: this function is
        # called from the ``lifespan`` startup path BEFORE uvicorn binds
        # the port, and from ``run_scraper`` on the event loop.  Inline,
        # a slow Sleeper (per-league round-trips with read timeouts)
        # blocked the socket bind past deploy verification's retry
        # budget — the root cause of the 2026-07-25 deploy failures /
        # auto-rollbacks — and stalled the loop at every scrape end.
        # The warm is a cache-priming optimization; requests that land
        # before it finishes simply fetch the overlay themselves via
        # the existing threadpool path in ``get_data``.
        _warm_overlays_in_background(contract_payload)

        if not contract_report.get("ok"):
            log.error(
                "API contract validation failed: %s",
                "; ".join((contract_report.get("errors") or [])[:5]),
            )

        raw = json.dumps(contract_payload, ensure_ascii=False, separators=(",", ":")).encode(
            "utf-8"
        )
        full_gzip = gzip.compress(raw, compresslevel=5)
        full_etag = hashlib.sha1(raw).hexdigest()

        # Runtime payload: keep canonical top-level data shape used by the live UI,
        # but remove heavyweight contract array duplication to reduce parse/transfer cost.
        runtime_payload = dict(contract_payload)
        runtime_payload.pop("playersArray", None)
        runtime_payload["payloadView"] = "runtime"
        runtime_raw = json.dumps(runtime_payload, ensure_ascii=False, separators=(",", ":")).encode(
            "utf-8"
        )
        runtime_gzip = gzip.compress(runtime_raw, compresslevel=5)
        runtime_etag = hashlib.sha1(runtime_raw).hexdigest()

        # Array payload: full contract minus the LEGACY ``players`` dict.
        # ``playersArray`` and the dict are parallel encodings; the array
        # is strictly richer (the dict's fields are underscore-mirrors +
        # flat per-source values the array carries in structured form)
        # and it is the branch ``buildRows`` prefers whenever present.
        # Desktop clients request ``?view=array`` and get the identical
        # board at roughly half the bytes/parse cost of the full view.
        array_payload = dict(contract_payload)
        array_payload.pop("players", None)
        array_payload["payloadView"] = "array"
        array_raw = json.dumps(array_payload, ensure_ascii=False, separators=(",", ":")).encode(
            "utf-8"
        )
        array_gzip = gzip.compress(array_raw, compresslevel=5)
        array_etag = hashlib.sha1(array_raw).hexdigest()

        # Startup payload: same contract shape, but strips heavyweight fields
        # not needed for first screen render so first data-visible is faster.
        startup_payload = build_api_startup_payload(runtime_payload)
        startup_raw = json.dumps(startup_payload, ensure_ascii=False, separators=(",", ":")).encode(
            "utf-8"
        )
        startup_gzip = gzip.compress(startup_raw, compresslevel=5)
        startup_etag = hashlib.sha1(startup_raw).hexdigest()

        # Compact payload: mobile / slow-network view (~90% smaller).
        # Precompute bytes + gzip + etag here so ``?view=compact`` serves
        # from the same fast path as the other views instead of running
        # ``compact_contract`` + ``json.dumps`` + gzip per request on the
        # event loop.
        compact_payload = None
        compact_raw = None
        compact_gzip = None
        compact_etag = None
        try:
            from src.api.compact_view import compact_contract

            compact_payload = compact_contract(contract_payload)
            compact_raw = json.dumps(
                compact_payload, ensure_ascii=False, separators=(",", ":")
            ).encode("utf-8")
            compact_gzip = gzip.compress(compact_raw, compresslevel=5)
            compact_etag = hashlib.sha1(compact_raw).hexdigest()
        except Exception as exc:  # noqa: BLE001
            # Non-fatal: the endpoint falls back to on-demand compaction
            # when the precomputed compact payload is unavailable.
            compact_payload = None
            compact_raw = None
            compact_gzip = None
            compact_etag = None
            log.warning("compact payload precompute failed: %s", exc)
    except Exception as e:
        # Failed build: publish the empty generation (same end state as
        # before — /api/data 503s, health invalid).  This also covers a
        # mid-encode failure, which previously could leave a contract
        # object published with no bytes behind it.
        _swap_to_empty()
        contract_health = {
            "ok": False,
            "status": "invalid",
            "errors": [f"contract build failed: {type(e).__name__}: {e}"],
            "warnings": [],
            "errorCount": 1,
            "warningCount": 0,
            "checkedAt": _utc_now_iso(),
            "contractVersion": API_DATA_CONTRACT_VERSION,
            "playerCount": 0,
        }
        log.error(f"Failed to pre-serialize latest payload: {e}")
        return

    # ── Stage 2: tight swap — publish the new generation.  Plain
    # reference assignments only; concurrent readers see the old
    # generation or the new one, never a torn one (each individual
    # response is built from ONE consistent (bytes, etag) pair grabbed
    # in a single statement server-side).
    latest_contract_data = contract_payload
    contract_health = contract_report
    served_source_coverage = new_coverage
    latest_data_bytes = raw
    latest_data_gzip_bytes = full_gzip
    latest_data_etag = full_etag
    latest_runtime_data = runtime_payload
    latest_runtime_data_bytes = runtime_raw
    latest_runtime_data_gzip_bytes = runtime_gzip
    latest_runtime_data_etag = runtime_etag
    latest_array_data = array_payload
    latest_array_data_bytes = array_raw
    latest_array_data_gzip_bytes = array_gzip
    latest_array_data_etag = array_etag
    latest_startup_data = startup_payload
    latest_startup_data_bytes = startup_raw
    latest_startup_data_gzip_bytes = startup_gzip
    latest_startup_data_etag = startup_etag
    latest_compact_data = compact_payload
    latest_compact_data_bytes = compact_raw
    latest_compact_data_gzip_bytes = compact_gzip
    latest_compact_data_etag = compact_etag
    # Old-generation memoized responses go with the old etag.
    _OVERRIDES_RESPONSE_CACHE.clear()
    _DRAFT_CAPITAL_CACHE.clear()


def load_from_disk() -> dict | None:
    """Load most recent dynasty_data_*.json from data/ directory."""
    json_files = sorted(DATA_DIR.glob("dynasty_data_*.json"), reverse=True)
    if not json_files:
        # Also check base dir for existing files from standalone scraper runs
        json_files = sorted(BASE_DIR.glob("dynasty_data_*.json"), reverse=True)
    if json_files:
        try:
            latest_path = json_files[0]
            with open(latest_path) as f:
                data = json.load(f)
            _set_latest_data_source(
                "disk_cache", str(latest_path), produced_at=data.get("scrapeTimestamp")
            )
            log.info(
                f"Loaded cached data from {latest_path.name} "
                f"({len(data.get('players', {}))} players)"
            )
            return data
        except Exception as e:
            log.error(f"Failed to load {json_files[0]}: {e}")
    return None


def _latest_file(directory: Path, pattern: str) -> Path | None:
    if not directory.exists():
        return None
    files = sorted(directory.glob(pattern), reverse=True)
    return files[0] if files else None


def _load_json_file(path: Path | None) -> dict | None:
    if path is None or not path.exists():
        return None
    try:
        with path.open("r", encoding="utf-8") as f:
            return json.load(f)
    except Exception as e:
        log.error(f"Failed to load scaffold json {path}: {e}")
        return None


async def run_scraper(trigger: str = "manual") -> dict | None:
    """
    Import and run the scraper, returning the dashboard JSON dict.
    Runs in the same event loop as the server.
    """
    global latest_data
    _reconcile_orphaned_running_state()
    if scrape_run_lock.locked():
        _record_scrape_event(
            "scrape_rejected_already_running",
            level="warning",
            message="run_scraper called while lock already held",
        )
        return latest_data

    async with scrape_run_lock:
        start = time.time()
        worker_id = _start_scrape_run(trigger=trigger)
        log.info("=" * 60)
        log.info("SCRAPE STARTING")
        log.info("=" * 60)

        try:
            _update_scrape_progress(
                step="bootstrap",
                source="import_scraper",
                step_index=1,
                step_total=4,
                event="phase_start",
                message="Importing scraper module",
            )

            # Import the scraper module from its exact file path
            # (importlib handles spaces in directory names that normal import can't)
            import importlib.util

            spec = importlib.util.spec_from_file_location("Dynasty_Scraper", str(SCRAPER_PATH))
            scraper = importlib.util.module_from_spec(spec)
            sys.modules["Dynasty_Scraper"] = scraper
            spec.loader.exec_module(scraper)

            # Override SCRIPT_DIR so output goes to our data/ folder
            scraper.SCRIPT_DIR = str(DATA_DIR)

            _update_scrape_progress(
                step="scrape",
                source="Dynasty Scraper.py",
                step_index=2,
                step_total=4,
                event="phase_start",
                message="Executing scraper.run()",
            )

            progress_callback = _build_scrape_progress_callback(worker_id)

            # Top-level run timeout guard so a wedged scraper cannot hold running=True forever.
            result = await asyncio.wait_for(
                scraper.run(progress_callback=progress_callback),
                timeout=SCRAPE_RUN_TIMEOUT_SECONDS,
            )

            _update_scrape_progress(
                step="validate",
                source="result_payload",
                step_index=3,
                step_total=4,
                event="phase_start",
                message="Validating scraper output",
            )

            if not result or not result.get("players"):
                raise RuntimeError("Scraper returned empty result")

            # Mirror fresh site_raw CSVs from the scraper's DATA_DIR output
            # path (data/exports/latest/site_raw/) back to the repo's
            # tracked CSVs/site_raw/ directory so that the CSV
            # enrichment in data_contract.py (which reads relative to repo
            # root) sees up-to-date values.  Without this, enrichment reads
            # permanently-stale CSVs from git history.  Only copies KTC and
            # IDPTradeCalc — DLF is a rank-signal file with a different
            # format maintained separately.
            try:
                import shutil as _sh

                src_raw = DATA_DIR / "exports" / "latest" / "site_raw"
                dst_raw = BASE_DIR / "CSVs" / "site_raw"
                if src_raw.exists() and dst_raw.exists():
                    for fname in ("ktc.csv", "idpTradeCalc.csv"):
                        src_file = src_raw / fname
                        dst_file = dst_raw / fname
                        if src_file.exists():
                            _sh.copy2(src_file, dst_file)
                    # Also mirror the full dynasty_data JSON so other
                    # consumers (tests, CLI tools) see the fresh file.
                    date_str = str(result.get("date") or "")
                    if date_str:
                        src_json = DATA_DIR / "exports" / "latest" / f"dynasty_data_{date_str}.json"
                        dst_json = BASE_DIR / "exports" / "latest" / f"dynasty_data_{date_str}.json"
                        if src_json.exists():
                            _sh.copy2(src_json, dst_json)
            except Exception as _mirror_err:
                log.warning(f"Post-scrape CSV mirror failed: {_mirror_err}")

            # Refresh Dynasty Nerds SF-TEP rankings.  The DN board is
            # inlined in the page HTML as a ``window.DR_DATA`` JS
            # constant — no Playwright required — so we run the plain
            # ``scripts/fetch_dynasty_nerds.py`` helper inline on every
            # scheduled scrape cycle.  Failure is logged and ignored so
            # a transient network error cannot fail the entire scrape.
            try:
                from scripts import fetch_dynasty_nerds as _dn_fetch

                rc = _dn_fetch.main(["--mirror-data-dir"])
                if rc == 2:
                    # Schema / row-count regression — surface loudly as
                    # a structured scrape event so /api/status shows the
                    # failure instead of burying it as a log line.
                    _record_scrape_event(
                        "dynasty_nerds_schema_regression",
                        level="error",
                        message=(
                            "Dynasty Nerds fetch exit=2 (DR_DATA shape changed or rows below floor)"
                        ),
                        exit_code=rc,
                    )
                elif rc != 0:
                    _record_scrape_event(
                        "dynasty_nerds_fetch_failed",
                        level="warning",
                        message=f"Dynasty Nerds fetch returned exit={rc}",
                        exit_code=rc,
                    )
            except Exception as _dn_err:
                _record_scrape_event(
                    "dynasty_nerds_fetch_exception",
                    level="warning",
                    message=f"Dynasty Nerds fetch raised: {_dn_err}",
                )

            # Refresh FantasyPros Dynasty Superflex (offense) rankings.
            # The dynasty-superflex page inlines an ``ecrData = {...}``
            # JS constant, so a plain ``requests.get`` with a browser
            # UA returns the full payload.  The fetch script extracts
            # QB/RB/WR/TE consensus ECR ranks and writes a rank-signal CSV.
            try:
                from scripts import fetch_fantasypros_offense as _fpoff_fetch

                rc = _fpoff_fetch.main(["--mirror-data-dir"])
                if rc == 2:
                    _record_scrape_event(
                        "fantasypros_offense_schema_regression",
                        level="error",
                        message=(
                            "FantasyPros Offense fetch exit=2 "
                            "(ecrData shape changed or rows below floor)"
                        ),
                        exit_code=rc,
                    )
                elif rc != 0:
                    _record_scrape_event(
                        "fantasypros_offense_fetch_failed",
                        level="warning",
                        message=f"FantasyPros Offense fetch returned exit={rc}",
                        exit_code=rc,
                    )
            except Exception as _fpoff_err:
                _record_scrape_event(
                    "fantasypros_offense_fetch_exception",
                    level="warning",
                    message=f"FantasyPros Offense fetch raised: {_fpoff_err}",
                )

            # Refresh FantasyPros Dynasty IDP rankings.  The combined
            # + DL/LB/DB pages inline their rankings in a JS
            # ``ecrData = {...}`` constant, so a plain ``requests.get``
            # with a browser UA returns the full payload.  The fetch
            # script derives per-player effective overall ranks via
            # anchor curves fit on the combined/individual overlap
            # and writes a rank-signal CSV.
            try:
                from scripts import fetch_fantasypros_idp as _fp_fetch

                rc = _fp_fetch.main(["--mirror-data-dir"])
                if rc == 2:
                    _record_scrape_event(
                        "fantasypros_idp_schema_regression",
                        level="error",
                        message=(
                            "FantasyPros IDP fetch exit=2 "
                            "(ecrData shape changed or rows below floor)"
                        ),
                        exit_code=rc,
                    )
                elif rc != 0:
                    _record_scrape_event(
                        "fantasypros_idp_fetch_failed",
                        level="warning",
                        message=f"FantasyPros IDP fetch returned exit={rc}",
                        exit_code=rc,
                    )
            except Exception as _fp_err:
                _record_scrape_event(
                    "fantasypros_idp_fetch_exception",
                    level="warning",
                    message=f"FantasyPros IDP fetch raised: {_fp_err}",
                )

            # Refresh The IDP Show (Adamidp) rankings.  The fetcher
            # reads cookies from ``idpshow_session.json`` at the repo
            # root — if the file is missing (e.g. fresh deploy before
            # the operator has pasted cookies) we skip silently.
            # When cookies have expired the fetcher returns non-zero
            # and we surface it as a warning so the stale-data banner
            # knows to prompt a cookie refresh.
            _idpshow_session = BASE_DIR / "idpshow_session.json"
            if _idpshow_session.exists():
                try:
                    from scripts import fetch_idpshow as _idpshow_fetch

                    rc = _idpshow_fetch.main([])
                    if rc != 0:
                        _record_scrape_event(
                            "idpshow_fetch_failed",
                            level="warning",
                            message=(
                                f"IDP Show fetch returned exit={rc}.  "
                                f"Session cookies may have expired — "
                                f"refresh idpshow_session.json."
                            ),
                            exit_code=rc,
                        )
                except Exception as _idpshow_err:
                    _record_scrape_event(
                        "idpshow_fetch_exception",
                        level="warning",
                        message=f"IDP Show fetch raised: {_idpshow_err}",
                    )
            else:
                log.info(
                    "IDP Show skipped — idpshow_session.json missing; "
                    "operator must paste cookies into that file to enable."
                )

            _update_scrape_progress(
                step="publish",
                source="api_cache",
                step_index=4,
                step_total=4,
                event="phase_start",
                message="Publishing data to in-memory cache",
            )

            elapsed = time.time() - start
            player_count = len(result.get("players", {}))
            site_count = len([s for s in result.get("sites", []) if s.get("playerCount", 0) > 0])
            total_sites = len(result.get("sites", []))

            # R-3 / audit O-3: Block partial scrape promotion.
            #
            # The ratio test below is NOT sufficient on its own and for a
            # long time was the whole guard.  ``result["sites"]`` is
            # populated from the legacy in-scraper SITES dict, which has
            # exactly TWO entries on live data (ktc, idpTradeCalc) against
            # a 21-source registry — measured on the pinned export
            # fixture.  So "fewer than half the sites" degenerates to
            # "fewer than one of two", i.e. it blocks only on TOTAL loss:
            # if KTC dies and IDPTradeCalc survives, 1 < 1 is false and
            # the board publishes without its own cross-market anchor.
            #
            # The payload already declares what it cannot do without.
            # ``coverageAudit.expectedSites`` names the anchor per asset
            # class ({"offense": ["ktc"], "idp": ["idpTradeCalc"]}), so a
            # missing anchor is detectable without inventing a threshold.
            # Losing one is not a "half the sites" condition — it is the
            # loss of a load-bearing input, and IDPTradeCalc in
            # particular is the IDP backbone and 90% of every IDP value
            # under alpha-shrinkage.
            missing_anchors = _missing_expected_sites(result)
            if missing_anchors or (total_sites > 0 and site_count < total_sites / 2):
                anchor_note = (
                    f"MISSING ANCHOR SOURCE(S): {', '.join(missing_anchors)}"
                    if missing_anchors
                    else f"only {site_count}/{total_sites} sites"
                )
                log.warning(
                    f"PARTIAL SCRAPE NOT PROMOTED — {anchor_note}; "
                    f"{player_count} players, {elapsed:.1f}s. Keeping last-known-good data."
                )
                send_alert(
                    f"PARTIAL SCRAPE NOT PROMOTED: {anchor_note}",
                    (
                        f"Players: {player_count}\n"
                        f"Sites with data: {site_count}/{total_sites}\n"
                        + (
                            f"Missing anchor sources: {', '.join(missing_anchors)}\n"
                            if missing_anchors
                            else ""
                        )
                        + f"Duration: {elapsed:.1f}s\n\n"
                        "Partial scrape data was NOT promoted to production.\n"
                        "The server continues serving last-known-good data.\n"
                        "Some sites may be down or blocking the scraper."
                    ),
                )
                # Audit O-2: this used to call _mark_scrape_success, filing a
                # REFUSED scrape as a successful one — so the 24h success
                # rate read 100% while every run was being thrown away.
                _mark_scrape_blocked(
                    f"Not promoted — {anchor_note}",
                    elapsed,
                    player_count,
                    site_count,
                    total_sites,
                )
                return latest_data  # Return existing data, not the partial result

            # R-10: Disk space guard — skip disk write if space is critically low.
            disk_ok, free_mb = _check_disk_space()
            if not disk_ok:
                log.error(
                    f"DISK SPACE LOW — only {free_mb}MB free (minimum {DISK_SPACE_MIN_MB}MB). "
                    "Scrape data will be served from memory but NOT written to disk."
                )
                send_alert(
                    f"DISK SPACE CRITICALLY LOW: {free_mb}MB free",
                    (
                        f"Available disk space: {free_mb}MB\n"
                        f"Minimum required: {DISK_SPACE_MIN_MB}MB\n\n"
                        "Scrape data was loaded into memory but NOT written to disk.\n"
                        "Please free disk space on the server."
                    ),
                )

            latest_data = result
            result_date = str(result.get("date") or "").strip()
            source_path = ""
            if result_date:
                candidate = DATA_DIR / f"dynasty_data_{result_date}.json"
                if candidate.exists():
                    source_path = str(candidate)
            _set_latest_data_source(
                "scrape_run", source_path, produced_at=result.get("scrapeTimestamp")
            )
            # Fresh scrape promotion — rank-history log gets a new
            # "today" entry.  Startup priming from cached disk data
            # (``_prime_latest_payload`` called in the lifespan hook)
            # leaves is_fresh_scrape=False so the history log stays
            # read-only until a real scrape lands.
            #
            # Offloaded to the threadpool: the contract build + 5×
            # multi-MB json/gzip encodes are seconds of CPU that used
            # to freeze the event loop at every scrape end.  The
            # function computes into locals and publishes via a tight
            # reference-assignment swap, so interleaved requests keep
            # serving the OLD generation until the new one is fully
            # encoded.  (The lifespan call site stays inline — it runs
            # before the port binds, so there is nothing to block.)
            await run_in_threadpool(_prime_latest_payload, result, is_fresh_scrape=True)

            _mark_scrape_success(elapsed, player_count, site_count, total_sites)

            log.info(
                f"SCRAPE COMPLETE — {player_count} players, "
                f"{site_count}/{total_sites} sites, {elapsed:.1f}s"
            )

            # Best-effort disk retention.  Regenerable raw/export
            # archives accumulate every scrape and would otherwise fill
            # the disk.  A prune failure must never fail the scrape.
            try:
                from src.maintenance.retention import prune_data_dir

                _ret = prune_data_dir(BASE_DIR)
                if _ret.total_deleted or _ret.total_errors:
                    log.info("retention: %s", _ret.summary())
            except Exception as _ret_exc:  # noqa: BLE001
                log.warning("retention prune skipped: %s", _ret_exc)

            return result
        except Exception as e:
            elapsed = time.time() - start
            _mark_scrape_failure(e, elapsed)
            error_trace = traceback.format_exc()
            log.error(f"SCRAPE FAILED after {elapsed:.1f}s: {e}")
            log.error(error_trace)
            send_alert(
                f"Scrape failed: {type(e).__name__}",
                f"Error: {e}\n\nDuration: {elapsed:.1f}s\n\n{error_trace[-1500:]}",
            )
            return None
        finally:
            _finalize_scrape_run(worker_id)


def check_uptime_once() -> tuple[bool, str | None, int | None]:
    """Run one synchronous uptime probe against the configured URL."""
    if not UPTIME_CHECK_URL:
        return False, "UPTIME_CHECK_URL is empty", None

    req = urllib.request.Request(
        UPTIME_CHECK_URL,
        headers={"User-Agent": "dynasty-uptime-watchdog/1.0"},
        method="GET",
    )
    try:
        with urllib.request.urlopen(req, timeout=UPTIME_CHECK_TIMEOUT_SEC) as resp:
            status_code = int(getattr(resp, "status", 200))
            if 200 <= status_code < 400:
                return True, None, status_code
            return False, f"Unexpected status code {status_code}", status_code
    except urllib.error.HTTPError as e:
        return False, f"HTTPError {e.code}", int(e.code)
    except Exception as e:
        return False, f"{type(e).__name__}: {e}", None


async def uptime_watchdog_loop():
    """Periodic external uptime checks + alerting."""
    if not UPTIME_CHECK_ENABLED:
        log.info("Uptime watchdog disabled (UPTIME_CHECK_ENABLED=false)")
        return
    if not UPTIME_CHECK_URL:
        log.warning("Uptime watchdog enabled but UPTIME_CHECK_URL is empty; watchdog disabled.")
        uptime_status["enabled"] = False
        return

    log.info(
        "Uptime watchdog enabled — url=%s interval=%ss threshold=%s",
        UPTIME_CHECK_URL,
        UPTIME_CHECK_INTERVAL_SEC,
        UPTIME_ALERT_FAIL_THRESHOLD,
    )
    while True:
        now_iso = datetime.now(timezone.utc).isoformat()
        ok, error, status_code = await asyncio.to_thread(check_uptime_once)
        uptime_status["last_check"] = now_iso
        uptime_status["last_http_status"] = status_code

        if ok:
            was_down = uptime_status["consecutive_failures"] >= UPTIME_ALERT_FAIL_THRESHOLD
            uptime_status["consecutive_failures"] = 0
            uptime_status["last_ok"] = now_iso
            uptime_status["last_error"] = None
            if was_down:
                send_alert(
                    "Uptime recovered",
                    f"Recovered successfully.\nURL: {UPTIME_CHECK_URL}\nChecked at: {now_iso}\nStatus: {status_code}",
                )
        else:
            uptime_status["consecutive_failures"] += 1
            uptime_status["last_error"] = error
            failures = uptime_status["consecutive_failures"]
            log.warning(
                "Uptime check failed (%s/%s): %s", failures, UPTIME_ALERT_FAIL_THRESHOLD, error
            )
            if failures >= UPTIME_ALERT_FAIL_THRESHOLD:
                send_alert(
                    f"Uptime check failing ({failures} consecutive)",
                    (
                        f"URL: {UPTIME_CHECK_URL}\n"
                        f"Consecutive failures: {failures}\n"
                        f"Last status code: {status_code}\n"
                        f"Last error: {error}\n"
                        f"Checked at: {now_iso}"
                    ),
                )

        await asyncio.sleep(max(30, UPTIME_CHECK_INTERVAL_SEC))


# ── SCHEDULER ───────────────────────────────────────────────────────────
async def scheduled_scrape():
    """Called by the background scheduler every SCRAPE_INTERVAL_HOURS."""
    log.info(f"Scheduled scrape triggered (every {SCRAPE_INTERVAL_HOURS}h)")
    await run_scraper(trigger="scheduled")
    # Update next scrape time
    from datetime import timedelta

    scrape_status["next_scrape"] = (
        datetime.now(timezone.utc) + timedelta(hours=SCRAPE_INTERVAL_HOURS)
    ).isoformat()


async def schedule_loop():
    """Simple async loop that runs the scraper on an interval."""
    from datetime import timedelta

    while True:
        scrape_status["next_scrape"] = (
            datetime.now(timezone.utc) + timedelta(hours=SCRAPE_INTERVAL_HOURS)
        ).isoformat()
        await asyncio.sleep(SCRAPE_INTERVAL_HOURS * 3600)
        await scheduled_scrape()


# ── APP LIFECYCLE ───────────────────────────────────────────────────────
@asynccontextmanager
async def lifespan(app: FastAPI):
    """Startup: load cached data + kick off first scrape + start scheduler."""
    global latest_data, _startup_checks_summary

    _metrics["server_start_time"] = _utc_now_iso()

    # 0. Run startup validation FIRST so misconfiguration surfaces
    # before any heavy work.  Never raises — logs individual check
    # results and stores the summary for /api/health.
    try:
        from src.api import startup_validation as _sv

        _startup_checks = _sv.run_all()
        _startup_checks_summary = _sv.summary(_startup_checks)
    except Exception as exc:  # noqa: BLE001
        log.error("startup_validation crashed: %s", exc)
        _startup_checks_summary = {
            "error": str(exc),
            "total": 0,
            "ok": 0,
            "failed": 1,
            "fatal": 0,
        }

    # 1. Load cached data immediately so the dashboard is usable right away
    latest_data = load_from_disk()
    _prime_latest_payload(latest_data)
    if latest_data:
        log.info("Dashboard ready with cached data")
    else:
        log.info("No cached data found — dashboard will show empty until first scrape completes")

    # 1b. Hydrate persisted auth sessions so users don't have to re-login
    # on every deploy.  Any failure here falls through to empty in-memory
    # sessions — the existing pre-persistence behavior — so a broken
    # session store can never brick auth entirely.
    try:
        from src.api import session_store as _ss

        hydrated = _ss.hydrate(allowlist=PRIVATE_APP_ALLOWED_USERNAMES)
        auth_sessions.update(hydrated)
        log.info("session_store: hydrated %d sessions from disk", len(hydrated))
    except Exception as exc:  # noqa: BLE001
        log.warning("session_store hydrate on startup failed: %s", exc)

    # 2. Start first scrape in background (don't block startup)
    async def initial_scrape():
        await asyncio.sleep(3)  # small delay to let server finish booting
        await run_scraper(trigger="startup")

    scrape_task = asyncio.create_task(initial_scrape())

    # 3. Start the recurring schedule
    scheduler_task = asyncio.create_task(schedule_loop())
    uptime_task = asyncio.create_task(uptime_watchdog_loop())
    # Public league snapshot warmup — kicks a background rebuild if
    # no persisted snapshot was loaded at boot.  Name is resolved at
    # call time (Python late-binding), so the fact that the function
    # is defined further down in the module is fine.
    try:
        _warmup_public_snapshot()
    except Exception as exc:  # noqa: BLE001
        log.warning("public_league warmup failed at startup: %s", exc)

    # Per-source value history backfill — if the snapshot log is
    # missing or empty, mine the historical ``data/dynasty_data_*.json``
    # exports so the PlayerPopup chart has ~28 days of per-source
    # history on day one.  Skipped when the log already has entries
    # (idempotent, safe to re-run).  Runs sync at boot because it's
    # fast (<2s) and the data is needed before the first
    # /api/data/player-source-history request lands.
    try:
        history_path = _source_history.HISTORY_PATH
        needs_backfill = not history_path.exists() or history_path.stat().st_size == 0
        if needs_backfill:
            exports = sorted((DATA_DIR).glob("dynasty_data_*.json"))
            if exports:
                written = _source_history.backfill_from_exports(exports)
                log.info(
                    "source_history: backfilled %d snapshots from %d exports", written, len(exports)
                )
    except Exception as exc:  # noqa: BLE001
        log.warning("source_history: startup backfill failed: %s", exc)

    log.info(f"Server started — scraping every {SCRAPE_INTERVAL_HOURS}h")
    log.info(f"Dashboard: http://localhost:{PORT}")

    yield  # app is running

    # Cleanup
    scrape_task.cancel()
    scheduler_task.cancel()
    uptime_task.cancel()
    log.info("Server shutting down")


# ── FASTAPI APP ─────────────────────────────────────────────────────────
app = FastAPI(
    title="Dynasty Trade Calculator",
    lifespan=lifespan,
)
app.add_middleware(GZipMiddleware, minimum_size=1024)

# Global exception handler — catches ANY unhandled exception from
# handlers, middleware, or dependency resolution.  Returns the
# standard error-envelope + logs with full context (requestId,
# path, method, IP, traceback).  Installed before any other
# middleware registers so it wraps everything.
from src.api.error_responses import install_exception_handler as _install_exc_handler  # noqa: E402

_install_exc_handler(app)

# ROS engine router.  Strict isolation: never mutates dynasty contract
# paths or trade-calculator math; reads/writes only ``data/ros/*``.
from src.ros.api import router as _ros_router  # noqa: E402

app.include_router(_ros_router)

# Consensus Edge router.  Strict isolation: reads ``latest_contract_data``
# and never mutates it, never writes ``rankDerivedValue``, and adds no
# behaviour to an existing route.  That isolation is what lets the
# ``consensus_edge`` feature flag — which gates every handler here and
# nothing else in the codebase — default **ON** since 2026-08-04 without
# being able to move a number that was already on screen.  Rollback:
# RISKIT_FEATURE_CONSENSUS_EDGE=0 + restart.
from src.consensus_edge.api import router as _consensus_edge_router  # noqa: E402

app.include_router(_consensus_edge_router)


@app.middleware("http")
async def _count_requests(request: Request, call_next):
    """R-9: Count all HTTP requests for metrics."""
    _metrics["request_count"] = _metrics.get("request_count", 0) + 1
    return await call_next(request)


# Roadmap 4.1: reject oversized request bodies before a handler buffers
# them (memory-exhaustion guard on the POST endpoints: trade / angle /
# rankings-overrides / custom-alerts / user-state / export-ktc).
# Generous default so no legitimate payload is ever 413'd; tunable
# without a redeploy via MAX_REQUEST_BYTES.  Only the declared
# Content-Length is checked (covers normal clients; no body buffering
# is added here).
try:
    _MAX_REQUEST_BYTES = int(os.environ.get("MAX_REQUEST_BYTES") or 2_097_152)
except (TypeError, ValueError):
    _MAX_REQUEST_BYTES = 2_097_152


@app.middleware("http")
async def _limit_request_size(request: Request, call_next):
    if request.method in ("POST", "PUT", "PATCH"):
        cl = request.headers.get("content-length")
        if cl is not None:
            try:
                if int(cl) > _MAX_REQUEST_BYTES:
                    return JSONResponse(
                        status_code=413,
                        content={
                            "ok": False,
                            "error": "Request body too large.",
                            "maxBytes": _MAX_REQUEST_BYTES,
                        },
                    )
            except (TypeError, ValueError):
                pass
    return await call_next(request)


# Paths under /api/* that do NOT require an authenticated session.
# Anything else gets 401'd by ``_private_api_gate`` below.  Closes
# the scrape risk: without this gate, ``curl /api/data`` from a
# stranger returns the full private rankings contract.
_PUBLIC_API_EXACT = frozenset(
    {
        "/api/health",
        "/api/status",
        "/api/uptime",
        "/api/metrics",
        "/api/leagues",
        "/api/rankings/sources",
        "/api/auth/status",
        "/api/auth/login",
        "/api/auth/logout",
        "/api/scaffold/status",
        # /league page is a public view — its draft-capital tab reads
        # this endpoint.  Payload is public Sleeper data (team names,
        # pick dollar values, owners) already viewable on Sleeper.
        #
        # It is public with a REDACTION, not because the raw payload is
        # safe: each pick also carries ``rookieName`` / ``rookiePos`` /
        # ``rookieKtcValue`` / ``rookieKtcDollar`` / ``rookieIdpDollar``,
        # filled from ``_our_rookie_pool()`` — our contract's
        # ``playersArray`` ordered by ``rankDerivedValue``.  That is the
        # proprietary rookie board, and an earlier version of this
        # comment asserted it wasn't here.  ``get_draft_capital`` strips
        # those fields for unauthenticated callers; only the private
        # /draft page consumes them.  See
        # ``_redact_draft_capital_for_public`` and
        # tests/api/test_draft_capital_public_redaction.py.
        "/api/draft-capital",
        # Aggregated public sports news (Sleeper trending + public
        # RSS/sitemap providers).  Zero league-private data — no
        # rosters, no rankings, no user state.  The public
        # /league/player/<id> journey page server-renders a "Recent
        # news" card from this endpoint, so it must be reachable
        # without a session.
        "/api/news",
    }
)
# Endpoints that handle their own auth (bearer token, etc.) — the
# session-cookie middleware must skip them so the endpoint's own
# check runs.
_SELF_AUTHED_API_EXACT = frozenset(
    {
        "/api/signal-alerts/run",
        # Same bearer-auth pattern as signal-alerts.  Bypass session
        # cookie gate so the systemd timer's curl call hits the
        # endpoint's own bearer check.
        "/api/custom-alerts/run",
        # E2E test-session bootstrap — handles its own bearer-token auth.
        # Returns 404 unless E2E_TEST_MODE + matching bearer secret are
        # both set, so having it bypass the session gate doesn't leak
        # anything in prod (env vars aren't set there).
        "/api/test/create-session",
        # Push public-key endpoint is read-only and stateless — no auth.
        "/api/push/public-key",
    }
)
_PUBLIC_API_PREFIXES = (
    "/api/public/league",
    # Article reads are public so the league can share /league/articles
    # links with people who don't have an account.  The data is already
    # public-safe: team names, scoring totals, manager display names —
    # nothing the public-league pipeline doesn't already expose at
    # /api/public/league.  Generation remains admin-only via the POST
    # endpoint's own _require_admin_session check.
    "/api/league/articles",
)


def _is_public_api_path(path: str) -> bool:
    if path in _PUBLIC_API_EXACT:
        return True
    if path in _SELF_AUTHED_API_EXACT:
        return True
    for prefix in _PUBLIC_API_PREFIXES:
        if path == prefix or path.startswith(prefix + "/"):
            return True
    return False


@app.middleware("http")
async def _private_api_gate(request: Request, call_next):
    """401 any /api/* call without a session, except the public
    allowlist.

    This is now the ONLY auth gate in this process.  It used to be
    half of a pair — page routes redirected via
    ``_require_auth_or_redirect`` — and #555 deleted that half along
    with the page routes themselves.  Pages are gated by
    ``frontend/middleware.js``, which is where they are actually
    served.

    Note the gate runs as MIDDLEWARE, ahead of routing, so it answers
    for unknown ``/api/*`` paths too: an unrecognised endpoint gets a
    401 rather than a 404, deliberately, so the surface is not
    enumerable.  Pinned by
    ``tests/api/test_private_auth.py::test_unknown_api_path_is_json_not_a_page``.

    Also applies rate limiting to public endpoints only — signed-
    in users on private endpoints aren't subject to the limit
    (they already paid the auth cost, and it's just Jason anyway).
    """
    path = request.url.path or ""
    # Rate limit public endpoints to protect against scraper abuse.
    if path.startswith("/api/") and _is_public_api_path(path):
        client_ip = _client_ip_from_request(request)
        try:
            from src.api import rate_limit as _rl

            limited, retry_after = _rl.should_rate_limit(client_ip)
        except Exception:  # noqa: BLE001 — never let rate-limiter break the gate
            limited, retry_after = False, 0
        if limited:
            return JSONResponse(
                status_code=429,
                content={
                    "error": "rate_limited",
                    "message": "Too many requests — slow down.",
                    "retryAfterSeconds": retry_after,
                },
                headers={
                    "Cache-Control": "no-store",
                    "Retry-After": str(retry_after),
                },
            )
    if path.startswith("/api/") and not _is_public_api_path(path):
        if not _is_authenticated(request):
            return JSONResponse(
                status_code=401,
                content={"error": "auth_required", "message": "Sign-in required."},
                headers={"Cache-Control": "no-store"},
            )
    return await call_next(request)


@app.middleware("http")
async def _request_context_middleware(request: Request, call_next):
    """Generate + propagate a per-request correlation ID.

    Registered AFTER ``_private_api_gate`` so it wraps the gate —
    every response (including 401/429 from the gate) gets an
    ``X-Request-Id`` header + the ContextVar is set for any log
    lines emitted during request handling.

    Accepts an incoming ``X-Request-Id`` header (e.g. from nginx
    or an uptime monitor) when present + sane (1-64 chars);
    otherwise mints a fresh token-urlsafe 12-char ID.
    """
    from src.utils import request_context as _rc

    incoming = str(request.headers.get("x-request-id") or "").strip()
    rid = incoming if (1 <= len(incoming) <= 64) else _rc.new_request_id()
    token = _rc.set_request_id(rid)
    try:
        response = await call_next(request)
    finally:
        _rc.reset_request_id(token)
    try:
        response.headers["X-Request-Id"] = rid
    except Exception:  # noqa: BLE001 — some response types reject mutations
        pass
    return response


def _client_ip_from_request(request: Request) -> str:
    """Prefer ``X-Forwarded-For`` (nginx sets it for us in prod);
    fall back to ``request.client.host``."""
    xff = request.headers.get("x-forwarded-for", "")
    if xff:
        # First entry in the chain is the original client.
        first = xff.split(",")[0].strip()
        if first:
            return first
    client = request.client
    return client.host if client else ""


def _overlay_encode_lock(cache_key) -> asyncio.Lock:
    """Return the per-key single-flight lock, creating it on first use."""
    lock = _OVERLAY_ENCODE_LOCKS.get(cache_key)
    if lock is None:
        lock = asyncio.Lock()
        _OVERLAY_ENCODE_LOCKS[cache_key] = lock
    return lock


def _evict_overlay_cache_if_oversized(keep_key) -> None:
    """Safety net for the stable-slot cache.

    Slots self-replace on refresh, so this should never fire in normal
    operation — the key space is bounded by the league registry.  It only
    guards against an unexpected key explosion (e.g. a registry reload
    mid-flight).  ``keep_key`` is the slot about to be written, so we
    never evict the entry this caller is in the middle of producing, and
    in-flight locks are preserved: dropping a HELD lock would let a
    second request create a fresh one and launch a duplicate encode.
    """
    if len(_OVERLAY_RESPONSE_CACHE) < _OVERLAY_RESPONSE_CACHE_MAX:
        return
    for k in [k for k in _OVERLAY_RESPONSE_CACHE if k != keep_key]:
        del _OVERLAY_RESPONSE_CACHE[k]
    # Safe without a guard: runs synchronously on the event-loop thread,
    # no await between the check and the delete.
    for k in [k for k, lk in _OVERLAY_ENCODE_LOCKS.items() if not lk.locked()]:
        del _OVERLAY_ENCODE_LOCKS[k]


async def _serialize_overlaid_response(request, scrubbed, headers, cache_key, overlay_version=None):
    """Serialize a live-overlay / cross-league ``/api/data`` response
    without blocking the event loop.

    The JSON encode (and gzip) of the multi-MB payload is offloaded to a
    worker thread, and — since the spliced overlay is stable within its
    ~15-min cache window — the encoded bytes are memoized under
    ``cache_key`` so repeat requests reuse them instead of re-encoding.
    Pass ``cache_key=None`` to skip the cache (still offloads the encode).
    Adds an ``ETag`` (with ``If-None-Match`` 304 support) and
    ``Vary: Accept-Encoding`` for the negotiated gzip/identity body.

    Concurrent misses for the same key are single-flighted on the event
    loop via a per-key ``asyncio.Lock``: only the first request offloads
    the multi-MB encode; the rest await the lock and read the cached
    result, so a burst after startup / overlay refresh can't fan out into
    N duplicate serializations that saturate the worker pool.

    Version info (overlay_version tuple) is stored inside the cache entry.
    On cache hit, stale versions are re-encoded and replace the prior
    generation in the same slot, bounding memory to at most one per slot.
    """

    def _encode():
        raw = json.dumps(scrubbed, ensure_ascii=False, separators=(",", ":")).encode("utf-8")
        etag = hashlib.sha1(raw).hexdigest()
        gz = gzip.compress(raw, compresslevel=5)
        return etag, raw, gz, overlay_version

    if cache_key is None:
        # Uncacheable (no freshness/version key) — nothing to coalesce on.
        etag, raw, gz, _ = await run_in_threadpool(_encode)
    else:
        entry = _OVERLAY_RESPONSE_CACHE.get(cache_key)
        # Check freshness: if cached version doesn't match current, treat as miss
        if entry is not None and entry[3] != overlay_version:
            entry = None
        if entry is None:
            async with _overlay_encode_lock(cache_key):
                # Re-check: another coroutine may have encoded this key
                # while we waited on the lock.
                entry = _OVERLAY_RESPONSE_CACHE.get(cache_key)
                if entry is not None and entry[3] != overlay_version:
                    entry = None
                if entry is None:
                    entry = await run_in_threadpool(_encode)
                    _evict_overlay_cache_if_oversized(cache_key)
                    _OVERLAY_RESPONSE_CACHE[cache_key] = entry
        etag, raw, gz, _ = entry

    headers["ETag"] = etag
    headers["Vary"] = "Accept-Encoding"
    incoming = request.headers.get("if-none-match", "").strip('"')
    if incoming and incoming == etag:
        return Response(status_code=304, headers=headers)
    accept_encoding = (request.headers.get("accept-encoding") or "").lower()
    if "gzip" in accept_encoding and gz:
        headers["Content-Encoding"] = "gzip"
        return Response(content=gz, media_type="application/json", headers=headers)
    return Response(content=raw, media_type="application/json", headers=headers)


def _overrides_encode_lock(cache_key) -> asyncio.Lock:
    """Per-key single-flight lock for the overrides response memo."""
    lock = _OVERRIDES_ENCODE_LOCKS.get(cache_key)
    if lock is None:
        lock = asyncio.Lock()
        _OVERRIDES_ENCODE_LOCKS[cache_key] = lock
    return lock


def _evict_overrides_cache_if_oversized(keep_key) -> None:
    """Bound the overrides memo.  Key space is normally tiny (one stock
    body × active leagues); the cap only guards against a client
    spraying distinct bodies.  Held locks are preserved for the same
    reason as ``_evict_overlay_cache_if_oversized``."""
    if len(_OVERRIDES_RESPONSE_CACHE) < _OVERRIDES_RESPONSE_CACHE_MAX:
        return
    for k in [k for k in _OVERRIDES_RESPONSE_CACHE if k != keep_key]:
        del _OVERRIDES_RESPONSE_CACHE[k]
    for k in [k for k, lk in _OVERRIDES_ENCODE_LOCKS.items() if not lk.locked()]:
        del _OVERRIDES_ENCODE_LOCKS[k]


# ── API ROUTES ──────────────────────────────────────────────────────────
@app.get("/api/data")
async def get_data(request: Request):
    """Return latest normalized/validated data contract JSON.

    Optional ``?leagueKey=...`` validates against the league registry.

    **Rankings are keyed by scoring profile, not by league.**  When
    two leagues share a profile, they share the rankings pipeline's
    output — the ``players`` / ``playersArray`` / ``sources`` blocks
    are identical and we serve them to any caller whose league
    resolves to the same profile.  Only the league-specific
    ``sleeper`` block (teams, rosters, owners) is per-league; when a
    different league is requested and we don't have that league's
    sleeper data loaded, the block is returned as ``None`` and
    ``meta.sleeperDataReady=false`` tells the client to show
    "no roster data yet" rather than rendering the default league's
    teams under the wrong name.

    503 ``data_not_ready`` only fires when the scoring profiles
    genuinely differ — i.e. the rankings themselves can't be reused.
    """
    # League validation comes first so a stale leagueKey returns 400
    # before we bother assembling the payload.  Skip the loaded-
    # contract check here and enforce below so the 503 path can
    # include the resolved league key in the response.
    try:
        league_cfg = _resolve_league_for_request(request)
    except LeagueResolutionError as err:
        return err.json_response()

    if latest_contract_data:
        loaded_meta = (
            latest_contract_data.get("meta") or {} if isinstance(latest_contract_data, dict) else {}
        )
        loaded_league = str(loaded_meta.get("leagueKey") or "")
        sleeper_matches = bool(loaded_league) and loaded_league == league_cfg.key

        # Scoring mismatch → genuinely different data; 503.  Decided by
        # the factual fingerprint, not the registry label, and unproven
        # fails closed (W18-F001).
        _scoring_err = _scoring_identity_error(latest_contract_data, league_cfg)
        if _scoring_err is not None:
            return _scoring_err

        view = (request.query_params.get("view") or "").strip().lower()
        startup_view = view in {"startup", "boot", "initial"}
        runtime_view = view in {"app", "runtime", "lite"}
        array_view = view in {"array", "desktop"}
        compact_view = view in {"compact", "slim"}

        payload_bytes = latest_data_bytes
        payload_gzip_bytes = latest_data_gzip_bytes
        payload_etag = latest_data_etag
        payload_obj = latest_contract_data
        payload_view_name = "full"

        if startup_view and latest_startup_data is not None:
            payload_bytes = latest_startup_data_bytes
            payload_gzip_bytes = latest_startup_data_gzip_bytes
            payload_etag = latest_startup_data_etag
            payload_obj = latest_startup_data
            payload_view_name = "startup"
        elif runtime_view and latest_runtime_data is not None:
            payload_bytes = latest_runtime_data_bytes
            payload_gzip_bytes = latest_runtime_data_gzip_bytes
            payload_etag = latest_runtime_data_etag
            payload_obj = latest_runtime_data
            payload_view_name = "runtime"
        elif array_view and latest_array_data is not None:
            # Desktop view: full contract minus the legacy ``players``
            # dict (a parallel encoding of ``playersArray``).  Identical
            # board, identical audit fields, ~half the bytes.
            payload_bytes = latest_array_data_bytes
            payload_gzip_bytes = latest_array_data_gzip_bytes
            payload_etag = latest_array_data_etag
            payload_obj = latest_array_data
            payload_view_name = "array"
        elif compact_view and latest_contract_data is not None:
            # Mobile / slow-network view.  Drops the legacy ``players``
            # dict (as ``array`` does) and prunes the three per-player
            # fields no frontend consumer reads.
            #
            # Measured 2026-08-18, 1,109-row contract, gzip level 6:
            # full 1,092.8 KB / array 631.8 KB / compact 491.0 KB.  The
            # comment here used to claim "~90% byte reduction" against
            # a view it was in fact 16.3% LARGER than — compact carried
            # both player encodings while array carried one.
            #
            # "Additive" was the other half of that wrong story: pruning
            # IS safe only while nothing reads a pruned field, and
            # fourteen of the pruned fields were read by
            # ``_materializePlayerArrayRow``.  The shape test could not
            # see it because it pinned the pruned payload's shape, not
            # the frontend's reads.  That gap is now
            # ``tests/api/test_compact_view_consumer_parity.py``.
            payload_view_name = "compact"
            if latest_compact_data_bytes is not None:
                # Fast path: precomputed at refresh time (bytes + gzip +
                # etag), so mobile requests skip the compaction + JSON
                # serialization + gzip that used to run on the event loop
                # for every request.
                payload_bytes = latest_compact_data_bytes
                payload_gzip_bytes = latest_compact_data_gzip_bytes
                payload_etag = latest_compact_data_etag
                payload_obj = latest_compact_data
            else:
                # Fallback: precompute unavailable (e.g. compaction raised
                # during refresh) — build on demand.
                from src.api.compact_view import compact_contract

                compact_obj = compact_contract(latest_contract_data)
                import json as _json

                payload_bytes = _json.dumps(compact_obj).encode("utf-8")
                payload_gzip_bytes = None  # regenerate-on-demand (no cached gzip)
                payload_etag = None
                payload_obj = compact_obj

        headers = {
            # Keep dashboard startup fast with a short cache window + conditional revalidation.
            "Cache-Control": "private, max-age=30, stale-while-revalidate=300",
            "X-Payload-View": payload_view_name,
        }

        # Live Sleeper overlay — applied to BOTH the loaded league
        # AND cross-league requests so post-trade roster moves reflect
        # within ~15 min instead of the ~2h scrape cadence.  The
        # overlay is in-process cached for 15 min (per league) so
        # steady-state cost is one Sleeper round-trip per league per
        # 15-min window.  Failures (Sleeper down, circuit breaker
        # open, empty roster) fall back to the existing paths:
        #   - sleeper_matches=True  → serve baked-in sleeper block
        #     from the cached payload bytes (preserves ETag/304/gzip
        #     fast path)
        #   - sleeper_matches=False → null the sleeper block + flag
        #     sleeperDataReady=false (existing cross-league fallback)
        # This unification is what makes /waivers, /trade, /rosters,
        # /draft etc. converge on the same 15-min ceiling regardless
        # of which league is "loaded."
        loaded_sleeper = (
            (latest_contract_data or {}).get("sleeper") or {}
            if isinstance(latest_contract_data, dict)
            else {}
        )
        id_to_player = loaded_sleeper.get("idToPlayer") or {}
        try:
            overlay = await run_in_threadpool(
                _sleeper_overlay.fetch_sleeper_overlay,
                sleeper_league_id=league_cfg.sleeper_league_id,
                id_to_player=id_to_player if isinstance(id_to_player, dict) else {},
            )
        except Exception as exc:  # noqa: BLE001
            log.warning(
                "sleeper_overlay fetch failed for %s: %s",
                league_cfg.key,
                exc,
            )
            overlay = None

        if overlay and overlay.get("teams"):
            # Splice live overlay onto the rankings payload.  Both
            # paths (loaded + cross-league) layer the overlay's
            # ``teams`` and ``trades`` over the baked sleeper block,
            # while preserving the NFL-wide maps the overlay doesn't
            # rebuild on its own.
            #
            # The overlay's ``trades`` are now produced in the same
            # ``[{leagueId, week, timestamp, sides[]}, ...]`` shape
            # that ``analyzeSleeperTradeHistory`` consumes on the
            # /trades page — see
            # ``src/api/sleeper_overlay.py::_build_trades_block``.
            # That parity is what makes /trades reflect Sleeper trade
            # activity within ~15 min instead of waiting on the next
            # 2h scrape.
            #
            # Strategy:
            #   sleeper_matches=True  → start from the baked sleeper
            #     block (keeps positions, idToPlayer, leagueSettings,
            #     rosterPositions, scoringSettings, etc.) and overlay
            #     teams + trades + tradeWindow* on top.  Tagged
            #     "live-merge" so ops can grep for the new path.
            #
            #   sleeper_matches=False → carry the NFL-wide maps from the
            #     loaded contract and apply the overlay on top, but take
            #     the league-SPECIFIC fields (scoringSettings,
            #     rosterPositions, leagueSettings) from the requested
            #     league's own config or leave them absent.  Readiness
            #     follows ownership — see
            #     ``sleeper_overlay.merge_cross_league_sleeper_block``
            #     (W18-F002).
            scrubbed = dict(payload_obj) if isinstance(payload_obj, dict) else {}
            cross_league_ready = True
            if sleeper_matches:
                overlay_full = {
                    **loaded_sleeper,
                    "teams": overlay["teams"],
                    "trades": overlay.get("trades") or [],
                    "waivers": overlay.get("waivers") or [],
                    "tradeWindowDays": overlay.get("tradeWindowDays")
                    or loaded_sleeper.get("tradeWindowDays"),
                    "tradeWindowStart": overlay.get("tradeWindowStart")
                    or loaded_sleeper.get("tradeWindowStart"),
                    "tradeWindowCutoffMs": overlay.get("tradeWindowCutoffMs")
                    or loaded_sleeper.get("tradeWindowCutoffMs"),
                    "overlaySource": "live-merge",
                    "overlayFetchedAt": overlay.get("overlayFetchedAt"),
                }
            else:
                overlay_full, cross_league_ready = (
                    _sleeper_overlay.merge_cross_league_sleeper_block(
                        loaded_sleeper=loaded_sleeper,
                        overlay=overlay,
                        requested_league_config=overlay.get("leagueConfig"),
                    )
                )
            scrubbed["sleeper"] = overlay_full
            # RE-STAMP THE LINEUP (C2-U1).  The overlay rebuilds
            # ``teams`` from scratch (``sleeper_overlay._build_teams_block``
            # emits no ``optimalLineup``), so the stamp taken at contract
            # build time is discarded here — on the NORMAL path, because
            # the overlay is warmed after every scrape and cached for
            # 15 minutes.  Without this the frontend fails closed and
            # /terminal, /rosters and the team-tier leaderboard all lose
            # their starter/bench split whenever Sleeper is REACHABLE,
            # which is the opposite of a degradation.
            #
            # Re-SOLVED, never copied: the overlay's rosters are fresher
            # than the baked ones, so a copied lineup could start a
            # player dropped ten minutes ago.  Values come from the baked
            # contract because some payload views strip ``playersArray``,
            # and they are scoring-profile scoped so they are identical
            # either way.  Degrades, never raises.
            try:
                _stamp_optimal_lineups_owner(
                    scrubbed,
                    rows=(latest_contract_data or {}).get("playersArray")
                    if isinstance(latest_contract_data, dict)
                    else None,
                )
            except Exception as exc:  # noqa: BLE001
                log.warning("optimalLineup re-stamp failed for %s: %s", league_cfg.key, exc)
            meta = dict(scrubbed.get("meta") or {})
            meta["leagueKey"] = league_cfg.key
            meta["scoringProfile"] = league_cfg.scoring_profile
            meta["sleeperDataReady"] = bool(sleeper_matches or cross_league_ready)
            meta["sleeperSource"] = "overlay"
            if not sleeper_matches:
                meta["sleeperLoadedLeagueKey"] = loaded_league or None
            scrubbed["meta"] = meta
            suffix = "overlay" if sleeper_matches else "cross-league-overlay"
            headers["X-Payload-View"] = f"{payload_view_name}-{suffix}"
            # Cache-key is stable across overlay refreshes: only the base
            # league/view context determines the key. Version info
            # (overlayFetchedAt + payloadETag) is stored inside the cache
            # entry and checked on hit; stale versions are re-encoded in
            # place, bounding memory to one generation per slot.
            overlay_fetched_at = overlay.get("overlayFetchedAt")
            overlay_cache_key = (
                (
                    "overlay",
                    league_cfg.key,
                    loaded_league or "",
                    payload_view_name,
                    bool(sleeper_matches),
                )
                if (overlay_fetched_at and payload_etag)
                else None
            )
            overlay_version = (overlay_fetched_at, payload_etag) if overlay_cache_key else None
            return await _serialize_overlaid_response(
                request, scrubbed, headers, overlay_cache_key, overlay_version
            )

        if not sleeper_matches:
            # Cross-league + overlay unavailable: null the sleeper
            # block so the UI falls back to the data-not-ready state
            # rather than rendering League A's teams under League B's
            # name.
            scrubbed = dict(payload_obj) if isinstance(payload_obj, dict) else {}
            meta = dict(scrubbed.get("meta") or {})
            meta["leagueKey"] = league_cfg.key
            meta["scoringProfile"] = league_cfg.scoring_profile
            meta["sleeperLoadedLeagueKey"] = loaded_league or None
            meta["sleeperDataReady"] = False
            scrubbed["sleeper"] = None
            scrubbed["meta"] = meta
            headers["X-Payload-View"] = f"{payload_view_name}-cross-league"
            # Same stable-slot scheme as the overlay path: the key is the
            # league/view context only, and the base ETag rides along as
            # the entry version, so each scrape refresh REPLACES the slot
            # instead of minting a new multi-MB generation beside it.
            xleague_cache_key = (
                ("xleague", league_cfg.key, loaded_league or "", payload_view_name)
                if payload_etag
                else None
            )
            xleague_version = (payload_etag,) if xleague_cache_key else None
            return await _serialize_overlaid_response(
                request, scrubbed, headers, xleague_cache_key, xleague_version
            )
        # sleeper_matches=True + overlay unavailable: fall through to
        # the cached payload-bytes fast path below — serves the baked
        # sleeper block from the most recent scrape.

        # This fast path serves gzip OR identity bytes off the same
        # publicly-cacheable URL depending on ``Accept-Encoding``.  Since
        # we set ``Content-Encoding`` by hand (bypassing GZipMiddleware,
        # which would otherwise add this), we must advertise the
        # negotiation so a shared/CDN cache keys the two encodings apart
        # instead of serving gzip to an identity client (or vice versa).
        headers["Vary"] = "Accept-Encoding"

        if payload_etag:
            headers["ETag"] = payload_etag
            incoming = request.headers.get("if-none-match", "").strip('"')
            if incoming and incoming == payload_etag:
                return Response(status_code=304, headers=headers)

        accept_encoding = (request.headers.get("accept-encoding") or "").lower()
        if "gzip" in accept_encoding and payload_gzip_bytes:
            headers["Content-Encoding"] = "gzip"
            return Response(
                content=payload_gzip_bytes, media_type="application/json", headers=headers
            )
        if payload_bytes:
            return Response(content=payload_bytes, media_type="application/json", headers=headers)
        return JSONResponse(content=payload_obj, headers=headers)
    return JSONResponse(
        status_code=503,
        content={"error": "No data available yet. First scrape may still be running."},
    )


@app.get("/api/dynasty-data")
async def get_dynasty_data_alias(request: Request):
    """Compatibility alias for frontend consumers expecting /api/dynasty-data."""
    return await get_data(request)


@app.get("/api/movers")
async def get_movers(request: Request):
    """Buy-low / sell-high signals from rank history.

    Query params:
      * ``window`` — days to compare against (default 14, max 90)
      * ``threshold`` — minimum absolute rank change to qualify
        (default 15)
      * ``limit`` — max entries per side (default 10)

    Returns::

        {
          "window": 2,
          "windowRequested": 14,
          "historyDepthDays": 2,
          "threshold": 15,
          "asOf": "2026-04-26",
          "risers": [
            {"name": "...", "team": "TEX", "position": "WR", "playerId": "...",
             "rankNow": 12, "rankThen": 32, "delta": 20,
             "rankNowDate": "2026-04-26", "rankThenDate": "2026-04-24",
             "spanDays": 2,
             "valueNow": 7421, "currentSourceRanks": {...}}
          ],
          "fallers": [...]
        }

    Risers = rank improved (number got smaller).  Fallers = rank got
    worse (number got bigger).

    ``window`` is the span we ACTUALLY measured, not the span that was
    asked for — ``windowRequested`` echoes the request.  The two differ
    whenever ``data/rank_history.jsonl`` is shallower than the request
    (it is only a few days deep today), and reporting the request in
    that case labelled a 2-day delta "90 days".  ``spanDays`` on each
    entry is the authoritative per-player span, since players enter the
    log at different dates.

    ``currentSourceRanks`` is each source's CURRENT rank for the player
    (from ``sourceOriginalRanks``), NOT a per-source delta — the log
    stores only the blended consensus rank, so per-source history to
    difference against does not exist.  The field was renamed from
    ``perSourceDelta`` for exactly this reason; this docstring lagged
    the rename and promised a breakdown of "which sources drove the
    move" that was never computed.
    """
    try:
        window = int(request.query_params.get("window", 14))
    except (TypeError, ValueError):
        window = 14
    window = max(2, min(90, window))
    try:
        threshold = int(request.query_params.get("threshold", 15))
    except (TypeError, ValueError):
        threshold = 15
    threshold = max(1, threshold)
    try:
        limit = int(request.query_params.get("limit", 10))
    except (TypeError, ValueError):
        limit = 10
    limit = max(1, min(50, limit))

    # Load full history; we need both endpoints of the window so we
    # ask for at least window+1 days.
    history = _rank_history.load_history(days=window + 1)
    if not history:
        return JSONResponse(
            content={
                "window": 0,
                "windowRequested": window,
                "historyDepthDays": 0,
                "threshold": threshold,
                "asOf": None,
                "risers": [],
                "fallers": [],
            },
            headers={"Cache-Control": "private, max-age=60, stale-while-revalidate=300"},
        )

    def _span_days(then_date: Any, now_date: Any) -> int | None:
        """Whole days between two ``YYYY-MM-DD`` stamps, or None."""
        if not isinstance(then_date, str) or not isinstance(now_date, str):
            return None
        try:
            a = datetime.strptime(then_date, "%Y-%m-%d")
            b = datetime.strptime(now_date, "%Y-%m-%d")
        except ValueError:
            return None
        return max(0, (b - a).days)

    # Resolve the comparison anchor by DATE, not by list position.
    # ``load_history(days=window + 1)`` only TRIMS the log — it cannot
    # extend it — so ``series[0]`` is just the oldest point on disk,
    # which today is ~2 days back regardless of what was requested.
    # Anchoring there and echoing ``window`` in the response labelled a
    # 2-day move "90 days".  Now: take the newest point at or before
    # ``asOf − window`` days when the log reaches that far, fall back to
    # the oldest point when it doesn't, and report the span we actually
    # measured either way.
    as_of = max((s[-1]["date"] for s in history.values() if s), default=None)
    # ``historyDepthDays`` must describe THE LOG, not the slice we just
    # took out of it.  ``history`` is ``load_history(days=window + 1)``,
    # so deriving depth from its oldest entry yields
    # ``min(true_depth, window)`` by construction — the field could never
    # say "the log is deeper than you asked", which is half of what a
    # depth field is for.  It read correctly only because the live log
    # happens to be shallower than any window, i.e. by the same accident
    # this whole finding was filed against.
    #
    # ``rank_history.coverage()`` reads the on-disk date range directly.
    # It falls back to the trimmed span if the log cannot be read, which
    # is the conservative direction: understating depth makes the window
    # look shorter, never longer.
    oldest_seen = min((s[0]["date"] for s in history.values() if s), default=None)
    # NOTE the unit: ``coverage()["spanDays"]`` is a COUNT of calendar
    # days covered (inclusive, so two snapshots two days apart span 3),
    # because it exists to compute ``missingDays``.  ``historyDepthDays``
    # sits beside ``window``, which is a lookback in days.  Take the
    # untrimmed FIRST DATE from coverage and measure it with the same
    # ``_span_days`` the window uses, so there is exactly one definition
    # of "days back" in this response rather than two that differ by one.
    try:
        _cov = _rank_history.coverage()
        _first_on_disk = _cov.get("firstDate") if isinstance(_cov, dict) else None
    except Exception:  # noqa: BLE001 — diagnostics must never break the route
        _first_on_disk = None
    history_depth_days = _span_days(
        _first_on_disk if isinstance(_first_on_disk, str) else oldest_seen, as_of
    )
    cutoff_date: str | None = None
    if isinstance(as_of, str):
        from datetime import timedelta

        try:
            cutoff_date = (datetime.strptime(as_of, "%Y-%m-%d") - timedelta(days=window)).strftime(
                "%Y-%m-%d"
            )
        except ValueError:
            cutoff_date = None
    # The response-level span is the span of ONE well-defined anchor
    # date — the newest snapshot at or before the cutoff, or the oldest
    # snapshot we have when the log is shallower than the request.  It
    # is not ``min(window, depth)``: the log is gap-tolerant, so with
    # sparse snapshots the nearest usable anchor can sit further back
    # than ``window`` and the honest number is that real span.  Players
    # who entered the log after the anchor date measure a shorter span;
    # ``spanDays`` on each entry is authoritative for that entry.
    anchor_date = oldest_seen
    if cutoff_date:
        newest_within = ""
        for s in history.values():
            # Series are date-sorted ascending, so the first point past
            # the cutoff ends this series' contribution — with a deep
            # log that is the second point, which keeps this pass cheap
            # instead of touching every point of every player.
            for p in s:
                d = p.get("date")
                if not isinstance(d, str) or d > cutoff_date:
                    break
                if d > newest_within:
                    newest_within = d
        if newest_within:
            anchor_date = newest_within
    measured_window = _span_days(anchor_date, as_of)
    if measured_window is None:
        measured_window = 0

    # Index the live contract by displayName so we can stitch in
    # team / position / current value + per-source rank metadata for
    # each mover.  Falls back to bare-name only when the player is
    # missing from the live board.
    contract = latest_contract_data or {}
    by_name: dict[str, dict] = {}
    by_name_lower: dict[str, dict] = {}
    by_name_scoped: dict[str, dict] = {}
    arr = contract.get("playersArray") or []
    if isinstance(arr, list):
        for row in arr:
            if not isinstance(row, dict):
                continue
            display = str(row.get("displayName") or row.get("canonicalName") or "")
            if not display:
                continue
            by_name[display] = row
            by_name_lower[display.lower()] = row
            asset = str(row.get("assetClass") or "").strip().lower()
            if asset:
                by_name_scoped[f"{display.lower()}::{asset}"] = row

    risers: list[dict] = []
    fallers: list[dict] = []
    for raw_key, series in history.items():
        if not series or len(series) < 2:
            continue
        # History keys are ``"{Display Name}::{asset_class}"`` (see
        # ``src/api/rank_history.py::_player_key``).  Split the key so
        # we can (a) look up the contract row by clean displayName,
        # (b) emit a clean name to the UI, and (c) carry assetClass
        # through to the response so the frontend can disambiguate
        # cross-universe name collisions (offense vs IDP).
        if "::" in raw_key:
            clean_name, asset_class = raw_key.split("::", 1)
        else:
            clean_name, asset_class = raw_key, ""
        clean_name = clean_name.strip()
        asset_class = asset_class.strip().lower()
        # Most-recent ``rank`` and the rank from ~``window`` days ago.
        # Series is already date-sorted ascending by load_history.
        latest = series[-1]
        anchor = series[0]
        if cutoff_date:
            for p in series:
                d = p.get("date")
                if isinstance(d, str) and d <= cutoff_date:
                    anchor = p
                else:
                    break
        try:
            r_now = int(latest.get("rank"))
            r_then = int(anchor.get("rank"))
        except (TypeError, ValueError):
            continue
        if r_now <= 0 or r_then <= 0:
            continue
        # Rank smaller = better; "delta" is positive when player rose.
        delta = r_then - r_now
        if abs(delta) < threshold:
            continue
        # Prefer a scope-matched contract row (handles offense/IDP
        # name collisions), then case-insensitive name match, then
        # exact-key fallback.
        row = (
            by_name_scoped.get(f"{clean_name.lower()}::{asset_class}")
            or by_name_lower.get(clean_name.lower())
            or by_name.get(raw_key)
            or {}
        )
        current_source_ranks: dict[str, int] = {}
        # Each source's CURRENT rank for this player — NOT a delta.
        # ``sourceOriginalRanks`` stamps the un-Hampel-filtered ranks
        # as of the live contract; the history log carries only the
        # blended consensus rank, so there is no per-source "then" to
        # difference against.  The local was named ``per_source_delta``
        # long after the response key was corrected — renamed here so
        # the code stops arguing with the payload.
        sor = row.get("sourceOriginalRanks") or {}
        if isinstance(sor, dict):
            for src_key, src_rank in sor.items():
                try:
                    current_source_ranks[str(src_key)] = int(src_rank)
                except (TypeError, ValueError):
                    continue
        # Prefer the contract's assetClass (most current) but fall
        # back to the history key's parsed asset_class so picks /
        # IDPs aren't mis-stamped as "?" when the contract has
        # rotated them off the board.
        resolved_asset = str(row.get("assetClass") or "").strip().lower() or asset_class or None
        entry = {
            "name": clean_name,
            "assetClass": resolved_asset,
            "playerId": str(row.get("playerId") or "") or None,
            "position": str(row.get("position") or "") or None,
            "team": str(row.get("team") or "") or None,
            "rankNow": r_now,
            "rankThen": r_then,
            "delta": delta,
            "rankNowDate": latest.get("date"),
            "rankThenDate": anchor.get("date"),
            "spanDays": _span_days(anchor.get("date"), latest.get("date")),
            "valueNow": int(row.get("rankDerivedValue") or 0) or None,
            "currentSourceRanks": current_source_ranks if current_source_ranks else None,
        }
        if delta > 0:
            risers.append(entry)
        else:
            fallers.append(entry)

    # Deterministic ordering.  The primary key is the move size, but
    # equal-delta ties used to fall through to ``history`` dict
    # insertion order — i.e. whatever order the JSONL happened to
    # stamp names in — so two runs over the same data could emit
    # different top-N slices.  Bigger current value wins the tie
    # (a 20-spot move by a top-30 asset is the more interesting one),
    # then name + assetClass as a total order.
    risers.sort(
        key=lambda e: (-e["delta"], -(e["valueNow"] or 0), e["name"], e["assetClass"] or "")
    )
    fallers.sort(
        key=lambda e: (e["delta"], -(e["valueNow"] or 0), e["name"], e["assetClass"] or "")
    )
    return JSONResponse(
        content={
            "window": measured_window,
            "windowRequested": window,
            "historyDepthDays": history_depth_days,
            "threshold": threshold,
            "asOf": as_of,
            "risers": risers[:limit],
            "fallers": fallers[:limit],
        },
        headers={"Cache-Control": "private, max-age=60, stale-while-revalidate=300"},
    )


@app.get("/api/data/rank-history")
async def get_rank_history(request: Request):
    """Per-player rank history series for the last ``days`` days.

    Every contract build appends the ranked board to a JSONL log
    (see ``src/api/rank_history.py``).  This endpoint reads the log
    and flips it into the per-player ``{name: [{date, rank}, ...]}``
    shape the frontend ``RankChangeGlyph`` consumes.

    Query params:
      * ``days`` — window in days (default
        ``_rank_history.DEFAULT_HISTORY_WINDOW_DAYS``, clamped to
        ``[1, _rank_history.MAX_SNAPSHOTS]`` — 1095, i.e. three years).
        This line used to say "max 180", which is the
        player-source-history window below, not this one.

    The log is already mirrored onto each row's ``rankHistory`` at
    contract build time, so most consumers don't need this endpoint
    — it exists for tools that want the raw series without fetching
    the full 4 MB contract.
    """
    try:
        requested = int(request.query_params.get("days", _rank_history.DEFAULT_HISTORY_WINDOW_DAYS))
    except (TypeError, ValueError):
        requested = _rank_history.DEFAULT_HISTORY_WINDOW_DAYS
    days = max(1, min(_rank_history.MAX_SNAPSHOTS, requested))
    history = _rank_history.load_history(days=days)
    return JSONResponse(
        content={"days": days, "history": history},
        headers={"Cache-Control": "private, max-age=60, stale-while-revalidate=300"},
    )


@app.get("/api/data/player-source-history")
async def get_player_source_history(request: Request):
    """Per-source value history for a single player.

    Returns the per-source and blended value timeline the PlayerPopup
    chart renders as multiple overlaid lines — one thin line per
    ranking source and one bold line for our blend.

    Query params:
      * ``name``        — player display name (required, case-insensitive)
      * ``days``        — window in days (default 180, max 180)
      * ``assetClass``  — optional disambiguator ("offense" / "idp" /
                          "pick") for cross-universe name collisions
    """
    name = (request.query_params.get("name") or "").strip()
    if not name:
        return JSONResponse(
            status_code=400,
            content={"error": "Missing required 'name' query param."},
        )
    try:
        requested = int(
            request.query_params.get("days", _source_history.DEFAULT_HISTORY_WINDOW_DAYS)
        )
    except (TypeError, ValueError):
        requested = _source_history.DEFAULT_HISTORY_WINDOW_DAYS
    days = max(1, min(_source_history.MAX_SNAPSHOTS, requested))
    asset_class = (request.query_params.get("assetClass") or "").strip() or None
    history = _source_history.load_player_history(
        name,
        days=days,
        asset_class=asset_class,
    )
    return JSONResponse(
        content={
            "name": name,
            "days": days,
            "assetClass": asset_class,
            **history,
        },
        headers={"Cache-Control": "private, max-age=120, stale-while-revalidate=600"},
    )


# ── Rankings override API ──────────────────────────────────────────
# These endpoints are the single authoritative path for custom-source
# configurations.  The frontend NEVER runs its own blended ranking
# engine when a user customizes source weights — instead it POSTs
# the override map here and receives either a full canonical
# contract or a compact delta payload re-computed by
# ``build_api_data_contract()`` / ``build_rankings_delta_payload()``
# with the overrides threaded into ``_compute_unified_rankings()``.


@app.get("/api/rankings/sources")
async def get_rankings_sources():
    """Return the canonical ranking-source registry.

    The frontend mirrors this registry statically in
    ``frontend/lib/dynasty-data.js::RANKING_SOURCES``; this endpoint
    exists so runtime tools, tests, and future builds can fetch the
    authoritative Python registry without reaching into module
    internals.  The shape matches the frontend entry exactly —
    ``assert_ranking_source_registry_parity()`` enforces that.
    """
    return JSONResponse(
        content={
            "sources": get_ranking_source_registry(),
            "contractVersion": API_DATA_CONTRACT_VERSION,
        }
    )


@app.post("/api/rankings/overrides")
async def post_rankings_overrides(request: Request):
    """Rebuild the canonical rankings with user-supplied source overrides.

    Accepts two equivalent body shapes:

      * ``{"enabled_sources": [...], "weights": {key: float, ...}}``
      * ``{"<source_key>": {"include": bool, "weight": float}, ...}``
        (legacy ``siteWeights`` shape from the frontend settings store)

    Response shape is controlled by the ``view`` query parameter:

      * ``view=full`` (default) — returns the full canonical
        contract (~4 MB uncompressed, identical shape to ``GET
        /api/data``).
      * ``view=delta`` (frontend default) — returns the compact
        delta payload (~70% smaller) containing only the
        override-sensitive fields per player.  The frontend merges
        the delta onto its cached base contract.
    """
    if not latest_data or not isinstance(latest_data, dict):
        return JSONResponse(
            status_code=503,
            content={
                "error": "No data available yet. First scrape may still be running.",
            },
        )

    try:
        body = await request.json()
    except Exception:
        body = None

    # Rankings follow scoring profile, not league key.  Validate the
    # key but only 503 when profiles actually differ — otherwise the
    # override pipeline can serve the same recomputed rankings to any
    # league that shares scoring.  The league-specific sleeper block
    # in the response is nulled below when the loaded contract was
    # built for a different league.
    try:
        league_cfg = _resolve_league_for_request(
            request,
            body=body if isinstance(body, dict) else None,
        )
    except LeagueResolutionError as err:
        return err.json_response()
    loaded_meta = (
        latest_contract_data.get("meta") or {} if isinstance(latest_contract_data, dict) else {}
    )
    loaded_league = str(loaded_meta.get("leagueKey") or "")
    sleeper_matches = bool(loaded_league) and loaded_league == league_cfg.key
    _scoring_err = _scoring_identity_error(latest_contract_data, league_cfg)
    if _scoring_err is not None:
        return _scoring_err

    overrides, warnings = normalize_source_overrides(body)
    tep_multiplier = normalize_tep_multiplier(body)
    tep_native_multiplier = normalize_tep_native_multiplier(body)

    view = (request.query_params.get("view") or "").strip().lower()
    delta_view = view in {"delta", "compact", "slim"}

    # ── Server-side valuation-mode composition ───────────────────────
    #
    # Custom source weights + the league-adjusted lens used to be
    # refused outright: the overlay's ranks are the ranks of
    # ``default_consensus x factor``, while the correct answer is the
    # rank of ``overridden_consensus x factor``, and the server had
    # never computed that board.  No client-side sequencing fixes it.
    # Computing it here is the fix.
    #
    # Note what this does to the endpoint's SCOPE.  Rankings follow the
    # scoring profile and are shared, which is why the profile check
    # above only 503s when profiles genuinely differ.  The valuation
    # overlay is league-scoped by necessity — ``lineupScarcity`` is
    # measured from THIS league's rosters — so the moment the caller
    # asks for it, this response stops being shareable across leagues
    # and has to 503 on a league mismatch exactly like
    # /api/valuation/league-adjusted.  Asking for the lens is what
    # narrows the scope; a plain override request is unaffected.
    valuation_mode = ""
    if isinstance(body, dict):
        raw_mode = body.get("valuation_mode") or body.get("valuationMode")
        valuation_mode = str(raw_mode or "").strip()
    want_league_adjusted = valuation_mode == "leagueAdjusted"

    # WITHDRAWN, on this path too (B9a).
    #
    # #822 rejected the league-aware methodology for promotion to
    # canonical and ruled it may no longer own a canonical field.  It
    # closed the engine gate (``_valuation_scoped_contract``) and moved
    # ``league_intel.overlay`` onto ``experimentalLeagueAdjusted*``
    # names — but this path was left composing factors into
    # ``rankDerivedValue`` itself, via
    # ``data_contract.apply_valuation_factors``.  The ±25% bound is
    # applied to the FACTOR and never to the PRODUCT, so the canonical
    # field left its own declared 1–9999 range: measured on the
    # 2026-08-14 board, the real factor set published 10,160 and the
    # factor cap published 12,471, both under the canonical field name.
    #
    # Ignored rather than refused, exactly as the engine gate does: a
    # stored ``leagueAdjusted`` on someone's phone must converge to the
    # canonical answer silently.  The composition helper is deleted
    # rather than left unreferenced, so there is no seam to re-thread by
    # accident; a future *validated* methodology re-opens one seam
    # deliberately, and renormalising onto the scale is part of what it
    # would have to earn.
    valuation_note: str | None = None
    if want_league_adjusted:
        valuation_note = "league_adjusted_withdrawn: not_canonical"

    # Snapshot the shared globals ONCE on the event loop.  The worker
    # thread below must not re-read them mid-build: a concurrent scrape
    # promotion swaps them, and a build that saw two generations would
    # emit a chimera payload.
    data_snapshot = latest_data
    source_snapshot = latest_data_source
    contract_version = latest_data_etag

    def _build_response_bytes() -> tuple[bytes, bytes]:
        """Full pipeline rebuild + meta stamping + JSON/gzip encode.

        Runs on a worker thread — this is multiple seconds of CPU on a
        real contract and previously blocked the whole event loop.  The
        raw bytes reproduce ``JSONResponse.render`` exactly
        (``ensure_ascii=False, allow_nan=False, separators=(",", ":")``)
        so memoizing bytes instead of returning JSONResponse is
        wire-invisible.
        """
        if delta_view:
            contract_payload = build_rankings_delta_payload(
                data_snapshot,
                data_source=source_snapshot,
                source_overrides=overrides if overrides else None,
                tep_multiplier=tep_multiplier,
                tep_native_multiplier=tep_native_multiplier,
            )
        else:
            contract_payload = build_api_data_contract(
                data_snapshot,
                data_source=source_snapshot,
                source_overrides=overrides if overrides else None,
                tep_multiplier=tep_multiplier,
                tep_native_multiplier=tep_native_multiplier,
            )

        if warnings:
            contract_payload.setdefault("warnings", []).extend(warnings)

        # Stamp meta fields so the frontend can assert compatibility:
        # ``leagueKey`` = resolved league; ``scoringProfile`` = which
        # rules the pipeline used; ``sleeperDataReady`` = whether the
        # ``sleeper`` block can be trusted for this league.  When the
        # loaded contract was built for a different league (same scoring
        # profile though — we'd have 503'd above if profiles differed),
        # null the sleeper block so callers don't render the wrong
        # league's teams.
        if isinstance(contract_payload, dict):
            meta = contract_payload.setdefault("meta", {})
            meta["leagueKey"] = league_cfg.key
            meta["scoringProfile"] = league_cfg.scoring_profile
            meta["sleeperDataReady"] = sleeper_matches
            if not sleeper_matches:
                contract_payload["sleeper"] = None
                meta["sleeperLoadedLeagueKey"] = loaded_league or None
            # Always stamp which lens produced this board. A client that
            # asked for leagueAdjusted and silently got market is the
            # failure mode this whole change exists to remove, so the
            # answer travels with the payload instead of being inferred.
            # Only one board is canonical now, so this is unconditional
            # rather than derived from whether factors were applied.
            meta["valuationMode"] = "market"
            if valuation_note:
                meta["valuationNote"] = valuation_note
                contract_payload.setdefault("warnings", []).append(valuation_note)

        raw = json.dumps(
            contract_payload,
            ensure_ascii=False,
            allow_nan=False,
            separators=(",", ":"),
        ).encode("utf-8")
        gz = gzip.compress(raw, compresslevel=5)
        return raw, gz

    # ── Response memo ────────────────────────────────────────────────
    # Cache key: everything that can change the response body, and
    # nothing that cannot.  The build consumes the NORMALIZED inputs —
    # the override map from ``normalize_source_overrides`` (always ``{}``
    # while ``_SOURCE_OVERRIDES_DISABLED``), the two tep knobs, and
    # ``valuation_mode`` (→ ``valuationNote``) — plus the ``warnings``
    # that get stamped onto the payload, so the key is derived from
    # exactly those.  Hashing the raw posted body instead (the previous
    # key) made every distinct source-toggle body pay a full pipeline
    # rebuild for a byte-identical response.  ``warnings`` rides the key
    # so this stays response-equivalence-complete even if the withdrawal
    # flag is ever flipped back (per-body unknown-key warnings would key
    # separately), and it distinguishes ``body=None`` (no withdrawal
    # warning — ``normalize_source_overrides`` early-returns before the
    # disabled check) from ``body={}`` (warning present).
    # ``valuation_mode`` requests used to be excluded because their
    # factors came from the gameplan module whose freshness this cache
    # cannot see; with the lens withdrawn there are no factors and
    # nothing unseeable left to exclude.  Entries are versioned on
    # ``latest_data_etag`` (the contract generation) and rebuilt in
    # place when stale; ``_prime_latest_payload`` clears the whole dict
    # on scrape promotion.
    cacheable = contract_version is not None
    cache_key = None
    if cacheable:
        normalized_inputs = json.dumps(
            {
                "overrides": overrides,
                "tep": tep_multiplier,
                "tepNative": tep_native_multiplier,
                "valuationMode": valuation_mode,
                "warnings": warnings,
            },
            sort_keys=True,
            separators=(",", ":"),
            default=str,
        )
        cache_key = (
            "overrides",
            hashlib.sha1(normalized_inputs.encode("utf-8")).hexdigest(),
            delta_view,
            league_cfg.key,
            sleeper_matches,
        )

    try:
        if cache_key is None:
            raw, gz = await run_in_threadpool(_build_response_bytes)
        else:
            entry = _OVERRIDES_RESPONSE_CACHE.get(cache_key)
            if entry is not None and entry[2] != contract_version:
                entry = None
            if entry is None:
                async with _overrides_encode_lock(cache_key):
                    entry = _OVERRIDES_RESPONSE_CACHE.get(cache_key)
                    if entry is not None and entry[2] != contract_version:
                        entry = None
                    if entry is None:
                        raw, gz = await run_in_threadpool(_build_response_bytes)
                        entry = (raw, gz, contract_version)
                        _evict_overrides_cache_if_oversized(cache_key)
                        _OVERRIDES_RESPONSE_CACHE[cache_key] = entry
            raw, gz, _ = entry
    except Exception as exc:
        log.exception("Failed to rebuild contract with overrides: %s", exc)
        return JSONResponse(
            status_code=500,
            content={
                "error": f"Failed to rebuild rankings with overrides: {exc}",
                "warnings": warnings,
            },
        )

    headers = {
        "Cache-Control": "no-store",
        "X-Payload-View": "rankings-overrides-delta" if delta_view else "rankings-overrides",
        "Vary": "Accept-Encoding",
    }
    accept_encoding = (request.headers.get("accept-encoding") or "").lower()
    if "gzip" in accept_encoding and gz:
        headers["Content-Encoding"] = "gzip"
        return Response(content=gz, media_type="application/json", headers=headers)
    return Response(content=raw, media_type="application/json", headers=headers)


# Cache of Sleeper league ``name`` fields.  Refreshed every
# ``_SLEEPER_NAME_TTL_SEC`` so a rename in Sleeper propagates to the
# UI without a deploy, but we don't hammer Sleeper on every
# /api/leagues request.
_SLEEPER_NAME_CACHE: dict[str, dict] = {}
_SLEEPER_NAME_TTL_SEC = 300


def _fetch_sleeper_league_name(sleeper_league_id: str) -> str | None:
    """Return the live ``name`` from ``/v1/league/<id>`` or None on
    any failure.  Cached per-league for 5 minutes.  Used by
    ``/api/leagues`` to label each league with its actual Sleeper
    name (e.g. "Risk It To Get The Brisket") instead of whatever
    operator-edited string lives in the registry's ``displayName``.
    """
    import time as _time

    sleeper_league_id = str(sleeper_league_id or "").strip()
    if not sleeper_league_id:
        return None
    now = _time.time()
    cached = _SLEEPER_NAME_CACHE.get(sleeper_league_id)
    if cached and (now - float(cached.get("fetched_at") or 0)) < _SLEEPER_NAME_TTL_SEC:
        return cached.get("name")

    try:
        url = f"https://api.sleeper.app/v1/league/{sleeper_league_id}"
        req = urllib.request.Request(url, headers={"User-Agent": "brisket-league-name/1.0"})
        with urllib.request.urlopen(req, timeout=5.0) as resp:
            body = resp.read()
        parsed = json.loads(body)
    except Exception:  # noqa: BLE001 — transient failures cache None so we don't refetch every call
        _SLEEPER_NAME_CACHE[sleeper_league_id] = {"name": None, "fetched_at": now}
        return None

    name = str((parsed or {}).get("name") or "").strip() or None
    _SLEEPER_NAME_CACHE[sleeper_league_id] = {"name": name, "fetched_at": now}
    return name


# Cache of (league_id, user_id) → {ownerId, teamName} so the user's
# default team in a given league can be auto-selected by
# ``/api/leagues`` without a second round-trip.  Same 5-min TTL as
# the league-name cache; a rename of the user's team in Sleeper
# propagates within that window.
_SLEEPER_USER_TEAM_CACHE: dict[tuple[str, str], dict] = {}


def _fetch_sleeper_user_team(
    sleeper_league_id: str,
    sleeper_user_id: str,
) -> dict | None:
    """Return ``{"ownerId", "teamName"}`` for the user in the given
    league, or ``None`` if the user isn't in the league or Sleeper
    is unreachable.

    Falls back gracefully: a single failure caches ``None`` for
    ``_SLEEPER_NAME_TTL_SEC`` so a flaky Sleeper doesn't force a
    refetch on every authed request.

    Used by ``/api/leagues`` so a user who isn't enumerated in a
    league's registry ``defaultTeamMap`` still gets their team
    auto-selected on that league (resolved via ownerId match from
    the session's ``sleeper_user_id``).
    """
    import time as _time

    sleeper_league_id = str(sleeper_league_id or "").strip()
    sleeper_user_id = str(sleeper_user_id or "").strip()
    if not sleeper_league_id or not sleeper_user_id:
        return None
    now = _time.time()
    cache_key = (sleeper_league_id, sleeper_user_id)
    cached = _SLEEPER_USER_TEAM_CACHE.get(cache_key)
    if cached and (now - float(cached.get("fetched_at") or 0)) < _SLEEPER_NAME_TTL_SEC:
        return cached.get("value")

    try:
        users_url = f"https://api.sleeper.app/v1/league/{sleeper_league_id}/users"
        req = urllib.request.Request(
            users_url,
            headers={"User-Agent": "brisket-user-team/1.0"},
        )
        with urllib.request.urlopen(req, timeout=5.0) as resp:
            users = json.loads(resp.read())
    except Exception:  # noqa: BLE001 — cache None on any transient failure
        _SLEEPER_USER_TEAM_CACHE[cache_key] = {"value": None, "fetched_at": now}
        return None

    # Find the authed user in the league.
    team_name = ""
    for u in users or []:
        if str(u.get("user_id") or "") == sleeper_user_id:
            team_name = (u.get("metadata") or {}).get("team_name") or u.get("display_name") or ""
            break
    if not team_name:
        _SLEEPER_USER_TEAM_CACHE[cache_key] = {"value": None, "fetched_at": now}
        return None

    value = {"ownerId": sleeper_user_id, "teamName": str(team_name).strip()}
    _SLEEPER_USER_TEAM_CACHE[cache_key] = {"value": value, "fetched_at": now}
    return value


@app.get("/api/leagues")
async def get_leagues(request: Request):
    """List every configured league.

    Public endpoint — the response contains no secrets (no Sleeper
    league IDs, no auth tokens).  The ``key`` is the stable identifier
    callers thread through the rest of the API when we eventually
    add ``?leagueId=`` parameters to league-scoped endpoints.

    ``displayName`` is pulled LIVE from Sleeper (``/v1/league/<id>``
    ``.name``) and cached for 5 minutes.  This way a league rename
    in Sleeper propagates to the switcher without a registry edit
    or redeploy.  Falls back to the registry's configured
    ``displayName`` when Sleeper is unreachable.

    Views:
      * Anonymous:     active leagues only.
      * Authenticated: active leagues + ``userDefaultKey`` (which
                       league the UI should land this user on by
                       default) + per-league ``userDefaultTeam``
                       entries from each league's ``defaultTeamMap``
                       so the frontend can auto-select the right
                       team in each league without a second round-
                       trip to user_kv.
    """
    session = _get_auth_session(request)
    active_cfgs = _league_registry.active_leagues()
    leagues = [cfg.public_dict() for cfg in active_cfgs]

    # Overlay the live Sleeper name for each league.  Run the
    # fetches in a threadpool so we don't block the event loop on
    # the Sleeper round-trip.  Cached 5 min so steady-state traffic
    # doesn't hammer Sleeper.
    def _fetch_names(cfgs):
        return [_fetch_sleeper_league_name(c.sleeper_league_id) for c in cfgs]

    sleeper_names = await run_in_threadpool(_fetch_names, active_cfgs)
    for i, live_name in enumerate(sleeper_names):
        if live_name:
            leagues[i]["displayName"] = live_name

    if session:
        username = (session.get("username") or "").strip().lower()
        sleeper_user_id = str(session.get("sleeper_user_id") or "").strip()

        # Stamp each league's entry with this user's default team
        # when the registry knows about one.  Only this authed user
        # sees their own default — we don't expose other usernames'
        # mappings even though the registry file holds them.
        #
        # Fallback: when the registry has no default_team_map entry
        # for this user on this league (common for newly-added
        # leagues), auto-resolve the user's team from Sleeper via
        # their ``sleeper_user_id`` → league users lookup.  Without
        # this fallback the team picker stays on "Pick your team"
        # forever for any league the registry hasn't been edited
        # for, forcing every dashboard block to sit at "Pick a team
        # to see..." until the user manually selects one.
        def _resolve_team(cfg):
            mapped = cfg.default_team_map.get(username) if username else None
            if mapped:
                return {
                    "ownerId": mapped.get("ownerId", "") or sleeper_user_id,
                    "teamName": mapped.get("teamName", ""),
                }
            if sleeper_user_id:
                return _fetch_sleeper_user_team(
                    cfg.sleeper_league_id,
                    sleeper_user_id,
                )
            return None

        resolved = await run_in_threadpool(lambda: [_resolve_team(c) for c in active_cfgs])
        for i, team in enumerate(resolved):
            if team:
                leagues[i]["userDefaultTeam"] = team

    body: dict[str, Any] = {
        "leagues": leagues,
        "defaultKey": _league_registry.default_league_key(),
    }
    if session:
        user_default = _league_registry.get_user_default_league(session.get("username") or "")
        body["userDefaultKey"] = user_default.key if user_default else None
    return JSONResponse(content=body, headers={"Cache-Control": "no-store"})


# ── League Comparison ─────────────────────────────────────────────────
# Compares one custom-scoring Sleeper league against a "standard"
# baseline league across multiple historical NFL seasons.  Read-only
# analysis tool — does NOT route through the league registry; the two
# league IDs come from config/league_comparison.json server-side, the
# UI never sends raw Sleeper IDs in requests.  See
# src/league_comparison/service.py for the full pipeline.
@app.get("/api/league-comparison")
async def get_league_comparison(request: Request):
    """Build the positional-balance comparison between the two
    configured leagues across the configured seasons.

    Query params:
      * ``refresh`` — if "1"/"true", bypass the 7-day disk cache and
        recompute (also bypasses the 1h scoring-settings cache so a
        commissioner edit propagates immediately).

    Errors:
      * 503 — Sleeper unreachable or scoring_settings missing for one
        of the configured leagues.
      * 500 — internal failure during computation.
    """
    refresh_raw = (request.query_params.get("refresh") or "").strip().lower()
    refresh = refresh_raw in ("1", "true", "yes", "on")
    try:
        from src.league_comparison import service as _lc_service

        payload = await run_in_threadpool(
            _lc_service.build_comparison,
            refresh=refresh,
        )
    except FileNotFoundError as exc:
        return JSONResponse(
            status_code=500,
            content={
                "error": "league_comparison_misconfigured",
                "detail": str(exc),
            },
        )
    except (urllib.error.URLError, urllib.error.HTTPError) as exc:
        return JSONResponse(
            status_code=503,
            content={
                "error": "sleeper_unreachable",
                "detail": str(exc),
            },
        )
    except ValueError as exc:
        return JSONResponse(
            status_code=503,
            content={
                "error": "league_data_unavailable",
                "detail": str(exc),
            },
        )
    except Exception as exc:  # noqa: BLE001
        logging.getLogger(__name__).exception("league_comparison_failed")
        return JSONResponse(
            status_code=500,
            content={
                "error": "league_comparison_failed",
                "detail": str(exc),
            },
        )
    return JSONResponse(
        content=payload,
        headers={"Cache-Control": "private, max-age=300, stale-while-revalidate=3600"},
    )


@app.get("/api/status")
async def get_status():
    """Return scraper status info."""
    status_payload = _scrape_status_payload()
    # Prefer full scrape payload for source-health truth (dlfImport/sourceRunSummary).
    # Contract payload is a compatibility fallback when full payload is unavailable.
    # ``served_source_coverage`` is the per-source contribution to the
    # board we are actually serving; the snapshot is the ANSWER built
    # from it plus the registry population (F-7).  Owner and input, not
    # two competing surfaces.
    source_health = _build_source_health_snapshot(
        latest_data or latest_contract_data, coverage=served_source_coverage
    )
    full_bytes = len(latest_data_bytes) if latest_data_bytes else 0
    runtime_bytes = len(latest_runtime_data_bytes) if latest_runtime_data_bytes else 0
    startup_bytes = len(latest_startup_data_bytes) if latest_startup_data_bytes else 0
    full_gzip_bytes = len(latest_data_gzip_bytes) if latest_data_gzip_bytes else 0
    runtime_gzip_bytes = (
        len(latest_runtime_data_gzip_bytes) if latest_runtime_data_gzip_bytes else 0
    )
    startup_gzip_bytes = (
        len(latest_startup_data_gzip_bytes) if latest_startup_data_gzip_bytes else 0
    )
    return JSONResponse(
        content={
            **status_payload,
            "contract": {
                "version": API_DATA_CONTRACT_VERSION,
                "health": contract_health,
                "value_authority": (latest_contract_data or {}).get("valueAuthority"),
            },
            "data_runtime": {
                # The name claims a REFRESH time, so it carries one: when
                # the board was produced.  It used to carry ``loadedAt`` —
                # when this process started holding it — which is a different
                # fact and is published beside it under its own name
                # (audit F-19).
                "last_data_refresh_at": latest_data_source.get("producedAt") or None,
                "last_payload_loaded_at": latest_data_source.get("loadedAt") or None,
                "active_data_source": latest_data_source,
                "payload_bytes_full": full_bytes,
                "payload_bytes_runtime": runtime_bytes,
                "payload_bytes_startup": startup_bytes,
                "payload_gzip_bytes_full": full_gzip_bytes,
                "payload_gzip_bytes_runtime": runtime_gzip_bytes,
                "payload_gzip_bytes_startup": startup_gzip_bytes,
                "runtime_payload_savings_bytes": max(0, full_bytes - runtime_bytes),
                "runtime_payload_savings_gzip_bytes": max(0, full_gzip_bytes - runtime_gzip_bytes),
                "startup_payload_savings_bytes": max(0, full_bytes - startup_bytes),
                "startup_payload_savings_gzip_bytes": max(0, full_gzip_bytes - startup_gzip_bytes),
            },
            "source_health": source_health,
            "backup_health": _backup_freshness(),
            "uptime": uptime_status,
            "has_data": latest_contract_data is not None,
            "player_count": int((latest_contract_data or {}).get("playerCount") or 0),
            "data_date": (latest_contract_data or {}).get("date"),
            # Real per-source coverage of the served board (see the
            # ``served_source_coverage`` global).  Monitoring asserts on
            # this, and it is now also the INPUT ``source_health`` above
            # counts from — so the two agree by construction rather than
            # being two answers to one question.  Until 2026-08-18
            # ``source_health`` took its population from the legacy
            # ``sites`` list (2 anchor rows, not the 3 this comment
            # claimed) and so could not detect a degraded board; F-7.
            "served_source_coverage": served_source_coverage,
            # R-4: Scrape success rate tracking
            "scrape_success_rate_24h": _scrape_success_rate_24h(),
            "last_n_scrapes": scrape_history[-20:],
            "leagues": _league_status_snapshot(),
            # 2026-04 upgrade observability — feature-flag state +
            # unified-mapper coverage.  All flags default off so this
            # is additive/informational; enabling a flag at runtime is
            # the operator's decision.
            "featureFlags": _feature_flag_snapshot_safe(),
            # The rank-history log is the only record of what the board
            # said on a past date, and its append is best-effort inside
            # a try at the contract-rebuild site.  A stall is therefore
            # silent; surfacing coverage here is what makes it visible
            # before a study needs the history and finds it absent.
            "rankHistoryCoverage": _rank_history_coverage_safe(),
            # The same argument one directory over, and the same defect
            # the line above exists to prevent.  `store.history_coverage`
            # was written so a halted retention timer is "visible before
            # a study needs the data rather than after it produces a
            # wrong answer" — and was then wired to nothing.  That is
            # what let the 2026-08-05 deploy install the producer, skip
            # the pusher, and leave the only symptom in a deploy log
            # nobody re-reads.
            "playerctxHistoryCoverage": _playerctx_history_coverage_safe(),
            "idMappingCoverage": _id_mapping_coverage_safe(),
            "nflDataProvider": _nfl_data_provider_status_safe(),
            "normalizationHealth": _normalization_health_safe(),
        },
        # Observability payload — the disk-probe helpers behind it are
        # memoized 30s server-side; let pollers (settings page, tools)
        # reuse it browser-side for the same window.
        headers={"Cache-Control": "private, max-age=15, stale-while-revalidate=30"},
    )


def _feature_flag_snapshot_safe() -> dict:
    """Return the feature-flag snapshot, tolerant of import errors
    so a malformed upgrade doesn't 500 /api/status.

    Reports ``gateStatus`` alongside ``enabled`` because "is it on?" is
    misleading on its own: 7 of 13 flags gate a module that nothing
    reachable from this file imports, so their value cannot change a
    response either way.  A reader who saw only ``enabled: true`` on
    ``espn_injury_feed`` would reasonably conclude the injury feed was
    running.  ``enabled`` keeps its old meaning and position so existing
    consumers are unaffected.
    """
    try:
        from src.api import feature_flags as _ff

        return _ff.effective_flags()
    except Exception as exc:  # noqa: BLE001
        log.warning("feature_flag snapshot failed: %s", exc)
        return {}


def _rank_history_coverage_safe() -> dict:
    """Rank-history log coverage, tolerant of read errors.

    Never raises: this is a diagnostic on a status endpoint, and a
    diagnostic that can take down the thing it reports on is worse than
    no diagnostic.
    """
    try:
        return _rank_history.coverage()
    except Exception as exc:  # noqa: BLE001
        log.warning("rank_history coverage failed: %s", exc)
        return {}


def _playerctx_pending_push(history_dir) -> dict:
    """Dated snapshots written locally but never committed.

    THE PRODUCER AND THE PUSHER FAIL INDEPENDENTLY, and only one of the
    two is visible in ``history_coverage``:

    * the producer stalls  → the directory goes stale → ``staleDays`` grows.
    * the pusher stalls    → the directory looks perfect and ``main`` gets
      nothing.

    The second is the one that actually happened on 2026-08-05, and
    coverage alone cannot see it — which is precisely why it went
    unnoticed until someone read a deploy log.  ``data/`` is gitignored
    repo-wide and retained snapshots reach the tree only by an explicit
    ``git add -f``, so "on disk but not tracked" is exactly "written but
    not pushed".

    Degrades to ``{}`` rather than raising or guessing: outside a git
    checkout there is no answer, and a fabricated zero would read as
    "nothing pending".
    """
    import subprocess

    try:
        proc = subprocess.run(
            ["git", "ls-files", "--", "data/playerctx/history"],
            cwd=str(BASE_DIR),
            capture_output=True,
            text=True,
            timeout=10,
        )
        if proc.returncode != 0:
            return {}
        tracked = {
            line.rsplit("/", 1)[-1].strip() for line in proc.stdout.splitlines() if line.strip()
        }
        pending = sorted(
            p.name for p in history_dir.glob("snapshot_*.json") if p.name not in tracked
        )
        out: dict = {"pendingPush": len(pending)}
        if pending:
            # Filename carries the date and sorts chronologically, which
            # is why store.history_path writes snapshot_YYYY-MM-DD.json.
            out["oldestPendingDate"] = pending[0].replace("snapshot_", "", 1).replace(".json", "")
        return out
    except Exception as exc:  # noqa: BLE001
        log.warning("playerctx pending-push probe failed: %s", exc)
        return {}


def _playerctx_history_coverage_safe() -> dict:
    """Retained playerctx snapshot coverage, tolerant of read errors.

    Same contract as ``_rank_history_coverage_safe`` — never raises —
    plus the pending-push augmentation, which lives here rather than in
    ``store`` so the store stays a pure filesystem module with no git
    dependency.
    """
    try:
        from src.playerctx import store as _pcx_store

        coverage = _pcx_store.history_coverage()
        if coverage.get("exists"):
            coverage.update(_playerctx_pending_push(_pcx_store.HISTORY_DIR))
        return coverage
    except Exception as exc:  # noqa: BLE001
        log.warning("playerctx history coverage failed: %s", exc)
        return {}


def _id_mapping_coverage_safe() -> dict:
    try:
        from src.identity import unified_mapper as _um

        return _um.mapping_coverage_snapshot()
    except Exception as exc:  # noqa: BLE001
        log.warning("id mapping coverage snapshot failed: %s", exc)
        return {}


def _nfl_data_provider_status_safe() -> dict:
    try:
        from src.nfl_data import ingest as _ing

        return _ing.provider_status()
    except Exception as exc:  # noqa: BLE001
        log.warning("nfl_data provider status failed: %s", exc)
        return {}


def _normalization_health_safe() -> dict:
    """Return the contract validation summary for /api/status.
    Runs on every status hit (not cached) — it's O(N) over the
    playersArray, ~5ms for ~1100 rows."""
    try:
        from src.canonical import normalization_validator as _nv

        return _nv.validate_contract(latest_contract_data or {})
    except Exception as exc:  # noqa: BLE001
        log.warning("normalization validator failed: %s", exc)
        return {}


def _league_status_snapshot() -> list[dict]:
    """Per-league data-health snapshot for /api/status.

    Reports for each ACTIVE league:
      * ``key`` / ``displayName``                — identity
      * ``source``                                — "primary-scrape" | "overlay" | "none"
      * ``teamCount``                             — rosters resolved
      * ``tradeCount``                            — trades loaded
      * ``overlayFetchedAt`` / ``overlayAgeSec``  — staleness (None when primary)
      * ``sleeperLeagueId``                       — raw id for correlation with logs

    This is the diagnostic surface for answering "does League B have
    fresh data, or am I serving the stale overlay again?".  When the
    source is ``none`` the UI surfaces the data-not-ready state.
    """
    import time as _time

    snapshot: list[dict[str, Any]] = []
    loaded = latest_contract_data or {}
    loaded_meta = loaded.get("meta") or {}
    loaded_key = loaded_meta.get("leagueKey")
    loaded_sleeper = loaded.get("sleeper") or {}
    for cfg in _league_registry.active_leagues():
        entry: dict[str, Any] = {
            "key": cfg.key,
            "displayName": cfg.display_name,
            "sleeperLeagueId": cfg.sleeper_league_id,
            "idpEnabled": cfg.idp_enabled,
            "scoringProfile": cfg.scoring_profile,
        }
        if cfg.key == loaded_key and isinstance(loaded_sleeper, dict):
            entry["source"] = "primary-scrape"
            entry["teamCount"] = len(loaded_sleeper.get("teams") or [])
            entry["tradeCount"] = len(loaded_sleeper.get("trades") or [])
            entry["overlayFetchedAt"] = None
            entry["overlayAgeSec"] = None
        else:
            cached = _sleeper_overlay._CACHE.get(cfg.sleeper_league_id) or {}
            payload = cached.get("payload") or {}
            fetched_at = float(cached.get("_cached_at") or 0)
            if payload.get("teams"):
                entry["source"] = "overlay"
                entry["teamCount"] = len(payload.get("teams") or [])
                entry["tradeCount"] = len(payload.get("trades") or [])
                entry["overlayFetchedAt"] = payload.get("overlayFetchedAt")
                entry["overlayAgeSec"] = (
                    round(_time.time() - fetched_at, 1) if fetched_at > 0 else None
                )
            else:
                entry["source"] = "none"
                entry["teamCount"] = 0
                entry["tradeCount"] = 0
                entry["overlayFetchedAt"] = None
                entry["overlayAgeSec"] = None
        snapshot.append(entry)
    return snapshot


@app.api_route("/api/health", methods=["GET", "HEAD"])
async def get_health():
    """Basic health endpoint for reverse proxy / uptime probes."""
    status_payload = _scrape_status_payload()

    # R-1: Data freshness check — flag stale if the BOARD is older than
    # SCRAPE_INTERVAL_HOURS * 3.  Measured from the payload's own production
    # time, never from when this process loaded it (audit F-19): the latter
    # reset to zero on every restart, so a stale board read fresh after every
    # deploy.  Unknown production time leaves both fields honest — age None,
    # not-stale — rather than asserting freshness we cannot support.
    raw_age = _board_age_hours()
    data_age_hours = None if raw_age is None else round(raw_age, 1)
    data_stale = data_age_hours is not None and data_age_hours > SCRAPE_INTERVAL_HOURS * 3

    # Session-cookie age surface.  Distinguishes AUTO-refreshing
    # sessions (scraper re-logs-in via stored credentials when the
    # cached cookies fail) from MANUAL-only sessions (operator
    # pastes browser cookies because the site blocks automated
    # login — currently just IDP Show, whose Substack paywall has
    # a captcha on password auth).  The frontend banner only alarms
    # on manual-only sessions since auto-refresh sessions fix
    # themselves on the next scrape.
    _session_ages: dict[str, dict] = {}

    _session_configs = {
        # Scraper POSTs DLF_USERNAME/PASSWORD to wp-login on failure,
        # so this file auto-refreshes.  Tracked for visibility only.
        "dlf_session.json": {"lifetimeDays": 14, "autoRefresh": True},
        # Scraper POSTs DRAFTSHARKS_EMAIL/PASSWORD on failure.
        "draftsharks_session.json": {"lifetimeDays": 30, "autoRefresh": True},
        # Substack captcha-gates password login — the ONLY way to
        # refresh these cookies is a manual browser dump.  Banner
        # alarms on this file specifically.
        "idpshow_session.json": {"lifetimeDays": 90, "autoRefresh": False},
    }
    for fname, cfg in _session_configs.items():
        lifetime_days = cfg["lifetimeDays"]
        auto_refresh = cfg["autoRefresh"]
        fpath = BASE_DIR / fname
        try:
            if not fpath.exists():
                _session_ages[fname] = {"present": False, "autoRefresh": auto_refresh}
                continue
            mtime_ts = fpath.stat().st_mtime
            age_days = round((datetime.now(timezone.utc).timestamp() - mtime_ts) / 86400, 1)
            days_remaining = max(0.0, round(lifetime_days - age_days, 1))
            # Only MANUAL sessions get the warnSoon flag — auto-refresh
            # sessions silently rotate when cached cookies expire, so
            # the banner shouldn't nag about them.
            warn_soon = not auto_refresh and days_remaining <= 14 and age_days > 0
            expired = (not auto_refresh) and days_remaining <= 0
            _session_ages[fname] = {
                "present": True,
                "autoRefresh": auto_refresh,
                "ageDays": age_days,
                "lifetimeDays": lifetime_days,
                "daysRemaining": days_remaining,
                "warnSoon": warn_soon,
                "expired": expired,
            }
        except Exception:
            _session_ages[fname] = {"present": False, "autoRefresh": auto_refresh}

    is_ok = (
        status_payload.get("last_error") in (None, "")
        and not status_payload.get("stalled")
        and not data_stale
        and bool(contract_health.get("ok", False))
    )
    status = "ok" if is_ok else "degraded"

    # Deeper health: startup checks + circuit breakers + session
    # store size.  Wrapped so a dependency import error can't bring
    # down the health endpoint itself.
    def _circuits_safe():
        try:
            from src.utils import circuit_breaker as _cb

            return _cb.snapshot_all()
        except Exception as exc:  # noqa: BLE001
            log.warning("health: circuit_breaker snapshot failed: %s", exc)
            return []

    def _sessions_safe():
        try:
            from src.api import session_store as _ss

            return {"persistedCount": _ss.count_active()}
        except Exception as exc:  # noqa: BLE001
            log.warning("health: session_store count failed: %s", exc)
            return {"persistedCount": None}

    circuits = _circuits_safe()
    # Any breaker in OPEN state flips overall status to degraded.
    any_breaker_open = any(c.get("state") == "open" for c in circuits)

    # Backup-freshness watchdog.  Backups run nightly (riskit-backup
    # .timer @ 02:00 UTC); anything older than 36h (or missing) means
    # the backup pipeline is broken — the failure mode that went
    # unnoticed for ~2 weeks.  send_alert is globally cooldown-rate-
    # limited and no-ops when SMTP isn't configured.  Backup staleness
    # does NOT flip the uptime status (the site is still serving).
    backup_health = _backup_freshness()
    _bage = backup_health.get("newestBackupAgeHours")
    if _bage is None or _bage > 36:
        send_alert(
            "Risk It: SQLite backups stale",
            f"Newest backup age: {_bage}h (dir={backup_health.get('backupDirUsed')}, "
            f"path={backup_health.get('newestBackupPath')}, dbCount="
            f"{backup_health.get('dbCount')}). Expected a fresh backup within 24h "
            f"from riskit-backup.timer (02:00 UTC). Check "
            f"deploy/backup_user_kv.sh and /var/log/riskit-backup.log.",
        )

    return JSONResponse(
        status_code=200 if is_ok else 503,
        content={
            "status": status,
            "service": "dynasty-server",
            "time_utc": datetime.now(timezone.utc).isoformat(),
            "has_data": latest_contract_data is not None,
            "data_stale": data_stale,
            "data_age_hours": data_age_hours,
            "last_scrape": status_payload.get("last_scrape"),
            "scrape_running": status_payload.get("is_running"),
            "scrape_stalled": status_payload.get("stalled"),
            # A worker that died mid-run. Distinct from ``scrape_stalled``
            # (running but not progressing) and from ``data_stale`` (old
            # data, healthy server): the last run ended without cleanup
            # and nothing will retry it on its own.
            "scrape_interrupted": bool(status_payload.get("interrupted")),
            "current_step": status_payload.get("current_step"),
            "current_source": status_payload.get("current_source"),
            "contract_version": API_DATA_CONTRACT_VERSION,
            "contract_ok": contract_health.get("ok"),
            "uptime_watchdog": {
                "enabled": uptime_status.get("enabled"),
                "target_url": uptime_status.get("target_url"),
            },
            "session_cookies": _session_ages,
            "sessions": _sessions_safe(),
            "circuitBreakers": circuits,
            "anyBreakerOpen": any_breaker_open,
            "startupChecks": _startup_checks_summary,
            "memberInMemorySessions": len(auth_sessions)
            if isinstance(auth_sessions, dict)
            else None,
            "backup_health": backup_health,
        },
    )


@app.get("/api/uptime")
async def get_uptime_status():
    """Detailed uptime watchdog state."""
    return JSONResponse(content=uptime_status)


@app.get("/api/metrics")
async def get_metrics():
    """R-9: Lightweight metrics endpoint for dashboards and monitoring."""
    now = datetime.now(timezone.utc)
    # Calculate data age
    # Board age, not process age — audit F-19.
    _age_h = _board_age_hours()
    data_age_seconds = None if _age_h is None else round(_age_h * 3600.0, 0)

    # Calculate uptime
    uptime_seconds = None
    if _metrics.get("server_start_time"):
        try:
            start_dt = datetime.fromisoformat(_metrics["server_start_time"])
            uptime_seconds = round((now - start_dt).total_seconds(), 0)
        except (ValueError, TypeError):
            pass

    disk_ok, free_mb = _check_disk_space()

    return JSONResponse(
        content={
            "server_start_time": _metrics.get("server_start_time"),
            "uptime_seconds": uptime_seconds,
            "request_count": _metrics.get("request_count", 0),
            "scrape_total": _metrics.get("scrape_total", 0),
            "scrape_failures": _metrics.get("scrape_failures", 0),
            "scrape_duration_seconds_last": _metrics.get("scrape_duration_seconds_last", 0),
            "data_age_seconds": data_age_seconds,
            "data_stale": (data_age_seconds or 0) > SCRAPE_INTERVAL_HOURS * 3 * 3600,
            "has_data": latest_contract_data is not None,
            "player_count": int((latest_contract_data or {}).get("playerCount") or 0),
            "disk_free_mb": free_mb,
            "disk_ok": disk_ok,
            "scrape_running": scrape_status.get("running", False),
        }
    )


# ── NEWS ───────────────────────────────────────────────────────────────
# Normalized news feed aggregating Sleeper trending + ESPN RSS.  The
# service layer owns caching, per-provider fault isolation, and the
# normalized NewsItem shape (see ``src/news/base.py``).  The route is
# a thin adapter: parse query args, delegate, and surface per-provider
# diagnostics so the frontend can distinguish "empty feed" from "all
# providers degraded".
@app.get("/api/news")
async def get_news(request: Request):
    limit_raw = request.query_params.get("limit")
    try:
        limit = int(limit_raw) if limit_raw else 50
    except ValueError:
        limit = 50
    limit = max(1, min(limit, 100))

    # Optional team-name filter.  Repeatable ?team=... or comma
    # separated.  Items that don't mention at least one of those
    # players are dropped from the response (client-side filtering
    # stays available for scope="roster"/"league" already).
    team_params = (
        request.query_params.getlist("team") if hasattr(request.query_params, "getlist") else []
    )
    team_names: list[str] = []
    for raw in team_params:
        for part in str(raw).split(","):
            if part.strip():
                team_names.append(part.strip())

    svc = _get_news_service()
    try:
        # One callable so the two full-contract scans run on the worker
        # thread too — as call arguments they'd execute on the event
        # loop before the threadpool hop.
        def _aggregate():
            return svc.aggregate(
                player_names=_live_player_names(),
                team_names=team_names or None,
                player_meta=_live_player_meta(),
            )

        aggregated = await run_in_threadpool(_aggregate)
    except Exception as exc:
        log.warning("/api/news aggregation failed: %s", exc)
        # Signal "temporarily unavailable" — the frontend surfaces an
        # explicit "news unavailable" state on 503 (there is no
        # client-side fixture fallback).
        return JSONResponse(
            status_code=503,
            content={
                "items": [],
                "providersUsed": [],
                "providerRuns": [],
                "error": f"{type(exc).__name__}",
                "generatedAt": datetime.now(timezone.utc).isoformat(),
            },
            headers={"Cache-Control": "no-store"},
        )

    payload = aggregated.to_dict()
    payload["source"] = "backend"
    payload["limit"] = limit
    if len(payload.get("items", [])) > limit:
        payload["items"] = payload["items"][:limit]
        payload["count"] = len(payload["items"])

    # Distinguish "providers worked, nothing trending" (legit 200
    # with empty items) from "every provider errored out" (503 —
    # the frontend renders its explicit "news unavailable" state).
    provider_runs = aggregated.provider_runs or []
    all_failed = bool(provider_runs) and not any(r.ok for r in provider_runs)
    if all_failed:
        return JSONResponse(
            status_code=503,
            content={
                **payload,
                "source": "backend",
                "error": "all_providers_failed",
            },
            headers={"Cache-Control": "no-store"},
        )
    # Public endpoint (see _PUBLIC_API_EXACT): a short shared-cache
    # TTL keeps repeat public hits off the aggregator, consistent
    # with the service's own 180s cache and the other public
    # endpoints' Cache-Control conventions.
    return JSONResponse(
        content=payload,
        headers={"Cache-Control": "public, max-age=60, stale-while-revalidate=180"},
    )


@app.get("/api/scaffold/status")
async def get_scaffold_status():
    """Return latest scaffold snapshot metadata for raw/canonical/league/report outputs."""
    raw_file = _latest_file(DATA_DIR / "raw_sources", "raw_source_snapshot_*.json")
    ingest_validation_file = _latest_file(DATA_DIR / "validation", "ingest_validation_*.json")
    league_file = _latest_file(DATA_DIR / "league", "league_snapshot_*.json")
    identity_file = _latest_file(DATA_DIR / "identity", "identity_resolution_*.json")
    if identity_file is None:
        identity_file = _latest_file(DATA_DIR / "identity", "identity_report_*.json")
    report_file = _latest_file(DATA_DIR / "reports", "ops_report_*.md")

    raw = _load_json_file(raw_file)
    ingest_validation = _load_json_file(ingest_validation_file)
    league = _load_json_file(league_file)
    identity = _load_json_file(identity_file)

    def _meta(path: Path | None) -> dict | None:
        if path is None or not path.exists():
            return None
        stat = path.stat()
        return {
            "name": path.name,
            "path": str(path),
            "mtime": datetime.fromtimestamp(stat.st_mtime, timezone.utc).isoformat(),
            "size_bytes": stat.st_size,
        }

    return JSONResponse(
        content={
            "raw_sources": {
                "file": _meta(raw_file),
                "source_count": len(raw.get("snapshots", [])) if raw else 0,
                "record_count": (
                    sum(len(s.get("records", [])) for s in raw.get("snapshots", [])) if raw else 0
                ),
            },
            "ingest_validation": {
                "file": _meta(ingest_validation_file),
                "status": ingest_validation.get("status", "missing")
                if ingest_validation
                else "missing",
                "missing_snapshot_field_count": ingest_validation.get(
                    "missing_snapshot_field_count", 0
                )
                if ingest_validation
                else 0,
                "missing_asset_field_count": ingest_validation.get("missing_asset_field_count", 0)
                if ingest_validation
                else 0,
            },
            "league": {
                "file": _meta(league_file),
                "asset_count": league.get("asset_count", 0) if league else 0,
            },
            "identity": {
                "file": _meta(identity_file),
                "master_player_count": identity.get("master_player_count", 0) if identity else 0,
                "single_source_count": identity.get("single_source_count", 0) if identity else 0,
                "conflict_count": identity.get("conflict_count", 0) if identity else 0,
            },
            "report": {
                "file": _meta(report_file),
            },
        }
    )


@app.get("/api/scaffold/raw")
async def get_scaffold_raw():
    file_path = _latest_file(DATA_DIR / "raw_sources", "raw_source_snapshot_*.json")
    payload = _load_json_file(file_path)
    if payload is None:
        return JSONResponse(status_code=404, content={"error": "No raw scaffold snapshot found"})
    return JSONResponse(content=payload)


@app.get("/api/scaffold/league")
async def get_scaffold_league():
    file_path = _latest_file(DATA_DIR / "league", "league_snapshot_*.json")
    payload = _load_json_file(file_path)
    if payload is None:
        return JSONResponse(status_code=404, content={"error": "No league scaffold snapshot found"})
    return JSONResponse(content=payload)


@app.get("/api/scaffold/identity")
async def get_scaffold_identity():
    """The newest identity-resolution report, LABELLED with its own age.

    C1-RET-07.  This endpoint serves ``_latest_file``, and identity
    collection has been halted since 2026-04-20 — so the response was a
    four-month-old report presented with nothing to distinguish it from
    one produced this morning.  A stale artifact served without its age
    is indistinguishable from a current one, which is the "stale =
    current" substitution the retention tranche exists to prevent.

    The repair is the honest label, not a fabricated refresh: the
    collector's return is separate REPAIR work and no response may
    imply data that was never collected.  ``evidenceFreshness`` is
    additive — every existing key of the report is untouched.
    """
    file_path = _latest_file(DATA_DIR / "identity", "identity_resolution_*.json")
    if file_path is None:
        file_path = _latest_file(DATA_DIR / "identity", "identity_report_*.json")
    payload = _load_json_file(file_path)
    if payload is None:
        return JSONResponse(status_code=404, content={"error": "No identity report found"})

    freshness: dict[str, Any] = {
        "sourceFile": getattr(file_path, "name", None),
        "observedAt": None,
        "ageHours": None,
        # "unknown", never a default of "fresh" -- an age we could not
        # measure must not read as an age within budget.
        "state": "unknown",
        "budgetHours": 48.0,
        "note": (
            "Age of the artifact on disk. This endpoint serves the newest "
            "identity report; it does not prove collection is still running."
        ),
    }
    try:
        from src.retention.health import age_hours, artifact_stamp  # noqa: PLC0415

        # Same owner as the retention health probe, and for the same
        # reason: the artifact's own dated filename beats mtime, which a
        # deploy or a restore rewrites.  Measured here — mtime put the
        # newest identity report at 4 days old; its filename puts it at
        # 116, which is the real halt.
        stamp, stamp_source = artifact_stamp(Path(file_path))
        age_h = age_hours(stamp)
        freshness["observedAt"] = stamp
        freshness["stampSource"] = stamp_source
        if age_h is not None:
            freshness["ageHours"] = round(age_h, 2)
            freshness["state"] = "fresh" if age_h <= freshness["budgetHours"] else "stale"
    except Exception:  # noqa: BLE001
        # The report itself still serves; only its label degrades, and
        # it degrades to "unknown" rather than to "fresh".
        pass

    if isinstance(payload, dict):
        body = dict(payload)
        body["evidenceFreshness"] = freshness
    else:
        body = {"report": payload, "evidenceFreshness": freshness}
    return JSONResponse(content=body)


@app.post("/api/waiver/suggestions")
async def post_waiver_suggestions(request: Request):
    """Generate waiver-wire suggestions for the requesting league.

    Players currently free-agent (not on any team's roster) ranked
    by consensus value.  Pre-draft window (Feb 1 – May 11) suppresses
    rookies.

    Request body (JSON):
      ``leagueKey``        optional — pin to a specific league
      ``minValue``         int — floor for ``rankDerivedValue``
      ``includeKicker``    bool — include K/DEF (default false)

    (The former ``applyScoringFit`` / ``scoringFitWeight`` params were
    documentation-only — the handler never read them and no code ever
    produced the adjusted values they promised (2026-07-25 calculation
    audit, F-1).  Unknown body fields are still accepted and ignored,
    so old callers keep working.)

    Returns ``{by_position, by_family, total, rookies_excluded,
    leagueKey}``.
    """
    if not latest_contract_data or not latest_contract_data.get("playersArray"):
        return JSONResponse(
            status_code=503,
            content={"error": "Live contract not loaded yet."},
        )
    try:
        body = await request.json()
    except Exception:
        body = {}
    if not isinstance(body, dict):
        body = {}

    try:
        league_cfg = _resolve_league_for_request(
            request,
            body=body,
            require_loaded_contract=True,
        )
    except LeagueResolutionError as err:
        return err.json_response()

    contract, valuation_mode, valuation_note = await _valuation_scoped_contract(
        request, body, league_cfg
    )
    sleeper = (latest_contract_data or {}).get("sleeper") or {}
    sleeper_teams = sleeper.get("teams") or []
    # Import the constant rather than re-typing its value: this endpoint
    # hardcoded 500 twice, so changing MIN_WAIVER_VALUE would have moved
    # the engine default while leaving the API's default behind it.
    from src.trade.waiver import MIN_WAIVER_VALUE as _MIN_WAIVER_VALUE  # noqa: PLC0415

    try:
        min_value = int(body.get("minValue", _MIN_WAIVER_VALUE))
    except (TypeError, ValueError):
        min_value = _MIN_WAIVER_VALUE
    include_kicker = bool(body.get("includeKicker"))
    try:
        faab_remaining = int(body.get("faabRemaining", 100))
    except (TypeError, ValueError):
        faab_remaining = 100

    from src.trade import waiver as _waiver  # noqa: PLC0415

    # The league's ORIGINAL budget sets the bid scale; the requesting
    # team's remaining balance is only a cap.  Passing the balance as
    # the budget (which this did until the FAAB engine landed) made a
    # player's worth shrink as the manager spent.
    _roster_settings = _league_registry.get_league_roster_settings(league_cfg.key) or {}
    _starters = _roster_settings.get("starters") or {}
    _league_budget = 100
    for _t in sleeper_teams:
        if isinstance(_t, dict) and isinstance(_t.get("faabBudget"), int) and _t["faabBudget"] > 0:
            _league_budget = _t["faabBudget"]
            break

    try:
        result = await run_in_threadpool(
            _waiver.find_waiver_targets,
            contract,
            sleeper_teams,
            min_value=min_value,
            include_kicker_def=include_kicker,
            user_faab_remaining=faab_remaining,
            league_budget=_league_budget,
            team_count=int(_roster_settings.get("teamCount") or len(sleeper_teams) or 12),
            starters_per_team=sum(
                int(v or 0) for k, v in _starters.items() if str(k).upper() != "K"
            )
            or 20,
        )
    except Exception as exc:  # noqa: BLE001
        log.error(f"Waiver suggestions failed: {exc}")
        return JSONResponse(status_code=500, content={"error": f"failed: {exc}"})

    if isinstance(result, dict):
        result["leagueKey"] = league_cfg.key
    _stamp_valuation_mode(result, valuation_mode, valuation_note)
    return JSONResponse(content=result)


@app.post("/api/waiver/faab-recommend")
async def post_waiver_faab_recommend(request: Request):
    """Recommend a FAAB bid for a single add/drop pair.

    Used by the manual add/drop calculator on /waivers.  Composes
    the existing _compute_faab_bid baseline with value-gain modifier,
    Sleeper trending kicker, league-historical analytics calibration,
    optional KTC crowd blend, and team FAAB cap.

    Request body (JSON):
      ``leagueKey``       optional — pin to a specific league
      ``addPlayerName``   required — display name of the add side
      ``dropPlayerName``  optional — display name of the drop side
      ``teamOwnerId``     optional — the requesting user's Sleeper
                          owner id.  When it matches a CURRENT
                          roster in the resolved league it unlocks
                          the team FAAB cap AND the FAAB v2
                          rival-contention model; missing OR
                          unmatched (stale/foreign) ids skip
                          contention with an explicit missing-factor
                          note — we never guess which team is the
                          user's.

    Returns the ``recommend_faab`` payload (see
    ``src/trade/faab_recommender.py``) with ``conservative``,
    ``standard``, ``aggressive``, ``max`` bids plus confidence
    breakdown + warnings + plain-English explanation.  FAAB v2
    adds ONLY new keys (backward compatible): ``contention``
    (clearing / topRival / perOpponent), ``inputsAsOf`` (rosters,
    leagueAnalytics, trending, intel timestamps) and
    ``staleInputs``.
    """
    if not latest_contract_data or not latest_contract_data.get("playersArray"):
        return JSONResponse(
            status_code=503,
            content={"error": "Live contract not loaded yet."},
        )
    try:
        body = await request.json()
    except Exception:
        body = {}
    if not isinstance(body, dict):
        body = {}

    try:
        league_cfg = _resolve_league_for_request(
            request,
            body=body,
            require_loaded_contract=True,
        )
    except LeagueResolutionError as err:
        return err.json_response()

    add_name = str(body.get("addPlayerName") or "").strip()
    drop_name = str(body.get("dropPlayerName") or "").strip()
    if not add_name:
        return JSONResponse(
            status_code=400,
            content={"error": "addPlayerName is required."},
        )

    # Resolve player rows + values from the live contract.
    contract, valuation_mode, valuation_note = await _valuation_scoped_contract(
        request, body, league_cfg
    )
    arr = (contract or {}).get("playersArray") or []

    # Loose trim+lowercase key (family 3 in the ``src/utils/name_clean``
    # registry) — both sides of this join are contract/display names
    # from the same vocabulary, so no stronger key is needed.  Local on
    # purpose: the byte-equal twins in ``src/trade/waiver.py`` and
    # ``src/api/source_history.py`` key unrelated domains.  Note this is
    # NOT the key the KTC crowd FAAB map uses — that one is
    # ``name_clean.compact_name_key`` and the recommender looks it up
    # itself.
    def _norm(s: str) -> str:
        return str(s or "").strip().lower()

    add_row: dict | None = None
    drop_row: dict | None = None
    add_target = _norm(add_name)
    drop_target = _norm(drop_name)
    for row in arr:
        if not isinstance(row, dict):
            continue
        rname = _norm(row.get("displayName") or row.get("name"))
        if rname == add_target:
            add_row = row
        elif drop_target and rname == drop_target:
            drop_row = row
    if add_row is None:
        return JSONResponse(
            status_code=404,
            content={"error": f"Player not found: {add_name!r}"},
        )

    add_value = float(add_row.get("rankDerivedValue") or 0)
    drop_value = float(drop_row.get("rankDerivedValue") or 0) if drop_row else 0.0
    add_position = add_row.get("position") or add_row.get("pos") or None

    # Rosters: prefer the live Sleeper TEAMS-ONLY overlay — its
    # ``teams`` carry ``faabRemaining`` (the baked scrape block does
    # not), which both the team cap and the rival-contention model
    # need.  Deliberately NOT the full ``fetch_sleeper_overlay``:
    # on a cold/expired cache that path also rebuilds the trades +
    # waivers history (dozens of serial transaction requests across
    # two seasons) and would stall this recommendation for many
    # seconds; ``fetch_sleeper_teams_overlay`` reuses a fresh full-
    # overlay cache when one exists and otherwise fetches only
    # rosters/users/settings behind its own short TTL.  Fall back to
    # the baked block when even that is down.
    sleeper = latest_contract_data.get("sleeper") or {}
    rosters_as_of: str | None = None
    sleeper_teams = sleeper.get("teams") or []
    try:
        _overlay_id_map = sleeper.get("idToPlayer") if isinstance(sleeper, dict) else {}
        overlay = await run_in_threadpool(
            lambda: _sleeper_overlay.fetch_sleeper_teams_overlay(
                sleeper_league_id=league_cfg.sleeper_league_id,
                id_to_player=_overlay_id_map if isinstance(_overlay_id_map, dict) else {},
            )
        )
    except Exception as exc:  # noqa: BLE001
        log.warning("faab-recommend overlay fetch failed for %s: %s", league_cfg.key, exc)
        overlay = None
    if overlay and overlay.get("teams"):
        sleeper_teams = overlay["teams"]
        rosters_as_of = overlay.get("overlayFetchedAt")

    # Who is rostered anywhere in the league.  Used below to split the
    # board into "on a roster" and "freely available", which is what
    # the engine's replacement anchor is measured from.
    #
    # The old ``top_value_in_pool`` / ``next_best_fa_value`` scan that
    # lived here is gone with the formula that needed it.  Audit finding
    # W-1 on main found the same defect independently and fixed it in
    # place: the scan filtered on "is this name on a roster" alone, and
    # draft picks are never on a Sleeper roster's ``players`` list, so
    # every pick counted as an available free agent and the best one set
    # the denominator every bid was divided by.  Measured there at ~2.4x
    # too low.  The new engine has no pool denominator at all — its
    # anchors come from league FORMAT — so the whole scan is redundant
    # rather than merely buggy.  The pick exclusion W-1 added survives
    # below, on the board/available split that DOES still feed the
    # replacement anchor.
    rostered_norms: set[str] = set()
    for t in sleeper_teams:
        for n in t.get("players") or []:
            rostered_norms.add(_norm(n))

    # League FAAB analytics — reuse the cached public-snapshot
    # path so this endpoint doesn't pay the multi-season fetch cost
    # on every recommend request.  When the snapshot isn't loaded
    # yet the recommender still works (it just degrades to
    # confidence=low and surfaces a "league analytics missing"
    # factor row).
    from src.api import faab_analytics  # noqa: PLC0415

    league_summary: dict | None = None
    league_analytics_as_of: str | None = None
    try:
        snap_obj = public_snapshot_store.load_snapshot()
        if snap_obj is not None:
            # League-scoping guard (CLAUDE.md: rosters/analytics are
            # per-league).  The public-league pipeline serves ONE
            # global snapshot today (the default league), so with two
            # active leagues sharing owners, another league's
            # aggression/median would corrupt every estimate here.
            # On mismatch, treat league analytics as UNAVAILABLE
            # (aggression → 1.0 + lowSample, env scaling skipped,
            # missing-input factor) rather than consume wrong-league
            # data.  Eventual fix: per-league analytics snapshots.
            snap_league_id = str(getattr(snap_obj, "root_league_id", "") or "")
            if snap_league_id and snap_league_id != str(league_cfg.sleeper_league_id):
                log.info(
                    "faab-recommend: public snapshot is for league %s but "
                    "request resolved to %s (%s) — league analytics "
                    "treated as unavailable",
                    snap_league_id,
                    league_cfg.sleeper_league_id,
                    league_cfg.key,
                )
            else:
                league_summary = faab_analytics.summarize_league_faab(snap_obj)
                league_analytics_as_of = getattr(snap_obj, "generated_at", None)
    except Exception as exc:  # noqa: BLE001
        log.warning(
            "faab analytics build failed for %s: %s",
            league_cfg.key,
            exc,
        )
        league_summary = None

    # Team FAAB remaining — pulled from the resolved league's
    # sleeper teams block.  When the active user has selected a
    # specific team, we use that team's remaining; otherwise the
    # endpoint cannot tell which team is asking and falls back to
    # the league budget as a soft cap.
    team_faab_remaining: int | None = None
    requested_team = body.get("teamOwnerId") or body.get("ownerId")
    requested_team_matched = False
    if requested_team:
        for t in sleeper_teams:
            if str(t.get("ownerId") or "") == str(requested_team):
                requested_team_matched = True
                rem = t.get("faabRemaining")
                if isinstance(rem, int):
                    team_faab_remaining = rem
                break

    # League budget resolution: analytics summary first, then the
    # overlay's per-team ``faabBudget`` (the resolved league's real
    # setting — critical when the league-scoping guard just discarded
    # the analytics: hard-coding 100 would halve every bid in a $200
    # league), then Sleeper's default 100.
    overlay_faab_budget: int | None = None
    for t in sleeper_teams:
        if isinstance(t, dict):
            fb = t.get("faabBudget")
            if isinstance(fb, int) and fb > 0:
                overlay_faab_budget = fb
                break
    league_budget = (
        (league_summary.get("leagueBudget") if isinstance(league_summary, dict) else None)
        or overlay_faab_budget
        or 100
    )

    # Sleeper trending adds — PRIMARY: the live TTL-cached adapter
    # (``src/adapters/sleeper_trending.py``, warmed post-scrape).
    # FALLBACK: the contract's ``sleeperTrending`` key, kept for
    # deployments that backfill it out-of-band (no producer writes
    # it in this repo today).  Missing both just lowers confidence.
    from src.adapters import sleeper_trending as _sleeper_trending  # noqa: PLC0415

    trending_for_player: dict | None = None
    trending_as_of: str | None = None
    add_player_id = str(add_row.get("playerId") or "").strip()
    try:
        trending_snapshot = await run_in_threadpool(_sleeper_trending.get_trending_adds)
    except Exception as exc:  # noqa: BLE001
        log.warning("faab-recommend trending fetch failed: %s", exc)
        trending_snapshot = None
    if (
        add_player_id
        and isinstance(trending_snapshot, dict)
        and isinstance(trending_snapshot.get("counts"), dict)
    ):
        trending_as_of = trending_snapshot.get("fetchedAt")
        # Data present AND the row has a resolvable Sleeper id: a
        # player absent from the board legitimately counts 0
        # (not-trending is signal, not a missing input).  Rows WITHOUT
        # a playerId can't be looked up here at all — treating that
        # impossible lookup as "0 trending" would mark the input as
        # present, inflate confidence, and bypass the name-based
        # fallback below, so they take the fallback path instead.
        #
        # ``asOf`` travels WITH the count.  The adapter serves the previous
        # snapshot when a fetch fails (deliberately, and with no absolute
        # cap), so age is the only thing separating "1,200 adds in the last
        # 24h" from a claim about a day that has long since passed.  Stamping
        # it into ``inputsAsOf`` alone was not enough: the factor row and the
        # confidence bucket never saw it and contradicted the sibling field.
        #
        # ABSENT FROM THE BOARD IS A REAL ZERO.  The board lists everyone
        # trending, so not appearing on it IS the observation — written as an
        # explicit default rather than ``or 0`` so it does not read as a
        # missing-data coercion, which is the distinction the coercion gate
        # exists to keep visible.  A MALFORMED count is a different thing
        # again: not "0 adds" but an unusable observation, so it becomes a
        # missing input rather than a fabricated zero.
        observed = trending_snapshot["counts"].get(add_player_id, 0)
        try:
            trending_count: int | None = int(observed)
        except (TypeError, ValueError):
            trending_count = None
        trending_for_player = {"count": trending_count, "asOf": trending_as_of}
    else:
        trending_block = latest_contract_data.get("sleeperTrending") or {}
        if isinstance(trending_block, dict):
            for pid, rec in trending_block.items():
                if not isinstance(rec, dict):
                    continue
                rname = _norm(rec.get("displayName") or rec.get("name"))
                if rname == add_target:
                    trending_for_player = rec
                    break

    from src.trade import faab_contention as _faab_contention  # noqa: PLC0415
    from src.trade.faab_recommender import (  # noqa: PLC0415
        _need_level,
        build_rivals as _build_rivals,
        compute_confidence,
        recommend_faab,
    )

    # ── Engine context ─────────────────────────────────────────────
    # The FAAB engine needs the league FORMAT (which sets the value
    # anchors), the point in the season, and the selected team's
    # roster shape.  All of it is resolved dynamically — nothing about
    # which team is asking is hard-coded.
    from src.trade import faab_engine as _faab_engine  # noqa: PLC0415
    from src.trade import faab_comparability as _faab_comparability  # noqa: PLC0415
    from src.trade.faab_history import (  # noqa: PLC0415
        build_crowd_market,
        crowd_evidence_for,
        crowd_refusal_reason,
        load_bid_history,
        load_crowd_history,
        summarize_bid_history,
    )

    roster_settings = _league_registry.get_league_roster_settings(league_cfg.key) or {}
    starters_map = roster_settings.get("starters") or {}
    # K is excluded: kickers are not on the valued board, so counting
    # their slots would push the all-in anchor down the board by one
    # slot per team for no corresponding player supply.
    starters_per_team = (
        sum(int(v or 0) for k, v in starters_map.items() if str(k).upper() != "K") or 20
    )
    team_count = int(roster_settings.get("teamCount") or len(sleeper_teams) or 12)
    roster_size = int(roster_settings.get("rosterSize") or 0)

    excluded_positions = set(
        _faab_engine.FaabConfig().get("anchors", "excludedPositions", []) or []
    )
    board_values: list[float] = []
    available_values: list[float] = []
    for row in arr:
        if not isinstance(row, dict):
            continue
        value = row.get("rankDerivedValue")
        if not isinstance(value, (int, float)) or value <= 0:
            continue
        # Two independent pick guards, deliberately both.  ``position``
        # is what the config names; ``assetClass`` is the signal audit
        # finding W-1 used when it found picks polluting the free-agent
        # pool.  A pick that somehow carries a player-ish position would
        # slip past the first check and land in the REPLACEMENT anchor,
        # which is the number every objective ceiling is measured from.
        if str(row.get("position") or "").upper() in excluded_positions:
            continue
        if str(row.get("assetClass") or "").lower() == "pick":
            continue
        board_values.append(float(value))
        if _norm(row.get("displayName") or row.get("name")) not in rostered_norms:
            available_values.append(float(value))

    current_week, in_season = _faab_engine.current_nfl_week()
    playoff_week_start = 15
    for team_row in sleeper_teams:
        if isinstance(team_row, dict) and isinstance(team_row.get("playoffWeekStart"), int):
            playoff_week_start = team_row["playoffWeekStart"]
            break

    # Market priors fitted from THIS league's real bid history, when a
    # snapshot exists (scripts/fetch_faab_history.py writes it).
    market_priors = summarize_bid_history(load_bid_history(league_cfg.key))

    # Cross-league crowd prices for THIS player, when we have them
    # (scripts/fetch_crowd_faab.py accumulates them).  A second,
    # independent read on how contested the claim will be — it moves
    # the expected clearing price and never the objective ceiling.
    #
    # Read through the comparability owner, which fails closed twice: a ledger
    # that has not been refreshed inside its budget is STALE and refused (a
    # snapshot proves when it was taken, not that it is still true), and a
    # position the retained population cannot price — measured, no external
    # league in this feed starts an individual defender — is refused rather
    # than answered from the wrong population.  Both refusals are reported.
    crowd_for_player: dict | None = None
    crowd_market_block: dict | None = None
    try:
        crowd_target = _faab_comparability.TargetFormat.from_league_config(
            league_cfg,
            roster_settings=roster_settings,
            original_budget=float(league_budget),
        )
        crowd_policy = _faab_comparability.ComparabilityPolicy.from_config(
            _faab_engine.FaabConfig()
        )
        # Both the disk read and the classification pass run off the event
        # loop — the ledger is a file and the pass is O(rows).
        crowd_market = await run_in_threadpool(
            lambda: build_crowd_market(
                load_crowd_history(league_cfg.key),
                target=crowd_target,
                policy=crowd_policy,
            )
        )
        crowd_for_player = crowd_evidence_for(crowd_market, add_name, add_position)
        crowd_market_block = crowd_market.to_dict()
        crowd_market_block["refusalReason"] = crowd_refusal_reason(crowd_market, add_position)
        crowd_market_block["playerHasEvidence"] = crowd_for_player is not None
    except Exception as exc:  # noqa: BLE001 — an optional signal must never 500
        log.warning("crowd bid lookup failed for %s: %s", league_cfg.key, exc)
        crowd_market_block = {"state": "missing", "refusalReason": "crowd_lookup_failed"}

    # The selected team's own roster shape.
    selected_team_row: dict | None = None
    if requested_team:
        for t in sleeper_teams:
            if isinstance(t, dict) and str(t.get("ownerId") or "") == str(requested_team):
                selected_team_row = t
                break
    own_players = (selected_team_row or {}).get("players") or []
    open_roster_spots = max(0, roster_size - len(own_players)) if roster_size else 0

    asset_pool = None
    try:
        asset_pool = await run_in_threadpool(
            _faab_contention.build_opponent_asset_pool, latest_contract_data
        )
    except Exception as exc:  # noqa: BLE001 — need analysis is optional
        log.warning("faab-recommend asset pool build failed: %s", exc)

    # Startable-depth need, resolved from the same board the values
    # come from.  Built once and shared with every rival so the user's
    # team and its opponents are judged by identical rules.
    faab_anchors = _faab_engine.resolve_anchors(
        board_values,
        _faab_engine.LeagueContext(
            original_budget=int(league_budget),
            team_count=team_count,
            starters_per_team=starters_per_team,
        ),
        available_values=available_values or None,
    )
    roster_index: dict[str, tuple[float, str]] = {}
    for row in arr:
        if not isinstance(row, dict):
            continue
        value = row.get("rankDerivedValue")
        if not isinstance(value, (int, float)):
            continue
        key = _norm(row.get("displayName") or row.get("name"))
        if key:
            roster_index[key] = (float(value), str(row.get("position") or ""))

    own_need = "neutral"
    if own_players:
        own_need = _need_level(
            add_position,
            own_players,
            asset_pool,
            anchors=faab_anchors,
            starters=starters_map,
            roster_index=roster_index,
        )

    risk_posture = str(body.get("riskPosture") or "balanced").strip().lower()
    if risk_posture not in ("conservative", "balanced", "aggressive"):
        risk_posture = "balanced"

    base_kwargs = dict(
        add_player_value=add_value,
        drop_player_value=drop_value,
        add_player_position=add_position,
        add_player_name=add_name,
        drop_player_name=drop_name or None,
        team_faab_remaining=team_faab_remaining,
        league_faab_summary=league_summary,
        sleeper_trending=trending_for_player,
        league_budget=int(league_budget),
        anchors=faab_anchors,
        board_values=board_values,
        available_values=available_values or None,
        team_count=team_count,
        starters_per_team=starters_per_team,
        current_week=current_week,
        playoff_week_start=playoff_week_start,
        in_season=in_season,
        open_roster_spots=open_roster_spots,
        need_level=own_need,
        risk_posture=risk_posture,
        league_key=league_cfg.key,
        crowd=crowd_for_player,
    )

    # First pass — value only, no rivals.  Kept so the response is
    # still meaningful when we cannot identify the user's team (and so
    # the objective ceiling, which is rival-independent by
    # construction, is always available).
    rec = recommend_faab(**base_kwargs)

    # Phase-5 intel snapshot — defensive plain-JSON read, no
    # ``src.intel`` import (may not be merged/deployed).  Loaded
    # regardless of contention so ``inputsAsOf.intel`` is honest.
    # Snapshots are LEAGUE-PARTITIONED (intel is roster-scoped →
    # league-scoped), so the resolved league's partition is read.
    # Threadpooled like every other I/O in this handler: it reads (and
    # JSON-parses) a snapshot file that grows with pool activity, and a
    # synchronous read here blocks the event loop for every other
    # request in flight.  D11 in docs/intel/AUDIT.md — the surrounding
    # calls were already wrapped and this one was simply missed.
    intel_snapshot = await run_in_threadpool(
        _faab_contention.load_intel_snapshot, league_key=league_cfg.key
    )
    intel_as_of = intel_snapshot.get("generatedAt") if isinstance(intel_snapshot, dict) else None

    # Rival contention (FAAB v2).  Requires a ``teamOwnerId`` that
    # matches a CURRENT roster in the resolved league — we never
    # guess which team is the user's, and a stale/foreign owner id
    # would exclude nobody from the opponent list, silently modeling
    # the user's own team as a rival and poisoning the clearing
    # price.
    contention: dict | None = None
    contention_skip_reason: str | None = None
    opponents: list[dict] = []
    if requested_team and requested_team_matched:
        opponents = [
            t
            for t in sleeper_teams
            if isinstance(t, dict) and str(t.get("ownerId") or "") != str(requested_team)
        ]
        # Usable-balance gate: when the roster fetch degraded (e.g.
        # the league-settings call failed) every ``faabRemaining``
        # can be absent, and the estimator treats a missing balance
        # as unverifiable — broke rivals would otherwise be modeled
        # at full bid, inflating the clearing price while reporting
        # skipped:false.  Require at least half the rivals to carry
        # an integer balance; below that, contention is skipped as
        # an explicitly-missing input.  (Individually missing
        # balances above the threshold are handled conservatively
        # inside ``estimate_rival_bids`` — flagged ``balanceUnknown``
        # and EXCLUDED from the topRival/clearing math.)
        usable_balances = sum(1 for t in opponents if isinstance(t.get("faabRemaining"), int))
        if not opponents:
            contention_skip_reason = "no opponent rosters available — rival contention skipped."
        elif usable_balances * 2 < len(opponents):
            contention_skip_reason = (
                "rival FAAB balances unavailable for most opponents — rival contention skipped."
            )
    if requested_team and requested_team_matched and contention_skip_reason is None:
        # Aggression is resolved per rival inside ``build_rivals``,
        # which looks each CURRENT opponent up by owner id — so a
        # departed manager's history is inert without a pre-filter,
        # and the filtered copy this block used to build was dead.
        intel_index = None
        if intel_snapshot is not None:
            intel_index = _faab_contention.build_intel_index(
                intel_snapshot,
                id_to_position=_faab_contention.player_position_map(latest_contract_data),
            )
        # Rival demand is TEAM-INDEPENDENT by construction in the
        # engine: it keys on the player's objective ceiling, which
        # knows nothing about the user's budget or drop side.  There
        # is no longer a second "neutral pass" to keep the two honest
        # — the separation is structural rather than procedural.
        try:
            rivals = await run_in_threadpool(
                lambda: _build_rivals(
                    opponents,
                    position=add_position,
                    asset_pool=asset_pool,
                    market_priors=market_priors,
                    league_summary=league_summary,
                    roster_size=roster_size,
                    anchors=faab_anchors,
                    starters=starters_map,
                    roster_index=roster_index,
                )
            )
        except Exception as exc:  # noqa: BLE001
            log.warning("faab-recommend rival build failed: %s", exc)
            rivals = []

        if rivals:
            # Intel: an owner who added THIS player (or this position)
            # in another league recently is more likely to contest.
            # It raises the probability they BID, never what the
            # player is worth.
            if intel_index:
                for rival in rivals:
                    factor, _level = _faab_contention.intel_factor(
                        intel_index, rival.owner_id, add_player_id or None, add_position
                    )
                    if factor > 1.0:
                        rival.aggression = float(rival.aggression) * factor

            contention = {"rivals": rivals}
            rec = recommend_faab(**base_kwargs, rivals=rivals)

    if contention is not None:
        rivals = contention["rivals"]
        demand = min(
            1.0,
            float(rec.get("objective", {}).get("pctOfOriginalBudget") or 0.0)
            / 100.0
            / max(1e-9, _faab_engine.FaabConfig().num("market", "demandSaturationBudgets", 2.5)),
        )
        per_opponent = _faab_engine.rival_expected_bids(
            rivals,
            demand_signal=demand,
            league=_faab_engine.LeagueContext(
                original_budget=int(league_budget),
                team_count=team_count,
                starters_per_team=starters_per_team,
            ),
        )
        unknown = sum(1 for r in per_opponent if r["balanceUnknown"])
        notes = [
            "Rival bids are estimates fitted from winning-bid history only — "
            "Sleeper never exposes losing bids, so selection bias is irreducible.",
        ]
        if unknown:
            notes.append(
                f"{unknown} opponent(s) have no visible FAAB balance — shown as estimates "
                "but excluded from the clearing price (an unverifiable rival must never "
                "raise your bid)."
            )
        low_sample_count = sum(1 for r in per_opponent if r["lowSample"])
        if low_sample_count:
            notes.append(
                f"{low_sample_count} opponent(s) below the winning-bid sample floor — "
                "their aggression defaulted to neutral (1.0)."
            )
        if market_priors.sample_size:
            notes.append(
                f"Fitted against {market_priors.sample_size} historical adds "
                f"({market_priors.zero_bid_share:.0%} of which cost $0)."
            )
        else:
            notes.append(
                "No bid history on file for this league — rival behaviour uses configured "
                "priors.  Run scripts/fetch_faab_history.py to improve this."
            )
        rec["contention"] = {
            "clearing": rec["bids"]["clearing"],
            "topRival": max(
                (r["expBid"] for r in per_opponent if not r["balanceUnknown"]), default=0
            ),
            "perOpponent": per_opponent,
            "estimateOnly": True,
            "notes": notes,
            "skipped": False,
        }
    else:
        if not requested_team:
            skip_reason = (
                "teamOwnerId not provided — rival contention skipped "
                "(we never guess which team is yours)."
            )
        elif not requested_team_matched:
            skip_reason = (
                "teamOwnerId does not match any current roster in this "
                "league — rival contention skipped."
            )
        elif contention_skip_reason is not None:
            skip_reason = contention_skip_reason
        else:
            skip_reason = "rival contention unavailable for this request."
        rec["factors"].append(
            {
                "label": "Rival contention",
                "contribution": "missing",
                "weight": 0.15,
                "missing": True,
            }
        )
        # The factor row above landed AFTER recommend_faab computed
        # its confidence — recompute so the reported bucket honestly
        # reflects the full factor set (a missing 0.15-weight factor
        # must be able to pull "high" down).
        rec["confidence"] = compute_confidence(rec["factors"])
        rec["contention"] = {
            "clearing": None,
            "topRival": None,
            "perOpponent": [],
            "estimateOnly": True,
            "notes": [skip_reason],
            "skipped": True,
        }

    rec["inputsAsOf"] = {
        "rosters": rosters_as_of,
        "leagueAnalytics": league_analytics_as_of,
        "trending": trending_as_of,
        "intel": intel_as_of,
    }
    rec["staleInputs"] = _faab_contention.stale_inputs(rec["inputsAsOf"])
    # Provenance for the external market lane.  "We have no crowd price for
    # this player" and "we refused to quote one" must not read the same, so
    # the state, the freshness, the tier composition, what was excluded and
    # why, and the refusal reason are all reported — including when the
    # answer is that nothing was usable.
    if crowd_market_block is not None:
        rec["crowdMarket"] = crowd_market_block
    rec["leagueKey"] = league_cfg.key
    rec["resolvedAddValue"] = add_value
    rec["resolvedDropValue"] = drop_value
    rec["resolvedAddPosition"] = add_position
    # ``resolvedAddValue`` is only interpretable against a named board —
    # the same player is worth a different number under each lens, and a
    # bid derived from one labelled as the other is exactly the
    # confusion this stamp removes.
    _stamp_valuation_mode(rec, valuation_mode, valuation_note)
    return JSONResponse(content=rec)


#: Cap on how many recent rows a market-ledger read returns.  These are
#: history browsers, not exports — an unbounded response would let one
#: league's whole lifetime of waiver claims or trades ride on a single
#: request, and no consumer needs more than a recent window at once.
_MARKET_LEDGER_RECENT_LIMIT = 100


@app.get("/api/market/waivers")
async def get_market_waivers(request: Request):
    """C4-WAIV-01 — this league's waiver-claim history.

    Read-only projection of the canonical acquisition ledger
    (``src.trade.waiver_ledger``); never triggers a Sleeper fetch. Most
    recent claims first (the ledger itself orders oldest-first with
    undated claims leading, so this endpoint reverses it — recency is
    what a reader wants here).

    Responses::

        200  {leagueKey, summary, recentClaims}
        400  unknown_league / inactive_league
        404  no_leagues_configured
    """
    try:
        league_cfg = _resolve_league_for_request(request)
    except LeagueResolutionError as err:
        return err.json_response()

    from src.trade import waiver_ledger as _waiver_ledger  # noqa: PLC0415

    summary = _waiver_ledger.waiver_ledger_summary(league_cfg.key)
    claims = _waiver_ledger.waiver_claims(league_cfg.key)
    recent = list(reversed(claims[-_MARKET_LEDGER_RECENT_LIMIT:]))
    return JSONResponse(
        content={
            "leagueKey": league_cfg.key,
            "summary": summary,
            "recentClaims": recent,
            "recentClaimsTruncated": len(claims) > len(recent),
        }
    )


@app.get("/api/market/trades")
async def get_market_trades(request: Request):
    """C4-MTL-01 — this league's own recorded trade history.

    Read-only projection of the canonical acquisition ledger
    (``src.trade.market_trade_ledger``), scoped to trades this league's
    own rosters made. This is NOT the broader cross-market ledger
    (``C4-MTL-02`` — external ingestion, permission-gated and not yet
    built); every row here is one of our own leagues' own trades. Most
    recent first.

    Responses::

        200  {leagueKey, summary, recentTrades}
        400  unknown_league / inactive_league
        404  no_leagues_configured
    """
    try:
        league_cfg = _resolve_league_for_request(request)
    except LeagueResolutionError as err:
        return err.json_response()

    from src.trade import market_trade_ledger as _market_trade_ledger  # noqa: PLC0415

    summary = _market_trade_ledger.market_ledger_summary(league_cfg.key)
    trades = _market_trade_ledger.market_trades(league_cfg.key)
    recent = list(reversed(trades[-_MARKET_LEDGER_RECENT_LIMIT:]))
    return JSONResponse(
        content={
            "leagueKey": league_cfg.key,
            "summary": summary,
            "recentTrades": recent,
            "recentTradesTruncated": len(trades) > len(recent),
        }
    )


@app.get("/api/valuation/league-adjusted")
async def get_league_adjusted_values(request: Request):
    """This league's league-adjusted value overlay (LI-9).

    Returns only the rows whose value the league adjustment MOVED,
    keyed by ``displayName``.  The client merges it over the consensus
    board; an unmentioned player is unchanged, not missing.  That keeps
    the payload small and makes "nothing to apply" expressible via
    ``isNoop`` rather than an ambiguous empty map.

    LEAGUE-SCOPED, and not merely by convention.  The adjustment is
    driven by ``lineupScarcity``, measured from this league's twelve
    rosters, so two leagues sharing a scoring profile get different
    values here.  That is precisely why these are NOT contract fields:
    the contract is scoring-profile-scoped and shared, and stamping a
    roster-derived number onto it would let one league's roster shape
    silently reprice another's board (CLAUDE.md's core split).  Hence
    the 503 on a league mismatch, matching ``/api/gameplan``.

    Query parameters::

        leagueKey       optional — standard resolver
        explanations    optional — ``1`` to include the full per-player
                        axis decomposition.  Off by default: it is a
                        debugging surface and multiplies the payload.

    Near-free on a warm league: it reuses ``/api/gameplan``'s cached
    bundle solely for its scarcity map and never forces a second solve.
    """
    try:
        league_cfg = _resolve_league_for_request(request)
    except LeagueResolutionError as err:
        body = {"error": err.code, "message": err.message}
        requested = (request.query_params.get("leagueKey") or "").strip()
        if requested:
            body["leagueKey"] = requested
        return JSONResponse(status_code=err.status, content=body)

    contract = latest_contract_data
    if not contract:
        return JSONResponse(
            status_code=503,
            content={
                "error": "data_not_ready",
                "message": "No data available yet. First scrape may still be running.",
                "leagueKey": league_cfg.key,
            },
        )

    loaded_meta = (contract.get("meta") or {}) if isinstance(contract, dict) else {}
    loaded_league = loaded_meta.get("leagueKey")
    if loaded_league and loaded_league != league_cfg.key:
        return JSONResponse(
            status_code=503,
            content={
                "error": "data_not_ready",
                "message": (
                    f"No data loaded for league {league_cfg.key!r} yet "
                    f"(server holds {loaded_league!r})."
                ),
                "leagueKey": league_cfg.key,
            },
        )

    want_explanations = (request.query_params.get("explanations") or "").strip() in {
        "1",
        "true",
        "yes",
    }

    try:
        payload = await run_in_threadpool(
            _gameplan.get_league_adjusted_values,
            league_cfg.key,
            league_cfg.scoring_profile,
            contract,
            include_explanations=want_explanations,
        )
    except _gameplan.GameplanUnavailable as exc:
        # The roster snapshot is missing, so scarcity is unmeasurable.
        # Serve an explicit empty overlay rather than a 503: the board
        # is still perfectly usable at consensus, and a hard failure
        # here would take the whole rankings page down for a missing
        # optional lens.
        return JSONResponse(
            content={
                "leagueKey": league_cfg.key,
                "isNoop": True,
                "adjustedCount": 0,
                "values": {},
                "unavailable": {"reason": exc.reason, "message": exc.detail},
            }
        )

    return JSONResponse(content=payload)


@app.get("/api/bdvm/values")
async def get_bdvm_values(request: Request):
    """BDVM fundamental dynasty values (feature-flagged, default OFF).

    A SECOND, INDEPENDENT value concept: projection-driven fundamental
    value per the Brisket Dynasty Valuation Model
    (docs/research/bdvm-v1/).  It runs beside the market board — it
    never reads or writes ``rankDerivedValue`` and the existing
    rankings/trade routes are untouched whether the flag is on or off.

    Query parameters::

        leagueKey     optional — standard resolver
        surplusMode   optional — ``option`` (default) | ``truncated`` |
                      ``plain``; exposes the §3.3 surplus ablation

    Responses::

        503 feature_disabled     flag ``bdvm_engine`` is off
        503 data_not_ready       no contract / wrong league loaded
        200 status=ok            fundamental values + market comparison
        200 status=no_projection_snapshot
                                 engine is live but no projection
                                 snapshot exists for the season — BDVM
                                 refuses to fabricate projections
    """
    from src.api import feature_flags as _ff  # noqa: PLC0415

    if not _ff.is_enabled("bdvm_engine"):
        return JSONResponse(
            status_code=503,
            content={"error": "feature_disabled", "flag": "bdvm_engine"},
        )
    try:
        league_cfg = _resolve_league_for_request(request)
    except LeagueResolutionError as err:
        return err.json_response()

    contract = latest_contract_data
    if not contract:
        return JSONResponse(
            status_code=503,
            content={
                "error": "data_not_ready",
                "message": "No data available yet. First scrape may still be running.",
                "leagueKey": league_cfg.key,
            },
        )
    loaded_meta = (contract.get("meta") or {}) if isinstance(contract, dict) else {}
    loaded_league = loaded_meta.get("leagueKey")
    if loaded_league and loaded_league != league_cfg.key:
        return JSONResponse(
            status_code=503,
            content={
                "error": "data_not_ready",
                "message": (
                    f"No data loaded for league {league_cfg.key!r} yet "
                    f"(server holds {loaded_league!r})."
                ),
                "leagueKey": league_cfg.key,
            },
        )

    surplus_mode = (request.query_params.get("surplusMode") or "option").strip().lower()
    if surplus_mode not in ("option", "truncated", "plain"):
        return JSONResponse(
            status_code=400,
            content={
                "error": "bad_request",
                "message": f"surplusMode must be option|truncated|plain, got {surplus_mode!r}",
            },
        )

    from src.api import bdvm_api as _bdvm_api  # noqa: PLC0415

    try:
        payload = await run_in_threadpool(
            _bdvm_api.get_bdvm_values,
            contract,
            league_cfg.key,
            surplus_mode=surplus_mode,
        )
    except Exception as exc:  # noqa: BLE001 — surface, never 500-crash the page
        return JSONResponse(
            status_code=503,
            content={
                "error": "bdvm_unavailable",
                "message": str(exc),
                "leagueKey": league_cfg.key,
            },
        )
    payload = dict(payload)
    payload["leagueKey"] = league_cfg.key
    return JSONResponse(content=payload)


def _bdvm_gate_and_league(request: Request, body: dict | None = None):
    """Shared gate for the BDVM family: flag check + league resolution +
    contract readiness.  Returns (league_cfg, contract, error_response).

    POST callers must parse their body FIRST and pass it here — the
    resolver only honors a body ``leagueKey`` when it receives the
    body (CLAUDE.md: query string for GET, body field for POST)."""
    from src.api import feature_flags as _ff  # noqa: PLC0415

    if not _ff.is_enabled("bdvm_engine"):
        return (
            None,
            None,
            JSONResponse(
                status_code=503,
                content={"error": "feature_disabled", "flag": "bdvm_engine"},
            ),
        )
    try:
        league_cfg = _resolve_league_for_request(request, body=body)
    except LeagueResolutionError as err:
        return None, None, err.json_response()
    contract = latest_contract_data
    if not contract:
        return (
            None,
            None,
            JSONResponse(
                status_code=503,
                content={
                    "error": "data_not_ready",
                    "message": "No data available yet. First scrape may still be running.",
                    "leagueKey": league_cfg.key,
                },
            ),
        )
    loaded_meta = (contract.get("meta") or {}) if isinstance(contract, dict) else {}
    loaded_league = loaded_meta.get("leagueKey")
    if loaded_league and loaded_league != league_cfg.key:
        return (
            None,
            None,
            JSONResponse(
                status_code=503,
                content={
                    "error": "data_not_ready",
                    "message": (
                        f"No data loaded for league {league_cfg.key!r} yet "
                        f"(server holds {loaded_league!r})."
                    ),
                    "leagueKey": league_cfg.key,
                },
            ),
        )
    return league_cfg, contract, None


@app.get("/api/bdvm/roster")
async def get_bdvm_roster(request: Request):
    """BDVM per-roster aggregates (feature-flagged, default OFF).

    Strategy capitals (contender/balanced/rebuilder/risk-neutral),
    now/future ratio, positional surplus vs. replacement, direction and
    the strategy currency each roster trades in.  Read-only over the
    live contract's sleeper.teams block.
    """
    league_cfg, contract, err = _bdvm_gate_and_league(request)
    if err is not None:
        return err
    from src.api import bdvm_api as _bdvm_api  # noqa: PLC0415

    try:
        payload = await run_in_threadpool(_bdvm_api.get_bdvm_roster, contract, league_cfg.key)
    except Exception as exc:  # noqa: BLE001
        return JSONResponse(
            status_code=503,
            content={"error": "bdvm_unavailable", "message": str(exc), "leagueKey": league_cfg.key},
        )
    payload = dict(payload)
    payload["leagueKey"] = league_cfg.key
    return JSONResponse(content=payload)


@app.get("/api/bdvm/trades")
async def get_bdvm_trades(request: Request):
    """BDVM double-positive trade scan (feature-flagged, default OFF).

    Finds packages that are positive for BOTH rosters in their own
    strategy currencies, gated by single-market external fairness.

    Query parameters::

        leagueKey   optional — standard resolver
        team        optional — owner id or roster name for side A;
                    omitted scans every pair
    """
    league_cfg, contract, err = _bdvm_gate_and_league(request)
    if err is not None:
        return err
    team = (request.query_params.get("team") or "").strip() or None
    from src.api import bdvm_api as _bdvm_api  # noqa: PLC0415

    try:
        payload = await run_in_threadpool(
            _bdvm_api.get_bdvm_trades, contract, league_cfg.key, team=team
        )
    except Exception as exc:  # noqa: BLE001
        return JSONResponse(
            status_code=503,
            content={"error": "bdvm_unavailable", "message": str(exc), "leagueKey": league_cfg.key},
        )
    payload = dict(payload)
    payload["leagueKey"] = league_cfg.key
    return JSONResponse(content=payload)


@app.post("/api/bdvm/trade-eval")
async def post_bdvm_trade_eval(request: Request):
    """CES evaluation of ONE specific trade in every strategy currency
    (feature-flagged, default OFF).

    Body::

        {
          "sideA": [{"playerId": "..."} | {"name": "..."} | "name", ...],
          "sideB": [...],
          "leagueKey": optional — standard resolver
        }

    Unresolvable refs are reported in ``unresolved`` per side, never
    silently priced at zero.  Package math is the display-layer CES,
    never a plain sum.
    """
    try:
        body = await request.json()
    except Exception:  # noqa: BLE001
        return JSONResponse(status_code=400, content={"error": "invalid_json"})
    # Body parses BEFORE the gate: the frontend sends leagueKey in the
    # JSON body (the POST convention), and the resolver only sees it
    # when the body is passed through.
    league_cfg, contract, err = _bdvm_gate_and_league(
        request, body=body if isinstance(body, dict) else None
    )
    if err is not None:
        return err
    side_a = body.get("sideA") if isinstance(body, dict) else None
    side_b = body.get("sideB") if isinstance(body, dict) else None
    if not isinstance(side_a, list) or not isinstance(side_b, list) or not (side_a or side_b):
        return JSONResponse(
            status_code=400,
            content={"error": "bad_request", "message": "sideA and sideB must be lists"},
        )
    from src.api import bdvm_api as _bdvm_api  # noqa: PLC0415

    try:
        payload = await run_in_threadpool(
            lambda: _bdvm_api.get_bdvm_trade_eval(
                contract, league_cfg.key, side_a=side_a, side_b=side_b
            )
        )
    except Exception as exc:  # noqa: BLE001
        return JSONResponse(
            status_code=503,
            content={"error": "bdvm_unavailable", "message": str(exc), "leagueKey": league_cfg.key},
        )
    payload = dict(payload)
    payload["leagueKey"] = league_cfg.key
    return JSONResponse(content=payload)


@app.get("/api/gameplan")
async def get_gameplan(request: Request):
    """Roster intelligence for one team: needs, window, targets, partners.

    The API surface over ``src/roster_intel/`` — ``analyze_roster``,
    the per-position profiles, the five-state competitive window, both
    target engines, the partner model, and (on request) the Trade
    Package Generator's Pareto frontier.  Assembly lives in
    ``src/api/gameplan.py``; this is the transport shell.

    Query parameters::

        leagueKey   optional — standard resolver (explicit key, else the
                    user's activeLeagueKey, else the registry default)
        team        optional — ownerId. Defaults to the session's
                    Sleeper user id, then the league's default_team_map.
        partner     optional — ownerId from the ``partners`` list. Its
                    PRESENCE is what turns the package generator on:
                    packages are per-counterparty and running all
                    eleven costs ~1.6 s.

    LEAGUE-SCOPED.  Rosters, starter slots, replacement levels, lineup
    solves, partner fit and packages all resolve through ``leagueKey``;
    only player ages and market prices follow the scoring profile.  So
    this endpoint 503s ``data_not_ready`` when the loaded contract is
    for a different league, matching ``/api/terminal`` and
    ``/api/trade/*``.

    Cost is real and is reported rather than hidden: a cold league
    build is ~1.35 s and a cold team view ~1.9 s on the 12-team league.
    Both are cached on a source stamp (snapshot mtimes + the contract's
    scrape timestamp) and run in a threadpool; every response stamps
    ``timing`` so the cost stays measurable in production.
    """
    try:
        league_cfg = _resolve_league_for_request(request)
    except LeagueResolutionError as err:
        # Echo the requested key so a caller can tell which league the
        # error is about, the way the other league-scoped routes do.
        body = {"error": err.code, "message": err.message}
        requested = (request.query_params.get("leagueKey") or "").strip()
        if requested:
            body["leagueKey"] = requested
        return JSONResponse(status_code=err.status, content=body)

    contract = latest_contract_data
    if not contract:
        return JSONResponse(
            status_code=503,
            content={
                "error": "data_not_ready",
                "message": "No data available yet. First scrape may still be running.",
                "leagueKey": league_cfg.key,
            },
        )

    loaded_meta = (contract.get("meta") or {}) if isinstance(contract, dict) else {}
    loaded_league = loaded_meta.get("leagueKey")
    if loaded_league and loaded_league != league_cfg.key:
        return JSONResponse(
            status_code=503,
            content={
                "error": "data_not_ready",
                "message": (
                    f"No data loaded for league {league_cfg.key!r} yet "
                    f"(server holds {loaded_league!r})."
                ),
                "leagueKey": league_cfg.key,
            },
        )

    params = request.query_params
    owner_id = (params.get("team") or params.get("ownerId") or "").strip()
    if not owner_id:
        session = _get_auth_session(request) or {}
        owner_id = str(session.get("sleeper_user_id") or "").strip()
        if not owner_id:
            username = str(session.get("username") or "").strip().lower()
            mapped = (league_cfg.default_team_map or {}).get(username) or {}
            owner_id = str(mapped.get("ownerId") or "").strip()
    if not owner_id:
        return JSONResponse(
            status_code=400,
            content={
                "error": "team_required",
                "message": (
                    "Pass ?team=<ownerId>. It could not be inferred: this session "
                    "carries no Sleeper user id and the league has no default team "
                    "mapping for it."
                ),
                "leagueKey": league_cfg.key,
            },
        )

    partner_owner_id = (params.get("partner") or "").strip() or None

    try:
        payload = await run_in_threadpool(
            _gameplan.get_team_gameplan,
            league_cfg.key,
            league_cfg.scoring_profile,
            contract,
            owner_id,
            partner_owner_id=partner_owner_id,
        )
    except _gameplan.GameplanUnavailable as exc:
        return JSONResponse(
            status_code=503,
            content={
                "error": "data_not_ready",
                "reason": exc.reason,
                "message": exc.detail,
                "leagueKey": league_cfg.key,
            },
        )
    except _gameplan.TeamNotInLeague:
        return JSONResponse(
            status_code=404,
            content={
                "error": "team_not_found",
                "message": f"No roster for owner {owner_id!r} in league {league_cfg.key!r}.",
                "leagueKey": league_cfg.key,
            },
        )

    # Private roster intelligence — never cached by a shared proxy.
    return JSONResponse(content=payload, headers={"Cache-Control": "no-store"})


@app.get("/api/roster/intelligence")
async def get_roster_intelligence(request: Request):
    """Canonical roster intelligence for one team: core, strength,
    weakness, age/value portfolio.

    The transport shell over ``src/roster_intel/`` via
    ``src/api/roster_intelligence.py``.  Everything it returns is
    computed by the canonical owners — meaningful core (C2-CORE-01),
    Team Strength (inventory row 1.1), Team Weakness (1.2) and the
    age-value portfolio (1.6) — so a consumer that reads this endpoint
    and a consumer that imports ``src.roster_intel`` get the same
    numbers by construction.

    Query parameters::

        leagueKey   optional — standard resolver (explicit key, else the
                    user's activeLeagueKey, else the registry default)
        team        optional — ownerId. Defaults to the session's
                    Sleeper user id, then the league's default_team_map.
        droppability  optional — "1"/"true" adds this team's canonical cut
                    ladder (C2-DROP-01) from ``src/draft/displacement.py``.
                    OFF by default: it re-runs the exact assignment solver
                    once per rung, which costs ~55 ms for one team against
                    ~69 ms for everything else on the live board.

    LEAGUE-SCOPED, and necessarily so: rosters, starter slots and every
    league-relative rank resolve through ``leagueKey``.  Only the
    canonical values and ages follow the scoring profile.  So this 503s
    ``data_not_ready`` on a contract loaded for a different league,
    matching ``/api/gameplan``, ``/api/terminal`` and ``/api/trade/*``.

    Reads the canonical contract directly.  It does NOT route through
    ``gameplan.get_league_bundle``: that bundle's rosters come from the
    ROS team-strength snapshot, which carries a 0-100 production index
    rather than canonical dynasty value and drops unpriced players
    before any consumer can see them.
    """
    try:
        league_cfg = _resolve_league_for_request(request)
    except LeagueResolutionError as err:
        body = {"error": err.code, "message": err.message}
        requested = (request.query_params.get("leagueKey") or "").strip()
        if requested:
            body["leagueKey"] = requested
        return JSONResponse(status_code=err.status, content=body)

    contract = latest_contract_data
    if not contract:
        return JSONResponse(
            status_code=503,
            content={
                "error": "data_not_ready",
                "message": "No data available yet. First scrape may still be running.",
                "leagueKey": league_cfg.key,
            },
        )

    loaded_meta = (contract.get("meta") or {}) if isinstance(contract, dict) else {}
    loaded_league = loaded_meta.get("leagueKey")
    if loaded_league and loaded_league != league_cfg.key:
        return JSONResponse(
            status_code=503,
            content={
                "error": "data_not_ready",
                "message": (
                    f"No data loaded for league {league_cfg.key!r} yet "
                    f"(server holds {loaded_league!r})."
                ),
                "leagueKey": league_cfg.key,
            },
        )

    owner_id = (request.query_params.get("team") or "").strip()
    if not owner_id:
        session = _get_auth_session(request) or {}
        owner_id = str(session.get("sleeper_user_id") or "").strip()
        if not owner_id:
            username = str(session.get("username") or "").strip().lower()
            mapped = (league_cfg.default_team_map or {}).get(username) or {}
            owner_id = str(mapped.get("ownerId") or "").strip()
    if not owner_id:
        return JSONResponse(
            status_code=400,
            content={
                "error": "team_required",
                "message": (
                    "Pass ?team=<ownerId>. It could not be inferred: this session "
                    "carries no Sleeper user id and the league has no default team "
                    "mapping for it."
                ),
                "leagueKey": league_cfg.key,
            },
        )

    # The league's DECLARED size, not the roster count: a snapshot
    # missing one roster must not shrink every weakness threshold.
    declared_teams = None
    try:
        settings = _league_registry.get_league_roster_settings(league_cfg.key) or {}
        raw = settings.get("teamCount")
        if isinstance(raw, int) and raw > 0:
            declared_teams = raw
    except Exception:  # noqa: BLE001 — the registry is optional here
        declared_teams = None

    want_drops = (request.query_params.get("droppability") or "").strip().lower() in {
        "1",
        "true",
        "yes",
    }

    try:
        payload = await run_in_threadpool(
            _roster_intelligence.get_team_roster_intelligence,
            contract,
            owner_id,
            team_count=declared_teams,
            include_droppability=want_drops,
        )
    except _roster_intelligence.TeamNotInLeague:
        return JSONResponse(
            status_code=404,
            content={
                "error": "team_not_found",
                "message": f"No roster for owner {owner_id!r} in league {league_cfg.key!r}.",
                "leagueKey": league_cfg.key,
            },
        )

    # Private roster intelligence — never cached by a shared proxy.
    return JSONResponse(content=payload, headers={"Cache-Control": "no-store"})


@app.post("/api/trade/suggestions")
async def post_trade_suggestions(request: Request):
    """Generate trade suggestions for a given roster.

    Accepts JSON body:
        {
          "roster": ["Josh Allen", "Bijan Robinson", ...],
          "league_rosters": [                              // optional
            {"team_name": "Team A", "players": ["Player1", ...]},
            ...
          ]
        }

    Requires canonical data to be loaded. Returns roster analysis
    and categorized trade suggestions with market-edge signals
    and optional opponent-fit labels.
    """
    # The suggestion engine now reads the live contract
    # (``playersArray``) directly — no offline canonical snapshot
    # required.  We only 503 if the live contract itself hasn't
    # loaded, which indicates a server-bootstrap problem rather than
    # a missing canonical build.
    if not latest_contract_data or not latest_contract_data.get("playersArray"):
        return JSONResponse(
            status_code=503,
            content={"error": "Live contract not loaded yet. Retry in a moment."},
        )
    try:
        body = await request.json()
    except Exception:
        return JSONResponse(status_code=400, content={"error": "Invalid JSON body"})

    # Validate leagueKey (body or query) against the registry.
    try:
        league_cfg = _resolve_league_for_request(
            request,
            body=body,
            require_loaded_contract=True,
        )
    except LeagueResolutionError as err:
        return err.json_response()

    roster = body.get("roster")
    if not isinstance(roster, list) or not roster:
        return JSONResponse(
            status_code=400,
            content={
                "error": "Request body must include 'roster' as a non-empty array of player names."
            },
        )

    from src.trade.suggestions import (
        build_asset_pool_from_contract,
        generate_suggestions_from_pool,
        starter_needs_for_league,
    )

    # Effective starter demand is a per-LEAGUE fact, not a constant: both
    # live leagues share a scoring profile but ``dynasty_main`` starts
    # 2 TE + 9 IDP while ``dynasty_new`` starts 1 TE and no IDP.  The
    # engine has always accepted ``starter_needs``; nothing ever passed
    # it, so every league got dynasty_main's lineup.  The derivation is
    # a no-op for dynasty_main by construction.
    starter_needs = starter_needs_for_league(getattr(league_cfg, "key", None))

    league_rosters = body.get("league_rosters")
    if league_rosters is not None and not isinstance(league_rosters, list):
        league_rosters = None

    # Optional ``ktc_top_n`` cap from the request body — the
    # /settings page exposes a slider (default 150, range 50-300)
    # so deeper-format leagues can include lower-ranked offense
    # players in suggestion candidacy.  Sanitize to integer in
    # [50, 300]; out-of-range or missing falls back to the engine's
    # default constant.
    #
    # The wire names stay ``ktc_top_n`` (request) / ``ktcTopNFilter``
    # (response) — they are the published contract.  The engine-side
    # constant is ``BOARD_TOP_N_FILTER``, because this gate ranks
    # against OUR blended board and never consulted KTC; the
    # ``KTC_TOP_N_FILTER`` alias is deprecated and not imported here.
    from src.trade.suggestions import BOARD_TOP_N_FILTER as _KTC_TOP_N_DEFAULT

    raw_ktc_top_n = body.get("ktc_top_n")
    try:
        ktc_top_n = int(raw_ktc_top_n) if raw_ktc_top_n is not None else _KTC_TOP_N_DEFAULT
    except (TypeError, ValueError):
        ktc_top_n = _KTC_TOP_N_DEFAULT
    ktc_top_n = max(50, min(300, ktc_top_n))

    # Build the asset pool directly from the live contract.  Every
    # field the suggestion engine needs already lives on the
    # ``playersArray`` rows (see ``build_asset_pool_from_contract``
    # docstring for the field map).  This replaces the old two-step
    # flow of (a) loading the offline canonical snapshot and
    # (b) overlaying live values on top — with the contract-native
    # path there's only one source of truth.
    contract, valuation_mode, valuation_note = await _valuation_scoped_contract(
        request, body, league_cfg
    )

    # C3-CON-01 — resolved on the request thread (it reads the session and the
    # user store) and handed to the generator, which applies it during
    # enumeration.
    constraints = _constraints_for_request(
        request, contract, league_cfg, surface="/api/trade/suggestions"
    )

    # Pool build + suggestion generation are heavy CPU passes over the
    # full playersArray × league rosters — run them on a worker thread
    # exactly like /api/trade/finder already does, instead of stalling
    # the event loop for every concurrent request.
    def _generate():
        pool = build_asset_pool_from_contract(
            contract,
            ktc_top_n=ktc_top_n,
        )
        # Roster capacity for the requesting team, built ONCE and reused
        # across every proposal — the expensive part (joining the roster to
        # the board, measuring waiver level league-wide, resolving starter
        # slots) does not depend on which trade is being scored.  Each
        # suggestion then reports what it would cost in forced releases;
        # nothing is filtered out on capacity grounds.
        # Suggestions are the product; capacity is an annotation on them, so a
        # failure here degrades to no annotation rather than no suggestions.
        capacity_context = _capacity_context_for(
            contract,
            league_cfg,
            {"players": list(roster)},
            surface="/api/trade/suggestions",
        )

        return generate_suggestions_from_pool(
            roster_names=roster,
            pool=pool,
            league_rosters=league_rosters,
            starter_needs=starter_needs,
            ktc_top_n=ktc_top_n,
            capacity_context=capacity_context,
            constraints=constraints,
        )

    try:
        result = await run_in_threadpool(_generate)
    except Exception as e:
        log.error(f"Trade suggestion generation failed: {e}")
        return JSONResponse(
            status_code=500, content={"error": f"Suggestion generation failed: {e}"}
        )

    if isinstance(result, dict):
        result["leagueKey"] = league_cfg.key
    _stamp_valuation_mode(result, valuation_mode, valuation_note)
    return JSONResponse(content=result)


@app.post("/api/trade/finder")
async def post_trade_finder(request: Request):
    """Find board-arbitrage trades: good for me on our model, plausible for them on KTC.

    Accepts JSON body:
        {
          "myTeam": "Team Name",
          "opponentTeams": ["Opponent 1", "Opponent 2"]   // or ["all"] for all teams
        }

    Requires live data to be loaded. Works against the production data payload
    (players dict with _rawComposite / _canonicalSiteValues fields).
    """
    if latest_contract_data is None:
        return JSONResponse(
            status_code=503,
            content={"error": "No data loaded. Trade Finder requires live player data."},
        )
    players = latest_contract_data.get("players")
    sleeper = latest_contract_data.get("sleeper") or {}
    sleeper_teams = sleeper.get("teams") or []
    if not players or not sleeper_teams:
        return JSONResponse(
            status_code=503,
            content={"error": "Player data or Sleeper rosters not available."},
        )

    try:
        body = await request.json()
    except Exception:
        return JSONResponse(status_code=400, content={"error": "Invalid JSON body"})

    # Validate leagueKey (body or query).
    try:
        league_cfg = _resolve_league_for_request(
            request,
            body=body,
            require_loaded_contract=True,
        )
    except LeagueResolutionError as err:
        return err.json_response()

    my_team = body.get("myTeam")
    if not my_team or not isinstance(my_team, str):
        return JSONResponse(
            status_code=400,
            content={"error": "Request body must include 'myTeam' as a team name string."},
        )

    opponent_teams = body.get("opponentTeams", [])
    if not isinstance(opponent_teams, list):
        return JSONResponse(status_code=400, content={"error": "'opponentTeams' must be a list."})

    # "all" means trade with every team except mine
    if opponent_teams == ["all"] or not opponent_teams:
        opponent_teams = [t["name"] for t in sleeper_teams if t.get("name") != my_team]

    # Wire name stays ``ktc_top_n``; the engine constant is
    # ``MARKET_TOP_N_FILTER`` (the gate is per-market — ktcSfTep for
    # offense + picks, idpTradeCalc for IDP — not KTC-only).  The
    # ``KTC_TOP_N_FILTER`` alias is deprecated and not imported here.
    from src.trade.finder import find_trades, MARKET_TOP_N_FILTER as _FINDER_KTC_TOP_N_DEFAULT

    raw_ktc_top_n_f = body.get("ktc_top_n")
    try:
        finder_ktc_top_n = (
            int(raw_ktc_top_n_f) if raw_ktc_top_n_f is not None else _FINDER_KTC_TOP_N_DEFAULT
        )
    except (TypeError, ValueError):
        finder_ktc_top_n = _FINDER_KTC_TOP_N_DEFAULT
    finder_ktc_top_n = max(50, min(300, finder_ktc_top_n))

    # Only OUR side of the arbitrage follows the lens.  The market
    # anchor is read from the raw ``players`` dict and stays the retail
    # board, which is the point: the finder measures the gap between
    # what an asset is worth to us and what the counterparty's
    # calculator says.  Adjusting both sides would close the gap it
    # exists to find.
    contract, valuation_mode, valuation_note = await _valuation_scoped_contract(
        request, body, league_cfg
    )

    # Roster capacity for the requesting team.  Built here rather than inside
    # the engine because it needs the league config, and reused across every
    # returned trade.  A failure here annotates nothing and breaks nothing.
    finder_capacity_context = _capacity_context_for(
        contract,
        league_cfg,
        next((t for t in sleeper_teams if t.get("name") == my_team), None),
        surface="/api/trade/finder",
    )

    # C3-CON-01 — the finder was the only engine that already accepted
    # `constraints`, and this route never passed any, so it ran unconstrained in
    # production while the parameter looked wired.
    finder_constraints = _constraints_for_request(
        request, contract, league_cfg, surface="/api/trade/finder"
    )

    try:
        result = await run_in_threadpool(
            find_trades,
            players=players,
            my_team=my_team,
            opponent_teams=opponent_teams,
            sleeper_teams=sleeper_teams,
            ktc_top_n=finder_ktc_top_n,
            capacity_context=finder_capacity_context,
            constraints=finder_constraints,
            # F-6 (audit finding K): the contract carries `playersArray`
            # with `rankDerivedValue` — the board the user actually sees.
            # Without this the finder arbitrages the raw scraper
            # composite, which no other engine and no UI surface reads.
            contract=contract,
        )
    except Exception as e:
        log.error(f"Trade Finder failed: {e}")
        return JSONResponse(status_code=500, content={"error": f"Trade Finder failed: {e}"})

    if isinstance(result, dict):
        result["leagueKey"] = league_cfg.key
    _stamp_valuation_mode(result, valuation_mode, valuation_note)
    return JSONResponse(content=result)


@app.post("/api/trade/import-ktc")
async def post_trade_import_ktc(request: Request):
    """Resolve a KeepTradeCut trade-calculator URL into ordered
    player lists the frontend can load into its sides.

    Body: ``{"url": "https://keeptradecut.com/trade-calculator?...&teamOne=1274&teamTwo=1555..."}``.

    Returns ``{sideOne, sideTwo, unresolved, sourceUrl}`` — see
    ``src/trade/ktc_import.py::resolve_trade_url`` for the shape.
    Public endpoint (same as the other /api/trade/* endpoints) so
    the drawer-less trade page works without re-authing.
    """
    from src.trade.ktc_import import resolve_trade_url  # noqa: PLC0415 — lazy import

    try:
        body = await request.json()
    except Exception:
        body = None
    if not isinstance(body, dict):
        return JSONResponse(
            status_code=400,
            content={"ok": False, "error": "JSON body required."},
        )
    url = str(body.get("url") or "").strip()
    if not url:
        return JSONResponse(
            status_code=400,
            content={"ok": False, "error": "Missing 'url' field."},
        )
    if "keeptradecut.com" not in url:
        return JSONResponse(
            status_code=400,
            content={
                "ok": False,
                "error": "URL must be a keeptradecut.com trade-calculator link.",
            },
        )

    # KTC HTML fetch + regex is blocking — run in threadpool so we
    # don't stall the event loop for other in-flight requests.
    try:
        result = await run_in_threadpool(resolve_trade_url, url)
    except ValueError as exc:
        return JSONResponse(
            status_code=400,
            content={"ok": False, "error": str(exc)},
        )
    except Exception as exc:  # noqa: BLE001 — surface upstream failures
        return JSONResponse(
            status_code=502,
            content={
                "ok": False,
                "error": "Failed to fetch KTC player map.",
                "detail": f"{type(exc).__name__}: {exc}",
            },
        )
    return JSONResponse(content={"ok": True, **result})


@app.post("/api/trade/export-ktc")
async def post_trade_export_ktc(request: Request):
    """Inverse of import-ktc: the two local sides -> a KTC URL.

    Body ``{"sideOne": ["Josh Allen", ...], "sideTwo": ["2026 Mid 1st", ...]}``.
    Returns ``{ok, url, unresolved:{sideOne,sideTwo}, resolvedCount}``.
    Reuses the cached KTC player map.  Public, like other /api/trade/*.
    """
    from src.trade.ktc_import import build_ktc_url  # noqa: PLC0415

    try:
        body = await request.json()
    except Exception:
        body = None
    if not isinstance(body, dict):
        return JSONResponse(status_code=400, content={"ok": False, "error": "JSON body required."})

    def _names(key):
        raw = body.get(key)
        if not isinstance(raw, list):
            return []
        return [str(x).strip() for x in raw if str(x or "").strip()]

    side_one, side_two = _names("sideOne"), _names("sideTwo")
    if not side_one and not side_two:
        return JSONResponse(status_code=400, content={"ok": False, "error": "No assets to export."})

    try:
        result = await run_in_threadpool(build_ktc_url, side_one, side_two)
    except Exception as exc:  # noqa: BLE001
        return JSONResponse(
            status_code=502,
            content={
                "ok": False,
                "error": "Failed to fetch KTC player map.",
                "detail": f"{type(exc).__name__}: {exc}",
            },
        )
    return JSONResponse(content={"ok": True, **result})


@app.post("/api/angle/find")
async def post_angle_find(request: Request):
    """Player-specific arbitrage: pick a player on your team, get
    targets on other teams where your rankings say win but KTC says
    fair-to-neutral (easy to pitch as "KTC says this is even").

    Accepts JSON body:
        {
          "ownerId": "472206636534984704",       // your sleeper ownerId
          "playerName": "Jayden Daniels",        // canonical name
          "targetTeamOwnerId": "729547...",      // optional — restrict
                                                  // candidates to this
                                                  // single opposing team
          "minMyGainPct": 5.0,                    // optional, default 5
          "maxMarketGainPct": 5.0,                // optional, default 5
          "limit": 50                             // optional, default 50
        }

    Market value is per-position: IDPTradeCalc for IDP (DL/LB/DB),
    KTC for everyone else. Legacy body key ``maxKtcGainPct`` is
    still accepted for backward compatibility.
    """
    if latest_contract_data is None:
        return JSONResponse(
            status_code=503,
            content={"error": "No data loaded. Angle requires live player data."},
        )
    players_array = (latest_contract_data or {}).get("playersArray")
    sleeper = latest_contract_data.get("sleeper") or {}
    sleeper_teams = sleeper.get("teams") or []
    if not players_array or not sleeper_teams:
        return JSONResponse(
            status_code=503,
            content={"error": "Player data or Sleeper rosters not available."},
        )
    try:
        body = await request.json()
    except Exception:
        return JSONResponse(status_code=400, content={"error": "Invalid JSON body"})

    # League routing: accept leagueKey in body or query, reject
    # unknown/inactive, 503 when the loaded contract is for a
    # different league.
    try:
        league_cfg = _resolve_league_for_request(
            request,
            body=body,
            require_loaded_contract=True,
        )
    except LeagueResolutionError as err:
        return err.json_response()

    owner_id = str(body.get("ownerId") or "").strip()
    player_name = str(body.get("playerName") or "").strip()
    if not owner_id or not player_name:
        return JSONResponse(
            status_code=400,
            content={"error": "Request body must include 'ownerId' and 'playerName'."},
        )

    target_team_owner_id = str(body.get("targetTeamOwnerId") or "").strip() or None
    if target_team_owner_id and target_team_owner_id == owner_id:
        return JSONResponse(
            status_code=400,
            content={"error": "targetTeamOwnerId must differ from ownerId."},
        )

    try:
        min_my = float(body.get("minMyGainPct", 5.0))
        # Accept new key, fall back to legacy for pre-rename clients.
        max_market = float(body.get("maxMarketGainPct", body.get("maxKtcGainPct", 5.0)))
        limit = int(body.get("limit", 50))
    except (TypeError, ValueError):
        return JSONResponse(
            status_code=400,
            content={"error": "minMyGainPct, maxMarketGainPct, limit must be numeric."},
        )

    from src.trade.angle import find_angles

    # Same asymmetry as the finder: our side follows the lens, the
    # market anchor the pitch rests on does not.
    contract, valuation_mode, valuation_note = await _valuation_scoped_contract(
        request, body, league_cfg
    )
    # Roster capacity for the REQUESTING team.  Angle already knows which team
    # is asking (``ownerId``), and each returned candidate reports what the
    # swap would cost in forced releases.  Nothing is filtered on it.
    angle_capacity_context = _capacity_context_for(
        contract,
        league_cfg,
        _team_block_by_owner_id(sleeper_teams, owner_id),
        surface="/api/angle/find",
    )
    try:
        result = await run_in_threadpool(
            find_angles,
            (contract or {}).get("playersArray") or players_array,
            player_name,
            owner_id,
            sleeper_teams,
            min_my_gain_pct=min_my,
            max_market_gain_pct=max_market,
            limit=limit,
            target_team_owner_id=target_team_owner_id,
            capacity_context=angle_capacity_context,
        )
    except Exception as exc:  # noqa: BLE001
        log.error(f"Angle find failed: {exc}")
        return JSONResponse(
            status_code=500,
            content={"error": f"Angle find failed: {exc}"},
        )
    if isinstance(result, dict):
        result["leagueKey"] = league_cfg.key
    _stamp_valuation_mode(result, valuation_mode, valuation_note)
    return JSONResponse(content=result)


@app.post("/api/angle/packages")
async def post_angle_packages(request: Request):
    """Multi-player variant of Angle. Two modes:

    * ``mode: "offer"`` (default) — offer a list of your players, get
      back counter-packages from other teams sized within ±1 of your
      offer that lean your way on my-value but look fair-or-better to
      the counterparty on market.
    * ``mode: "acquire"`` — pick players on opposing rosters you want
      to acquire; get back offer-side packages from YOUR OWN roster
      (size within ±1 of the desired count) that satisfy the same
      arbitrage math. Lets you skip picking your own players upfront.

    Body (offer mode):
        {
          "mode": "offer",                      // optional, default
          "ownerId": "472206636534984704",
          "playerNames": ["Jayden Daniels", ...],
          "minMyGainPct": 5.0,
          "maxMarketGainPct": 5.0,
          "limit": 50,
          "candidatePoolPerTeam": 25
        }

    Body (acquire mode):
        {
          "mode": "acquire",
          "ownerId": "472206636534984704",
          "acquirePlayerNames": ["Ja'Marr Chase", "Bijan Robinson"],
          "minMyGainPct": 5.0,
          "maxMarketGainPct": 5.0,
          "limit": 50,
          "candidatePoolPerTeam": 25
        }

    Market value is per-position: IDPTradeCalc for IDP (DL/LB/DB),
    KTC for everyone else. Legacy body key ``maxKtcGainPct`` is
    still accepted.
    """
    if latest_contract_data is None:
        return JSONResponse(
            status_code=503,
            content={"error": "No data loaded. Angle requires live player data."},
        )
    players_array = (latest_contract_data or {}).get("playersArray")
    sleeper = latest_contract_data.get("sleeper") or {}
    sleeper_teams = sleeper.get("teams") or []
    if not players_array or not sleeper_teams:
        return JSONResponse(
            status_code=503,
            content={"error": "Player data or Sleeper rosters not available."},
        )
    try:
        body = await request.json()
    except Exception:
        return JSONResponse(status_code=400, content={"error": "Invalid JSON body"})

    # League routing (same pattern as /api/angle/find).
    try:
        league_cfg = _resolve_league_for_request(
            request,
            body=body,
            require_loaded_contract=True,
        )
    except LeagueResolutionError as err:
        return err.json_response()

    owner_id = str(body.get("ownerId") or "").strip()
    mode = str(body.get("mode") or "offer").strip().lower()
    if mode not in ("offer", "acquire"):
        return JSONResponse(
            status_code=400,
            content={"error": "'mode' must be 'offer' or 'acquire'."},
        )

    # In offer mode the user builds an offer from their roster and the
    # search returns counter-packages from other teams. In acquire
    # mode the user picks players on opposing rosters they want to
    # acquire and the search returns offer-side packages from their
    # own roster.
    if mode == "acquire":
        names_key = "acquirePlayerNames"
        names = body.get(names_key) or body.get("playerNames") or []
    else:
        names_key = "playerNames"
        names = body.get(names_key) or []
    if not owner_id or not isinstance(names, list) or not names:
        return JSONResponse(
            status_code=400,
            content={
                "error": (
                    f"Request body must include 'ownerId' and a non-empty {names_key!r} list."
                )
            },
        )
    player_names = [str(n).strip() for n in names if str(n).strip()]

    try:
        min_my = float(body.get("minMyGainPct", 5.0))
        # Accept renamed key; fall back to legacy for pre-rename clients.
        max_market = float(body.get("maxMarketGainPct", body.get("maxKtcGainPct", 5.0)))
        limit = int(body.get("limit", 50))
        pool = int(body.get("candidatePoolPerTeam", 25))
        per_team = int(body.get("perTeamLimit", 4))
        min_player = float(body.get("minPlayerMyValue", 0.0))
    except (TypeError, ValueError):
        return JSONResponse(
            status_code=400,
            content={"error": "numeric params must be numeric."},
        )
    positions_req = body.get("positions") or []
    if not isinstance(positions_req, list):
        positions_req = []
    positions_req = [str(p).strip() for p in positions_req if str(p).strip()]

    include_idp_raw = body.get("includeIdp", False)
    include_idp = bool(include_idp_raw) and include_idp_raw not in ("false", "0", "")
    # Back-compat: if the caller explicitly requested an IDP position
    # via ``positions`` but didn't set ``includeIdp`` (e.g. legacy
    # scripts predating the toggle), treat that as an implicit opt-in.
    # Otherwise ``positions=["DL"]`` alone would filter the pool down
    # to zero candidates, which silently breaks those callers.
    from src.trade.angle import _IDP_POSITIONS as _ANGLE_IDP_POSITIONS

    if not include_idp and any(
        str(p).strip().upper() in _ANGLE_IDP_POSITIONS for p in positions_req
    ):
        include_idp = True

    # Same asymmetry as /api/angle/find: our side follows the lens, the
    # market anchor the pitch rests on does not.
    contract, valuation_mode, valuation_note = await _valuation_scoped_contract(
        request, body, league_cfg
    )
    lens_rows = (contract or {}).get("playersArray") or players_array

    # One context for both modes — the requesting team is the same either way,
    # and only which SIDE it appears on changes.
    angle_capacity_context = _capacity_context_for(
        contract,
        league_cfg,
        _team_block_by_owner_id(sleeper_teams, owner_id),
        surface="/api/angle/packages",
    )

    if mode == "acquire":
        from src.trade.angle import find_acquisition_packages

        # C3-CON-01.  Acquire mode is the one Angle mode that CHOOSES outgoing
        # assets — the other two enumerate what we would receive against a send
        # side the user typed — so it is the one that carries constraints.
        angle_constraints = _constraints_for_request(
            request, contract, league_cfg, surface="/api/angle/packages"
        )

        try:
            result = await run_in_threadpool(
                find_acquisition_packages,
                lens_rows,
                player_names,
                owner_id,
                sleeper_teams,
                min_my_gain_pct=min_my,
                max_market_gain_pct=max_market,
                limit=limit,
                candidate_pool=pool,
                positions=positions_req or None,
                min_player_my_value=min_player,
                include_idp=include_idp,
                capacity_context=angle_capacity_context,
                constraints=angle_constraints,
            )
        except Exception as exc:  # noqa: BLE001
            log.error(f"Angle acquire failed: {exc}")
            return JSONResponse(
                status_code=500,
                content={"error": f"Angle acquire failed: {exc}"},
            )
        result = {"mode": "acquire", **result, "leagueKey": league_cfg.key}
        _stamp_valuation_mode(result, valuation_mode, valuation_note)
        return JSONResponse(content=result)

    target_teams_req = body.get("targetTeamOwnerIds") or []
    if not isinstance(target_teams_req, list):
        target_teams_req = []
    target_teams_req = [str(t).strip() for t in target_teams_req if str(t).strip()]

    seeds_req = body.get("seedPlayerNames") or []
    if not isinstance(seeds_req, list):
        seeds_req = []
    seeds_req = [str(s).strip() for s in seeds_req if str(s).strip()]

    from src.trade.angle import find_angle_packages

    try:
        result = await run_in_threadpool(
            find_angle_packages,
            lens_rows,
            player_names,
            owner_id,
            sleeper_teams,
            min_my_gain_pct=min_my,
            max_market_gain_pct=max_market,
            limit=limit,
            candidate_pool_per_team=pool,
            per_team_limit=per_team,
            positions=positions_req or None,
            min_player_my_value=min_player,
            target_team_owner_ids=target_teams_req or None,
            seed_player_names=seeds_req or None,
            include_idp=include_idp,
            capacity_context=angle_capacity_context,
        )
    except Exception as exc:  # noqa: BLE001
        log.error(f"Angle packages failed: {exc}")
        return JSONResponse(
            status_code=500,
            content={"error": f"Angle packages failed: {exc}"},
        )
    result = {"mode": "offer", **result, "leagueKey": league_cfg.key}
    _stamp_valuation_mode(result, valuation_mode, valuation_note)
    return JSONResponse(content=result)


@app.get("/api/scaffold/validation")
async def get_scaffold_validation():
    ingest_file = _latest_file(DATA_DIR / "validation", "ingest_validation_*.json")
    canonical_file = _latest_file(DATA_DIR / "validation", "canonical_validation_*.json")
    ingest = _load_json_file(ingest_file)
    canonical = _load_json_file(canonical_file)
    return JSONResponse(
        content={
            "ingest_validation_file": str(ingest_file) if ingest_file else None,
            "canonical_validation_file": str(canonical_file) if canonical_file else None,
            "ingest": ingest or {},
            "canonical": canonical or {},
        }
    )


@app.get("/api/scaffold/report")
async def get_scaffold_report():
    file_path = _latest_file(DATA_DIR / "reports", "ops_report_*.md")
    if file_path is None or not file_path.exists():
        return JSONResponse(status_code=404, content={"error": "No scaffold report found"})
    return FileResponse(file_path, media_type="text/markdown")


# ── DRAFT CAPITAL ──────────────────────────────────────────────────────
# Pick dollar values from CSV, rookie rankings from KTC (live) or CSV (fallback).
# Uses a decay curve to fill/extrapolate KTC values to all 72 picks.
DRAFT_DATA_XLSX = Path(__file__).parent / "CSVs" / "Draft Data.xlsx"
DRAFT_DATA_CSV = Path(__file__).parent / "CSVs" / "draft_data.csv"


def _sleeper_league_id_for_draft(league_key: str | None = None) -> str:
    """Resolve the Sleeper league ID for draft endpoints via the
    league registry.  Previously a module-level constant read from
    ``SLEEPER_LEAGUE_ID`` env var; now routes through
    ``league_registry.get_sleeper_league_id()`` which itself falls
    back to the env var when no registry.json is configured.

    If ``league_key`` is provided, that specific league's Sleeper ID
    is returned (after validation).  ``None`` resolves to the
    default league — the existing single-league behavior.

    Returns empty string when no league is configured at all so
    callers can short-circuit instead of making a Sleeper call to
    ``/league/``.
    """
    sid = _league_registry.get_sleeper_league_id(league_key)
    return sid or ""


class LeagueResolutionError(Exception):
    """Raised when a request references a league the server can't
    serve.  Carries an HTTP status + body so route handlers can
    ``except`` once and return a uniform error response.

    Status codes:
      * 400 — ``leagueKey`` is present but unknown or inactive
      * 503 — requested league is valid but no contract is loaded
              for it yet (single-league instance, or scrape in progress)
      * 404 — no leagues configured at all (fresh dev machine)
    """

    def __init__(self, status: int, code: str, message: str):
        self.status = status
        self.code = code
        self.message = message
        super().__init__(message)

    def json_response(self) -> "JSONResponse":
        return JSONResponse(
            status_code=self.status,
            content={"error": self.code, "message": self.message},
        )


def _resolve_league_for_request(
    request: "Request",
    *,
    body: dict | None = None,
    require_loaded_contract: bool = False,
) -> "_league_registry.LeagueConfig":
    """Pick the right league for this request.

    Resolution order:
      1. Explicit ``leagueKey`` in the query string
      2. Explicit ``leagueKey`` in the request body (when provided)
      3. The authenticated user's ``activeLeagueKey`` (from user_kv)
      4. The registry's default league

    Passes 1 + 2 go through ``get_league_by_key`` which also accepts
    aliases.  Passes 3 + 4 always resolve to a canonical key.  An
    inactive league at passes 1–2 raises 400 so a stale frontend
    can't accidentally keep hitting a retired league.

    When ``require_loaded_contract=True`` and the resolved league
    doesn't match the league that built ``latest_contract_data``,
    raises 503 ``data_not_ready``.  This is the guard that keeps
    single-instance deployments from returning garbage for a league
    they haven't scraped yet.
    """
    # 1 + 2: explicit leagueKey in query or body.
    explicit = (request.query_params.get("leagueKey") or "").strip()
    if not explicit and isinstance(body, dict):
        explicit = str(body.get("leagueKey") or "").strip()
    if explicit:
        cfg = _league_registry.get_league_by_key(explicit)
        if cfg is None:
            raise LeagueResolutionError(
                400,
                "unknown_league",
                f"Unknown leagueKey {explicit!r}",
            )
        if not cfg.active:
            raise LeagueResolutionError(
                400,
                "inactive_league",
                f"League {cfg.key!r} is not active",
            )
    else:
        # 3: user's saved preference.
        cfg = None
        session = _get_auth_session(request)
        if session:
            username = str(session.get("username") or "").strip()
            if username:
                try:
                    state = _user_kv.get_user_state(username) or {}
                    saved = (state.get("activeLeagueKey") or "").strip()
                    if saved:
                        candidate = _league_registry.get_league_by_key(saved)
                        if candidate is not None and candidate.active:
                            cfg = candidate
                except Exception:  # noqa: BLE001
                    cfg = None
        # 4: registry default.
        if cfg is None:
            cfg = _league_registry.get_default_league()
        if cfg is None:
            raise LeagueResolutionError(
                404,
                "no_leagues_configured",
                "No leagues configured on this server",
            )

    if require_loaded_contract:
        loaded_key = None
        try:
            loaded_key = (latest_contract_data or {}).get("meta", {}).get("leagueKey")
        except Exception:  # noqa: BLE001
            loaded_key = None
        if loaded_key and loaded_key != cfg.key:
            raise LeagueResolutionError(
                503,
                "data_not_ready",
                f"No data loaded for league {cfg.key!r} yet",
            )
    return cfg


def _requested_valuation_mode(request: "Request", body: dict | None = None) -> str:
    """Which board this request is asking for: ``market`` or
    ``leagueAdjusted``.

    Body first, then query string, so a POST engine and a GET surface
    can ask the same way.  Anything unrecognised reads as ``market``:
    a typo must degrade to the board the server always has, never to an
    error and never to a lens the caller did not name.
    """
    raw = ""
    if isinstance(body, dict):
        raw = str(body.get("valuation_mode") or body.get("valuationMode") or "").strip()
    if not raw:
        raw = (request.query_params.get("valuationMode") or "").strip()
    return "leagueAdjusted" if raw == "leagueAdjusted" else "market"


async def _valuation_scoped_contract(
    request: "Request",
    body: dict | None,
    league_cfg: "_league_registry.LeagueConfig",
    *,
    base: dict | None = None,
) -> tuple[dict | None, str, str | None]:
    """The contract this request should be answered from.

    Returns ``(contract, mode, note)``.  ``mode`` is what the caller
    must stamp on its response — the mode ACTUALLY served, which is not
    always the mode requested.

    WHY THE ENGINES NEEDED THIS.  ``/api/valuation/league-adjusted``
    publishes factors for the *client* to multiply on.  Every trade,
    angle, waiver and terminal engine runs server-side off
    ``latest_contract_data`` and never sees them.  So switching the
    board used to change ``/rankings`` and nothing else: adjusted
    rankings, market-priced trade advice, no way to tell which was
    which.  Handing the engines a pre-adjusted contract fixes all of
    them at once, because every engine reads exactly one value —
    ``rankDerivedValue`` on ``playersArray`` rows.

    DEGRADES, NEVER FAILS.  Scarcity is unmeasurable without a roster
    snapshot, and the adjustment can produce an incoherent board.  In
    both cases the market board is served with a ``note`` naming the
    reason.  Refusing outright would take down working engines to
    protect an optional lens; serving the market board *silently* under
    an adjusted label is the failure this whole path exists to remove.
    Callers stamp ``mode``, so the answer travels with the payload.

    Callers must have resolved the league first.  An adjusted contract
    is valid for exactly one league — ``lineupScarcity`` comes from that
    league's rosters — so this is never shareable across leagues even
    when they share a scoring profile.

    ``base`` overrides the contract the lens is applied to, for the two
    callers that do not answer from ``latest_contract_data`` verbatim:
    /api/terminal and /api/trade/simulate both splice a foreign league's
    Sleeper overlay onto the loaded rankings, and that splice must
    survive.  The factors themselves still come from the loaded
    contract's rows, which are the same rows — only the ``sleeper``
    block differs.
    """
    source = base if base is not None else latest_contract_data

    # WITHDRAWN 2026-08-14 — one canonical methodology, server-owned.
    #
    # This is the single place the league-adjusted lens ever reached an
    # engine, so it is the single place to close it.  The lens was
    # evaluated for promotion to canonical and rejected under the
    # outcome-evidence bar (see
    # ``docs/valuation/LEAGUE_AWARE_METHODOLOGY_REJECTION.md``), and an
    # unvalidated methodology may not answer an engine request.
    #
    # Closing it HERE rather than at the nine call sites is deliberate:
    # every caller keeps its ``valuation_mode`` plumbing and keeps
    # stamping the mode it was actually served, so responses stay
    # self-describing and a future validated methodology has one seam to
    # re-open instead of nine to re-thread.
    #
    # The request is IGNORED, not refused.  A stored ``leagueAdjusted``
    # on someone's phone must converge to the canonical answer silently —
    # refusing would turn an obsolete localStorage value into a broken
    # page for a user who never chose anything.
    requested = _requested_valuation_mode(request, body)
    if requested == "leagueAdjusted":
        return source, "market", "league_adjusted_withdrawn: not_canonical"
    return source, "market", None


def _stamp_valuation_mode(result: Any, mode: str, note: str | None) -> None:
    """Record which board answered, on the response the caller returns.

    Unconditional — including for ``market`` — because "this is the
    market board" and "this field is missing" must not read the same to
    a client deciding what label to show next to a number.
    """
    if not isinstance(result, dict):
        return
    result["valuationMode"] = mode
    if note:
        result["valuationNote"] = note
        warnings = result.get("warnings")
        if isinstance(warnings, list):
            warnings.append(note)
        else:
            result["warnings"] = [note]


def _constraints_for_request(
    request: "Request",
    contract: dict | None,
    league_cfg: Any,
    *,
    surface: str,
) -> Any:
    """Recommendation constraints for one (user, league), or the fail-closed set.

    ONE place, because C3-CON-01 is consumed by every automatically generated
    trade surface and §2.3 forbids page-local copies.  The canonical owner is
    ``src/trade/constraints``; this only resolves its inputs.

    **Read-only plumbing.**  It reads a stored per-(user, league) protection
    block if one exists and never writes one — the storage service and the UI
    that would populate it are ``C3-CON-02`` / ``C3-CON-03``, separate rows.
    Until those land this resolves to "nothing configured" for every user, which
    is a legitimate answer and NOT a failure: §7 acceptance 12 and the owner's
    own ``test_no_configured_preference_is_not_a_failure`` both turn on that
    distinction.

    Fail-closed is reserved for genuinely not knowing.  An anonymous request has
    no protections, which is knowable; a store that RAISES is not, and returns
    ``UNRESOLVED`` so every generator refuses rather than recommending under
    constraints it could not read.
    """
    from src.trade.constraints import UNRESOLVED, resolve_constraints

    league_key = getattr(league_cfg, "key", None)
    try:
        persistent: dict | None = None
        session = _get_auth_session(request)
        if session:
            username = str(session.get("username") or "").strip()
            if username:
                state = _user_kv.get_user_state(username) or {}
                by_league = state.get("tradeConstraintsByLeague") or {}
                if isinstance(by_league, dict) and league_key:
                    block = by_league.get(league_key)
                    if isinstance(block, dict):
                        persistent = block
        return resolve_constraints(contract=contract, persistent=persistent)
    except Exception as exc:  # noqa: BLE001 — unknown must not become unconstrained
        log.warning("recommendation constraints unresolved for %s: %s", surface, exc)
        return UNRESOLVED


def _capacity_context_for(
    contract: dict | None,
    league_cfg: Any,
    team_block: dict | None,
    *,
    surface: str,
) -> Any:
    """Roster-capacity context for one team, or ``None`` with a log line.

    ONE place, because capacity is an ANNOTATION on four different trade
    products (suggestions, finder, angle offer, angle acquire) and a failure to
    compute it must never take any of them down.  The canonical owner is
    ``src/trade/roster_capacity``; this resolves the arguments and swallows
    nothing else.
    """

    if team_block is None:
        return None
    try:
        from src.trade.roster_capacity import build_capacity_context

        return build_capacity_context(
            contract,
            getattr(league_cfg, "key", None),
            team_block,
            roster_settings=dict(getattr(league_cfg, "roster_settings", None) or {}),
        )
    except Exception as exc:  # noqa: BLE001
        log.warning("roster capacity unavailable for %s: %s", surface, exc)
        return None


def _team_block_by_owner_id(sleeper_teams: list, owner_id: str) -> dict | None:
    """The Sleeper team dict for one ``ownerId``, or ``None``."""

    wanted = str(owner_id or "").strip()
    if not wanted:
        return None
    for team in sleeper_teams or []:
        if isinstance(team, dict) and str(team.get("ownerId") or "") == wanted:
            return team
    return None


_KTC_TOTAL_PICKS = 72  # fill rookie data for all 6 rounds (12 teams × 6 rounds)
DRAFT_TOTAL_BUDGET = 1200  # $100 × 12 teams

# Cache for KTC live data: {"rookies": [...], "fetched_at": timestamp}
_ktc_cache = {"rookies": None, "fetched_at": 0}

# /api/draft-capital result cache — see the endpoint for the full
# cache contract (key = league key, TTL 300s, ?refresh=1 busts,
# cleared on scrape promotion because the non-default path prices
# picks off the live contract).
_DRAFT_CAPITAL_CACHE: dict[str, tuple[float, dict]] = {}
_DRAFT_CAPITAL_TTL_SEC = 300.0
_DRAFT_CAPITAL_LOCKS: dict[str, asyncio.Lock] = {}

# Per-pick fields that carry OUR board, not public Sleeper data.  They
# are filled from ``_our_rookie_pool()``, which reads
# ``latest_contract_data['playersArray']`` ordered by
# ``rankDerivedValue`` — the same field the public-league payload guard
# blocklists outright (src/public_league/public_contract.py).  So an
# anonymous caller was receiving the full ordered top-72 proprietary
# rookie board plus per-rookie derived dollars.
_DRAFT_CAPITAL_PRIVATE_PICK_FIELDS = (
    "rookieName",
    "rookiePos",
    "rookieKtcValue",
    "rookieKtcDollar",
    "rookieIdpDollar",
    # The raw board value the dollar ladder is derived from — the single
    # most proprietary number on the payload, and blocklisted outright by
    # the public-league guard.  Added with the field itself, not after.
    "rookieBoardValue",
    # Per-rookie board trust diagnostics (source dispersion, single-source
    # flag).  Board internals, not Sleeper facts, so they redact with the
    # rest of the rookie board rather than leaking how confident our
    # pipeline is in each row.  Added with the fields, not after.
    "rookieDispersionCV",
    "rookieSingleSource",
)


def _redact_draft_capital_for_public(result):
    """Strip the proprietary rookie board from a draft-capital payload.

    ``/api/draft-capital`` is deliberately public: the public /league
    page's draft-capital tab reads it for team names, pick ownership and
    pick dollar values — all of it visible on Sleeper already.  What was
    NOT public-safe is the rookie ranking stapled onto each pick.

    Returns a COPY.  The caller must never mutate the cached object:
    ``_DRAFT_CAPITAL_CACHE`` is shared across sessions, so redacting in
    place would strip the fields from the authenticated /draft page too
    for the rest of the TTL — a cache-poisoning bug with the same shape
    as the leak it fixes.  Only the ``picks`` list is rebuilt; every
    other branch is shared by reference and never written to.

    The only consumer of these fields is the private /draft page
    (frontend/app/draft/page.jsx); the public league section never reads
    them, so redaction costs no functionality.
    """
    if not isinstance(result, dict):
        return result
    picks = result.get("picks")
    if not isinstance(picks, list):
        return result
    redacted = dict(result)
    redacted["picks"] = [
        {k: v for k, v in pick.items() if k not in _DRAFT_CAPITAL_PRIVATE_PICK_FIELDS}
        if isinstance(pick, dict)
        else pick
        for pick in picks
    ]
    redacted["rookieBoardRedacted"] = True
    return redacted


_KTC_CACHE_TTL = 6 * 3600  # 6 hours


import re  # noqa: E402 — KTC import block sits mid-file with its route group


def _ktc_decay_curve(known_rookies, total_picks=72):
    """Extend rookie KTC values to `total_picks` using an exponential decay curve.

    Fits an exponential decay  value = A * e^(-k * pick)  to the known data points,
    then extrapolates for any missing picks beyond what KTC provides.
    If fewer than `total_picks` rookies exist from KTC, synthetic entries are
    generated with the curve values and placeholder names.
    """
    if not known_rookies:
        return known_rookies

    # Already have enough rookies
    if len(known_rookies) >= total_picks:
        return known_rookies[:total_picks]

    # Fit exponential decay: ln(value) = ln(A) - k * pick
    # Use first and last known data points for a robust fit
    v1 = known_rookies[0]["value"]
    vn = known_rookies[-1]["value"]
    n = len(known_rookies)

    if v1 <= 0 or vn <= 0 or n < 2:
        return known_rookies

    # k = (ln(v1) - ln(vn)) / (n - 1)
    k = (math.log(v1) - math.log(vn)) / (n - 1)
    A = v1 * math.exp(k)  # A = v1 / e^(-k*0) adjusted so pick index 0 → v1

    extended = list(known_rookies)
    for i in range(n, total_picks):
        projected_value = max(1, int(round(A * math.exp(-k * i))))
        extended.append(
            {
                "name": f"Rookie #{i + 1}",
                "pos": "—",
                "value": projected_value,
            }
        )
    return extended


def _fetch_ktc_rookies_live():
    """Try to scrape KTC rookie rankings from keeptradecut.com.

    Parses the HTML for player entries with class 'onePlayer'.
    Returns list of {"name", "pos", "value"} or None on failure.
    """
    import html.parser

    url = "https://keeptradecut.com/dynasty-rankings/rookie-rankings"
    headers = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36",
        "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
        "Accept-Language": "en-US,en;q=0.5",
    }

    try:
        req = urllib.request.Request(url, headers=headers)
        resp = urllib.request.urlopen(req, timeout=15)
        raw = resp.read()
        charset = resp.headers.get_content_charset() or "utf-8"
        page_html = raw.decode(charset, errors="replace")
    except Exception as e:
        logging.info(f"KTC live fetch failed: {e}")
        return None

    # Parse player data from HTML — KTC uses divs with class "onePlayer"
    # Each player has: .player-name (a tag text), .position, .value
    rookies = []

    class KTCParser(html.parser.HTMLParser):
        def __init__(self):
            super().__init__()
            self._in_player = False
            self._in_name = False
            self._in_pos = False
            self._in_value = False
            self._current = {}

        def handle_starttag(self, tag, attrs):
            cls = dict(attrs).get("class", "")
            if "onePlayer" in cls:
                self._in_player = True
                self._current = {}
            elif self._in_player:
                if "player-name" in cls:
                    self._in_name = True
                elif cls.strip() == "position":
                    self._in_pos = True
                elif cls.strip() == "value":
                    self._in_value = True

        def handle_data(self, data):
            text = data.strip()
            if not text:
                return
            if self._in_name:
                self._current["name"] = self._current.get("name", "") + text
            elif self._in_pos:
                self._current["pos"] = text
            elif self._in_value:
                self._current["value_str"] = text

        def handle_endtag(self, tag):
            if self._in_name and tag == "a":
                self._in_name = False
            elif self._in_pos and tag in ("span", "div", "p"):
                self._in_pos = False
            elif self._in_value and tag in ("span", "div", "p"):
                self._in_value = False
            elif self._in_player and tag == "div":
                # Try to finalize this player
                name = self._current.get("name", "").strip()
                pos = self._current.get("pos", "").strip()
                val_str = self._current.get("value_str", "").strip().replace(",", "")
                if name and val_str:
                    # Clean team suffix from name (e.g. "Player NAMEnyj" -> "Player NAME")
                    # KTC appends 2-3 letter team codes or "FA"/"RFA"/"R" suffix
                    clean_name = re.sub(r"\s+(FA|RFA|R|[A-Z]{2,3})$", "", name)
                    try:
                        value = int(val_str)
                        if value > 0:
                            # Filter to fantasy-relevant positions
                            pos_upper = pos.upper()
                            if any(p in pos_upper for p in ("QB", "RB", "WR", "TE")):
                                rookies.append(
                                    {"name": clean_name or name, "pos": pos, "value": value}
                                )
                    except ValueError:
                        pass
                self._in_player = False

    try:
        parser = KTCParser()
        parser.feed(page_html)
    except Exception as e:
        logging.warning(f"KTC HTML parse failed: {e}")
        return None

    if len(rookies) < 5:
        logging.info(f"KTC parse returned only {len(rookies)} rookies, likely blocked")
        return None

    # Sort by value descending (should already be, but ensure)
    rookies.sort(key=lambda r: -r["value"])
    logging.info(
        f"KTC live: fetched {len(rookies)} rookies (top: {rookies[0]['name']} = {rookies[0]['value']})"
    )
    return rookies


def _get_ktc_rookies():
    """Get KTC rookie rankings: try live fetch (cached 6h), fall back to CSV."""
    now = time.time()

    # Return cache if fresh
    if _ktc_cache["rookies"] is not None and (now - _ktc_cache["fetched_at"]) < _KTC_CACHE_TTL:
        return _ktc_cache["rookies"]

    # Try live fetch
    live = _fetch_ktc_rookies_live()
    if live:
        _ktc_cache["rookies"] = live
        _ktc_cache["fetched_at"] = now
        return live

    # Fall back to CSV
    csv_rookies = _parse_csv_rookies()
    if csv_rookies:
        logging.info(f"Using CSV fallback: {len(csv_rookies)} rookies")
        return csv_rookies

    return []


def _our_rookie_pool(top_n: int = 72) -> list[dict]:
    """Return our top-N rookies from the live canonical contract.

    Reads ``latest_contract_data['playersArray']`` (the post-build view
    that carries ``rankDerivedValue`` — the rank-curve-smoothed,
    market-corridor-clamped value the rest of the app uses).  Filters
    to ``rookie=True`` AND has a Sleeper ``playerId`` so college /
    undrafted KTC entries (e.g. Trinidad Chambliss) get dropped.
    Sorted by ``rankDerivedValue`` descending — matches the rankings
    page and user-facing rookie ordering exactly.

    Returns a list of ``{name, pos, value, ktcRaw, idpRaw}`` dicts.
    ``value`` is the board's blended raw value (not yet on the $1200
    scale; conversion happens in ``_rookie_dollars_from_values``).
    ``ktcRaw`` / ``idpRaw`` are the per-vendor raw values used to
    derive each rookie's KTC and IDPTradeCalc dollar equivalents on
    the same $1200 scale (see ``_vendor_dollars_for_rookies``).

    ``dispersionCV`` / ``singleSource`` are trust diagnostics carried
    for Perfect Draft's confidence bootstrap, which needs a per-player
    sigma and was previously falling back to one flat constant for the
    whole board.  Read the CV's ZERO carefully: the scraper's
    ``_coeff_var`` returns 0.0 whenever it has fewer than two
    comparable site values, so ``0.0`` means *dispersion unobserved*,
    not *perfect agreement*.  Measured on the 2026-08-04 board that is
    31 of the top 72 rookies, and they are exactly the thinnest-covered
    rows — handing them a literal sigma of zero would present the least
    trustworthy values as the most certain.  The client treats
    non-positive as unobserved for that reason.
    """
    contract = latest_contract_data or {}
    if not isinstance(contract, dict):
        return []
    pa = contract.get("playersArray")
    if not isinstance(pa, list) or not pa:
        return []

    # Selection lives in src/draft/rookie_pool.py because the Perfect Draft
    # roster context needs the SAME set for the opposite purpose — to exclude
    # it from the free-agent pool.  Two copies of "who is in this auction"
    # would eventually disagree, and the failure would be silent.
    from src.draft.rookie_pool import select_rookie_rows  # noqa: PLC0415

    out: list[dict] = []
    for p in select_rookie_rows(contract, top_n):
        val = p.get("rankDerivedValue")
        if val is None:
            vals = p.get("values") or {}
            val = vals.get("overall") if isinstance(vals, dict) else None
        csv = p.get("canonicalSiteValues") or {}
        if not isinstance(csv, dict):
            csv = {}
        ktc_raw = csv.get("ktcSfTep") or csv.get("ktc")
        idp_raw = csv.get("idpTradeCalc")
        cv = p.get("marketDispersionCV")
        out.append(
            {
                "name": str(p.get("canonicalName") or p.get("displayName") or ""),
                "pos": (str(p.get("position") or "").upper() or None),
                "value": float(val),
                "ktcRaw": float(ktc_raw)
                if isinstance(ktc_raw, (int, float)) and ktc_raw > 0
                else None,
                "idpRaw": float(idp_raw)
                if isinstance(idp_raw, (int, float)) and idp_raw > 0
                else None,
                "assetClass": p.get("assetClass"),
                "dispersionCV": float(cv) if isinstance(cv, (int, float)) and cv > 0 else None,
                "singleSource": bool(p.get("isSingleSource")),
            }
        )
    return out


def _vendor_dollars_for_rookies(
    rookies: list[dict],
    total: int = DRAFT_TOTAL_BUDGET,
) -> tuple[dict[str, float], dict[str, float]]:
    """Derive per-rookie KTC and IDPTradeCalc dollar values on the same
    $1200 scale as our board, so ``nominationCandidates`` and
    ``bestValueOnBoard`` can compute apples-to-apples vendor-vs-our gaps.

    For each vendor:
      1. Take rookies that have a positive raw value for that vendor.
      2. Sort them by that vendor's raw value desc (the vendor's own
         ranking, not ours).
      3. Run ``_rookie_dollars_from_values`` on those sorted values to
         get the vendor's per-rookie dollar amounts.
      4. Map back: rookie at vendor's rank N gets dollars[N].

    Returns ``(ktc_by_name, idp_by_name)`` — both maps keyed by lowercase
    name (case-insensitive) for robust frontend joining.
    """

    def _dollars_by_vendor(key: str) -> dict[str, float]:
        eligible = [r for r in rookies if r.get(key) and r[key] > 0]
        if not eligible:
            return {}
        eligible.sort(key=lambda r: -r[key])
        # Pad to top_n so the formula stays calibrated against the same
        # 72-row Hill curve the picks use.  Vendors that don't rank a
        # full 72 just leave the tail empty.
        values = [r[key] for r in eligible]
        dollars = _rookie_dollars_from_values(values, total)
        return {r["name"].lower(): d for r, d in zip(eligible, dollars)}

    return _dollars_by_vendor("ktcRaw"), _dollars_by_vendor("idpRaw")


# mtime-keyed cache of the Draft Data workbook.  The workbook was
# re-parsed by openpyxl on EVERY /api/draft-capital request (twice per
# request, in fact — spine read + full parse).  Key: (path, mtime_ns,
# size); invalidation: file replacement (the only way the sheet
# changes); no TTL.  Cached workbooks are never close()d — data_only
# workbooks are fully materialized in memory after load.
_DRAFT_WB_CACHE: dict[str, tuple[tuple, Any]] = {}
_DRAFT_WB_LOCK = threading.Lock()


def _load_draft_workbook():
    """Return the (cached) Draft Data workbook, or None if unavailable."""
    try:
        import openpyxl
    except ImportError:
        return None
    if not DRAFT_DATA_XLSX.exists():
        return None
    try:
        stat = DRAFT_DATA_XLSX.stat()
        key = (str(DRAFT_DATA_XLSX), stat.st_mtime_ns, stat.st_size)
    except OSError:
        return None
    with _DRAFT_WB_LOCK:
        cached = _DRAFT_WB_CACHE.get("wb")
        if cached and cached[0] == key:
            return cached[1]
        try:
            wb = openpyxl.load_workbook(DRAFT_DATA_XLSX, data_only=True)
        except Exception as exc:  # noqa: BLE001
            logging.warning(f"Could not open {DRAFT_DATA_XLSX}: {exc}")
            return None
        _DRAFT_WB_CACHE["wb"] = (key, wb)
        return wb


def _read_sheet_spine_and_floor(n: int = 72) -> tuple[list[float], int]:
    """Return (spine[E2:E_n+1], r5_bonus_total) from the workbook so
    ``_rookie_dollars_from_values`` faithfully matches the sheet's
    formula.  Spine is the per-pick historical anchor (median + mean
    of 2023–2025 L values divided by 2).  R5-bonus total is the $12
    floor injection for picks 5.01–5.12 (== 12 in a 12-team league).

    Falls back to a flat spine and zero R5 bonus if the workbook is
    unreadable so callers don't blow up.
    """
    try:
        wb = _load_draft_workbook()
        if wb is None:
            return [1.0] * n, 0
        ws = wb["Draft Data"]
        spine: list[float] = []
        for r in range(2, 2 + n):
            v = ws.cell(r, 5).value  # E = column 5
            spine.append(float(v) if v is not None else 1.0)
    except Exception:
        return [1.0] * n, 0
    return spine, n  # R5 bonus = +$1 × 12 picks = $12 total floor add


def _rookie_dollars_from_values(
    values: list[float], total: int = DRAFT_TOTAL_BUDGET
) -> list[float]:
    """Convert raw rookie values to dollar amounts using the sheet's
    Hill-curve formula (column C → I → J → K → D in Draft Data.xlsx),
    summing to ``total``.

    Faithful reproduction of the Google Sheet formula:

        decay_rate = clip(0.6 × stdev(B[0..5]) / mean(B[0..5]), 0.03, 0.08)
        weight[i]  = exp(-decay_rate × i) × (1 − 0.2 × exp(−0.12 × i))
        term1[i]   = B[i] × weight[i] / Σ B[j] × weight[j]
        spine[i]   = sheet E[i]  (median+mean of 2023–2025 L values / 2)
        excess[i]  = max(0, spine[i] − 1)
        term2[i]   = excess[i] / Σ excess[j]
        r5_bonus[i] = 1 if 48 ≤ i < 60 else 0     (round-5 floor add)
        floor[i]    = 1 + r5_bonus[i]              (per-row guaranteed)
        C[i]       = floor[i] + (total − Σ floor) × (0.6 × term1[i] + 0.4 × term2[i])
        I[i]       = floor(C[i] × 2)
        J[i]       = C[i] × 2 − I[i]
        K[i]       = I[i] + (1 if rank(J[i]) ≤ total × 2 − Σ I[j])
        L[i]       = K[i] / 2 + (R1 carryover distributed across first M)

    Spine and R5-bonus are draft-slot anchors lifted from the
    workbook — keeping them lets the formula match the sheet exactly
    when fed the sheet's own B values, and gives rookies in the R5-
    equivalent range (positions 49–60) a small kicker the same way
    real R5 picks are kept above R6.
    """
    import math

    n = len(values)
    if n == 0:
        return []
    M = 12

    spine, r5_bonus_count = _read_sheet_spine_and_floor(n)
    if len(spine) < n:
        spine = list(spine) + [1.0] * (n - len(spine))

    head = values[: min(6, n)]
    head_mean = sum(head) / len(head) if head else 0.0
    if head_mean > 0 and len(head) >= 2:
        var = sum((x - head_mean) ** 2 for x in head) / (len(head) - 1)
        head_std = math.sqrt(var)
        decay_rate = max(0.03, min(0.08, 0.6 * head_std / head_mean))
    else:
        decay_rate = 0.05

    weights = [math.exp(-decay_rate * i) * (1 - 0.2 * math.exp(-0.12 * i)) for i in range(n)]
    denom_decay = sum(values[i] * weights[i] for i in range(n)) or 1.0
    term1 = [values[i] * weights[i] / denom_decay for i in range(n)]

    excess = [max(0.0, spine[i] - 1) for i in range(n)]
    denom_spine = sum(excess) or 1.0
    term2 = [e / denom_spine for e in excess]

    # Per-row floor: $1 every row + $1 for round-5-equivalent rows.
    floors_per_row = [1.0 + (1.0 if 48 <= i < 60 else 0.0) for i in range(n)]
    floor_pool = sum(floors_per_row)
    pool = float(total) - floor_pool
    c_vals = [floors_per_row[i] + pool * (0.6 * term1[i] + 0.4 * term2[i]) for i in range(n)]

    # Monotone half-dollar rounding (sheet uses ×2 / /2).
    doubled = [c * 2 for c in c_vals]
    i_vals = [math.floor(x) for x in doubled]
    fracs = [doubled[i] - i_vals[i] for i in range(n)]
    target = total * 2
    # Monotone bonus: top (target − Σ I) fractional values get +1.
    bonus_count = max(0, target - sum(i_vals))
    rank_order = sorted(range(n), key=lambda i: (-fracs[i], i))
    bonus_idx = set(rank_order[:bonus_count])
    monotone_bonus = [1 if i in bonus_idx else 0 for i in range(n)]

    # K column with caps + monotone-non-increasing constraint
    # (sheet K2 has no cap; K14+ caps at round-5 / round-6 floors and
    # forces K[i] ≤ K[i-1] so the dollar curve never re-rises).
    def cap_for(i: int) -> int:
        if 48 <= i < 60:
            return 4  # R5 cap → D ≤ $2
        if 60 <= i < 72:
            return 2  # R6 cap → D ≤ $1
        return 10**9

    k_vals: list[int] = []
    for i in range(n):
        candidate = i_vals[i] + monotone_bonus[i]
        if i == 0:
            k_vals.append(candidate)
        else:
            k_vals.append(min(candidate, cap_for(i), k_vals[i - 1]))

    # R1 carryover: D = (K + carryover) / 2 for the first M rookies.
    # Distributes the residual (target − Σ K) over the first M rows so
    # the totals still add to ``total`` after the K caps shave value
    # off R5/R6.
    carry_total = target - sum(k_vals)
    base_carry = carry_total // M if M else 0
    extra = carry_total % M if M else 0
    dollars = []
    for i in range(n):
        carry = (base_carry + (1 if i < extra else 0)) if i < M else 0
        dollars.append((k_vals[i] + carry) / 2.0)
    return dollars


def _parse_csv_rookies():
    """Parse rookie rankings from the draft data CSV (cols 22-25)."""
    import csv

    if not DRAFT_DATA_CSV.exists():
        return []
    try:
        with open(DRAFT_DATA_CSV, newline="", encoding="utf-8") as f:
            rows = list(csv.reader(f))
    except Exception:
        return []

    rookies = []
    rank_header_idx = None
    for i, row in enumerate(rows):
        if len(row) > 22 and row[22].strip() == "Rank":
            rank_header_idx = i
            break
    if rank_header_idx is not None:
        for row in rows[rank_header_idx + 1 :]:
            if len(row) < 26:
                continue
            rank_str = row[22].strip() if row[22] else ""
            player = row[23].strip() if row[23] else ""
            pos = row[24].strip() if row[24] else ""
            value_str = row[25].strip() if row[25] else ""
            if not rank_str or not player:
                continue
            try:
                int(rank_str)
                value = int(value_str)
            except (ValueError, TypeError):
                continue
            if value > 0:
                rookies.append({"name": player, "pos": pos, "value": value})
    return rookies


def _parse_draft_xlsx():
    """Read the Draft Data workbook (.xlsx) directly for exact decimal
    values.  Returns (pick_dollars, workbook_picks, slot_to_original, wb_team_totals)
    or None if unavailable.

    Cell references (1-indexed Excel columns):
        P2:AA7   — round/pick value grid (raw per-slot values)
        Q45:Q116 — final per-pick dollar values (post-expansion-averaging)
        R45:R116 — final per-pick owners
        O30:R42  — standings (slot → original owner)
        T63:U74  — team totals (authoritative)
    """
    wb = _load_draft_workbook()
    if wb is None:
        return None

    ws = wb["Draft Data"]

    # ── Raw per-slot values from the grid P2:AA7 (pre-expansion) ──
    pick_dollars: list[float] = []
    for row in range(2, 8):
        for col in range(16, 28):  # P=16 .. AA=27
            v = ws.cell(row, col).value
            pick_dollars.append(float(v) if v is not None else 0.0)

    # ── Final per-pick dollar values L2:L73 ──
    # The "Final Dollar Per Pick" column is the sheet's curve-smoothed,
    # monotone-rounded, half-dollar-precision value per pick (sums to
    # exactly $1200 across 72 picks).  It is NOT expansion-pair-averaged
    # like Q45:Q116, so R5 picks read as $1.50 instead of being floored
    # to $1 by downstream budget-balancing.  This is the user's chosen
    # source of truth for displayed pick values.
    pick_values_l: list[float] = []
    for row in range(2, 74):
        v = ws.cell(row, 12).value  # L = column 12
        pick_values_l.append(float(v) if v is not None else 0.0)

    # ── Final pick assignments Q45:R116 ──
    workbook_picks: list[dict] = []
    for row in range(45, 117):
        rnd = ws.cell(row, 15).value  # O
        pk = ws.cell(row, 16).value  # P
        val = ws.cell(row, 17).value  # Q
        own = ws.cell(row, 18).value  # R
        if rnd is None or pk is None or val is None or own is None:
            continue
        workbook_picks.append(
            {
                "round": int(rnd),
                "pick": int(pk),
                "value": float(val),
                "owner": str(own).strip(),
            }
        )

    # ── Standings O30:R42 — slot → original owner ──
    slot_to_original_owner: dict[int, str] = {}
    for row in range(30, 43):
        owner = ws.cell(row, 16).value  # P = Owner
        slot = ws.cell(row, 18).value  # R = Pick #
        if owner and slot is not None:
            try:
                slot_to_original_owner[int(slot)] = str(owner).strip()
            except (ValueError, TypeError):
                pass

    # ── Team totals T63:U74 ──
    workbook_team_totals: dict[str, float] = {}
    for row in range(63, 75):
        team = ws.cell(row, 20).value  # T
        val = ws.cell(row, 21).value  # U
        if team and val is not None:
            workbook_team_totals[str(team).strip()] = float(val)

    # NOTE: no wb.close() — the workbook is shared via _DRAFT_WB_CACHE.
    return pick_dollars, workbook_picks, slot_to_original_owner, workbook_team_totals, pick_values_l


def _parse_draft_csv_fallback():
    """Parse the draft data CSV (legacy fallback when .xlsx is unavailable).
    Returns (pick_dollars, workbook_picks, slot_to_original, wb_totals) or None.
    """
    import csv

    if not DRAFT_DATA_CSV.exists():
        return None
    try:
        with open(DRAFT_DATA_CSV, newline="", encoding="utf-8") as f:
            rows = list(csv.reader(f))
    except Exception:
        return None

    pick_dollars: list[float] = []
    for row in rows[1:]:
        if len(row) < 12:
            continue
        pick_str = row[0].strip() if row[0] else ""
        val_str = row[11].strip() if row[11] else ""
        if not pick_str or not val_str:
            break
        try:
            pick_dollars.append(float(val_str))
        except (ValueError, TypeError):
            break

    workbook_picks: list[dict] = []
    slot_to_original_owner: dict[int, str] = {}
    in_picks, in_standings = False, False
    for row in rows:
        if len(row) <= 17:
            continue
        c14, c15, c16, c17 = [(row[i].strip() if row[i] else "") for i in (14, 15, 16, 17)]
        if c14 == "Round" and c15 == "Pick" and c16 == "Value" and c17 == "Owner":
            in_picks, in_standings = True, False
            continue
        if c14 == "Standings":
            in_standings, in_picks = True, False
            continue
        if in_standings and c14 and c15 and c17:
            try:
                slot_to_original_owner[int(c17)] = c15
            except (ValueError, TypeError):
                pass
            continue
        if in_picks:
            if not c14:
                in_picks = False
                continue
            try:
                rnd, pk, val = int(c14), int(c15), float(c16)
                if rnd >= 1 and pk >= 1 and c17:
                    workbook_picks.append({"round": rnd, "pick": pk, "value": val, "owner": c17})
            except (ValueError, TypeError):
                continue

    wb_totals: dict[str, float] = {}
    in_team = False
    for row in rows:
        if len(row) <= 20:
            continue
        c19 = row[19].strip() if row[19] else ""
        c20 = row[20].strip() if row[20] else ""
        if c19 == "Team" and c20.startswith("Auction"):
            in_team = True
            continue
        if in_team:
            if not c19 or not c20:
                in_team = False
                continue
            try:
                wb_totals[c19] = float(c20)
            except (ValueError, TypeError):
                in_team = False

    # CSV fallback has no separate "Final Dollar Per Pick" column —
    # use the grid-derived pick_dollars as the L surrogate so the
    # downstream pipeline still gets a length-72 list.
    return pick_dollars, workbook_picks, slot_to_original_owner, wb_totals, list(pick_dollars)


def _parse_draft_data():
    """Read draft capital data from the workbook (.xlsx preferred) or CSV.
    Returns (pick_dollars, workbook_picks, slot_to_original, wb_team_totals,
    rookies, pick_values_l).
    """
    result = _parse_draft_xlsx()
    if result is None:
        result = _parse_draft_csv_fallback()
    if result is None:
        return [], [], {}, {}, [], []

    pick_dollars, workbook_picks, slot_to_original, wb_totals, pick_values_l = result

    rookies = _get_ktc_rookies()
    rookies = _ktc_decay_curve(rookies, _KTC_TOTAL_PICKS)

    return pick_dollars, workbook_picks, slot_to_original, wb_totals, rookies, pick_values_l


def _round_to_budget(values: list[float], budget: int = 1200) -> list[int]:
    """Round a list of floats to integers that sum to exactly *budget*.

    Uses largest-remainder rounding: floor each value, then distribute
    the deficit to the values with the largest fractional parts.
    """
    import math

    floors = [math.floor(v) for v in values]
    remainders = [(v - math.floor(v), i) for i, v in enumerate(values)]
    deficit = budget - sum(floors)
    # Sort by fractional part descending; break ties by index
    remainders.sort(key=lambda x: (-x[0], x[1]))
    for k in range(int(deficit)):
        floors[remainders[k][1]] += 1
    return floors


def _fetch_draft_capital(league_key: str | None = None, *, apply_sleeper_trades: bool = True):
    """Compute draft capital per team.

    The Draft Data workbook is the authoritative source for pick
    VALUES (Q45:Q116) and the slot→original-owner standings
    (O30:R42).  Pick OWNERSHIP is overlaid live from Sleeper's
    ``/traded_picks`` endpoint when ``apply_sleeper_trades=True``
    (default), so trades made on Sleeper between scheduled
    refreshes are reflected immediately in per-team dollar totals.

    If the Sleeper join fails or no trades are reported, the
    workbook's R45:R116 ownership column is used verbatim.

    ``league_key`` selects which league's Sleeper IDs to use for the
    team-name join.  None resolves to the registry default.
    ``apply_sleeper_trades=False`` preserves the legacy workbook-
    only behavior (used by tests that assert sheet-derived totals).
    """
    pick_dollars, workbook_picks, slot_to_original, wb_team_totals, rookies, pick_values_l = (
        _parse_draft_data()
    )
    if not workbook_picks:
        return {"error": "Draft data workbook not found or empty"}

    current_year = datetime.now(timezone.utc).year
    # Default to calendar year; overwritten with the actual draft
    # season once the Sleeper drafts response is read (below).  This
    # distinction matters around Dec→Jan when the league is still on
    # the prior season's rookie draft.
    league_season = current_year
    num_teams = len(slot_to_original) or 12
    draft_rounds = max(1, len(workbook_picks) // num_teams)
    # Per-pick dollar values come from the workbook's L2:L73 column
    # ("Final Dollar Per Pick") — the sheet's curve-smoothed,
    # half-dollar-precision figures that already sum to $1200 across
    # 72 picks.  Using L directly (instead of re-rounding the Q-column
    # expansion-averaged values) preserves the $1.50 separation between
    # round 5 and round 6 picks that the user maintains in the sheet.
    pick_values_per_slot = (
        list(pick_values_l)
        if pick_values_l and len(pick_values_l) >= len(workbook_picks)
        else [wp["value"] for wp in workbook_picks]
    )

    # ── First-name → Sleeper team-name mapping ──
    # Built by joining sheet standings (slot → first name) with
    # Sleeper's draft slot_to_roster_id (slot → roster id → team name).
    first_name_to_team: dict[str, str] = {}
    all_team_names: list[str] = []
    # (round, slot) → new-owner first name from Sleeper traded_picks.
    # Empty when Sleeper is unreachable or apply_sleeper_trades=False.
    sleeper_trade_overrides: dict[tuple[int, int], str] = {}
    # Parallel to ``sleeper_trade_overrides`` but carrying the stable
    # roster_id of the pick's new owner.  ``isTraded`` keys off this
    # instead of round-tripping the new owner's first name back
    # through ``first_name_to_rid`` — that reverse map is ambiguous
    # whenever two rosters share a workbook first name (Codex P1:
    # ``first_name_to_rid.setdefault`` silently keeps only the first
    # roster, so the other roster's traded picks resolve to the
    # wrong id and isTraded flips).
    sleeper_trade_override_rids: dict[tuple[int, int], int] = {}
    # slot → display name for the team originally assigned that slot.
    # Populated inside the try block; overridden by live standings when
    # at least one team has non-zero fppts for the current season.
    roster_fppts: dict[int, float] = {}
    slot_to_origin_display: dict[int, str] = {}
    # ``effective_slot_to_rid``, ``live_standings_active``, and
    # ``first_name_to_rid`` are populated inside the try block
    # alongside the Sleeper joins; initialise them here so the
    # picks loop below (outside the try) never NameErrors when
    # the Sleeper fetch raises early.
    effective_slot_to_rid: dict[int, int] = {}
    live_standings_active: bool = False
    first_name_to_rid: dict[str, int] = {}
    try:
        _league_id_for_draft = _sleeper_league_id_for_draft(league_key)
        if not _league_id_for_draft:
            # No league configured at all — skip the Sleeper joins and
            # leave the mapping empty; downstream code renders draft
            # capital without a team-name column.
            raise RuntimeError("no_sleeper_league_configured")
        # League metadata first — its ``season`` is the canonical
        # anchor for the active-pick-season window.  Filtering trades
        # by /drafts alone breaks during the offseason gap when the
        # league has rolled over but Sleeper hasn't yet created the
        # new season's draft object: in that window /drafts returns
        # only the prior season, while /league/{id}.season already
        # reports the active one and /traded_picks contains the
        # active-season trades that need to apply to the workbook.
        try:
            league_resp = urllib.request.urlopen(
                f"https://api.sleeper.app/v1/league/{_league_id_for_draft}", timeout=15
            )
            league_meta = json.loads(league_resp.read())
            league_meta_season = int(league_meta.get("season"))
        except Exception as exc:  # noqa: BLE001
            logging.warning(f"Sleeper league-meta fetch failed: {exc}")
            league_meta_season = None
        if league_meta_season:
            league_season = league_meta_season

        rosters_resp = urllib.request.urlopen(
            f"https://api.sleeper.app/v1/league/{_league_id_for_draft}/rosters", timeout=15
        )
        rosters = json.loads(rosters_resp.read())

        users_resp = urllib.request.urlopen(
            f"https://api.sleeper.app/v1/league/{_league_id_for_draft}/users", timeout=15
        )
        user_map: dict[str, str] = {}
        for u in json.loads(users_resp.read()):
            uid = u.get("user_id")
            user_map[uid] = (
                u.get("metadata", {}).get("team_name") or u.get("display_name") or f"Team {uid}"
            )

        roster_name_by_id: dict[int, str] = {}
        owner_to_roster_id: dict[str, int] = {}
        for r in rosters:
            rid = r.get("roster_id")
            if rid is None:
                continue
            rid = int(rid)
            oid = r.get("owner_id", "")
            if oid:
                owner_to_roster_id[str(oid)] = rid
            roster_name_by_id[rid] = user_map.get(oid, f"Team {rid}")
        all_team_names = list(roster_name_by_id.values())

        for _r in rosters:
            _rid = _r.get("roster_id")
            if _rid is None:
                continue
            _settings = _r.get("settings") or {}
            try:
                # Sleeper exposes points-for as fpts (integer) +
                # fpts_decimal (0-99, the sub-point decimal part).
                _fpts_int = int(_settings.get("fpts", 0) or 0)
                _fpts_dec = int(_settings.get("fpts_decimal", 0) or 0)
                roster_fppts[int(_rid)] = _fpts_int + _fpts_dec / 100
            except (TypeError, ValueError):
                roster_fppts[int(_rid)] = 0.0

        drafts_resp = urllib.request.urlopen(
            f"https://api.sleeper.app/v1/league/{_league_id_for_draft}/drafts", timeout=15
        )
        # /league/{id}.season is the primary anchor for the active
        # season; /drafts.season is only consulted as a fallback when
        # the league-meta call failed.  This pair handles both
        # boundary cases: the Dec→Jan calendar-year flip AND the
        # offseason window where the league has rolled over but its
        # new draft object hasn't been created yet.
        all_drafts = [d for d in json.loads(drafts_resp.read()) if isinstance(d, dict)]
        draft_seasons: list[int] = []
        for d in all_drafts:
            try:
                draft_seasons.append(int(d.get("season")))
            except (TypeError, ValueError):
                continue
        if not league_meta_season and draft_seasons:
            league_season = max(draft_seasons)
        # Once the current-season rookie draft is complete, the
        # workbook is hand-maintained to track NEXT year's picks
        # (slot dollar values, ownership column, projected standings).
        # Bump ``league_season`` so the display label and the Sleeper
        # ``traded_picks`` filter both target the upcoming draft year.
        # The slot_to_roster join below intentionally keeps using the
        # most-recent COMPLETE draft (slot order from the prior year's
        # final standings) because the workbook's standings carry the
        # same placeholder ordering until the new NFL season finishes.
        current_draft_complete = any(
            isinstance(d, dict)
            and str(d.get("status", "")).lower() == "complete"
            and (str(d.get("season", "")) == str(league_season))
            for d in all_drafts
        )
        slot_lookup_season = league_season
        if current_draft_complete:
            league_season = league_season + 1
        # The slot-to-roster map MUST come from a real draft.  Use the
        # most-recent-completed draft (``slot_lookup_season``) so the
        # first_name → roster_id bridge keeps working into the
        # offseason while we track next-year picks.  This is safe as
        # long as the workbook's standings reflect the same prior-year
        # final order that produced this draft's slot map — once the
        # upcoming NFL season is played, the user updates both the
        # workbook standings AND a new Sleeper draft will exist.
        slot_to_roster: dict[int, int] = {}
        for draft in all_drafts:
            try:
                season = int(draft.get("season"))
            except (TypeError, ValueError):
                continue
            draft_id = draft.get("draft_id")
            if season != slot_lookup_season or not draft_id:
                continue
            try:
                detail_resp = urllib.request.urlopen(
                    f"https://api.sleeper.app/v1/draft/{draft_id}", timeout=15
                )
                draft_detail = json.loads(detail_resp.read())
            except Exception:
                draft_detail = {}
            slot_map = draft_detail.get("slot_to_roster_id") or draft.get("slot_to_roster_id") or {}
            if isinstance(slot_map, dict):
                for slot, rid_val in slot_map.items():
                    try:
                        s, rv = int(slot), int(rid_val)
                    except (TypeError, ValueError):
                        continue
                    if s > 0 and rv in roster_name_by_id:
                        slot_to_roster[s] = rv
            if not slot_to_roster:
                draft_order = draft_detail.get("draft_order") or draft.get("draft_order") or {}
                if isinstance(draft_order, dict):
                    for uid, slot in draft_order.items():
                        rid = owner_to_roster_id.get(str(uid))
                        try:
                            s = int(slot)
                        except (TypeError, ValueError):
                            continue
                        if rid in roster_name_by_id and s > 0:
                            slot_to_roster[s] = rid

        for slot, first_name in slot_to_original.items():
            rid = slot_to_roster.get(int(slot))
            if rid is not None and first_name:
                first_name_to_team[str(first_name).strip()] = roster_name_by_id[rid]

        # ── Effective slot → roster_id mapping ──────────────────────
        # Pre-season we trust the most-recent completed draft's
        # slot→roster_id map (also the workbook's slot ordering).
        # Once any team has non-zero points (= the season has
        # started), we reorder so the team with the fewest points
        # occupies slot 1, second-fewest slot 2, etc.
        #
        # This single mapping is the source of truth for EVERY
        # slot-keyed downstream view:
        #   • slot_to_origin_display  (originalOwner per slot)
        #   • roster_id_to_slot       (used to key trade overrides)
        #   • the per-slot default owner_first inside the picks loop
        #
        # Driving all three off the same dict prevents the desyncs
        # Codex flagged: if standings shift mid-season, an untraded
        # pick must NOT be reported as ``isTraded`` (currentOwner
        # has to follow the same slot reshuffle as originalOwner),
        # and a traded pick must follow the traded roster to its
        # NEW slot — i.e. the override is keyed by where the
        # original roster sits AFTER the standings shuffle, not by
        # its historical workbook slot.
        #
        # Both fppts AND a complete slot_to_roster bridge are
        # required.  ``roster_id_to_first_name`` below is built by
        # joining ``slot_to_original`` against ``slot_to_roster``;
        # if the bridge is empty or only partially populated, any
        # slot whose roster isn't in the join stays unmapped and
        # owner_first falls back to the workbook value while
        # ``slot_to_origin_display`` still flips to live names —
        # those slots' picks then read as ``isTraded: true``
        # spuriously.  Require the bridge to cover at least every
        # workbook slot before activating the override; partial
        # coverage falls back to the workbook ordering the same
        # way a totally-missing bridge does, preserving the
        # invariant that an untraded pick has originalOwner ==
        # currentOwner.
        #
        # Additionally, restrict the live-fpts reshuffle to the
        # roster IDs that are actually in ``slot_to_roster`` —
        # Sleeper sometimes exposes more rosters than the workbook
        # tracks (expansion/inactive rosters), and pulling one of
        # those into ``effective_slot_to_rid`` would leave the
        # affected slot without a first_name bridge and re-trigger
        # the isTraded false-positive on untraded picks at that
        # slot.  Filtering to the bridged set guarantees every
        # slot in ``effective_slot_to_rid`` has a corresponding
        # entry in ``roster_id_to_first_name`` after the join.
        _bridged_rids: set[int] = {int(rid) for rid in slot_to_roster.values()}
        _bridged_fppts: dict[int, float] = {
            int(rid): pts for rid, pts in roster_fppts.items() if int(rid) in _bridged_rids
        }
        # Codex P1: a pure count gate (``len(slot_to_roster) >=
        # len(slot_to_original)``) still activates when Sleeper's
        # ``slot_to_roster_id`` has enough entries but keyed on slots
        # that don't cover the workbook's slot set (e.g. 12 entries
        # keyed 1..10,13,14 against a 1..12 workbook).  The
        # ``slot_to_original`` ↔ ``slot_to_roster`` join below then
        # leaves the uncovered workbook slots' rosters out of
        # ``roster_id_to_first_name``, so originalOwner (live mapping)
        # and currentOwner (stale workbook fallback) disagree and
        # untraded picks at those slots read isTraded=false with
        # originalOwner != currentOwner.  Require the workbook slot
        # set to be a SUBSET of the bridged slot keys — that
        # guarantees every workbook slot joins and the first-name
        # bridge is complete for every roster the reshuffle can place.
        _workbook_slots: set[int] = {int(s) for s in slot_to_original}
        _bridged_slot_keys: set[int] = {int(s) for s in slot_to_roster}
        live_standings_active = (
            any(v > 0 for v in _bridged_fppts.values())
            and bool(slot_to_original)
            and _workbook_slots.issubset(_bridged_slot_keys)
            and len(_bridged_fppts) >= len(slot_to_original)
        )
        effective_slot_to_rid: dict[int, int] = dict(slot_to_roster)
        if live_standings_active:
            _sorted_rids = sorted(_bridged_fppts, key=lambda r: _bridged_fppts[r])
            effective_slot_to_rid = {
                int(i): int(rid)
                for i, rid in enumerate(_sorted_rids[: len(slot_to_original)], start=1)
            }

        # slot → display name, derived from the effective mapping.
        for slot, rid in effective_slot_to_rid.items():
            slot_to_origin_display[int(slot)] = roster_name_by_id.get(rid, f"Team {rid}")
        # Slots in the workbook that the effective mapping didn't
        # cover (e.g. a 14-team workbook joined against a 12-roster
        # Sleeper league) fall back to the workbook's first-name →
        # team-name join.
        for slot, fn in slot_to_original.items():
            slot_to_origin_display.setdefault(
                int(slot),
                first_name_to_team.get(str(fn).strip(), str(fn).strip()),
            )

        # roster_id → slot (inverse of effective_slot_to_rid) and
        # roster_id → workbook first-name.  Both bridges are needed
        # to translate Sleeper traded_picks (keyed by roster_id)
        # into the workbook's (round, slot, first-name) space.
        #
        # roster_id_to_slot follows ``effective_slot_to_rid`` so a
        # traded pick is applied to the slot the original roster
        # occupies AFTER the live-standings reshuffle — Codex P1:
        # otherwise the override mutates the wrong slot and
        # corrupts currentOwner for unrelated picks.
        roster_id_to_slot: dict[int, int] = {
            int(rid): int(slot) for slot, rid in effective_slot_to_rid.items()
        }
        # roster_id_to_first_name stays anchored to the HISTORICAL
        # slot_to_roster: the workbook stamps a first_name per
        # historical slot, and that name is a stable label per
        # roster regardless of where standings move them.
        roster_id_to_first_name: dict[int, str] = {}
        for slot, first_name in slot_to_original.items():
            rid = slot_to_roster.get(int(slot))
            if rid is not None and first_name:
                roster_id_to_first_name[int(rid)] = str(first_name).strip()
        # Inverse: workbook first_name → roster_id.  Used by the
        # picks loop to resolve currentOwner's roster_id (whether
        # from the workbook hand-entry, the live-standings remap,
        # or a Sleeper trade override) so isTraded can be computed
        # against stable identifiers instead of display strings —
        # Sleeper doesn't enforce unique ``display_name``/
        # ``team_name`` across rosters, so a display-only compare
        # silently hides genuine trades between two rosters that
        # happen to share the same rendered name.
        # Collision-aware: if two distinct rosters carry the same
        # workbook first name, the name is NOT a usable identifier —
        # mark it ambiguous (None) so the picks loop falls back to the
        # safe display compare instead of silently attributing one
        # roster's picks to the other (Codex P1: the prior
        # ``setdefault`` kept only the first roster and mis-flagged
        # isTraded for every duplicate-name roster's untraded picks).
        # Common-case owner resolution (live-standings reshuffle and
        # Sleeper trade overrides) no longer touches this map at all —
        # those branches carry the roster_id directly.
        first_name_to_rid: dict[str, int | None] = {}
        for rid, fn in roster_id_to_first_name.items():
            if fn in first_name_to_rid and first_name_to_rid[fn] != rid:
                first_name_to_rid[fn] = None  # ambiguous → unresolvable
            else:
                first_name_to_rid.setdefault(fn, rid)

        if apply_sleeper_trades:
            try:
                traded_resp = urllib.request.urlopen(
                    f"https://api.sleeper.app/v1/league/{_league_id_for_draft}/traded_picks",
                    timeout=15,
                )
                traded = json.loads(traded_resp.read())
            except Exception as exc:  # noqa: BLE001
                logging.warning(f"Sleeper traded_picks fetch failed: {exc}")
                traded = []
            if isinstance(traded, list):
                for t in traded:
                    if not isinstance(t, dict):
                        continue
                    try:
                        season = int(t.get("season"))
                        round_n = int(t.get("round"))
                        original_rid = int(t.get("roster_id"))
                        new_rid = int(t.get("owner_id"))
                    except (TypeError, ValueError):
                        continue
                    if season != league_season:
                        continue
                    original_slot = roster_id_to_slot.get(original_rid)
                    new_first = roster_id_to_first_name.get(new_rid)
                    if original_slot is None or not new_first:
                        continue
                    sleeper_trade_overrides[(round_n, original_slot)] = new_first
                    # Stable id for the isTraded compare — never derived
                    # from the (ambiguous) first name.
                    sleeper_trade_override_rids[(round_n, original_slot)] = new_rid

    except Exception as e:
        logging.warning(f"Sleeper API failed for draft capital team-name mapping: {e}")

    def display(first_name) -> str:
        fn = str(first_name).strip() if first_name else ""
        return first_name_to_team.get(fn, fn) if fn else "Unknown"

    # ── Build pick list + team totals from sheet ownership ──
    all_picks: list[dict] = []
    team_totals_decimal: dict[str, float] = {}

    # Seed every known team at $0 so teams that own no picks still
    # show up in the output (the /draft dashboard relies on this to
    # render the full 12-team roster).  We pre-seed with Sleeper
    # team names ONLY when the first-name → team-name bridge is
    # populated — otherwise ``display()`` will fall back to raw
    # first names for the picks below, and pre-seeding Sleeper
    # names would produce duplicate logical rows (e.g. "Russini
    # Panini" $0 + "Jason" $XX).  In the rollover gap where the
    # bridge is empty, seed from workbook first names instead so
    # the seeded keys match the picks-loop keys.
    if all_team_names and first_name_to_team:
        for t in all_team_names:
            team_totals_decimal.setdefault(t, 0.0)
    else:
        for first_name in slot_to_original.values():
            team_totals_decimal.setdefault(display(first_name), 0.0)

    for overall_idx, wp in enumerate(workbook_picks):
        rnd = wp["round"]
        slot = wp["pick"]
        val = wp["value"]
        owner_first = wp["owner"]
        origin_first = slot_to_original.get(slot, owner_first)
        # When live standings have reshuffled the slot order, the
        # workbook's R45:R116 owner column refers to the slot's
        # PRE-shuffle occupant — keeping it would mis-flag untraded
        # picks as ``isTraded`` (currentOwner ≠ originalOwner)
        # because only originalOwner followed the reshuffle.  Reset
        # owner_first to the roster now occupying this slot under
        # the effective mapping; the Sleeper traded-picks override
        # below still has the final say for actually-traded picks.
        # ``origin_rid`` is the roster the slot's pick belongs to
        # under the effective (post-reshuffle) mapping; an untraded
        # pick's owner is, by definition, that same roster — so seed
        # ``owner_rid`` from it and only move it when a trade actually
        # reassigns the pick.  This keeps the common untraded case
        # off the ambiguous first-name reverse map entirely.
        origin_rid = effective_slot_to_rid.get(slot)
        owner_rid: int | None = origin_rid
        if live_standings_active:
            _slot_rid = effective_slot_to_rid.get(slot)
            if _slot_rid is not None:
                owner_first = roster_id_to_first_name.get(_slot_rid, owner_first)
                owner_rid = _slot_rid
        elif owner_first and origin_first and str(owner_first).strip() != str(origin_first).strip():
            # Workbook-recorded (hand-entered) trade with no live
            # reshuffle: the only path that must resolve owner →
            # roster_id through the name map.  ``first_name_to_rid``
            # is collision-aware (ambiguous names map to None), so a
            # duplicate workbook first name yields owner_rid=None and
            # the safe display compare below — never a silent
            # mis-attribution to the wrong roster.
            owner_rid = first_name_to_rid.get(str(owner_first).strip())
        # Sleeper traded_picks wins over the workbook's R45:R116
        # column when both are available — Sleeper is the system of
        # record for trades, the workbook is hand-edited and lags.
        # The new owner's roster_id comes straight from the Sleeper
        # payload (``sleeper_trade_override_rids``), never re-derived
        # from the new owner's first name.
        sleeper_owner = sleeper_trade_overrides.get((rnd, slot))
        if sleeper_owner:
            owner_first = sleeper_owner
            owner_rid = sleeper_trade_override_rids.get((rnd, slot))

        owner_team = display(owner_first)
        origin_team = slot_to_origin_display.get(slot) or display(origin_first)

        # ``isTraded`` compares stable roster_ids when both sides
        # resolve through the workbook bridge — Sleeper doesn't
        # enforce unique display names across rosters, so a
        # display-string compare silently misses a trade between
        # two rosters that happen to share the same rendered
        # team_name/display_name.  Fall back to the display-name
        # comparison only when an id can't be derived (e.g.
        # pre-season + workbook-only flow where the Sleeper
        # bridge is empty, or an ambiguous duplicate workbook
        # first name), since that's the only correctness-
        # preserving signal available in that path.
        if origin_rid is not None and owner_rid is not None:
            is_traded = origin_rid != owner_rid
        else:
            is_traded = origin_team != owner_team

        # L2:L73 ("Final Dollar Per Pick") is the authoritative per-pick
        # dollar from the sheet — half-dollar precision preserved.
        dollar = (
            float(pick_values_per_slot[overall_idx])
            if overall_idx < len(pick_values_per_slot)
            else float(val)
        )

        all_picks.append(
            {
                "pick": f"{rnd}.{str(slot).zfill(2)}",
                "round": rnd,
                "pickInRound": slot,
                "overallPick": overall_idx + 1,
                "dollarValue": dollar,
                "adjustedDollarValue": dollar,
                # Same source as dollarValue.  The legacy distinction
                # between expansion-averaged Q and unaveraged grid no
                # longer applies — L is unaveraged by construction.
                "originalDollarValue": dollar,
                "originalOwner": origin_team,
                "currentOwner": owner_team,
                "isTraded": is_traded,
                "isExpansion": slot <= 2,
                "rookieName": None,
                "rookiePos": None,
                "rookieKtcValue": None,
            }
        )
        team_totals_decimal.setdefault(owner_team, 0.0)
        team_totals_decimal[owner_team] += float(val)

    # Round team totals to integers summing to exactly 1200, matching the
    # workbook's SUMIF-over-decimals approach.
    team_names = sorted(team_totals_decimal, key=lambda t: -team_totals_decimal[t])
    team_decimal_list = [team_totals_decimal[t] for t in team_names]
    team_int_list = _round_to_budget(team_decimal_list, DRAFT_TOTAL_BUDGET)
    team_totals = {t: v for t, v in zip(team_names, team_int_list)}
    total_budget = sum(team_int_list)

    # Rookie pool — prefer our top-72 from the live contract (filtered
    # to rostered rookies via Sleeper ID), fall back to KTC if the
    # contract isn't loaded yet.  Each rookie's dollar value comes from
    # the sheet's Hill-curve formula applied to OUR values, so the pool
    # totals to $1200 and tail rookies bottom out at $1.
    our_rookies = _our_rookie_pool(_KTC_TOTAL_PICKS)
    rookie_source = "ours_filtered"
    rookie_dollar_overrides: list[float] = []
    ktc_by_name: dict[str, float] = {}
    idp_by_name: dict[str, float] = {}
    if our_rookies:
        raw_values = [r["value"] for r in our_rookies]
        rookie_dollar_overrides = _rookie_dollars_from_values(
            raw_values,
            DRAFT_TOTAL_BUDGET,
        )
        # Vendor $ on the same $1200 scale (KTC sorted by KTC raw,
        # IDPTC sorted by IDPTC raw) so the frontend's gap math is
        # honest dollar-vs-dollar instead of dollar-vs-raw-thousand.
        ktc_by_name, idp_by_name = _vendor_dollars_for_rookies(
            our_rookies,
            DRAFT_TOTAL_BUDGET,
        )
        rookies = [
            {
                "name": r["name"],
                "pos": r["pos"],
                "value": d,
                "ktcDollar": ktc_by_name.get(r["name"].lower()),
                "idpTradeCalcDollar": idp_by_name.get(r["name"].lower()),
                # The board value the dollar ladder was derived FROM.
                # ``_our_rookie_pool`` already carries it; it used to be
                # discarded here.  Perfect Draft needs it because net roster
                # value is measured in rankDerivedValue units against the
                # displaced player, and the $ ladder is not convertible back.
                "boardValue": r["value"],
                # Per-rookie trust, for the confidence bootstrap.  ``None``
                # on the CV means dispersion was unobservable (one site),
                # NOT that the sources agreed — see ``_our_rookie_pool``.
                "dispersionCV": r.get("dispersionCV"),
                "singleSource": r.get("singleSource"),
            }
            for r, d in zip(our_rookies, rookie_dollar_overrides)
        ]
    else:
        rookie_source = "ktc_fallback"

    # Fill rookie rankings into picks (rookie i → pick i; pick i is the
    # i-th overall slot in draft order).  ``rookieKtcValue`` retains its
    # legacy field name for back-compat but now carries our derived
    # dollar value when the contract-sourced pool is in play.  New
    # ``rookieKtcDollar`` / ``rookieIdpDollar`` fields are the per-rookie
    # vendor dollar values used by the "Good to nominate" + "Best value"
    # panels.
    for i, pick in enumerate(all_picks):
        if i < len(rookies):
            pick["rookieName"] = rookies[i]["name"]
            pick["rookiePos"] = rookies[i]["pos"]
            pick["rookieKtcValue"] = rookies[i]["value"]
            if "ktcDollar" in rookies[i]:
                pick["rookieKtcDollar"] = rookies[i]["ktcDollar"]
                pick["rookieIdpDollar"] = rookies[i]["idpTradeCalcDollar"]
            if "boardValue" in rookies[i]:
                pick["rookieBoardValue"] = rookies[i]["boardValue"]
            if "dispersionCV" in rookies[i]:
                pick["rookieDispersionCV"] = rookies[i]["dispersionCV"]
                pick["rookieSingleSource"] = rookies[i]["singleSource"]

    sorted_teams = sorted(team_totals.items(), key=lambda x: -x[1])

    # KTC data source info — only meaningful when we fell back to KTC.
    ktc_source = (
        "live"
        if (
            _ktc_cache["rookies"] is not None
            and (time.time() - _ktc_cache["fetched_at"]) < _KTC_CACHE_TTL
        )
        else "csv"
    )
    ktc_count = len([r for r in rookies if not r["name"].startswith("Rookie #")]) if rookies else 0

    return {
        "picks": all_picks,
        "teamTotals": [{"team": t, "auctionDollars": v} for t, v in sorted_teams],
        "totalBudget": total_budget,
        "numTeams": num_teams,
        "draftRounds": draft_rounds,
        "season": league_season,
        "ktcSource": ktc_source,
        "ktcRookieCount": ktc_count,
        "ktcTotalFilled": len(rookies),
        "rookieSource": rookie_source,
    }


@app.get("/api/draft-capital")
async def get_draft_capital(request: Request, refresh: str = ""):
    """Return draft capital breakdown per team using Sleeper pick ownership
    and the pick value curve from the draft data spreadsheet.

    Accepts ``?leagueKey=...`` to scope the Sleeper roster + users +
    drafts calls to a specific league; absent, falls through to the
    user's saved pref and then the registry default.  Unknown or
    inactive keys return 400.

    The pick-value Excel workbook (``CSVs/Draft Data.xlsx``) is
    wired to the default league's draft — per-team budgets,
    carry-over balances, and standings all reflect that league's
    actual data.  Angle-finder + roster picks still work across
    leagues via the Sleeper overlay; only the workbook-sourced budget
    column is league-specific.

    Non-default leagues take the Sleeper-derived fallback, which needs
    the canonical contract for pick values and so 503s
    ``data_not_ready`` when NO contract is loaded.  A contract belonging
    to a *different* league is still served — that is Defect D-2, an
    open decision, not settled here.

    (This paragraph previously claimed non-default leagues "501 with
    ``not_configured_for_league``".  No such branch has existed since
    the fallback landed; the docstring described a design that was
    replaced and not updated.)

    Pass ``?refresh=1`` to force a fresh KTC fetch."""
    try:
        league_cfg = _resolve_league_for_request(request)
    except LeagueResolutionError as err:
        return err.json_response()
    default_cfg = _league_registry.get_default_league()
    is_default_league = default_cfg and league_cfg.key == default_cfg.key

    # CLAUDE.md lists /api/draft-capital among the endpoints that must
    # 503 ``data_not_ready`` when the loaded contract is not this
    # league's.  It never did, and the failure was not a blank board:
    # the Sleeper-derived path reads pick values via
    # ``_pick_value_from_contract``, which falls through to a HARDCODED
    # table — 7000 / 4000 / 2000 / 1200 by round — when the contract
    # cannot answer.  With no contract loaded the endpoint therefore
    # returned 200 and a full board of invented numbers that a caller
    # cannot distinguish from the Hill-curve-calibrated real ones.
    #
    # Guard only the non-default path: the workbook path reads
    # ``CSVs/Draft Data.xlsx`` and does not consult the contract at all,
    # so 503-ing it on a cold contract would break the one league that
    # never needed it.
    #
    # SCOPE, DELIBERATELY NARROW.  This fires only when there is NO
    # contract, not when the contract belongs to a different league.
    # The mismatch case is Defect D-2 in docs/python-coverage-audit.md —
    # an OPEN product decision between "503 per CLAUDE.md's table" and
    # "keep the cross-league fallback and fix the doc" — and
    # ``tests/api/test_league_isolation_invariants.py`` pins today's
    # behaviour explicitly rather than pre-empting it.  Serving a
    # foreign league its OWN Sleeper rosters is real functionality;
    # taking it away is the operator's call, not a bug fix.
    #
    # No-contract-at-all has no such tension: nobody wants the
    # fabricated values described above, and refusing them removes no
    # capability.
    if not is_default_league and not latest_contract_data:
        return JSONResponse(
            status_code=503,
            content={
                "error": "data_not_ready",
                "message": "No data available yet. First scrape may still be running.",
                "leagueKey": league_cfg.key,
            },
        )

    # ── Result cache ─────────────────────────────────────────────
    # This endpoint used to run ~4s of synchronous work — an openpyxl
    # parse plus up to six blocking Sleeper calls — directly on the
    # event loop, per request, with /draft calling it three times per
    # page load.  Cache: key = league key; TTL = 300s (same freshness
    # class as the Sleeper overlay: pick ownership moves on trades,
    # which the 15-min overlay already bounds — 5 min here is
    # strictly fresher); invalidation = ``?refresh=1`` (which also
    # busts the KTC cache, as before) or TTL expiry; stale data can
    # be shown for ≤5 min, matching the rest of the Sleeper-derived
    # surfaces.  Per-league asyncio single-flight so concurrent
    # misses (the /draft triple-fetch) coalesce onto one build, which
    # itself runs in the threadpool — never on the loop.
    now = time.time()
    # The cache stores the FULL payload, including the proprietary rookie
    # board.  Redaction is applied per-response, on a copy, so an
    # anonymous request can never strip the fields from the cached object
    # that the authenticated /draft page reads next.
    viewer_is_authed = _is_authenticated(request)
    cache_slot = _DRAFT_CAPITAL_CACHE.get(league_cfg.key)
    if not refresh and cache_slot and (now - cache_slot[0]) < _DRAFT_CAPITAL_TTL_SEC:
        cached = cache_slot[1]
        return JSONResponse(
            content=cached if viewer_is_authed else _redact_draft_capital_for_public(cached),
            headers={"Cache-Control": "private, max-age=60, stale-while-revalidate=300"},
        )

    if refresh:
        _ktc_cache["fetched_at"] = 0  # invalidate cache

    def _compute():
        if is_default_league:
            # Workbook path — rich per-pick values pinned to League A's
            # rookie pool.
            return _fetch_draft_capital(league_cfg.key)
        # Sleeper-derived fallback for non-default leagues.
        # Uses the canonical contract's pick values (Hill-curve-
        # calibrated) + Sleeper's traded_picks.  Clearly labeled
        # in the UI so users see "Sleeper-derived" vs.
        # "workbook-calibrated".
        from src.api.draft_capital_fallback import build_sleeper_derived

        # ``draft_rounds`` used to be omitted here, so every non-default league
        # was built as a 4-round draft while the default runs 6 — and the $1200
        # budget is normalized across whatever picks that loop produces, so the
        # wrong count silently redistributed every team's auction dollars.
        # ``num_teams`` is gone: it was declared, never referenced, and shadowed
        # by the roster feed's own count, which is exactly what made the missing
        # parameter next to it look wired.
        # The rookie board follows the SCORING PROFILE, not the league key
        # (CLAUDE.md, "Rankings vs league context"), so a league sharing the
        # loaded contract's profile can be served the same rookies — which is
        # what stops /draft there falling back to a hardcoded list. A league on
        # a different profile gets no rookie fields rather than another
        # profile's board.
        #
        # Profile match is necessary but NOT sufficient. Both live leagues are
        # ``superflex_tep15_ppr1``, yet ``dynasty_main`` starts DL/LB/DB and
        # ``dynasty_new`` starts none — ``idpEnabled`` and ``rosterSettings``
        # are leagueKey properties, not profile properties. Serving the shared
        # board verbatim would put defenders nobody can start onto a non-IDP
        # league's draft board, at real dollar values, ahead of offensive
        # rookies it can. So IDP rows are dropped for a league that does not
        # use them and the dollar ladder is rebuilt over what remains.
        rookie_rows = None
        try:
            loaded_key = ((latest_contract_data or {}).get("meta") or {}).get("leagueKey")
            # Same scoring-identity question as every other cross-league
            # gate, so the same owner answers it (W18-F001).  The rookie
            # pool is a set of VALUES, so serving it for a league whose
            # scoring was never proven to match is the defect in
            # miniature — priced rookies under the wrong rules.
            if loaded_key and _scoring_identity_error(latest_contract_data, league_cfg) is None:
                from src.trade.angle import _IDP_POSITIONS as _ANGLE_IDP_POSITIONS  # noqa: PLC0415

                pool = _our_rookie_pool(_KTC_TOTAL_PICKS)
                if pool and not getattr(league_cfg, "idp_enabled", True):
                    pool = [
                        r
                        for r in pool
                        if str(r.get("assetClass") or "").lower() != "idp"
                        and str(r.get("pos") or "").upper() not in _ANGLE_IDP_POSITIONS
                    ]
                if pool:
                    dollars = _rookie_dollars_from_values(
                        [r["value"] for r in pool], DRAFT_TOTAL_BUDGET
                    )
                    ktc_by_name, idp_by_name = _vendor_dollars_for_rookies(pool, DRAFT_TOTAL_BUDGET)
                    rookie_rows = [
                        {
                            "name": r["name"],
                            "pos": r["pos"],
                            "dollar": d,
                            "boardValue": r["value"],
                            "ktcDollar": ktc_by_name.get(r["name"].lower()),
                            "idpTradeCalcDollar": idp_by_name.get(r["name"].lower()),
                            "dispersionCV": r.get("dispersionCV"),
                            "singleSource": r.get("singleSource"),
                        }
                        for r, d in zip(pool, dollars)
                    ]
        except Exception:  # noqa: BLE001 — a missing rookie board must not 500 the page
            logging.warning("draft-capital fallback: rookie board unavailable", exc_info=True)
            rookie_rows = None

        return build_sleeper_derived(
            league_cfg.sleeper_league_id,
            latest_contract_data or {},
            current_season=datetime.now(timezone.utc).year,
            declared_draft_rounds=(league_cfg.roster_settings or {}).get("draftRounds"),
            rookies=rookie_rows,
        )

    lock = _DRAFT_CAPITAL_LOCKS.get(league_cfg.key)
    if lock is None:
        lock = asyncio.Lock()
        _DRAFT_CAPITAL_LOCKS[league_cfg.key] = lock
    try:
        async with lock:
            # Re-check after the wait — a concurrent request may have
            # populated the slot while we queued.
            cache_slot = _DRAFT_CAPITAL_CACHE.get(league_cfg.key)
            if (
                not refresh
                and cache_slot
                and (time.time() - cache_slot[0]) < _DRAFT_CAPITAL_TTL_SEC
            ):
                result = cache_slot[1]
            else:
                result = await run_in_threadpool(_compute)
                if isinstance(result, dict):
                    result["leagueKey"] = league_cfg.key
                    _DRAFT_CAPITAL_CACHE[league_cfg.key] = (time.time(), result)
        return JSONResponse(
            content=result if viewer_is_authed else _redact_draft_capital_for_public(result),
            headers={"Cache-Control": "private, max-age=60, stale-while-revalidate=300"},
        )
    except Exception as e:
        logging.error(f"Draft capital computation failed: {e}")
        return JSONResponse(
            status_code=500, content={"error": f"Draft capital computation failed: {str(e)}"}
        )


@app.get("/api/draft/roster-context")
async def get_draft_roster_context(
    request: Request,
    ownerId: str = "",
    rosterId: str = "",
    teamName: str = "",
):
    """Roster context the Perfect Draft optimizer runs against.

    Everything here is static for the duration of a draft — rosters, board
    values, waiver levels and the cut ladder do not move when a rookie sells —
    which is why the budget solve itself runs on the client against live
    ``localStorage`` state instead of round-tripping the whole draft workspace
    on every recorded pick.  See ``src/draft/context.py``.

    Called with no team identifier this returns just the team list, so the UI
    can populate its selector in the same request it will later use to fetch a
    context.

    Query parameters::

        leagueKey  optional; falls back to the user's active league, then the
                   registry default
        ownerId    preferred team handle (stable across a team rename)
        rosterId   secondary handle
        teamName   display-name fallback, matched case-insensitively

    Responses::

        200  {teams, context|null, leagueKey}
        400  bad_request      — team identifier supplied but no such team
        503  feature_disabled — perfect_draft flag is off
        503  data_not_ready   — no contract, or none for the requested league
        503  draft_context_unavailable — the build raised
    """
    from src.api import feature_flags as _ff  # noqa: PLC0415

    if not _ff.is_enabled("perfect_draft"):
        return JSONResponse(
            status_code=503,
            content={"error": "feature_disabled", "flag": "perfect_draft"},
        )
    try:
        league_cfg = _resolve_league_for_request(request)
    except LeagueResolutionError as err:
        return err.json_response()

    contract = latest_contract_data
    if not contract:
        return JSONResponse(
            status_code=503,
            content={
                "error": "data_not_ready",
                "message": "No data available yet. First scrape may still be running.",
                "leagueKey": league_cfg.key,
            },
        )
    # The league-match gate is what scopes this feature to the league whose
    # rosters are actually loaded.  Every figure the optimizer needs — open
    # spots, the cut ladder, waiver levels — is derived from THIS league's
    # rosters, and the server holds one league's at a time, so serving another
    # league's would be wrong rather than merely incomplete.
    #
    # This comment used to justify the gate by saying a league served only by
    # the Sleeper-derived draft-capital fallback "has no genuine rookie pool on
    # /draft, and would also fail this check".  The first half stopped being
    # true when ``_serialize_pick`` began stapling the real rookie board onto
    # that path; the gate is unchanged and still correct, because it turns on
    # rosters and never on rookie fields.
    loaded_meta = (contract.get("meta") or {}) if isinstance(contract, dict) else {}
    loaded_league = loaded_meta.get("leagueKey")
    if loaded_league and loaded_league != league_cfg.key:
        return JSONResponse(
            status_code=503,
            content={
                "error": "data_not_ready",
                "message": (
                    f"No rosters loaded for league {league_cfg.key!r} yet "
                    f"(server holds {loaded_league!r})."
                ),
                "leagueKey": league_cfg.key,
            },
        )

    from src.api import draft_optimizer_api as _draft_api  # noqa: PLC0415

    def _compute():
        teams = _draft_api.get_draft_teams(contract)
        ctx = None
        if ownerId or rosterId or teamName:
            ctx = _draft_api.get_roster_context(
                contract,
                league_cfg.key,
                owner_id=ownerId or None,
                roster_id=rosterId or None,
                team_name=teamName or None,
            )
        return {"teams": teams, "context": ctx}

    try:
        payload = await run_in_threadpool(_compute)
    except ValueError as exc:
        # ``unknown_team`` / ``no_rosters_loaded`` — never fall back to some
        # other team's numbers; every figure in this payload is roster-specific.
        code = str(exc) or "bad_request"
        status = 503 if code == "no_rosters_loaded" else 400
        return JSONResponse(
            status_code=status,
            content={"error": code, "leagueKey": league_cfg.key},
        )
    except Exception as exc:  # noqa: BLE001 — surface, never 500-crash the board
        logging.exception("Perfect Draft roster context failed")
        return JSONResponse(
            status_code=503,
            content={
                "error": "draft_context_unavailable",
                "message": str(exc),
                "leagueKey": league_cfg.key,
            },
        )

    payload["leagueKey"] = league_cfg.key
    return JSONResponse(content=payload)


@app.get("/api/sleeper/draft/picks")
async def get_sleeper_draft_picks(
    request: Request,
    afterPickNo: int = 0,
):
    """Live auction-draft picks stream for the /draft companion.

    The /draft page polls this every 2-3 seconds while the user has
    live sync toggled ON.  Each response carries only the picks with
    ``pickNo > afterPickNo`` (the client passes its latest known
    cursor) so steady-state polls return tiny payloads.

    Accepts ``?leagueKey=...`` — picks are league-scoped (rosters/
    draft data), so this follows the CLAUDE.md leagueKey routing
    pattern rather than scoring-profile routing.  Auto-discovers the
    in-progress draft for the league via ``/v1/league/{id}/drafts``;
    no draft id needs to be stored in the registry.

    Response shape::

        {
            "leagueKey":    str,
            "draftId":      str,
            "status":       "drafting" | "pre_draft" | "complete" | "unknown",
            "latestPickNo": int,
            "picks":        [{
                "pickNo":  int,
                "playerId": str,   # Sleeper player_id
                "amount":  int,    # $ paid (0 for snake)
                "ownerId": str,    # Sleeper user_id of winner
                "rosterId": int|null,
                "round":   int|null,
                "pickedAt": int|null,
            }, ...],
            "fetchedAt":    iso-str,
        }

    Error responses (clean JSON, never 500):

    * 400 ``unknown_league`` / ``inactive_league`` — bad ``leagueKey``
    * 404 ``no_active_draft`` — league has no current-season draft
    * 503 ``sleeper_unavailable`` — Sleeper fetch failed (circuit
      breaker open, network error)
    """
    try:
        league_cfg = _resolve_league_for_request(request)
    except LeagueResolutionError as err:
        return err.json_response()

    try:
        snapshot = _sleeper_overlay.fetch_live_draft_picks(
            league_cfg.sleeper_league_id,
            after_pick_no=int(afterPickNo or 0),
        )
    except Exception as e:  # noqa: BLE001 — never 500 a live-poll endpoint
        logging.warning("sleeper draft picks fetch failed: %s", e)
        return JSONResponse(
            status_code=503,
            content={
                "error": "sleeper_unavailable",
                "message": "Sleeper draft fetch failed; try again shortly.",
            },
        )

    # Stamp each pick with the canonical display name so the
    # frontend can match it to a workspace row (which is keyed by
    # ``playerSlug(displayName)``) without re-fetching the full
    # 4MB /api/data contract.  Reads ``latest_contract_data``'s in-
    # process ``playersArray`` — already loaded, zero extra network
    # cost.  When the contract isn't loaded (cold boot), the field
    # is left null and the frontend falls back to surfacing the raw
    # Sleeper player_id so the user can record manually.
    if isinstance(snapshot, dict) and snapshot.get("picks"):
        id_to_name: dict[str, str] = {}
        try:
            pa = (latest_contract_data or {}).get("playersArray") or []
            if isinstance(pa, list):
                for p in pa:
                    if not isinstance(p, dict):
                        continue
                    pid = str(p.get("playerId") or "").strip()
                    name = str(p.get("displayName") or p.get("canonicalName") or "").strip()
                    if pid and name:
                        id_to_name[pid] = name
        except Exception:  # noqa: BLE001
            id_to_name = {}
        for pick in snapshot["picks"]:
            pid = str(pick.get("playerId") or "")
            pick["playerName"] = id_to_name.get(pid) or None

    if snapshot is None:
        # Two distinguishable cases collapse here: no active draft, or
        # a Sleeper fetch failure inside ``fetch_live_draft_picks``.
        # The frontend treats both as "live sync unavailable" — it
        # auto-stops polling and shows an amber dot.  Probing
        # ``/v1/league/{id}/drafts`` again to split the cases would
        # cost a round-trip on every cold poll, which isn't worth it
        # for a diagnostic the user can't act on.
        return JSONResponse(
            status_code=404,
            content={
                "error": "no_active_draft",
                "message": (
                    "No active draft for this league, or Sleeper is unreachable right now."
                ),
                "leagueKey": league_cfg.key,
            },
        )

    snapshot["leagueKey"] = league_cfg.key
    return JSONResponse(content=snapshot)


# ── PUBLIC LEAGUE ROUTES ───────────────────────────────────────────────
# The /api/public/league* endpoints serve the public /league page.
# They are intentionally fork-isolated from the private canonical
# pipeline — no dependence on latest_data / latest_contract_data, no
# private ranking / valuation signals.  The public contract is
# assembled in src/public_league/public_contract.py and runs through
# an allowlist guard before it is serialized.
from src.public_league import (  # noqa: E402 — grouped after route block above
    PUBLIC_SECTION_KEYS,
    build_public_contract,
    build_public_snapshot,
    build_section_payload,
)
from src.public_league.public_contract import (  # noqa: E402 — grouped with public-league block
    assert_public_payload_safe,
    is_private_intelligence_section,
)
from src.public_league.sleeper_client import PUBLIC_MAX_SEASONS  # noqa: E402 — grouped with public-league block
from src.public_league import snapshot_store as public_snapshot_store  # noqa: E402 — grouped with public-league block
from src.public_league import csv_export as public_csv_export  # noqa: E402 — grouped with public-league block
from src.public_league import matchup_recap as public_matchup_recap  # noqa: E402 — grouped with public-league block
from src.public_league import player_journey as public_player_journey  # noqa: E402 — grouped with public-league block

_PUBLIC_LEAGUE_CACHE_TTL_SECONDS = int(os.getenv("PUBLIC_LEAGUE_CACHE_TTL", "300"))
_PUBLIC_LEAGUE_PERSIST = _env_bool("PUBLIC_LEAGUE_PERSIST_SNAPSHOT", True)
_PUBLIC_LEAGUE_WARMUP = _env_bool("PUBLIC_LEAGUE_WARMUP_AT_STARTUP", True)


from src.api.public_activity_valuation import (  # noqa: E402 — grouped with public-league block
    build_asof_valuation as _build_asof_valuation,
)
from src.history import store as _history_store  # noqa: E402 — grouped with public-league block


def _build_public_activity_valuation():
    """Resolver FACTORY for the public activity trade feed.

    Returns a callable matching ``activity._ResolverFactory`` — given
    every ``(asset, instant)`` pair a feed build needs, it resolves each
    against the canonical temporal ledger (``src.history.asof`` — C1-U4)
    AS OF the trade's own instant, never against today's board (V1-97 /
    C3-REPLAY-01: a trade graded with tomorrow's evidence is a hindsight
    leak).  ``build_asof_valuation`` itself does no I/O until the
    returned factory is actually called with a feed's request list, so
    unlike the old contract-parsing builder this needs no per-generation
    memo — returning the plain function reference is already free.

    Returns ``None`` only when the ledger file does not exist at all
    (fresh server, nothing ever recorded) — the public activity feed
    then ships without grade annotations, the same graceful-degradation
    contract as before.  When the ledger exists but simply has no
    observation for a given asset/instant, that is handled per-asset by
    the resolver itself (an honest "insufficient historical evidence"
    side, never a missing feature).
    """
    if not _history_store.DB_PATH.exists():
        return None
    return _build_asof_valuation


# ── Public contract response memo ────────────────────────────────────
# ``GET /api/public/league`` used to reassemble ALL 16 public sections
# + recursively safety-walk + ``json.dumps`` the multi-MB result on
# EVERY request (2.5-4s of TTFB), even though the snapshot it builds
# from is SWR-cached for 5 minutes and the identical contract is
# already assembled — and then discarded — during every snapshot
# rebuild's persist step.
#
# Cache contract:
#   key:  (snapshot.root_league_id, snapshot.generated_at,
#          latest_data_etag) — ``generated_at`` is re-stamped on every
#          snapshot rebuild (natural invalidation), and the private
#          contract etag is in the key because the activity-feed trade
#          grades derive from the private board's generation.
#   value: the encoded response bytes (byte-compatible with
#          ``JSONResponse.render``).
#   TTL:  none (generation-keyed); bound 4 entries.
#   staleness: none beyond the existing 300s snapshot SWR window —
#          identical payload freshness to the uncached path.
_PUBLIC_CONTRACT_BYTES_CACHE: dict = {}
_PUBLIC_CONTRACT_BYTES_LOCK = threading.Lock()
_PUBLIC_CONTRACT_BYTES_MAX = 4


def _public_contract_cache_key(snapshot):
    return (
        getattr(snapshot, "root_league_id", ""),
        getattr(snapshot, "generated_at", ""),
        latest_data_etag,
    )


def _cached_public_contract_bytes(snapshot):
    with _PUBLIC_CONTRACT_BYTES_LOCK:
        return _PUBLIC_CONTRACT_BYTES_CACHE.get(_public_contract_cache_key(snapshot))


def _store_public_contract_bytes(snapshot, contract) -> bytes:
    """Encode ``contract`` exactly like ``JSONResponse.render`` and
    memoize the bytes under the snapshot's generation key."""
    raw = json.dumps(contract, ensure_ascii=False, allow_nan=False, separators=(",", ":")).encode(
        "utf-8"
    )
    key = _public_contract_cache_key(snapshot)
    with _PUBLIC_CONTRACT_BYTES_LOCK:
        if len(_PUBLIC_CONTRACT_BYTES_CACHE) >= _PUBLIC_CONTRACT_BYTES_MAX:
            for k in [k for k in _PUBLIC_CONTRACT_BYTES_CACHE if k != key]:
                del _PUBLIC_CONTRACT_BYTES_CACHE[k]
        _PUBLIC_CONTRACT_BYTES_CACHE[key] = raw
    return raw


_public_league_cache: dict = {
    "snapshot": None,
    "snapshot_league_id": None,
    "fetched_at": 0.0,
    "refreshing": False,
    # Set when an upstream rebuild raises.  Only ``fetched_at`` advances
    # on success, so without a *failure* stamp the post-lock re-check in
    # ``_rebuild_public_snapshot`` can never be satisfied while the
    # vendor is down, and every queued waiter re-attempts in turn.
    "last_failure_at": 0.0,
    "last_failure_error": None,
}
_public_league_refresh_lock = threading.Lock()

# ── Rebuild storm containment ────────────────────────────────────────
# Every ``/api/public/league*`` handler resolves its snapshot inside
# ``run_in_threadpool(_build)``, so a caller waiting on a rebuild holds
# an AnyIO worker token for the whole wait.  Those tokens come from the
# process-wide default limiter shared by every sync endpoint and every
# other ``run_in_threadpool`` call in this file — so enough waiters
# starve unrelated endpoints, ``/api/health`` included.
#
# The lock alone does not bound this.  It stops duplicate *work* on the
# success path (measured: 8 concurrent force-refreshes against a 0.5s
# builder produce 1 upstream call), but on the *failure* path the
# re-check it guards is unsatisfiable, so the same 8 callers produced 8
# serial upstream attempts — 4.01s wall and 18.02 thread-seconds for
# 0.5s of nominal work, growing without limit in N.  That burst arrives
# exactly when the vendor is down and users are mashing refresh.
#
# Two bounds, both required:
#   * a failure cooldown, so a known-down upstream is reported from
#     memory instead of re-probed once per caller;
#   * a wait ceiling, so no caller holds a worker token indefinitely
#     behind an in-flight rebuild.
#
# This is the hazard the ``playoffOdds`` single-flight cache below was
# built for; the snapshot rebuild never got the equivalent treatment.

#: How long a failed rebuild suppresses further upstream attempts.
#: Deliberately short — a cooldown without an expiry turns one vendor
#: blip into a permanently dead endpoint.
_PUBLIC_LEAGUE_FAILURE_COOLDOWN_SECONDS: float = float(
    os.getenv("PUBLIC_LEAGUE_FAILURE_COOLDOWN_SECONDS", "30")
)

#: How long a caller will wait for an in-flight rebuild before giving
#: up.  On timeout it serves the stale snapshot when one exists and
#: raises (-> 503) when none does; either way it releases its worker.
_PUBLIC_LEAGUE_REFRESH_WAIT_SECONDS: float = float(
    os.getenv("PUBLIC_LEAGUE_REFRESH_WAIT_SECONDS", "10")
)


class PublicSnapshotUnavailable(RuntimeError):
    """Raised when a rebuild is refused rather than attempted.

    Handlers already map any exception out of the build closure to a 503,
    so this needs no special casing — it exists so the refusal is
    distinguishable from a genuine upstream error in logs and tests.
    """


# ── Heavy-section single-flight cache ────────────────────────────────
# ``playoffOdds`` ALWAYS runs a 10,000-run, pure-Python (GIL-bound)
# Monte Carlo — it has no precomputed artifact to fall back on (unlike
# ``rosPlayoffOdds`` / ``rosChampionship``, which prefer a file written
# by the scheduled ROS job and only simulate on a cache miss).  Offloaded
# naively, a burst of concurrent ``playoffOdds`` requests would each
# launch an independent simulation and saturate the shared threadpool.
#
# ``archives`` has the same SHAPE for a different reason (added
# 2026-07-30).  It is the single most expensive builder in the contract:
# ``src/public_league/archives.py`` rebuilds four other sections
# (history, activity, draft, awards) before its own five walks, and
# ``assert_public_payload_safe`` then recurses the whole ~800 KB result.
# That is ~1.4s of GIL-bound Python per request, and it ran fresh on
# EVERY request — measured 34.3s / 19.3s / 2.9s TTFB on production
# against a ~0.53s baseline for every other section, because concurrent
# requests each launched their own build and held an AnyIO worker token
# for the duration.  The 2 MB aggregate contract is *faster* than this
# 737 KB subset of it precisely because the aggregate is memoized
# (``_store_public_contract_bytes``) and the section route threw the
# identical work away.
#
# So we single-flight + memoize these two:
#   * Coordination happens on the EVENT LOOP via a per-section
#     ``asyncio.Lock`` (see ``_get_heavy_section_payload``), so waiters
#     ``await`` on the loop instead of occupying AnyIO worker tokens.
#     Exactly one request offloads the simulation to ``run_in_threadpool``;
#     the rest wake to the cached result.
#   * The result is keyed by the snapshot's identity + freshness
#     (``root_league_id`` + ``generated_at``).  Both are derived purely
#     from the snapshot (no external files), so that key is complete — a
#     snapshot refresh mints a new ``generated_at`` and transparently
#     invalidates the entry (bounded to one payload per section).
#     Freshness is therefore unchanged by memoizing: the 300s SWR window
#     on the snapshot still governs how old the data can be.
#
# One asymmetry worth knowing before adding a third key: the JSON route
# passes ``activity_valuation`` and the CSV route does not, while the
# cache key includes neither.  That is safe for both current members —
# ``playoffOdds`` resolves through ``_LAZY_SECTION_BUILDERS`` and
# ``archives`` through ``_SECTION_BUILDERS``, and neither branch of
# ``build_section_payload`` forwards the kwarg (only ``activity`` and
# the aggregate walk do).  A section that DOES consume it must not be
# added here without putting it in the cache key.
#
# The file-backed ROS sections are deliberately NOT cached here: they are
# cheap file reads in the common case, and caching them by snapshot
# identity would hide fresh results the ROS publisher writes between
# snapshot refreshes.  They read their artifact fresh on every request.
_HEAVY_SECTION_KEYS = frozenset({"playoffOdds", "archives"})
_heavy_section_cache: dict = {}
_heavy_section_async_locks: dict = {}


def _heavy_section_async_lock(section: str) -> asyncio.Lock:
    """Return the per-section ``asyncio.Lock``, creating it on first use.

    Only ever touched from async handlers running on the single event-loop
    thread, so the dict access needs no additional synchronization.
    """
    lock = _heavy_section_async_locks.get(section)
    if lock is None:
        lock = asyncio.Lock()
        _heavy_section_async_locks[section] = lock
    return lock


async def _get_heavy_section_payload(snapshot, section, *, activity_valuation=None):
    """Single-flight + per-snapshot memoization for a heavy section,
    coordinated on the event loop.

    Callers must have already resolved ``snapshot`` (its ``generated_at``
    is the cache key).  Waiters block on an ``asyncio.Lock`` on the loop —
    NOT inside the threadpool — so they don't hold worker tokens hostage;
    only the winner offloads ``build_section_payload`` to the threadpool.
    The cached payload is already safety-checked inside
    ``build_section_payload`` and is only ever read by callers, so sharing
    one instance across concurrent requests is safe.
    """
    cache_key = (
        getattr(snapshot, "root_league_id", ""),
        getattr(snapshot, "generated_at", ""),
    )
    cached = _heavy_section_cache.get(section)
    if cached is not None and cached[0] == cache_key:
        return cached[1]
    async with _heavy_section_async_lock(section):
        # Re-check: another coroutine may have computed this section for
        # the same snapshot while we were waiting on the lock.
        cached = _heavy_section_cache.get(section)
        if cached is not None and cached[0] == cache_key:
            return cached[1]
        payload = await run_in_threadpool(
            build_section_payload,
            snapshot,
            section,
            activity_valuation=activity_valuation,
        )
        _heavy_section_cache[section] = (cache_key, payload)
        return payload


# Observability counters for the public-league snapshot cache.  Logged
# at every serve path via ``_log_public_league_event`` so the uptime
# watchdog + log-scraping tooling can track cold-fetch regressions, the
# cache hit ratio, and thundering-herd refresh suppression.
_public_league_metrics: dict = {
    "cache_hit": 0,
    "cache_stale_served": 0,
    "cache_miss_cold_rebuild": 0,
    "force_refresh": 0,
    "background_refresh_started": 0,
    "background_refresh_suppressed": 0,
    "rebuild_count": 0,
    "rebuild_failures": 0,
    # Storm containment.  A rising ``rebuild_cooldown_*`` pair means the
    # upstream is down and the cooldown is absorbing the retry burst
    # rather than forwarding it; a rising ``refresh_wait_timeout_*`` pair
    # means rebuilds are outlasting the wait ceiling and the ceiling —
    # or the rebuild — needs attention.
    "rebuild_cooldown_served_stale": 0,
    "rebuild_cooldown_refused": 0,
    "refresh_wait_timeout_served_stale": 0,
    "refresh_wait_timeout_refused": 0,
    "total_rebuild_seconds": 0.0,
    "last_rebuild_seconds": None,
    "last_rebuild_iso": None,
    "last_contract_bytes": None,
    "last_season_count": None,
    "last_manager_count": None,
}


def _log_public_league_event(event: str, **fields) -> None:
    """Emit a single structured log line for a public_league event.

    Keeps the shape ``public_league_event=<name> key=value ...`` so a
    log shipper can ingest it directly without regex-wrangling.  All
    values are JSON-stringified for safety.
    """
    parts = [f"public_league_event={event}"]
    for key, value in fields.items():
        try:
            rendered = json.dumps(value, default=str)
        except (TypeError, ValueError):
            rendered = json.dumps(str(value))
        parts.append(f"{key}={rendered}")
    logging.info(" ".join(parts))


def _public_league_metrics_snapshot() -> dict:
    """Copy of the metrics dict safe to ship out of the process."""
    snap = dict(_public_league_metrics)
    # Derived fields.
    total = snap["cache_hit"] + snap["cache_stale_served"] + snap["cache_miss_cold_rebuild"]
    snap["total_served"] = total
    snap["cache_hit_ratio"] = round(snap["cache_hit"] / total, 4) if total else None
    snap["avg_rebuild_seconds"] = (
        round(snap["total_rebuild_seconds"] / snap["rebuild_count"], 4)
        if snap["rebuild_count"]
        else None
    )
    return snap


# Best-effort: load the most recent persisted snapshot at process
# start so a cold-started server can still serve the public /league
# page while the first Sleeper rebuild is running in the background.
try:
    _persisted = public_snapshot_store.load_snapshot()
    if _persisted is not None and _persisted.seasons:
        _public_league_cache["snapshot"] = _persisted
        _public_league_cache["snapshot_league_id"] = _persisted.root_league_id
        _public_league_cache["fetched_at"] = 0.0  # forces refresh on next hit
        logging.info(
            "Loaded persisted public_league snapshot for league %s (%d seasons)",
            _persisted.root_league_id,
            len(_persisted.seasons),
        )
except Exception as _exc:  # noqa: BLE001
    logging.warning("Public league snapshot load at startup failed: %s", _exc)


def _is_truthy_flag(value: Any) -> bool:
    """``?flag=1|true|yes|on`` — and nothing else.

    ``bool("0")`` is True, which is why this exists.
    """
    return str(value or "").strip().lower() in {"1", "true", "yes", "on"}


def _authorized_force_refresh(request: Request, refresh: Any) -> bool:
    """May THIS caller force a public-snapshot rebuild? (B8)

    Two defects in one line.  ``force_refresh=bool(refresh)`` treated any
    non-empty string as true, so ``?refresh=0`` — the spelling a caller
    uses to say NO — forced a rebuild.  And nothing checked who was
    asking, so an anonymous caller could skip the cache on every request
    and drive an unbounded number of full snapshot rebuilds, each one an
    O(seasons x weeks) walk plus Sleeper round-trips.  The only remaining
    bound was the rate limiter, whose key is caller-controlled.

    So: parse the flag honestly, and require a session to act on it.
    An anonymous ``?refresh=1`` is IGNORED rather than refused — the
    public page keeps working, it just reads the cache like everyone
    else.  Refusing would break a public URL to fix an abuse surface.

    This is deliberately a REQUEST-level authorization check and not a
    new auth system.  The future ``Sync Sleeper / Refresh League Data``
    action on /league is exactly the authorized caller this admits: it
    will hold a session, so it gets a real refresh through the same
    ``_get_public_snapshot(force_refresh=True)`` path, with the existing
    single-flight and failure-cooldown dedupe still doing their work.
    Nothing here forecloses that feature; it is what makes it safe.
    """
    if not _is_truthy_flag(refresh):
        return False
    return bool(_is_authenticated(request))


def _public_section_access_error(section: str, request: Request) -> JSONResponse | None:
    """``None`` when ``section`` may be served to THIS caller.

    Registering a builder makes a section buildable, not public.  Three
    of them are per-manager decision intelligence (see
    ``public_contract.PRIVATE_INTELLIGENCE_SECTIONS``) and need a
    session; the rest are league-wide facts and stay open.

    One predicate for every representation.  The ``.csv`` route serves
    the same payload through a different door, and a boundary enforced
    on one door is not a boundary.
    """
    if not is_private_intelligence_section(section):
        return None
    if _is_authenticated(request):
        return None
    return JSONResponse(
        status_code=401,
        content={
            "error": "auth_required",
            "message": (
                f"{section!r} contains manager-specific intelligence and " "requires a session."
            ),
            "section": section,
        },
    )


def _public_section_league_error(section: str, league_key: str) -> JSONResponse | None:
    """``None`` when a ``leagueKey`` may be honoured for this request.

    The section routes accepted no ``leagueKey`` at all and always
    resolved the registry default, so a caller asking for a second
    league got the FIRST league's payload, byte for byte, with nothing
    on the response saying so — measured on production for
    ``faabAnalytics``, which ``ManualAddDrop.jsx`` requests with an
    explicit ``?leagueKey=``.

    The public-league product is single-league today, so the honest
    answer to "give me the other league" is a refusal, not a silent
    substitution.  Unknown keys 400 like every other league-aware route;
    a known-but-not-public league is told so explicitly.
    """
    requested = str(league_key or "").strip()
    if not requested:
        return None
    # Alias-aware, and returns None rather than raising on an unknown
    # key — the same resolver every other league-aware route uses.
    cfg = _league_registry.get_league_by_key(requested)
    if cfg is None:
        return JSONResponse(
            status_code=400,
            content={
                "error": "unknown_league",
                "message": f"Unknown leagueKey: {requested!r}.",
            },
        )
    public_id = str(_league_registry.get_sleeper_league_id() or "").strip()
    if public_id and str(cfg.sleeper_league_id).strip() != public_id:
        return JSONResponse(
            status_code=404,
            content={
                "error": "league_not_public",
                "message": (
                    f"League {cfg.key!r} has no public-league snapshot. The "
                    "public league product covers one league; refusing rather "
                    "than serving another league's payload."
                ),
                "leagueKey": cfg.key,
            },
        )
    return None


def _public_league_id(league_key: str | None = None) -> str:
    """Return the public-facing Sleeper league id.

    Routes through the league registry (``league_registry``) so that
    the ID comes from ``config/leagues/registry.json`` when present
    and from ``SLEEPER_LEAGUE_ID`` env var as a back-compat fallback.
    Returns empty string when no league is configured — callers that
    immediately hit Sleeper should treat that as "no snapshot
    available" rather than calling ``/league/`` with an empty path.

    ``league_key`` (B8) resolves a SPECIFIC league instead of the
    default, alias-aware, empty string when unknown.  The public-league
    routes took no league at all and always resolved the default, which
    is how a request for one league was answered with another league's
    payload; making the identity nameable here is what lets the callers
    above refuse rather than silently substitute.
    """
    if league_key:
        cfg = _league_registry.get_league_by_key(league_key)
        return str(getattr(cfg, "sleeper_league_id", "") or "").strip()
    sid = _league_registry.get_sleeper_league_id()
    return (sid or "").strip()


def _recent_rebuild_failure(league_id: str) -> str | None:
    """Return the recorded error if a rebuild failed inside the cooldown.

    Split out of ``_rebuild_public_snapshot`` so both the pre-lock fast
    path and the post-lock re-check consult exactly one predicate.
    """
    last_failure_at = float(_public_league_cache.get("last_failure_at") or 0.0)
    if not last_failure_at:
        return None
    if (time.time() - last_failure_at) >= _PUBLIC_LEAGUE_FAILURE_COOLDOWN_SECONDS:
        return None
    if _public_league_cache.get("snapshot_league_id") not in (None, league_id):
        # A different league's outage says nothing about this one.
        return None
    return str(_public_league_cache.get("last_failure_error") or "upstream unavailable")


def _rebuild_public_snapshot(league_id: str, *, trigger: str = "sync"):
    """Synchronously rebuild the public snapshot for ``league_id``.

    Guarded by ``_public_league_refresh_lock`` so a burst of requests
    while the background refresh is running doesn't multiply work — and,
    on the failure path, by the cooldown and wait ceiling documented at
    ``_PUBLIC_LEAGUE_FAILURE_COOLDOWN_SECONDS``.  The lock alone bounds
    duplicate work only while the upstream is healthy; see that comment
    for the measured storm it does not contain.
    """
    # Fast path, taken WITHOUT the lock: a known-down upstream is
    # reported from memory.  Checking here rather than only after
    # acquiring is the point — a waiter that queues for the lock is
    # holding an AnyIO worker token the whole time.
    failure = _recent_rebuild_failure(league_id)
    if failure is not None:
        cached = _public_league_cache.get("snapshot")
        if cached is not None and _public_league_cache.get("snapshot_league_id") == league_id:
            _public_league_metrics["rebuild_cooldown_served_stale"] += 1
            _log_public_league_event(
                "rebuild_cooldown_served_stale",
                trigger=trigger,
                league_id=league_id,
                error=failure,
            )
            return cached
        _public_league_metrics["rebuild_cooldown_refused"] += 1
        _log_public_league_event(
            "rebuild_cooldown_refused",
            trigger=trigger,
            league_id=league_id,
            error=failure,
        )
        raise PublicSnapshotUnavailable(f"public league upstream unavailable: {failure}")

    acquired = _public_league_refresh_lock.acquire(timeout=_PUBLIC_LEAGUE_REFRESH_WAIT_SECONDS)
    if not acquired:
        # Someone else has been rebuilding for longer than we are willing
        # to hold a worker token for.  Serve what we have; refuse if we
        # have nothing.
        cached = _public_league_cache.get("snapshot")
        if cached is not None and _public_league_cache.get("snapshot_league_id") == league_id:
            _public_league_metrics["refresh_wait_timeout_served_stale"] += 1
            _log_public_league_event(
                "refresh_wait_timeout_served_stale",
                trigger=trigger,
                league_id=league_id,
            )
            return cached
        _public_league_metrics["refresh_wait_timeout_refused"] += 1
        _log_public_league_event(
            "refresh_wait_timeout_refused",
            trigger=trigger,
            league_id=league_id,
        )
        raise PublicSnapshotUnavailable(
            "public league snapshot rebuild is already running; try again shortly"
        )

    try:
        now = time.time()
        cached = _public_league_cache.get("snapshot")
        cached_id = _public_league_cache.get("snapshot_league_id")
        fetched_at = float(_public_league_cache.get("fetched_at") or 0.0)
        # If another thread just refreshed while we were waiting on the
        # lock, reuse that work.
        if (
            cached is not None
            and cached_id == league_id
            and (now - fetched_at) < _PUBLIC_LEAGUE_CACHE_TTL_SECONDS
        ):
            _log_public_league_event(
                "refresh_deduped",
                trigger=trigger,
                league_id=league_id,
            )
            return cached
        # Re-check under the lock: the thread we queued behind may have
        # just discovered the upstream is down.  Without this, the whole
        # queue still attempts serially — the original defect.
        failure = _recent_rebuild_failure(league_id)
        if failure is not None:
            if cached is not None and cached_id == league_id:
                _public_league_metrics["rebuild_cooldown_served_stale"] += 1
                return cached
            _public_league_metrics["rebuild_cooldown_refused"] += 1
            raise PublicSnapshotUnavailable(f"public league upstream unavailable: {failure}")
        started = time.time()
        snapshot = None
        try:
            snapshot = build_public_snapshot(league_id, max_seasons=PUBLIC_MAX_SEASONS)
        except Exception as exc:  # noqa: BLE001
            _public_league_metrics["rebuild_failures"] += 1
            _public_league_cache["last_failure_at"] = time.time()
            _public_league_cache["last_failure_error"] = str(exc)
            _log_public_league_event(
                "rebuild_failed",
                trigger=trigger,
                league_id=league_id,
                error=str(exc),
            )
            raise
        finally:
            _public_league_cache["refreshing"] = False

        # A zero-season snapshot is a FAILURE, not a result.
        #
        # ``walk_league_chain`` returns ``[]`` on any Sleeper miss
        # (``src/public_league/sleeper_client.py`` — ``if not league:
        # break``, 8s timeout), and ``build_public_snapshot`` then
        # returns a snapshot with no seasons rather than raising. Every
        # line below treats that as success: it clears the failure
        # cooldown, caches the empty snapshot with a fresh
        # ``fetched_at`` so it is served for the full 300s TTL plus
        # stale-while-revalidate, and skips the persist block on
        # ``and snapshot.seasons`` — which leaves
        # ``last_contract_bytes`` frozen at its last healthy value.
        #
        # The result is a total content outage that every signal calls
        # healthy. Observed live on 2026-07-30 at 16:18 UTC, on the
        # deploy restart that shipped this file's own PR:
        # ``/api/public/league`` served 7,403 bytes with
        # ``leagueName: ""``, ``managers: 0``, ``seasonsCovered: []`` —
        # HTTP 200, ``rebuild_failures: 0``, and
        # ``last_contract_bytes: 2005444``. A monitor keyed on payload
        # size (``deploy/grafana/public-league-dashboard.json``) reads
        # 2 MB and green while the public page has nothing on it. Only
        # ``last_season_count``/``last_manager_count`` told the truth,
        # and nothing gated on them.
        #
        # So take the same path a raised exception takes: count the
        # failure, arm the cooldown, and serve the last good snapshot
        # if we have one. If we do not, refuse — a 503 is a true
        # statement about a page we cannot render, and a 200 carrying
        # empty sections is not.
        #
        # This deliberately makes a genuinely season-less league
        # unserveable rather than served-empty. That is the right
        # trade: there is no such league here (every chain in
        # ``config/leagues/registry.json`` has seasons), and a real one
        # would surface loudly on the first request instead of looking
        # like a healthy empty page forever.
        if not snapshot.seasons:
            _public_league_metrics["rebuild_failures"] += 1
            _public_league_metrics["last_season_count"] = 0
            _public_league_metrics["last_manager_count"] = 0
            _public_league_cache["last_failure_at"] = time.time()
            _public_league_cache["last_failure_error"] = (
                "empty snapshot: zero seasons (upstream league chain unreachable)"
            )
            _log_public_league_event(
                "rebuild_failed",
                trigger=trigger,
                league_id=league_id,
                error="empty snapshot: zero seasons",
            )
            if cached is not None and cached_id == league_id:
                _public_league_metrics["rebuild_cooldown_served_stale"] += 1
                return cached
            raise PublicSnapshotUnavailable("public league snapshot came back with zero seasons")

        # A success clears the cooldown, so recovery needs no operator
        # action and no waiting out the remaining window.
        _public_league_cache["last_failure_at"] = 0.0
        _public_league_cache["last_failure_error"] = None

        elapsed = round(time.time() - started, 4)
        _public_league_cache["snapshot"] = snapshot
        _public_league_cache["snapshot_league_id"] = league_id
        _public_league_cache["fetched_at"] = time.time()
        _public_league_metrics["rebuild_count"] += 1
        _public_league_metrics["total_rebuild_seconds"] += elapsed
        _public_league_metrics["last_rebuild_seconds"] = elapsed
        _public_league_metrics["last_rebuild_iso"] = _utc_now_iso()
        _public_league_metrics["last_season_count"] = len(snapshot.seasons)
        _public_league_metrics["last_manager_count"] = len(snapshot.managers.by_owner_id)

        contract_bytes = None
        if _PUBLIC_LEAGUE_PERSIST and snapshot.seasons:
            try:
                contract = build_public_contract(
                    snapshot,
                    activity_valuation=_build_public_activity_valuation(),
                )
                public_snapshot_store.persist_snapshot(snapshot, contract=contract)
                # Seed the response-bytes memo with this build — the
                # contract used to be assembled here and then THROWN
                # AWAY while every request rebuilt it from scratch.
                contract_bytes = len(_store_public_contract_bytes(snapshot, contract))
                _public_league_metrics["last_contract_bytes"] = contract_bytes
            except Exception as exc:  # noqa: BLE001
                logging.warning("Failed to persist public_league snapshot: %s", exc)

        _log_public_league_event(
            "rebuild_complete",
            trigger=trigger,
            league_id=league_id,
            elapsed_seconds=elapsed,
            seasons=len(snapshot.seasons),
            managers=len(snapshot.managers.by_owner_id),
            contract_bytes=contract_bytes,
        )
        return snapshot
    finally:
        _public_league_refresh_lock.release()


def _kick_background_refresh(league_id: str, *, trigger: str = "stale-while-revalidate"):
    """Start a daemon thread that refreshes the public snapshot in the
    background.  No-op if another refresh is already running."""
    if _public_league_cache.get("refreshing"):
        _public_league_metrics["background_refresh_suppressed"] += 1
        return
    _public_league_cache["refreshing"] = True
    _public_league_metrics["background_refresh_started"] += 1
    _log_public_league_event(
        "background_refresh_started",
        trigger=trigger,
        league_id=league_id,
    )

    def _worker():
        try:
            _rebuild_public_snapshot(league_id, trigger=trigger)
        except Exception as exc:  # noqa: BLE001
            logging.warning("Background public_league refresh failed: %s", exc)
        finally:
            _public_league_cache["refreshing"] = False

    threading.Thread(
        target=_worker,
        name="public-league-warmup",
        daemon=True,
    ).start()


def _get_public_snapshot(force_refresh: bool = False):
    """Return (possibly cached) public snapshot for the current league.

    Stale-while-revalidate behavior: if a cached snapshot exists but
    has passed TTL, we still return the stale payload immediately and
    kick a background refresh.  The NEXT request gets the fresh data.
    First-request latency is therefore bounded by whatever the client
    already has on disk, not by the Sleeper fetch time.

    ``force_refresh`` bypasses this and blocks on a fresh fetch —
    used by the manual ``?refresh=1`` query and the warmup path.
    """
    league_id = _public_league_id()
    now = time.time()
    cached = _public_league_cache.get("snapshot")
    cached_id = _public_league_cache.get("snapshot_league_id")
    fetched_at = float(_public_league_cache.get("fetched_at") or 0.0)
    fresh = (
        cached is not None
        and cached_id == league_id
        and (now - fetched_at) < _PUBLIC_LEAGUE_CACHE_TTL_SECONDS
    )
    if fresh and not force_refresh:
        _public_league_metrics["cache_hit"] += 1
        return cached
    if force_refresh:
        _public_league_metrics["force_refresh"] += 1
        return _rebuild_public_snapshot(league_id, trigger="force-refresh")
    # Stale-but-serveable: return the cached payload and refresh in
    # the background so subsequent requests get fresh data.
    if cached is not None and cached_id == league_id:
        _public_league_metrics["cache_stale_served"] += 1
        _kick_background_refresh(league_id)
        return cached
    # Cold start — block on a sync rebuild.
    _public_league_metrics["cache_miss_cold_rebuild"] += 1
    return _rebuild_public_snapshot(league_id, trigger="cold-start")


def _warmup_public_snapshot():
    """Kick a background snapshot rebuild at startup when no warm cache
    was loaded from disk.  Bounded by the same lock as the request-path
    refresher so the first request still benefits.

    Invoked from the FastAPI ``lifespan`` contextmanager (see the
    ``lifespan`` function earlier in this file); do not register it as
    an ``@app.on_event`` handler — that API is deprecated.
    """
    if not _PUBLIC_LEAGUE_WARMUP:
        return
    league_id = _public_league_id()
    if not league_id:
        return
    cached = _public_league_cache.get("snapshot")
    cached_id = _public_league_cache.get("snapshot_league_id")
    needs_refresh = (
        cached is None
        or cached_id != league_id
        or float(_public_league_cache.get("fetched_at") or 0.0) == 0.0
    )
    if not needs_refresh:
        return
    _kick_background_refresh(league_id, trigger="startup-warmup")


_PUBLIC_LEAGUE_CACHE_CONTROL = (
    f"public, max-age=60, stale-while-revalidate={_PUBLIC_LEAGUE_CACHE_TTL_SECONDS}"
)


@app.get("/api/public/league/metrics")
async def get_public_league_metrics():
    """Small, public-safe observability endpoint for the snapshot cache.

    Exposes the counters that ``_log_public_league_event`` has been
    emitting: cache hit ratio, rebuild wall-clock, contract byte-size,
    last rebuild timestamp.  Useful for the uptime watchdog, for
    external dashboards, and for smoke-testing cold-fetch regressions.

    NOTE: no private data — just aggregated counters for the cache.
    """
    snap = _public_league_metrics_snapshot()
    # Diagnostic: is the valuation pipeline wired up right now?  This
    # only surfaces the boolean — never any private values — and lets
    # us answer "why are no grades showing on /league activity?" by
    # hitting one URL.  ``valuationReady=False`` means the public
    # activity feed will ship without grade badges (no asset value
    # source available), which is the documented graceful degradation
    # path; the page itself does not break.
    valuation_ready = _build_public_activity_valuation() is not None
    return JSONResponse(
        content={
            "leagueId": _public_league_id(),
            "cacheTtlSeconds": _PUBLIC_LEAGUE_CACHE_TTL_SECONDS,
            "warmupEnabled": _PUBLIC_LEAGUE_WARMUP,
            "persistEnabled": _PUBLIC_LEAGUE_PERSIST,
            "tradeGrading": {
                "valuationReady": valuation_ready,
                "privateContractLoaded": latest_contract_data is not None,
                "privateContractPlayerCount": int(
                    (latest_contract_data or {}).get("playerCount") or 0
                ),
            },
            "metrics": snap,
        },
        headers={"Cache-Control": "no-store"},
    )


@app.get("/api/public/league")
async def get_public_league(request: Request, refresh: str = ""):
    """Full public league contract — every section + league header.

    This endpoint is intentionally separate from /api/data.  It never
    reads the private canonical pipeline, never exposes private
    rankings / edge signals, and runs through an allowlist guard
    before serialization.
    """

    def _build():
        # Snapshot fetch (blocking network I/O) AND contract assembly
        # (potentially heavy per-section CPU) both run in the worker so
        # the event loop is never starved — see the ``run_in_threadpool``
        # note on the section endpoint below.
        snapshot = _get_public_snapshot(force_refresh=_authorized_force_refresh(request, refresh))
        # Serve pre-encoded bytes for the current (snapshot, private
        # contract) generation — see _PUBLIC_CONTRACT_BYTES_CACHE.
        # ``?refresh=1`` bypasses the read (still repopulates).
        cached = None if refresh else _cached_public_contract_bytes(snapshot)
        if cached is not None:
            return cached
        payload = build_public_contract(
            snapshot,
            activity_valuation=_build_public_activity_valuation(),
        )
        # NOTE: no assert_public_payload_safe here — build_public_contract
        # runs the full recursive walk internally before returning
        # (public_contract.py); the second walk was pure duplicate cost
        # over a multi-MB tree.
        return _store_public_contract_bytes(snapshot, payload)

    try:
        raw = await run_in_threadpool(_build)
        return Response(
            content=raw,
            media_type="application/json",
            headers={"Cache-Control": _PUBLIC_LEAGUE_CACHE_CONTROL},
        )
    except AssertionError as exc:
        logging.error("Public league contract tripped safety assert: %s", exc)
        return JSONResponse(
            status_code=500,
            content={"error": "Public league contract safety violation."},
        )
    except Exception as exc:  # noqa: BLE001
        logging.error("Public league contract build failed: %s", exc)
        return JSONResponse(
            status_code=503,
            content={"error": f"Public league data unavailable: {exc}"},
        )


@app.get("/api/public/league/matchup/{season}/{week}/{matchup_id}")
async def get_public_league_matchup(
    season: str,
    week: int,
    matchup_id: int,
    request: Request,
    refresh: str = "",
):
    """Per-matchup public recap — full lineups, scoring, pre-week standings.

    ``season`` is the season year string (e.g. ``"2025"``).
    Runs through the same safety allowlist as the rest of the contract.
    """

    def _build():
        snapshot = _get_public_snapshot(force_refresh=_authorized_force_refresh(request, refresh))
        recap = public_matchup_recap.build_matchup_recap(
            snapshot,
            season,
            int(week),
            int(matchup_id),
        )
        if recap is None:
            return None
        payload = {
            "contractVersion": "public-league-matchup/2026-04-17.v1",
            "league": {
                "rootLeagueId": snapshot.root_league_id,
                "currentLeagueId": snapshot.current_season.league_id
                if snapshot.current_season
                else "",
                "leagueName": str((snapshot.current_season.league or {}).get("name") or "")
                if snapshot.current_season
                else "",
                "managers": snapshot.managers.to_public_list(),
                "seasonsCovered": snapshot.season_ids,
                "generatedAt": snapshot.generated_at,
            },
            "matchup": recap,
        }
        assert_public_payload_safe(payload)
        return payload

    try:
        payload = await run_in_threadpool(_build)
        if payload is None:
            return JSONResponse(
                status_code=404,
                content={
                    "error": f"No matchup found at season={season} week={week} matchup_id={matchup_id}",
                },
            )
        return JSONResponse(
            content=payload,
            headers={"Cache-Control": _PUBLIC_LEAGUE_CACHE_CONTROL},
        )
    except AssertionError as exc:
        logging.error("Matchup recap tripped safety assert: %s", exc)
        return JSONResponse(
            status_code=500,
            content={"error": "Public league contract safety violation."},
        )
    except Exception as exc:  # noqa: BLE001
        logging.error("Matchup recap build failed: %s", exc)
        return JSONResponse(
            status_code=503,
            content={"error": f"Matchup recap unavailable: {exc}"},
        )


@app.get("/api/public/league/matchups")
async def list_public_league_matchups(request: Request, refresh: str = ""):
    """Index endpoint — every (season, week, matchup_id) that has a
    scored pair.  Useful for sitemap generation + the index landing."""

    def _build():
        snapshot = _get_public_snapshot(force_refresh=_authorized_force_refresh(request, refresh))
        payload = {
            "seasonsCovered": snapshot.season_ids,
            "matchups": public_matchup_recap.list_matchups(snapshot),
            "generatedAt": snapshot.generated_at,
        }
        assert_public_payload_safe(payload)
        return payload

    try:
        payload = await run_in_threadpool(_build)
        return JSONResponse(
            content=payload,
            headers={"Cache-Control": _PUBLIC_LEAGUE_CACHE_CONTROL},
        )
    except Exception as exc:  # noqa: BLE001
        logging.error("Matchup index failed: %s", exc)
        return JSONResponse(
            status_code=503,
            content={"error": f"Matchup index unavailable: {exc}"},
        )


@app.get("/api/public/league/player/{player_id}")
async def get_public_league_player(player_id: str, request: Request, refresh: str = ""):
    """Public player-journey view: every trade, waiver, weekly starter
    slot, per-manager scoring summary for a given Sleeper player_id."""

    def _build():
        snapshot = _get_public_snapshot(force_refresh=_authorized_force_refresh(request, refresh))
        journey = public_player_journey.build_player_journey(snapshot, player_id)
        if journey is None:
            return None
        payload = {
            "contractVersion": "public-league-player/2026-04-17.v1",
            "league": {
                "rootLeagueId": snapshot.root_league_id,
                "leagueName": str((snapshot.current_season.league or {}).get("name") or "")
                if snapshot.current_season
                else "",
                "managers": snapshot.managers.to_public_list(),
                "seasonsCovered": snapshot.season_ids,
                "generatedAt": snapshot.generated_at,
            },
            "player": journey,
        }
        assert_public_payload_safe(payload)
        return payload

    try:
        payload = await run_in_threadpool(_build)
        if payload is None:
            return JSONResponse(
                status_code=404,
                content={"error": f"No public journey data for player_id={player_id!r}"},
            )
        return JSONResponse(
            content=payload,
            headers={"Cache-Control": _PUBLIC_LEAGUE_CACHE_CONTROL},
        )
    except AssertionError as exc:
        logging.error("Player journey tripped safety assert: %s", exc)
        return JSONResponse(
            status_code=500,
            content={"error": "Public league contract safety violation."},
        )
    except Exception as exc:  # noqa: BLE001
        logging.error("Player journey build failed: %s", exc)
        return JSONResponse(
            status_code=503,
            content={"error": f"Player journey unavailable: {exc}"},
        )


@app.get("/api/public/league/players")
async def list_public_league_players(request: Request, refresh: str = ""):
    """Index endpoint — every player who appears on a roster or in a
    transaction in the 2-season window.  Lightweight so the frontend
    can build a player-autocomplete."""

    def _build():
        snapshot = _get_public_snapshot(force_refresh=_authorized_force_refresh(request, refresh))
        payload = {
            "seasonsCovered": snapshot.season_ids,
            "players": public_player_journey.list_players_with_activity(snapshot),
            "generatedAt": snapshot.generated_at,
        }
        assert_public_payload_safe(payload)
        return payload

    try:
        payload = await run_in_threadpool(_build)
        return JSONResponse(
            content=payload,
            headers={"Cache-Control": _PUBLIC_LEAGUE_CACHE_CONTROL},
        )
    except Exception as exc:  # noqa: BLE001
        logging.error("Players index failed: %s", exc)
        return JSONResponse(
            status_code=503,
            content={"error": f"Players index unavailable: {exc}"},
        )


@app.get("/api/public/league/{section}.csv")
async def get_public_league_section_csv(
    section: str,
    request: Request,
    owner: str = "",
    kind: str = "",
    refresh: str = "",
    leagueKey: str = "",
):
    """CSV download for any public-league section.

    Matches the JSON endpoint at ``/api/public/league/{section}`` but
    serializes the underlying payload as CSV via ``csv_export``.
    Supports the same ``owner`` qualifier for franchise and a ``kind``
    qualifier for archives (``trades|waivers|weeklyMatchups|rookieDrafts|
    seasonResults|managers``).

    The CSV is generated from the same safety-checked JSON payload the
    /api/public/league route serves, so no new leak surface is added.

    Registered BEFORE the generic /{section} handler so FastAPI's path
    matching resolves the ``.csv`` suffix first.
    """
    if section == "hall_of_fame":
        # Hall of Fame is a derived projection of the history section.
        def _build_hof():
            snapshot = _get_public_snapshot(
                force_refresh=_authorized_force_refresh(request, refresh)
            )
            history_payload = build_section_payload(snapshot, "history")
            assert_public_payload_safe(history_payload)
            return public_csv_export.export_hall_of_fame(history_payload["data"])

        try:
            filename, text = await run_in_threadpool(_build_hof)
            return Response(
                content=text,
                media_type="text/csv; charset=utf-8",
                headers={
                    "Content-Disposition": f'attachment; filename="{filename}"',
                    "Cache-Control": _PUBLIC_LEAGUE_CACHE_CONTROL,
                },
            )
        except Exception as exc:  # noqa: BLE001
            logging.error("CSV export hall_of_fame failed: %s", exc)
            return JSONResponse(
                status_code=503,
                content={"error": f"CSV export unavailable: {exc}"},
            )

    if section not in PUBLIC_SECTION_KEYS:
        return JSONResponse(
            status_code=404,
            content={
                "error": f"Unknown public league section: {section!r}",
                "availableSections": list(PUBLIC_SECTION_KEYS) + ["hall_of_fame"],
            },
        )
    _league_err = _public_section_league_error(section, leagueKey)
    if _league_err is not None:
        return _league_err
    _access_err = _public_section_access_error(section, request)
    if _access_err is not None:
        return _access_err

    try:
        snapshot = await run_in_threadpool(
            _get_public_snapshot, force_refresh=_authorized_force_refresh(request, refresh)
        )
        if section in _HEAVY_SECTION_KEYS and not kind and not owner:
            # Reuse the single-flighted / cached payload, then serialize to
            # CSV in the worker.
            #
            # The qualifier check is load-bearing, not defensive padding:
            # this branch never FORWARDS a qualifier, which was invisible
            # while ``playoffOdds`` (which takes none) was the only heavy
            # section.  ``archives`` DOES take ``?kind=`` — and
            # ``public_csv_export.export_section`` falls back to trades for
            # a missing kind, so without this a request for
            # ``archives.csv?kind=waivers`` would return a trades CSV with
            # a 200 and nothing saying a qualifier had been ignored.  A
            # qualified request drops to the uncached branch below, which
            # honours it; that is the rarer path, and correctness beats the
            # memo there.  ``owner`` is included for the same reason even
            # though no heavy section reads it today — the predicate should
            # match the stated rule, or the next section added here
            # reintroduces the bug.
            payload = await _get_heavy_section_payload(snapshot, section)

            def _export():
                assert_public_payload_safe(payload)
                return public_csv_export.export_section(section, payload["data"])

            filename, text = await run_in_threadpool(_export)
        else:

            def _build_csv():
                payload = build_section_payload(snapshot, section)
                assert_public_payload_safe(payload)
                kwargs = {}
                if section == "franchise" and owner:
                    kwargs["owner_id"] = str(owner).strip()
                if section == "archives" and kind:
                    kwargs["kind"] = str(kind).strip()
                return public_csv_export.export_section(section, payload["data"], **kwargs)

            filename, text = await run_in_threadpool(_build_csv)
        return Response(
            content=text,
            media_type="text/csv; charset=utf-8",
            headers={
                "Content-Disposition": f'attachment; filename="{filename}"',
                "Cache-Control": _PUBLIC_LEAGUE_CACHE_CONTROL,
            },
        )
    except AssertionError as exc:
        logging.error("CSV export safety violation in section %s: %s", section, exc)
        return JSONResponse(
            status_code=500,
            content={"error": "Public league contract safety violation."},
        )
    except Exception as exc:  # noqa: BLE001
        logging.error("CSV export for section %s failed: %s", section, exc)
        return JSONResponse(
            status_code=503,
            content={"error": f"CSV export unavailable: {exc}"},
        )


@app.get("/api/public/league/{section}")
async def get_public_league_section(
    section: str,
    request: Request,
    owner: str = "",
    refresh: str = "",
    leagueKey: str = "",
    lens: str = "",
):
    """Single public-league section JSON payload.

    ``section`` must be one of ``PUBLIC_SECTION_KEYS``.  When the
    ``franchise`` section is requested with ``?owner=<owner_id>`` we
    also include a narrowed ``franchiseDetail`` block so the frontend
    can render a single franchise page without downloading every
    franchise's detail dict.

    ``lens`` (V1-52) selects which of the canonical power engine's two
    lenses the ``rosPower`` section answers with —
    ``power_v2.LENS_FORWARD_LOOKING`` (default, matches the
    aggregate-contract behavior) or ``power_v2.LENS_RESULTS_ONLY``.
    Ignored for every other section.  Rejected outright rather than
    silently falling back to the default: a typo'd lens value silently
    answering the wrong question is worse than a 400.

    NOTE: the ``.csv`` variant above MUST remain registered before this
    route — FastAPI otherwise matches ``/{section}`` against
    ``history.csv`` with ``section="history.csv"``.
    """
    if section not in PUBLIC_SECTION_KEYS:
        return JSONResponse(
            status_code=404,
            content={
                "error": f"Unknown public league section: {section!r}",
                "availableSections": list(PUBLIC_SECTION_KEYS),
            },
        )
    if lens and section == "rosPower":
        from src.ros import power_v2  # noqa: PLC0415

        _valid_lenses = (power_v2.LENS_FORWARD_LOOKING, power_v2.LENS_RESULTS_ONLY)
        if lens not in _valid_lenses:
            return JSONResponse(
                status_code=400,
                content={
                    "error": f"Unknown power lens: {lens!r}",
                    "availableLenses": list(_valid_lenses),
                },
            )
    _league_err = _public_section_league_error(section, leagueKey)
    if _league_err is not None:
        return _league_err
    _access_err = _public_section_access_error(section, request)
    if _access_err is not None:
        return _access_err

    try:
        # Snapshot fetch is blocking (network I/O) — offload it.  Its
        # ``generated_at`` is the cache key for the heavy path below.
        snapshot = await run_in_threadpool(
            _get_public_snapshot, force_refresh=_authorized_force_refresh(request, refresh)
        )
        if section in _HEAVY_SECTION_KEYS:
            # playoffOdds: single-flight + memoize, coordinated on the loop
            # so concurrent waiters don't occupy threadpool workers.  Heavy
            # sections are never ``franchise``, so no owner-detail step.
            payload = await _get_heavy_section_payload(
                snapshot,
                section,
                activity_valuation=_build_public_activity_valuation(),
            )
        else:
            # Every other section still runs its build in the worker so a
            # heavier-than-expected builder can't block the event loop.
            def _build():
                payload = build_section_payload(
                    snapshot,
                    section,
                    activity_valuation=_build_public_activity_valuation(),
                )
                if section == "franchise" and owner:
                    detail_map = payload.get("data", {}).get("detail") or {}
                    payload["franchiseDetail"] = detail_map.get(str(owner).strip())
                if section == "rosPower" and lens:
                    # ``build_section_payload`` above already ran the
                    # default (forward-looking) lens; only recompute when
                    # a non-default lens was explicitly requested, so the
                    # common case pays no extra cost.
                    from src.ros import power_v2  # noqa: PLC0415

                    if lens != power_v2.LENS_FORWARD_LOOKING:
                        payload["data"] = power_v2.build_section(snapshot, lens=lens)
                assert_public_payload_safe(payload)
                return payload

            payload = await run_in_threadpool(_build)
        return JSONResponse(
            content=payload,
            headers={"Cache-Control": _PUBLIC_LEAGUE_CACHE_CONTROL},
        )
    except AssertionError as exc:
        logging.error("Public section %s tripped safety assert: %s", section, exc)
        return JSONResponse(
            status_code=500,
            content={"error": "Public league contract safety violation."},
        )
    except Exception as exc:  # noqa: BLE001
        logging.error("Public league section %s failed: %s", section, exc)
        return JSONResponse(
            status_code=503,
            content={"error": f"Public league section unavailable: {exc}"},
        )


@app.post("/api/scrape")
async def trigger_scrape(request: Request, background_tasks: BackgroundTasks):
    """Manually trigger a scrape. Returns immediately; scrape runs in background.

    Accepts optional ``?leagueKey=...`` — today the scraper runs the
    registry's default league regardless, but we still validate the
    key so a multi-league-aware frontend can't accidentally ask for a
    retired league.  A non-default key currently returns 501
    ``not_implemented`` because multi-league scraping isn't wired up
    yet (that's a future refactor of Dynasty Scraper.py).
    """
    # Validate the key first.  Non-default leagues don't run the
    # full ranking scrape (the pipeline is single-league) — instead
    # they refresh the on-demand Sleeper overlay (rosters + trades
    # + pick ownership) so the UI picks up new trades / roster
    # moves without waiting for the 15-min cache to expire.  Same
    # shape of response either way.
    try:
        league_cfg = _resolve_league_for_request(request)
    except LeagueResolutionError as err:
        return err.json_response()
    default_cfg = _league_registry.get_default_league()
    if default_cfg and league_cfg.key != default_cfg.key:
        # Invalidate + rewarm the overlay for this league.  Returns
        # immediately with the refreshed team/trade counts so the
        # UI can show "X trades loaded" right away.
        loaded_sleeper = (
            latest_contract_data.get("sleeper") or {}
            if isinstance(latest_contract_data, dict)
            else {}
        )
        id_map = loaded_sleeper.get("idToPlayer") if isinstance(loaded_sleeper, dict) else {}
        _sleeper_overlay.invalidate_overlay_cache(league_cfg.sleeper_league_id)
        try:
            overlay = await run_in_threadpool(
                _sleeper_overlay.fetch_sleeper_overlay,
                sleeper_league_id=league_cfg.sleeper_league_id,
                id_to_player=id_map if isinstance(id_map, dict) else {},
                force_refresh=True,
            )
        except Exception as exc:  # noqa: BLE001
            return JSONResponse(
                status_code=502,
                content={
                    "error": "sleeper_overlay_fetch_failed",
                    "message": f"Sleeper overlay refresh failed: {exc}",
                    "leagueKey": league_cfg.key,
                },
            )
        if not overlay:
            return JSONResponse(
                status_code=502,
                content={
                    "error": "sleeper_overlay_empty",
                    "message": "Overlay fetch succeeded but returned no data.",
                    "leagueKey": league_cfg.key,
                },
            )
        return JSONResponse(
            content={
                "message": f"Sleeper overlay refreshed for {league_cfg.key!r}.",
                "leagueKey": league_cfg.key,
                "teamCount": len(overlay.get("teams") or []),
                "tradeCount": len(overlay.get("trades") or []),
                "waiverCount": len(overlay.get("waivers") or []),
                "overlayFetchedAt": overlay.get("overlayFetchedAt"),
            }
        )

    status_payload = _scrape_status_payload()
    if status_payload.get("is_running") or scrape_run_lock.locked():
        _record_scrape_event(
            "scrape_request_rejected",
            level="warning",
            message="Manual trigger rejected because scrape is already active",
            stalled=status_payload.get("stalled"),
            current_step=status_payload.get("current_step"),
            current_source=status_payload.get("current_source"),
        )
        return JSONResponse(
            status_code=409,
            content={"error": "Scrape already in progress", "status": status_payload},
        )

    # Run in background so the API returns immediately
    _record_scrape_event(
        "scrape_requested", message="Manual scrape trigger accepted", trigger="manual_api"
    )
    background_tasks.add_task(run_scraper, "manual_api")
    return JSONResponse(
        content={
            "message": "Scrape started in background",
            "status": _scrape_status_payload(),
        }
    )


@app.post("/api/test-alert")
async def test_alert():
    """Send a test alert email to verify configuration."""
    if not ALERT_ENABLED:
        return JSONResponse(
            status_code=400,
            content={"error": "Alerts not enabled. Set environment variable ALERT_ENABLED=true"},
        )
    try:
        send_alert("Test Alert", "If you're reading this, email alerts are working!")
        return JSONResponse(content={"message": f"Test alert sent to {ALERT_TO}"})
    except Exception as e:
        return JSONResponse(status_code=500, content={"error": f"Failed: {str(e)}"})


# ── AUTH + ENTRY GATE ROUTES ────────────────────────────────────────────
@app.get("/api/auth/status")
async def auth_status(request: Request):
    session = _get_auth_session(request)
    if not session:
        return JSONResponse(content={"authenticated": False})
    return JSONResponse(
        content={
            "authenticated": True,
            "username": session.get("username"),
            "displayName": session.get("display_name") or session.get("username"),
            "sleeperUserId": session.get("sleeper_user_id") or None,
            "avatar": session.get("avatar") or None,
            "authMethod": session.get("auth_method") or "password",
            # Lets the shell hide operator surfaces (/admin, /tools/*)
            # from users who would only get a 403 from them.  This is a
            # UI affordance, NOT the access control: every admin
            # endpoint still runs _require_admin_session independently,
            # so a client that lies to itself about this flag gains
            # nothing.
            "isAdmin": str(session.get("username") or "").lower() in PRIVATE_APP_ALLOWED_USERNAMES,
        }
    )


@app.post("/api/auth/login")
async def auth_login(request: Request):
    payload: dict = {}
    try:
        raw = await request.json()
        if isinstance(raw, dict):
            payload = raw
    except Exception:
        payload = {}

    username = str(payload.get("username") or "").strip()
    password = str(payload.get("password") or "")
    # Post-login default lands users on "/" (the Brisket Home
    # dashboard — Team Value + Top Movers + Risers/Fallers).  An
    # explicit ``next`` from the form preserves deep-link return,
    # but the bare default goes home, not to the legacy /app shell.
    next_path = _sanitize_next_path(payload.get("next"), "/")

    # Owner login: full session, max-age = JASON_AUTH_COOKIE_MAX_AGE.
    if username == JASON_LOGIN_USERNAME and password == JASON_LOGIN_PASSWORD:
        session_id = _create_auth_session(username, auth_method="password")
        response = JSONResponse(content={"ok": True, "redirect": next_path})
        response.set_cookie(
            key=JASON_AUTH_COOKIE_NAME,
            value=session_id,
            path="/",
            httponly=True,
            samesite="lax",
            secure=JASON_AUTH_COOKIE_SECURE,
            max_age=JASON_AUTH_COOKIE_MAX_AGE,
        )
        return response

    # Guest-pass fall-through: any password (regardless of username
    # field) that matches an active guest pass mints a time-bounded
    # session.  Cookie max-age is capped at the pass's remaining
    # lifetime so the session expires alongside the pass.  ``username``
    # is left as the literal ``"guest"`` so downstream admin gates
    # (which check username == JASON_LOGIN_USERNAME) reject guests
    # automatically — a guest can browse the private surface but
    # cannot trigger admin actions or generate more passes.
    pass_row = await run_in_threadpool(_guest_passes.validate, password)
    if pass_row is not None:
        remaining = max(0.0, pass_row.expires_at_epoch - time.time())
        if remaining < 1.0:
            # Expired during the millisecond between fetch and now —
            # treat as invalid.
            return JSONResponse(
                status_code=401,
                content={"ok": False, "error": "Invalid username or password."},
            )
        session_id = _create_auth_session(
            "guest",
            auth_method="guest_pass",
            expires_at_epoch=pass_row.expires_at_epoch,
            guest_pass_id=pass_row.id,
        )
        response = JSONResponse(
            content={
                "ok": True,
                "redirect": next_path,
                "guest": True,
                "expiresAtEpoch": pass_row.expires_at_epoch,
            }
        )
        # Cookie max-age is the smaller of the pass's remaining
        # lifetime or the standard ceiling.  Browser may still keep
        # the cookie longer if the user fiddles with system time;
        # ``_get_auth_session`` enforces ``expires_at_epoch`` server-
        # side either way.
        response.set_cookie(
            key=JASON_AUTH_COOKIE_NAME,
            value=session_id,
            path="/",
            httponly=True,
            samesite="lax",
            secure=JASON_AUTH_COOKIE_SECURE,
            max_age=int(min(remaining, JASON_AUTH_COOKIE_MAX_AGE)),
        )
        return response

    return JSONResponse(
        status_code=401,
        content={"ok": False, "error": "Invalid username or password."},
    )


@app.post("/api/auth/logout")
async def auth_logout(request: Request):
    _clear_auth_session(request)
    response = JSONResponse(content={"ok": True})
    response.delete_cookie(key=JASON_AUTH_COOKIE_NAME, path="/")
    return response


# ``GET /logout`` used to live here, next to its POST sibling above. It
# went with the page proxy (#555), and the recon that scoped that deletion
# deliberately left this as a judgement call rather than guessing. Three
# reasons to take it:
#
#   * it is a PAGE-namespace path on a backend that now serves no pages,
#     and it redirected to "/" — a route that no longer exists here, so
#     keeping it means clearing your session and landing on a 404
#   * zero callers, in or out of the tree. The UI logs out via
#     ``POST /api/auth/logout`` (useAuth.js:185, through the Next bridge
#     at frontend/app/api/auth/logout/route.js), and there is no
#     frontend/app/logout page, so nginx already 404s this path in
#     production
#   * it is a GET that destroys session state, i.e. CSRF-able by a bare
#     ``<img src="…:8000/logout">``. That was an acceptable trade for a
#     convenience redirect with real users; it is not one for a handler
#     with no callers and a dead destination
#
# The POST sibling is the real logout and is untouched.


# ── USER PREFERENCE PERSISTENCE (AUTH-GATED) ────────────────────────────
# Durable per-user state that follows the authenticated session across
# devices.  Backed by SQLite at ``data/user_kv.sqlite`` (see
# ``src/api/user_kv.py``).  Anonymous requests get 401 — the frontend
# hook falls back to a localStorage-only path when unauthenticated, so
# a logged-out visitor still sees defaults without polluting the
# shared store.


@app.get("/api/user/state")
async def get_user_state_api(request: Request):
    session = _get_auth_session(request)
    if not session:
        return JSONResponse(status_code=401, content={"error": "auth_required"})
    username = str(session.get("username") or "").strip()
    state = await run_in_threadpool(_user_kv.get_user_state, username)
    return JSONResponse(
        content={"username": username, "state": state},
        headers={"Cache-Control": "no-store"},
    )


@app.put("/api/user/state")
async def put_user_state_api(request: Request):
    session = _get_auth_session(request)
    if not session:
        return JSONResponse(status_code=401, content={"error": "auth_required"})
    username = str(session.get("username") or "").strip()
    try:
        body = await request.json()
    except Exception:
        body = None
    if not isinstance(body, dict):
        return JSONResponse(status_code=400, content={"error": "invalid_body"})
    patch: dict = {}
    if "selectedTeam" in body:
        sel = body.get("selectedTeam")
        if sel is None:
            patch["selectedTeam"] = None
        elif isinstance(sel, dict):
            patch["selectedTeam"] = {
                "ownerId": str(sel.get("ownerId") or ""),
                "name": str(sel.get("name") or ""),
            }
    if "watchlist" in body:
        wl = body.get("watchlist")
        if wl is None:
            patch["watchlist"] = None
        elif isinstance(wl, list):
            patch["watchlist"] = [str(x) for x in wl if isinstance(x, (str, int))]
    if "dismissedSignals" in body:
        ds = body.get("dismissedSignals")
        if ds is None:
            patch["dismissedSignals"] = None
        elif isinstance(ds, dict):
            clean: dict[str, int] = {}
            for k, v in ds.items():
                try:
                    clean[str(k)] = int(v)
                except (TypeError, ValueError):
                    continue
            patch["dismissedSignals"] = clean
    if "dismissalAliases" in body:
        da = body.get("dismissalAliases")
        if da is None:
            patch["dismissalAliases"] = None
        elif isinstance(da, dict):
            patch["dismissalAliases"] = {
                str(k): str(v)
                for k, v in da.items()
                if isinstance(k, str) and isinstance(v, (str, int))
            }
    if "notificationsEmail" in body:
        # Optional email for signal-alert digests.  Stored per-user
        # in user_kv under the ``notificationsEmail`` key so the
        # alert loop can resolve it without a separate table.  We
        # only accept plausible-looking addresses — no MX check,
        # just format validation.
        ne = body.get("notificationsEmail")
        if ne is None or ne == "":
            patch["notificationsEmail"] = None
        elif isinstance(ne, str):
            s = ne.strip()
            if "@" in s and "." in s.split("@")[-1] and len(s) <= 254:
                patch["notificationsEmail"] = s
    if "notificationsEnabled" in body:
        patch["notificationsEnabled"] = bool(body.get("notificationsEnabled"))
    if "activeLeagueKey" in body:
        # The user's preferred league from the registry.  Validate
        # against the live registry so a stale/typo value can't
        # silently land a user on a nonexistent league on next load.
        # ``None`` / empty string clears the preference → callers fall
        # back to the registry's default league on next read.
        raw = body.get("activeLeagueKey")
        if raw is None or raw == "":
            patch["activeLeagueKey"] = None
        elif isinstance(raw, str):
            candidate = raw.strip()
            cfg = _league_registry.get_league_by_key(candidate)
            if cfg is not None and cfg.active:
                patch["activeLeagueKey"] = cfg.key  # canonicalize via alias lookup
            # Unknown or inactive league → silently drop.  The
            # frontend will notice the server didn't echo the key back
            # and fall through to the default.
    if "selectedTeamsByLeague" in body:
        # Per-league selected team map.  Accepts
        #   {"leagueKey": {"ownerId": "...", "teamName": "...",
        #                  "rosterId": <int|str>, "managerName": "..."}}
        # Each entry's leagueKey must resolve against the registry
        # (aliases canonicalized); unknown / inactive leagues are
        # dropped.  Null/"" clears the entire map.
        raw = body.get("selectedTeamsByLeague")
        if raw is None or raw == "":
            patch["selectedTeamsByLeague"] = None
        elif isinstance(raw, dict):
            clean: dict[str, dict[str, object]] = {}
            for lkey, spec in raw.items():
                if not isinstance(lkey, str) or not isinstance(spec, dict):
                    continue
                cfg = _league_registry.get_league_by_key(lkey.strip())
                if cfg is None or not cfg.active:
                    continue
                owner_id = str(spec.get("ownerId") or "").strip()
                team_name = str(spec.get("teamName") or "").strip()
                if not owner_id and not team_name:
                    # An empty entry clears that league's selection
                    # (distinct from not touching the map at all).
                    clean[cfg.key] = {"ownerId": "", "teamName": ""}
                    continue
                entry: dict[str, object] = {
                    "ownerId": owner_id,
                    "teamName": team_name,
                }
                roster_id = spec.get("rosterId")
                if roster_id is not None:
                    entry["rosterId"] = str(roster_id)
                manager_name = str(spec.get("managerName") or "").strip()
                if manager_name:
                    entry["managerName"] = manager_name
                clean[cfg.key] = entry
            patch["selectedTeamsByLeague"] = clean
    state = await run_in_threadpool(_user_kv.merge_user_state, username, patch)
    return JSONResponse(
        content={"username": username, "state": state},
        headers={"Cache-Control": "no-store"},
    )


# ── Web Push subscriptions ────────────────────────────────────────
@app.get("/api/push/public-key")
async def get_push_public_key():
    """Return the VAPID public key the browser uses when calling
    ``pushManager.subscribe({applicationServerKey})``.  Stateless,
    no auth — the public key is, by definition, public."""
    if not _push_delivery.is_configured():
        return JSONResponse(
            status_code=503,
            content={"error": "push_not_configured"},
            headers={"Cache-Control": "no-store"},
        )
    return JSONResponse(
        content={"publicKey": _push_delivery.public_key()},
        headers={"Cache-Control": "public, max-age=3600"},
    )


@app.post("/api/push/subscribe")
async def post_push_subscribe(request: Request):
    session = _get_auth_session(request)
    if not session:
        return JSONResponse(status_code=401, content={"error": "auth_required"})
    username = str(session.get("username") or "").strip()
    try:
        body = await request.json()
    except Exception:
        body = None
    if not isinstance(body, dict):
        return JSONResponse(status_code=400, content={"error": "invalid_body"})
    try:
        state = await run_in_threadpool(_user_kv.get_user_state, username) or {}
        new_subs = _push_delivery.upsert_subscription(state, body)
    except ValueError as exc:
        return JSONResponse(status_code=400, content={"error": str(exc)})
    await run_in_threadpool(
        _user_kv.set_user_field,
        username,
        "pushSubscriptions",
        new_subs,
    )
    return JSONResponse(
        content={"ok": True, "count": len(new_subs)},
        headers={"Cache-Control": "no-store"},
    )


@app.post("/api/push/unsubscribe")
async def post_push_unsubscribe(request: Request):
    session = _get_auth_session(request)
    if not session:
        return JSONResponse(status_code=401, content={"error": "auth_required"})
    username = str(session.get("username") or "").strip()
    try:
        body = await request.json()
    except Exception:
        body = None
    endpoint = ""
    if isinstance(body, dict):
        endpoint = str(body.get("endpoint") or "").strip()
    if not endpoint:
        return JSONResponse(status_code=400, content={"error": "endpoint_required"})
    state = await run_in_threadpool(_user_kv.get_user_state, username) or {}
    new_subs = _push_delivery.remove_subscription(state, endpoint)
    await run_in_threadpool(
        _user_kv.set_user_field,
        username,
        "pushSubscriptions",
        new_subs,
    )
    return JSONResponse(
        content={"ok": True, "count": len(new_subs)},
        headers={"Cache-Control": "no-store"},
    )


# ── Custom alerts (user-defined value/rank watchers) ──────────────
@app.get("/api/custom-alerts")
async def get_custom_alerts(request: Request):
    session = _get_auth_session(request)
    if not session:
        return JSONResponse(status_code=401, content={"error": "auth_required"})
    username = str(session.get("username") or "").strip()
    state = await run_in_threadpool(_user_kv.get_user_state, username) or {}
    rules = state.get("customAlerts") or []
    if not isinstance(rules, list):
        rules = []
    return JSONResponse(
        content={"rules": rules},
        headers={"Cache-Control": "no-store"},
    )


@app.put("/api/custom-alerts")
async def put_custom_alerts(request: Request):
    session = _get_auth_session(request)
    if not session:
        return JSONResponse(status_code=401, content={"error": "auth_required"})
    username = str(session.get("username") or "").strip()
    try:
        body = await request.json()
    except Exception:
        body = None
    if not isinstance(body, dict):
        return JSONResponse(status_code=400, content={"error": "invalid_body"})
    raw_rules = body.get("rules")
    if not isinstance(raw_rules, list):
        return JSONResponse(status_code=400, content={"error": "rules_array_required"})
    if len(raw_rules) > 50:
        return JSONResponse(status_code=400, content={"error": "too_many_rules"})

    cleaned: list[dict[str, object]] = []
    for entry in raw_rules:
        try:
            cleaned.append(_custom_alerts.validate_rule(entry))
        except ValueError as exc:
            return JSONResponse(
                status_code=400,
                content={"error": "invalid_rule", "detail": str(exc)},
            )

    state = await run_in_threadpool(_user_kv.get_user_state, username) or {}
    prior_state = state.get("customAlertsState") or {}
    if not isinstance(prior_state, dict):
        prior_state = {}
    new_ids = {r["id"] for r in cleaned}
    pruned_state = {
        k: v
        for k, v in prior_state.items()
        if isinstance(k, str) and k.split("::", 1)[0] in new_ids
    }

    await run_in_threadpool(
        _user_kv.merge_user_state,
        username,
        {"customAlerts": cleaned, "customAlertsState": pruned_state},
    )
    return JSONResponse(
        content={"rules": cleaned, "count": len(cleaned)},
        headers={"Cache-Control": "no-store"},
    )


@app.post("/api/custom-alerts/run")
async def run_custom_alerts(request: Request):
    """Cron-driven evaluator.  Walks every user's rules, fires any
    that hit, delivers via email and/or push, updates cooldown state.

    Same bearer-token auth model as ``/api/signal-alerts/run``.
    """
    cron_auth_ok = False
    if SIGNAL_ALERT_CRON_TOKEN:
        header = (request.headers.get("authorization") or "").strip()
        if header.lower().startswith("bearer "):
            presented = header.split(None, 1)[1].strip()
            if hmac.compare_digest(presented, SIGNAL_ALERT_CRON_TOKEN):
                cron_auth_ok = True
    if not cron_auth_ok:
        session = _get_auth_session(request)
        if not session or session.get("auth_method") != "password":
            return JSONResponse(
                status_code=401,
                content={"error": "admin_auth_required"},
            )

    if not latest_contract_data:
        return JSONResponse(status_code=503, content={"error": "no_live_contract"})
    players_array = (latest_contract_data or {}).get("playersArray") or []

    db = await run_in_threadpool(_user_kv.all_user_states)
    summary: dict[str, dict[str, int]] = {}

    for username, state in (db or {}).items():
        if not isinstance(state, dict):
            continue
        rules = state.get("customAlerts") or []
        if not isinstance(rules, list) or not rules:
            continue
        cooldown_state = state.get("customAlertsState") or {}
        if not isinstance(cooldown_state, dict):
            cooldown_state = {}

        try:
            hits = _custom_alerts.evaluate_alerts(
                rules,
                players_array,
                state=cooldown_state,
            )
        except Exception as exc:
            log.warning("custom-alerts evaluation failed for %s: %s", username, exc)
            continue
        if not hits:
            continue

        notify_email = ""
        if isinstance(state.get("notificationsEmail"), str):
            notify_email = state["notificationsEmail"].strip()

        delivered_email = 0
        delivered_push = 0
        endpoints_to_prune: list[str] = []
        new_state = dict(cooldown_state)

        for hit in hits:
            email_ok = False
            push_ok = False

            if "email" in hit.channels and notify_email:
                try:
                    email_ok = _deliver_email_smtp(
                        notify_email,
                        f"[Brisket alert] {hit.title}",
                        f"{hit.body}\n\n— Brisket custom alert ({hit.kind})",
                    )
                except Exception as exc:
                    log.warning("custom-alert email failed for %s: %s", username, exc)
                if email_ok:
                    delivered_email += 1

            if "push" in hit.channels:
                count, prune = _push_delivery.fanout(
                    state,
                    title=f"Brisket: {hit.title}",
                    body=hit.body,
                    url="/rankings",
                    tag=hit.rule_id,
                )
                push_ok = count > 0
                delivered_push += count
                endpoints_to_prune.extend(prune)

            if email_ok or push_ok:
                new_state = _custom_alerts.mark_fired(new_state, hit)

        patch: dict[str, object] = {}
        if new_state != cooldown_state:
            patch["customAlertsState"] = new_state
        if endpoints_to_prune:
            kept = [
                s
                for s in _push_delivery.list_subscriptions(state)
                if s.get("endpoint") not in set(endpoints_to_prune)
            ]
            patch["pushSubscriptions"] = kept
        if patch:
            await run_in_threadpool(_user_kv.merge_user_state, username, patch)

        summary[username] = {
            "hits": len(hits),
            "email": delivered_email,
            "push": delivered_push,
        }

    return JSONResponse(
        content={"ok": True, "users": summary},
        headers={"Cache-Control": "no-store"},
    )


@app.post("/api/user/signals/dismiss")
async def dismiss_signal_api(request: Request):
    session = _get_auth_session(request)
    if not session:
        return JSONResponse(status_code=401, content={"error": "auth_required"})
    username = str(session.get("username") or "").strip()
    try:
        body = await request.json()
    except Exception:
        body = None
    if not isinstance(body, dict):
        return JSONResponse(status_code=400, content={"error": "invalid_body"})
    signal_key = str(body.get("signalKey") or "").strip()
    if not signal_key:
        return JSONResponse(
            status_code=400,
            content={"error": "signalKey_required"},
        )
    try:
        ttl_ms = int(body.get("ttlMs") or 7 * 24 * 3600 * 1000)
    except (TypeError, ValueError):
        ttl_ms = 7 * 24 * 3600 * 1000
    alias_sid = str(body.get("aliasSleeperId") or "").strip() or None
    alias_name = str(body.get("aliasDisplayName") or "").strip() or None
    # Scope the dismissal to the active league.  Validated against
    # the registry; unknown/inactive keys fall through to legacy flat
    # dismissal.  See user_kv.dismiss_signal docstring.
    raw_league = str(body.get("leagueKey") or "").strip()
    scoped_key: str | None = None
    if raw_league:
        cfg = _league_registry.get_league_by_key(raw_league)
        if cfg is not None and cfg.active:
            scoped_key = cfg.key
    state = await run_in_threadpool(
        _user_kv.dismiss_signal,
        username,
        signal_key,
        ttl_ms=ttl_ms,
        alias_sleeper_id=alias_sid,
        alias_display_name=alias_name,
        league_key=scoped_key,
    )
    return JSONResponse(
        content={"username": username, "state": state},
        headers={"Cache-Control": "no-store"},
    )


@app.post("/api/user/signals/restore")
async def restore_signal_api(request: Request):
    session = _get_auth_session(request)
    if not session:
        return JSONResponse(status_code=401, content={"error": "auth_required"})
    username = str(session.get("username") or "").strip()
    try:
        body = await request.json()
    except Exception:
        body = None
    if not isinstance(body, dict):
        return JSONResponse(status_code=400, content={"error": "invalid_body"})
    signal_key = str(body.get("signalKey") or "").strip()
    if not signal_key:
        return JSONResponse(
            status_code=400,
            content={"error": "signalKey_required"},
        )
    raw_league = str(body.get("leagueKey") or "").strip()
    scoped_key: str | None = None
    if raw_league:
        cfg = _league_registry.get_league_by_key(raw_league)
        if cfg is not None and cfg.active:
            scoped_key = cfg.key
    state = await run_in_threadpool(
        _user_kv.undismiss_signal,
        username,
        signal_key,
        league_key=scoped_key,
    )
    return JSONResponse(
        content={"username": username, "state": state},
        headers={"Cache-Control": "no-store"},
    )


# ── TERMINAL AGGREGATION ENDPOINT ───────────────────────────────────────
#
# Server-side aggregate of everything the landing page needs: team
# aggregates, market movers, signals, news, portfolio.  See
# ``src/api/terminal.py`` for the builder.  Two modes:
#
#   * Authenticated users get the full payload including signals,
#     portfolio, watchlist, and roster-aware team aggregates.
#   * Anonymous users get a public slice (league + top150 movers,
#     news for top-150 players) — enough for an at-a-glance "market
#     pulse" without leaking private identifiers or roster state.
#
# Availability:
#   * When the live contract is loaded: serve the fresh aggregation.
#   * When the live contract hasn't loaded yet (cold start): fall
#     back to the most recent cached dynasty_data_*.json export
#     from disk.  The frontend sees a ``stale: true`` flag and a
#     ``staleAs`` date so it can surface "last good data from
#     YYYY-MM-DD" instead of spinning forever.
#   * If even the cached export is absent, surface a 503 with the
#     same shape so the frontend error UI can render a coherent
#     message.


def _latest_cached_contract_from_disk() -> tuple[dict | None, str | None]:
    """Return the most recent on-disk ``dynasty_data_*.json`` export
    parsed as a contract, plus the date string it was stamped with.
    Used when ``latest_contract_data`` hasn't been primed yet (cold
    start between process-restart and first scrape).
    """
    try:
        candidates = sorted(
            DATA_DIR.glob("dynasty_data_*.json"),
            key=lambda p: p.stat().st_mtime,
            reverse=True,
        )
    except OSError:
        return None, None
    for candidate in candidates:
        try:
            with candidate.open("r", encoding="utf-8") as f:
                raw = json.load(f)
        except (OSError, json.JSONDecodeError):
            continue
        if not isinstance(raw, dict):
            continue
        # Older exports are pre-contract-builder and lack
        # ``playersArray``; the terminal builder handles that path
        # via the legacy dict, so we can still serve.
        if not raw.get("players") and not raw.get("playersArray"):
            continue
        return raw, raw.get("date") or candidate.stem.replace("dynasty_data_", "")
    return None, None


@app.get("/api/terminal")
async def get_terminal(request: Request):
    session = _get_auth_session(request)
    authed = bool(session)
    username = str((session or {}).get("username") or "").strip() if authed else ""

    # League routing — validate the key, but DON'T require a loaded
    # contract yet (we want to fall through to the disk cache below
    # for the default league).  The loaded-contract check fires
    # after the in-memory contract is available.
    try:
        league_cfg = _resolve_league_for_request(request)
    except LeagueResolutionError as err:
        return err.json_response()

    contract = latest_contract_data
    stale = False
    stale_as = None
    if not contract:
        # 503 fallback (Item 3 from the TODO list): try the most
        # recent cached export before giving up.  Disk cache is for
        # the default league only — if a non-default league is
        # requested we skip the cache and return 503 cleanly.
        default_cfg = _league_registry.get_default_league()
        if default_cfg and league_cfg.key == default_cfg.key:
            cached, cached_date = _latest_cached_contract_from_disk()
            if cached:
                contract = cached
                stale = True
                stale_as = cached_date
        if not contract:
            return JSONResponse(
                status_code=503,
                content={
                    "error": "No data available yet. First scrape may still be running.",
                    "stale": False,
                    "leagueKey": league_cfg.key,
                },
            )

    # Cross-league request: the loaded contract is for a different
    # league than the one requested.  If the scoring is PROVEN to match,
    # splice in a live Sleeper overlay (rosters + trades) so the
    # terminal + team widgets actually have data to render.  Only
    # 503 when the overlay fetch fails completely — Sleeper
    # unreachable, invalid league ID, etc.
    loaded_meta = (contract.get("meta") or {}) if isinstance(contract, dict) else {}
    loaded_league = loaded_meta.get("leagueKey")
    if loaded_league and loaded_league != league_cfg.key:
        # Rankings incompatible; the /api/data 503 path explains this
        # shape.  Terminal can't do anything.
        _scoring_err = _scoring_identity_error(contract, league_cfg)
        if _scoring_err is not None:
            return _scoring_err
        loaded_sleeper = contract.get("sleeper") or {}
        id_map = loaded_sleeper.get("idToPlayer") if isinstance(loaded_sleeper, dict) else None
        try:
            overlay = await run_in_threadpool(
                _sleeper_overlay.fetch_sleeper_overlay,
                sleeper_league_id=league_cfg.sleeper_league_id,
                id_to_player=id_map if isinstance(id_map, dict) else {},
            )
        except Exception as exc:  # noqa: BLE001
            log.warning(
                "terminal overlay fetch failed for %s: %s",
                league_cfg.key,
                exc,
            )
            overlay = None
        if not overlay or not overlay.get("teams"):
            return JSONResponse(
                status_code=503,
                content={
                    "error": "data_not_ready",
                    "message": f"Sleeper overlay for league {league_cfg.key!r} unavailable.",
                    "leagueKey": league_cfg.key,
                },
            )
        # Build a hybrid contract: global rankings + per-league
        # sleeper.  The NFL-wide maps carry forward so the terminal
        # builder resolves positions/IDs the same as for the primary
        # league; the league-SPECIFIC fields come from the requested
        # league or not at all (W18-F002).
        hybrid_sleeper, cross_league_ready = _sleeper_overlay.merge_cross_league_sleeper_block(
            loaded_sleeper=loaded_sleeper if isinstance(loaded_sleeper, dict) else {},
            overlay=overlay,
            requested_league_config=overlay.get("leagueConfig"),
        )
        contract = {
            **contract,
            "sleeper": hybrid_sleeper,
            "meta": {
                **loaded_meta,
                "leagueKey": league_cfg.key,
                "sleeperDataReady": bool(cross_league_ready),
                "sleeperSource": "overlay",
                "sleeperLoadedLeagueKey": loaded_league,
            },
        }

    # Apply the requested lens on top of whatever contract the routing
    # above settled on — including the cross-league hybrid, whose
    # spliced ``sleeper`` block must survive.
    contract, valuation_mode, valuation_note = await _valuation_scoped_contract(
        request, None, league_cfg, base=contract
    )

    params = request.query_params
    team_owner_id = (params.get("team") or params.get("ownerId") or "").strip()
    team_name = (params.get("teamName") or "").strip()
    try:
        window_days = int(params.get("windowDays") or 30)
    except (TypeError, ValueError):
        window_days = 30
    window_days = max(7, min(180, window_days))

    user_state: dict = {}
    if authed and username:
        try:
            user_state = await run_in_threadpool(_user_kv.get_user_state, username)
        except Exception as exc:  # noqa: BLE001
            log.warning("/api/terminal user_kv read failed: %s", exc)

    resolved_team = None
    if authed:
        # Anonymous callers get the public slice even if they pass a
        # ``team`` param — we never expose per-roster state without
        # authentication.  ``resolved_team=None`` is enforced below.
        resolved_team = _terminal.resolve_team(
            contract,
            owner_id=team_owner_id,
            name=team_name,
        )
        # Auto-resolve via the authenticated Sleeper user id when the
        # client didn't pass an explicit team.  This is the "Sleeper
        # login → your team lights up on first page load" path — no
        # manual team picker needed when the authed user owns a team
        # in this league.
        if resolved_team is None and not team_owner_id and not team_name:
            session_sleeper_id = str((session or {}).get("sleeper_user_id") or "").strip()
            if session_sleeper_id:
                resolved_team = _terminal.resolve_team(
                    contract,
                    owner_id=session_sleeper_id,
                    name=None,
                )

    try:
        # ONE callable so every heavy piece runs on the worker thread.
        # Python evaluates call arguments eagerly, so the previous
        # ``run_in_threadpool(build, ..., news_items=gather_news_items(...))``
        # shape executed the news aggregation (up to ~11 RSS providers)
        # and two full-contract scans ON THE EVENT LOOP before the
        # threadpool hop — stalling every concurrent request.
        def _build_terminal():
            news_items = _terminal.gather_news_items(
                lambda: _get_news_service(),
                _live_player_names(),
                (resolved_team or {}).get("name") if resolved_team else None,
                player_meta=_live_player_meta(),
            )
            return _terminal.build_terminal_payload(
                contract,
                resolved_team=resolved_team,
                window_days=window_days,
                news_items=news_items,
                user_state=user_state,
                public_mode=not authed,
                # Scope dismissals to the active league so a dismissal
                # on league A doesn't silence the same player's signal
                # on league B.  See terminal.build_terminal_payload +
                # user_kv.active_dismissals docstrings.
                league_key=league_cfg.key if league_cfg else None,
            )

        payload = await run_in_threadpool(_build_terminal)
    except Exception as exc:
        log.exception("/api/terminal build failed: %s", exc)
        return JSONResponse(
            status_code=500,
            content={"error": f"terminal_build_failed: {type(exc).__name__}"},
        )

    # Stale-data stamp: consumers can render a "last good data from
    # YYYY-MM-DD" banner when the live scrape hasn't caught up.
    payload["stale"] = stale
    payload["staleAs"] = stale_as
    payload["authenticated"] = authed
    _stamp_valuation_mode(payload, valuation_mode, valuation_note)

    cache_control = (
        "public, max-age=60, stale-while-revalidate=600"
        if not authed
        else "private, max-age=30, stale-while-revalidate=120"
    )
    return JSONResponse(
        content=payload,
        headers={"Cache-Control": cache_control},
    )


@app.post("/api/trade/simulate")
async def post_trade_simulate(request: Request):
    """Pure-function what-if: apply a hypothetical trade to the
    authenticated user's team and return the delta payload.

    Body::

        {
          "team":       "<ownerId>" (optional — defaults to session
                        owner when signed in via Sleeper),
          "teamName":   "<teamName>" (optional fallback lookup),
          "playersIn":  ["Ja'Marr Chase", ...],   # inbound players
          "playersOut": ["Drake London", ...],     # outbound players
          "picksIn":    ["2026 1.04", ...],        # inbound picks
          "picksOut":   ["2027 2.08", ...]         # outbound picks
        }

    Response shape matches ``trade_simulator.simulate_trade``.
    No persistence — the live contract is never mutated.
    """
    session = _get_auth_session(request)
    if not session:
        return JSONResponse(status_code=401, content={"error": "auth_required"})
    if not latest_contract_data:
        return JSONResponse(
            status_code=503,
            content={"error": "No data available yet."},
        )
    try:
        body = await request.json()
    except Exception:
        body = None
    if not isinstance(body, dict):
        return JSONResponse(status_code=400, content={"error": "invalid_body"})

    # Validate leagueKey but don't require the loaded contract's
    # leagueKey to match — when the user is on a non-default league
    # we can splice in a live Sleeper overlay (same trick the
    # /api/terminal endpoint uses).  Without this, every trade
    # simulation on League B returns data_not_ready.
    try:
        league_cfg = _resolve_league_for_request(request, body=body)
    except LeagueResolutionError as err:
        return err.json_response()

    # Build the contract this trade sim runs against.  When the
    # request league matches the loaded contract, just use it.
    # When they differ but the scoring is PROVEN to match, splice in
    # the per-league Sleeper overlay so the resolver can find the
    # user's team in this league's rosters.
    contract = latest_contract_data
    loaded_meta = (contract.get("meta") or {}) if isinstance(contract, dict) else {}
    loaded_league = loaded_meta.get("leagueKey")
    if loaded_league and loaded_league != league_cfg.key:
        # Trade simulation needs matching rankings.
        _scoring_err = _scoring_identity_error(contract, league_cfg)
        if _scoring_err is not None:
            return _scoring_err
        # Splice in a live overlay for the requested league.
        loaded_sleeper = contract.get("sleeper") or {}
        id_map = loaded_sleeper.get("idToPlayer") if isinstance(loaded_sleeper, dict) else None
        try:
            overlay = await run_in_threadpool(
                _sleeper_overlay.fetch_sleeper_overlay,
                sleeper_league_id=league_cfg.sleeper_league_id,
                id_to_player=id_map if isinstance(id_map, dict) else {},
            )
        except Exception as exc:  # noqa: BLE001
            log.warning(
                "trade-simulate overlay fetch failed for %s: %s",
                league_cfg.key,
                exc,
            )
            overlay = None
        if not overlay or not overlay.get("teams"):
            return JSONResponse(
                status_code=503,
                content={
                    "error": "data_not_ready",
                    "message": f"Sleeper overlay for league {league_cfg.key!r} unavailable.",
                    "leagueKey": league_cfg.key,
                },
            )
        hybrid_sleeper, _ = _sleeper_overlay.merge_cross_league_sleeper_block(
            loaded_sleeper=loaded_sleeper if isinstance(loaded_sleeper, dict) else {},
            overlay=overlay,
            requested_league_config=overlay.get("leagueConfig"),
        )
        contract = {**latest_contract_data, "sleeper": hybrid_sleeper}

    # Lens on top of whatever contract the routing above settled on,
    # keeping any spliced cross-league ``sleeper`` block.
    contract, valuation_mode, valuation_note = await _valuation_scoped_contract(
        request, body, league_cfg, base=contract
    )

    team_owner_id = str(body.get("team") or "").strip()
    team_name = str(body.get("teamName") or "").strip()

    # Session auto-resolve: if the user didn't pass a team, use
    # the Sleeper user_id attached to their session.
    if not team_owner_id and not team_name:
        team_owner_id = str(session.get("sleeper_user_id") or "").strip()

    resolved_team = _terminal.resolve_team(
        contract,
        owner_id=team_owner_id,
        name=team_name,
    )
    if resolved_team is None:
        return JSONResponse(
            status_code=404,
            content={"error": "team_not_found", "leagueKey": league_cfg.key},
        )

    def _str_list(key):
        vs = body.get(key) or []
        if not isinstance(vs, list):
            return []
        return [str(x) for x in vs if isinstance(x, (str, int)) and str(x).strip()]

    result = await run_in_threadpool(
        _trade_simulator.simulate_trade,
        contract,
        resolved_team=resolved_team,
        players_in=_str_list("playersIn"),
        players_out=_str_list("playersOut"),
        picks_in=_str_list("picksIn"),
        picks_out=_str_list("picksOut"),
        roster_settings=dict(league_cfg.roster_settings or {}),
        league_key=league_cfg.key,
    )
    result["leagueKey"] = league_cfg.key
    _stamp_valuation_mode(result, valuation_mode, valuation_note)
    return JSONResponse(
        content=result,
        headers={"Cache-Control": "no-store"},
    )


def _deliver_email_smtp(to: str, subject: str, body: str) -> bool:
    """SMTP delivery bound to the existing ALERT_* env vars.

    Returns True on successful send, False on any error.  Errors
    are logged but never raised — the alert runner catches
    exceptions itself and we want deliver-per-user to be isolated.

    Tries STARTTLS on port 587 first (port 465 is firewall-blocked
    on the production VPS as of 2026-04-26).  Falls back to SMTP_SSL
    on 465 if 587 fails — covers environments where 587 is the
    blocked path instead.
    """
    if not ALERT_ENABLED or not ALERT_FROM or not ALERT_PASSWORD or not to:
        return False
    msg = MIMEMultipart("alternative")
    msg["Subject"] = subject
    msg["From"] = ALERT_FROM
    msg["To"] = to
    msg.attach(MIMEText(body, "plain"))
    try:
        with smtplib.SMTP("smtp.gmail.com", 587, timeout=15) as s:
            s.starttls()
            s.login(ALERT_FROM, ALERT_PASSWORD)
            s.send_message(msg)
        return True
    except Exception as exc587:  # noqa: BLE001
        log.warning(
            "signal-alert SMTP 587 STARTTLS failed to %s: %s — falling back to 465",
            to,
            exc587,
        )
        try:
            with smtplib.SMTP_SSL("smtp.gmail.com", 465, timeout=15) as s:
                s.login(ALERT_FROM, ALERT_PASSWORD)
                s.send_message(msg)
            return True
        except Exception as exc465:  # noqa: BLE001
            log.warning("signal-alert SMTP 465 SSL also failed to %s: %s", to, exc465)
            return False


# ── Admin endpoints (Phase 11 follow-ons) ────────────────────────
#
# Gated on both: (1) a valid session AND (2) an explicit admin
# username check against PRIVATE_APP_ALLOWED_USERNAMES.  Every
# admin action is logged with username + action for audit.


def _require_admin_session(request: Request):
    """Returns the session dict on success; returns a JSONResponse
    error on failure.  Caller pattern:

        session_or_err = _require_admin_session(request)
        if isinstance(session_or_err, JSONResponse):
            return session_or_err
        session = session_or_err
    """
    session = _get_auth_session(request)
    if not session:
        return JSONResponse(status_code=401, content={"error": "auth_required"})
    username = str(session.get("username") or "").strip().lower()
    if not PRIVATE_APP_ALLOWED_USERNAMES or username not in PRIVATE_APP_ALLOWED_USERNAMES:
        return JSONResponse(
            status_code=403,
            content={"error": "admin_required", "message": "Allowlisted users only."},
        )
    return session


@app.post("/api/test/create-session")
async def post_test_create_session(request: Request):
    """E2E-only session bootstrap — gated behind two env vars.

    Returns 404 (NOT 401) unless BOTH:
      * ``E2E_TEST_MODE=1`` (or true/yes/on)
      * ``E2E_TEST_SECRET`` matches the caller's ``Authorization:
        Bearer <secret>`` header

    In prod neither var is set, so this endpoint is invisible.

    ``E2E_TEST_USERNAME`` is a THIRD requirement, and it fails
    closed: there is no default.  An earlier revision fell back to
    the operator's real username, so enabling E2E mode without
    naming a test user minted a session for a real (allowlisted,
    admin-capable) account.  The identity a test session assumes
    must be stated explicitly, never inherited.
    """
    mode_raw = os.getenv("E2E_TEST_MODE", "").strip().lower()
    if mode_raw not in ("1", "true", "yes", "on"):
        return JSONResponse(status_code=404, content={"error": "not_found"})
    expected = os.getenv("E2E_TEST_SECRET", "").strip()
    auth = str(request.headers.get("authorization", "")).strip()
    provided = auth[len("Bearer ") :].strip() if auth.lower().startswith("bearer ") else ""
    if not expected or provided != expected:
        return JSONResponse(status_code=404, content={"error": "not_found"})
    # Misconfiguration, reported only to a caller that already proved
    # it holds the secret — so the message can be actionable without
    # leaking the endpoint's existence to anyone else.
    username = (os.getenv("E2E_TEST_USERNAME") or "").strip().lower()
    if not username:
        log.error(
            "/api/test/create-session refused: E2E_TEST_MODE is on but "
            "E2E_TEST_USERNAME is unset — refusing rather than defaulting "
            "to a real account."
        )
        return JSONResponse(
            status_code=500,
            content={
                "error": "e2e_username_not_configured",
                "message": (
                    "E2E_TEST_USERNAME must be set to the throwaway username "
                    "test sessions should assume.  There is no default."
                ),
            },
        )
    session_id = _create_auth_session(
        username=username,
        sleeper_user_id=os.getenv("E2E_TEST_SLEEPER_USER_ID", "").strip() or None,
        display_name=username,
        auth_method="e2e_test",
    )
    res = JSONResponse(
        content={
            "ok": True,
            "username": username,
            "sessionId": session_id,
        }
    )
    res.set_cookie(
        JASON_AUTH_COOKIE_NAME,
        session_id,
        max_age=3600,
        httponly=True,
        samesite="lax",
    )
    return res


@app.post("/api/admin/nfl-data/flush")
async def post_admin_nfl_data_flush(request: Request):
    """Flush every nfl_data cache entry (forces next fetch to go
    upstream).  Use when an upstream schema change is suspected
    and cached parquet is stale.
    """
    session_or_err = _require_admin_session(request)
    if isinstance(session_or_err, JSONResponse):
        return session_or_err
    session = session_or_err
    from src.nfl_data import cache as _nflc

    cache_dir = _nflc._default_cache_dir()  # noqa: SLF001
    deleted = 0
    try:
        if cache_dir.exists():
            for p in cache_dir.iterdir():
                try:
                    p.unlink()
                    deleted += 1
                except OSError:
                    pass
    except Exception as exc:  # noqa: BLE001
        return JSONResponse(
            status_code=500,
            content={"error": "flush_failed", "message": str(exc)},
        )
    log.info(
        "admin action: nfl_data flush by %s — %d entries evicted",
        session.get("username"),
        deleted,
    )
    return JSONResponse(content={"ok": True, "evicted": deleted})


@app.post("/api/admin/sessions/force-logout-all")
async def post_admin_force_logout_all(request: Request):
    """Emergency: sign-out-everyone hammer.  Wipes both the in-memory
    dict AND the persistent store so a stolen session / compromise
    can be remediated without a deploy."""
    session_or_err = _require_admin_session(request)
    if isinstance(session_or_err, JSONResponse):
        return session_or_err
    session = session_or_err
    in_mem_count = len(auth_sessions)
    auth_sessions.clear()
    persisted = 0
    try:
        from src.api import session_store as _ss

        persisted = _ss.force_clear_all()
    except Exception as exc:  # noqa: BLE001
        log.warning("force_clear_all session_store: %s", exc)
    log.warning(
        "admin action: FORCE-LOGOUT-ALL by %s — %d in-memory + %d persisted",
        session.get("username"),
        in_mem_count,
        persisted,
    )
    return JSONResponse(
        content={
            "ok": True,
            "inMemoryCleared": in_mem_count,
            "persistedCleared": persisted,
        }
    )


@app.post("/api/admin/guest-pass")
async def post_admin_guest_pass_create(request: Request):
    """Generate a time-bounded guest password.

    Body::

        {"durationHours": 12, "note": "for Brent"}

    Returns the plaintext token in the response — this is the ONLY
    time it's exposed.  ``list_passes`` and ``status``/audit calls
    return the GuestPass shape WITHOUT the plaintext (only its sha256
    hash lives in the DB).  The owner shares the token with their
    guest; the guest types it into the login form's password field
    (any username, or empty).  Validation lives in
    ``src/api/guest_passes.py::validate``.
    """
    session_or_err = _require_admin_session(request)
    if isinstance(session_or_err, JSONResponse):
        return session_or_err
    session = session_or_err
    payload: dict[str, Any] = {}
    try:
        raw = await request.json()
        if isinstance(raw, dict):
            payload = raw
    except Exception:
        payload = {}
    raw_hours = payload.get("durationHours")
    note = str(payload.get("note") or "")
    try:
        pass_row, token = await run_in_threadpool(
            _guest_passes.create,
            duration_hours=float(raw_hours) if raw_hours is not None else 12.0,
            note=note,
            created_by=str(session.get("username") or ""),
        )
    except ValueError as exc:
        return JSONResponse(
            status_code=400,
            content={"error": "invalid_duration", "message": str(exc)},
        )
    except Exception as exc:  # noqa: BLE001
        log.warning("guest_pass create failed: %s", exc)
        return JSONResponse(
            status_code=500,
            content={"error": "create_failed"},
        )
    log.info(
        "admin action: guest_pass minted id=%d by %s expires=%.0f note=%r",
        pass_row.id,
        session.get("username"),
        pass_row.expires_at_epoch,
        pass_row.note[:40],
    )
    return JSONResponse(
        content={
            "ok": True,
            "token": token,  # plaintext — shown ONCE, never again retrievable
            "pass": pass_row.to_dict(),
        }
    )


@app.get("/api/admin/guest-passes")
async def get_admin_guest_passes(request: Request):
    """List recent guest passes.  Query param ``activeOnly=1`` filters
    to currently-valid passes only.  Plaintext tokens are NEVER
    returned — only the hash-derived metadata.
    """
    session_or_err = _require_admin_session(request)
    if isinstance(session_or_err, JSONResponse):
        return session_or_err
    active_only = request.query_params.get("activeOnly") == "1"
    rows = await run_in_threadpool(
        _guest_passes.list_passes,
        include_inactive=not active_only,
    )
    return JSONResponse(
        content={
            "passes": [r.to_dict() for r in rows],
        }
    )


@app.post("/api/admin/guest-pass/{pass_id:int}/revoke")
async def post_admin_guest_pass_revoke(pass_id: int, request: Request):
    """Mark a guest pass revoked.  Active sessions minted from the
    pass remain valid until their next request hits
    ``_get_auth_session`` and finds the pass's per-session
    ``expires_at_epoch`` already past — but a revoke specifically
    flips the ``validate`` gate False, so NO new sessions can be
    minted.  For an immediate boot, follow with
    ``POST /api/admin/sessions/force-logout-all``.
    """
    session_or_err = _require_admin_session(request)
    if isinstance(session_or_err, JSONResponse):
        return session_or_err
    session = session_or_err
    ok = await run_in_threadpool(_guest_passes.revoke, pass_id)
    log.info(
        "admin action: guest_pass revoke id=%d by %s ok=%s",
        pass_id,
        session.get("username"),
        ok,
    )
    return JSONResponse(content={"ok": ok, "id": pass_id})


@app.post("/api/admin/signal-state/migrate")
async def post_admin_signal_state_migrate(request: Request):
    """One-shot migration: legacy ``signalAlertState`` →
    ``signalAlertStateByLeague[defaultLeagueKey]`` for every user.
    Idempotent."""
    session_or_err = _require_admin_session(request)
    if isinstance(session_or_err, JSONResponse):
        return session_or_err
    session = session_or_err
    from src.api import signal_state_migration as _mig

    default_cfg = _league_registry.get_default_league()
    if default_cfg is None:
        return JSONResponse(
            status_code=500,
            content={"error": "no_default_league"},
        )
    result = _mig.migrate_all(default_league_key=default_cfg.key)
    log.info(
        "admin action: signal-state migrate by %s — counts=%s",
        session.get("username"),
        result.get("counts"),
    )
    return JSONResponse(content=result)


@app.get("/api/player/{sleeper_id}/realized")
async def get_player_realized(sleeper_id: str, request: Request):
    """Return realized weekly fantasy points for a player against the
    authed user's active league scoring settings.

    Gated on ``realized_points_api``, which defaults **ON**
    (``src/api/feature_flags.py``).  This docstring previously said
    "default OFF"; it was wrong, and the difference mattered — it made a
    live defect read as dormant.  ``nfl_data_ingest`` is also needed to
    fetch stats, which is why this returns an empty weeks list with a
    clear ``reason`` rather than 500-ing when none are available.

    What kept the row-filter bug below invisible was not the flag but
    the absence of a caller: nothing in the frontend requests this.
    """
    from src.api import feature_flags as _ff

    session = _get_auth_session(request)
    if not session:
        return JSONResponse(status_code=401, content={"error": "auth_required"})
    if not _ff.is_enabled("realized_points_api"):
        return JSONResponse(
            status_code=503,
            content={
                "error": "feature_disabled",
                "flag": "realized_points_api",
            },
        )
    try:
        league_cfg = _resolve_league_for_request(request)
    except LeagueResolutionError as err:
        return err.json_response()

    # Scoring settings for the REQUESTED league.
    #
    # This used to read ``latest_contract_data["sleeper"]["scoringSettings"]``
    # unconditionally — the LOADED league's card — so asking about a
    # player in a second league scored their weeks under the first
    # league's rules and stamped the answer with the requested
    # ``leagueKey``. That is the W18-F002 shape (requested-league
    # identity over another league's config) on a route B6 did not
    # enumerate.
    #
    # The per-league snapshot is the right owner. It is deliberately not
    # freshness-gated here: freshness governs whether one league's
    # RANKINGS may be reused for another, and this is not that question
    # — it is "what are this league's own scoring rules", where a stored
    # card is the best available answer and a stale one is still that
    # league's. The contract is used only when it demonstrably belongs
    # to the requested league.
    scoring_settings = _league_registry.scoring_settings_for_league(league_cfg) or {}
    if not scoring_settings:
        _loaded_meta = (latest_contract_data or {}).get("meta") or {}
        if str(_loaded_meta.get("leagueKey") or "") == league_cfg.key:
            scoring_settings = ((latest_contract_data or {}).get("sleeper") or {}).get(
                "scoringSettings"
            ) or {}
    if not scoring_settings:
        return JSONResponse(
            content={
                "sleeperId": sleeper_id,
                "leagueKey": league_cfg.key,
                "reason": "no_scoring_settings_for_league",
                "weeks": [],
                "totalPoints": 0.0,
                "weekCount": 0,
            }
        )

    # Fetch weekly stats via nfl_data_ingest (already flag-gated —
    # returns [] when nfl_data_ingest is off).  We scope to the
    # current season for freshness + prior season for comparison.
    from src.nfl_data import ingest as _ing
    from src.nfl_data import pbp_weekly as _pbp_weekly
    from src.nfl_data import realized_points as _rp

    now_year = datetime.now(timezone.utc).year
    years = [now_year - 1, now_year]
    weekly = _ing.fetch_weekly_stats(years)

    if not weekly:
        return JSONResponse(
            content={
                "sleeperId": sleeper_id,
                "leagueKey": league_cfg.key,
                "reason": "no_stats_available",
                "weeks": [],
                "totalPoints": 0.0,
                "weekCount": 0,
            }
        )

    # Find this player's GSIS via the unified mapper, then filter.
    from src.identity import unified_mapper as _um

    # The MASTER Sleeper player directory — ``{player_id: {gsis_id,
    # full_name, position, ...}}``, which is what ``resolve_player``
    # indexes.
    #
    # This used to read ``sleeper_block["players"]`` / ``["playerDict"]``.
    # The contract's sleeper block has NEITHER key (it carries
    # ``idToPlayer`` / ``playerIds`` / ``positions``), so ``players_dir``
    # was always ``None`` and every request returned
    # ``reason: "unmapped_player"`` — a well-formed 200 that answered
    # nothing, for every player, always. ``idToPlayer`` could not have
    # substituted either: it is ``{id: name}`` and carries no
    # ``gsis_id``, which is the join key the weekly stat rows use.
    #
    # ``fetch_nfl_players`` is the existing process-cached full dump
    # (~5 MB once per process, ``{}`` on any failure), so this adds no
    # new download path and no second cache.
    from src.public_league.sleeper_client import fetch_nfl_players as _fetch_nfl_players

    players_dir = _fetch_nfl_players()
    resolved = _um.resolve_player(players_dir, sleeper_id=str(sleeper_id))
    if resolved is None or not resolved.gsis_id:
        return JSONResponse(
            content={
                "sleeperId": sleeper_id,
                "leagueKey": league_cfg.key,
                "reason": "unmapped_player",
                "weeks": [],
            }
        )
    # ``player_id``, not ``player_id_gsis``.
    #
    # ``player_id_gsis`` is the field name on the WeeklyStatRow
    # DATACLASS; the raw nflverse rows ``fetch_weekly_stats`` returns
    # use ``player_id``. Filtering on the dataclass name matched ZERO
    # rows for every player, so this endpoint returned a well-formed
    # 200 with an empty ``weeks`` list — for everyone, always.
    #
    # Measured 2026-07-27 on the real 2025 file: the old expression
    # matched 0 rows for a GSIS id the correct one matched 17.
    #
    # It went unnoticed because nothing CALLS this endpoint — the flag
    # defaults ON, so the route has been live and answering wrongly.
    # (An earlier version of this comment said the flag was off. It is
    # not, and believing so made a live defect read as dormant.)
    #
    # Both keys are accepted so a caller passing normalized
    # dataclass-shaped rows still works.
    target_gsis = str(resolved.gsis_id or "").strip()
    player_rows = [
        r
        for r in weekly
        if target_gsis
        and str(r.get("player_id") or r.get("player_id_gsis") or "").strip() == target_gsis
    ]
    # The six reception bands and the player special-teams rules have no
    # column on the weekly feed; without this the response is a total that
    # silently omits them.  With it, a season the artifact does not cover
    # is reported in ``unscored`` rather than scored as zero.
    _pbp_index = _pbp_weekly.SeasonPbpIndex()
    cumulative = _rp.compute_cumulative_points(
        [_pbp_index.attach(r) for r in player_rows],
        scoring_settings,
        position=resolved.position,
    )
    return JSONResponse(
        content={
            "sleeperId": sleeper_id,
            "gsisId": resolved.gsis_id,
            "fullName": resolved.full_name,
            "position": resolved.position,
            "leagueKey": league_cfg.key,
            **cumulative,
        }
    )


@app.post("/api/trade/simulate-mc")
async def post_trade_simulate_mc(request: Request):
    """Monte Carlo trade simulator (Phase 9 of the 2026-04 upgrade).

    Uses the consensus-band distribution from Phase 4 (`valueBand`
    on each player row) to produce a probabilistic view of trade
    outcomes: win probability, delta distribution, range of outcomes.

    Lives alongside `/api/trade/simulate` — the existing endpoint
    is unchanged.  This is additive, behind the `monte_carlo_trade`
    feature flag.  When the flag is off this endpoint returns 503
    `feature_disabled` so clients can fall back cleanly.

    Body::

        {
          "sideA": [{"name": "...", "rankDerivedValue": N,
                     "team": "...", "pos": "...",
                     "valueBand": {"p10":, "p50":, "p90":}}],
          "sideB": [...],
          "nSims": 50000,                        # optional, default 50000
          "sameTeamRho": 0.25,                   # optional correlation knob
          "samePosGroupRho": 0.10,               # optional correlation knob
          "seed": 42,                            # optional for reproducible runs
          "applyConsolidationAdjustment": false  # optional, default false
        }

    Response (on success)::

        {
          "winProbA": 0.62,
          "winProbB": 0.38,
          "meanDelta": 450.2,
          "stdDelta": 1240.5,
          "deltaRange": {"p10": ..., "p50": ..., "p90": ...},
          "nSims": 50000,
          "method": "consensus_based_win_rate",
          "labelHint": "consensus_based_win_rate",
          "disclaimer": "..."
        }

    The ``labelHint`` + ``disclaimer`` fields are PART OF THE
    CONTRACT — the frontend MUST render the disclaimer somewhere
    visible so users don't mis-read win probability as real-world
    odds.
    """
    from src.api import feature_flags as _ff
    from src.trade import monte_carlo as _mc

    session = _get_auth_session(request)
    if not session:
        return JSONResponse(status_code=401, content={"error": "auth_required"})
    if not _ff.is_enabled("monte_carlo_trade"):
        return JSONResponse(
            status_code=503,
            content={
                "error": "feature_disabled",
                "flag": "monte_carlo_trade",
                "message": "Monte Carlo simulator is not yet enabled.",
            },
        )
    try:
        body = await request.json()
    except Exception:
        body = None
    if not isinstance(body, dict):
        return JSONResponse(status_code=400, content={"error": "invalid_body"})
    side_a_raw = body.get("sideA") or []
    side_b_raw = body.get("sideB") or []
    if not isinstance(side_a_raw, list) or not isinstance(side_b_raw, list):
        return JSONResponse(status_code=400, content={"error": "sides_must_be_lists"})
    side_a = [tp for tp in (_mc.build_trade_player(r) for r in side_a_raw) if tp is not None]
    side_b = [tp for tp in (_mc.build_trade_player(r) for r in side_b_raw) if tp is not None]
    try:
        n_sims = int(body.get("nSims") or 50000)
    except (TypeError, ValueError):
        n_sims = 50000
    # Guardrail — clamp to a sane range.  The upper bound bounds the
    # worst-case compute time (see SIMULATE_MC_MAX_SIMS); 50k sims gives
    # a tight enough win-probability estimate that more is just wasted
    # CPU on the event loop's behalf.
    n_sims = max(1000, min(SIMULATE_MC_MAX_SIMS, n_sims))
    try:
        rho_t = float(body.get("sameTeamRho", 0.25))
        rho_p = float(body.get("samePosGroupRho", 0.10))
    except (TypeError, ValueError):
        rho_t, rho_p = 0.25, 0.10
    seed = body.get("seed")
    try:
        seed = int(seed) if seed is not None else None
    except (TypeError, ValueError):
        seed = None
    # Strict boolean check: only JSON `true` enables this.  Strings like
    # "false" or "0" must not accidentally activate consolidation adjustment.
    _raw_ca = body.get("applyConsolidationAdjustment", False)
    apply_ca = _raw_ca is True
    # Symmetrize the direction (A→B averaged with B→A) so ordering
    # never biases the result — critical invariant per the Phase 11
    # integration pass.  Then enrich with the decision-layer fields
    # (valueDelta / adjustedDelta / winPct / riskLevel / tierImpact)
    # the trade calculator UI consumes.
    from src.trade import symmetrize as _sym

    def _run_mc() -> dict:
        base = _sym.simulate_symmetric(
            side_a,
            side_b,
            n_sims=n_sims,
            same_team_rho=rho_t,
            same_pos_group_rho=rho_p,
            seed=seed,
            apply_consolidation_adjustment=apply_ca,
        )
        return _sym.enrich_with_decision_shape(base, side_a, side_b)

    # The Monte Carlo is CPU-bound pure Python.  Run it in the thread
    # pool so it never blocks the event loop (health checks + every
    # other request), with a hard wall-clock backstop.  On timeout the
    # worker thread can't be cancelled mid-loop, but it's bounded by
    # SIMULATE_MC_MAX_SIMS so it finishes within a few seconds anyway —
    # the client gets a clean 504 instead of an open-ended hang.
    try:
        enriched = await asyncio.wait_for(
            run_in_threadpool(_run_mc),
            timeout=SIMULATE_MC_TIMEOUT_SECONDS,
        )
    except asyncio.TimeoutError:
        log.warning(
            "simulate-mc timed out after %ss (n_sims=%d)",
            SIMULATE_MC_TIMEOUT_SECONDS,
            n_sims,
        )
        return JSONResponse(
            status_code=504,
            content={
                "error": "timeout",
                "message": (
                    "Monte Carlo simulation exceeded the time budget. "
                    "Retry with fewer players or sims."
                ),
            },
        )
    return JSONResponse(content=enriched)


@app.post("/api/signal-alerts/run")
async def run_signal_alerts(request: Request):
    """Trigger a signal-alert sweep for every user with email
    notifications enabled.  Admin-only — requires the password
    session method; a Sleeper-login session is NOT authorized
    since that auth method is trust-on-first-use and shouldn't be
    able to trigger mass-email sends.

    The alert runner:
      1. Loads every user from user_kv.
      2. For each with ``notificationsEnabled: true`` AND a valid
         ``notificationsEmail``, builds a terminal payload for that
         user (resolving their team by ``sleeper_user_id`` when
         present) and runs ``signal_alerts.process_user_alerts``.
      3. Returns a summary.

    Wire this to a cron / systemd timer for automated daily digests.
    Cron clients authenticate via the shared ``SIGNAL_ALERT_CRON_TOKEN``
    as a Bearer token; browser clients still need a password session.
    """
    # Two auth paths: (1) a password-session admin from the browser,
    # (2) an opaque bearer token for cron / systemd timers.  Either is
    # sufficient on its own.  Failure mode: if no session AND no token
    # (or token mismatch), reject.  Short-circuit the token check
    # before touching cookies so an unset token can never authorize.
    cron_auth_ok = False
    if SIGNAL_ALERT_CRON_TOKEN:
        header = (request.headers.get("authorization") or "").strip()
        if header.lower().startswith("bearer "):
            presented = header.split(None, 1)[1].strip()
            # Constant-time compare to avoid timing leaks.
            if hmac.compare_digest(presented, SIGNAL_ALERT_CRON_TOKEN):
                cron_auth_ok = True
    if not cron_auth_ok:
        session = _get_auth_session(request)
        if not session or session.get("auth_method") != "password":
            return JSONResponse(
                status_code=401,
                content={"error": "admin_auth_required"},
            )
    if not latest_contract_data:
        return JSONResponse(
            status_code=503,
            content={"error": "no_live_contract"},
        )

    # Walk every user_kv row.  No pagination — at current scale
    # (dozens of users tops) this is fine.  For each user, loop
    # over every active league: build a league-specific terminal
    # payload (via Sleeper overlay for non-default leagues) and
    # run the alert detector with the league_key scoped.  Cooldowns
    # are now nested per league so a SELL in league A doesn't
    # silently eat a SELL in league B for the same player.
    def _run_sweep() -> dict[str, object]:
        db = _user_kv.all_user_states()
        summary: list[dict] = []

        # BDVM market-signal transitions piggyback on this same sweep
        # (flag-gated; no extra timer).  The whole-league board + the
        # ownerId → rostered-playerIds map are computed ONCE per league
        # and shared across users — the engine run is the expensive
        # part, the per-user slice is a set lookup.  Any failure here
        # degrades to a "bdvm": {"error": ...} field, never a failed
        # sweep.
        bdvm_enabled = False
        try:
            from src.api import feature_flags as _ff  # noqa: PLC0415

            bdvm_enabled = _ff.is_enabled("bdvm_engine")
        except Exception as exc:  # noqa: BLE001
            log.warning("bdvm flag check failed in signal sweep: %s", exc)
        bdvm_league_cache: dict[str, tuple[dict | None, dict[str, set]]] = {}

        # News → structured-events ingest rides the same daily sweep
        # (no extra timer).  Runs BEFORE the per-user loop so freshly
        # ingested events feed this very sweep's valuations — the
        # events-file fingerprint in bdvm_api's cache key makes the
        # write invalidate any earlier cached board.  Auto events land
        # in the §7 speculation lane (confidence < 0.5): they can only
        # widen uncertainty, never move a mean.
        bdvm_news_events_summary: dict | None = None
        if bdvm_enabled:
            try:
                from src.bdvm import news_events as _bdvm_news_events  # noqa: PLC0415
                from src.utils.name_clean import normalize_player_name  # noqa: PLC0415

                _news_svc = _get_news_service()
                _aggregated = _news_svc.aggregate(
                    player_names=_live_player_names(),
                    player_meta=_live_player_meta(),
                )
                _season = int(
                    (latest_contract_data or {}).get("currentDraftYear")
                    or datetime.now(timezone.utc).year
                )
                bdvm_news_events_summary = _bdvm_news_events.ingest_news_events(
                    [it.to_dict() for it in _aggregated.items],
                    season=_season,
                    name_normalizer=normalize_player_name,
                )
            except Exception as exc:  # noqa: BLE001
                log.warning("bdvm news-events ingest failed: %s", exc)
                bdvm_news_events_summary = {"ok": False, "error": str(exc)}

        def _bdvm_league_data(cfg_key: str, contract: dict):
            if cfg_key in bdvm_league_cache:
                return bdvm_league_cache[cfg_key]
            values: dict | None = None
            assets_by_owner: dict[str, set] = {}
            try:
                from src.api import bdvm_api as _bdvm_api  # noqa: PLC0415

                values = _bdvm_api.get_bdvm_values(contract, cfg_key)
                roster = _bdvm_api.get_bdvm_roster(contract, cfg_key)
                if roster.get("status") == "ok":
                    for r in roster.get("rosters") or []:
                        oid = str(r.get("ownerId") or "")
                        if oid:
                            assets_by_owner[oid] = {
                                str(a.get("playerId") or "") for a in (r.get("assets") or [])
                            }
            except Exception as exc:  # noqa: BLE001
                log.warning("bdvm signal sweep data failed for %s: %s", cfg_key, exc)
                values = None
            bdvm_league_cache[cfg_key] = (values, assets_by_owner)
            return values, assets_by_owner

        loaded_league = (
            (latest_contract_data or {}).get("meta", {}).get("leagueKey")
            if isinstance(latest_contract_data, dict)
            else None
        )
        loaded_sleeper = (
            latest_contract_data.get("sleeper") or {}
            if isinstance(latest_contract_data, dict)
            else {}
        )
        active_leagues = _league_registry.active_leagues()
        for username, state in db.items():
            if not isinstance(state, dict):
                continue
            if not state.get("notificationsEnabled"):
                continue
            email = str(state.get("notificationsEmail") or "").strip()
            if not email:
                continue
            owner_id = str((state.get("selectedTeam") or {}).get("ownerId") or "")
            selected_teams = state.get("selectedTeamsByLeague") or {}
            if not isinstance(selected_teams, dict):
                selected_teams = {}
            user_summary: list[dict] = []
            for cfg in active_leagues:
                # Skip leagues the user isn't in — if there's no
                # team-map entry and the contract doesn't resolve a
                # team, they have nothing to alert on here.
                league_entry = selected_teams.get(cfg.key) or {}
                league_owner_id = str(league_entry.get("ownerId") or "").strip() or owner_id
                # Build the league-specific contract.  For the
                # loaded league this is just latest_contract_data;
                # for other active leagues we splice in the overlay.
                if cfg.key == loaded_league:
                    contract = latest_contract_data
                elif _scoring_identity_error(latest_contract_data, cfg) is None:
                    id_map = (
                        loaded_sleeper.get("idToPlayer") if isinstance(loaded_sleeper, dict) else {}
                    )
                    try:
                        overlay = _sleeper_overlay.fetch_sleeper_overlay(
                            sleeper_league_id=cfg.sleeper_league_id,
                            id_to_player=id_map if isinstance(id_map, dict) else {},
                        )
                    except Exception as exc:  # noqa: BLE001
                        log.warning(
                            "signal-alerts overlay failed for %s / %s: %s",
                            username,
                            cfg.key,
                            exc,
                        )
                        overlay = None
                    if not overlay or not overlay.get("teams"):
                        # No data for this league — skip, not a
                        # failure (e.g. Sleeper transient error).
                        continue
                    hybrid_sleeper, _ = _sleeper_overlay.merge_cross_league_sleeper_block(
                        loaded_sleeper=loaded_sleeper if isinstance(loaded_sleeper, dict) else {},
                        overlay=overlay,
                        requested_league_config=overlay.get("leagueConfig"),
                    )
                    contract = {**latest_contract_data, "sleeper": hybrid_sleeper}
                else:
                    # Scoring not proven to match — rankings aren't
                    # comparable, skip this league for this run.
                    continue

                team = _terminal.resolve_team(
                    contract,
                    owner_id=league_owner_id,
                    name=None,
                )
                if team is None:
                    # User isn't in this league — nothing to alert on.
                    continue
                try:
                    payload = _terminal.build_terminal_payload(
                        contract,
                        resolved_team=team,
                        window_days=30,
                        user_state=state,
                        league_key=cfg.key,
                    )
                except Exception as exc:  # noqa: BLE001
                    user_summary.append(
                        {
                            "leagueKey": cfg.key,
                            "ok": False,
                            "reason": f"build_error:{type(exc).__name__}",
                        }
                    )
                    continue
                result = _signal_alerts.process_user_alerts(
                    username,
                    signals=payload.get("signals") or [],
                    display_name=username,
                    email=email,
                    delivery=_deliver_email_smtp,
                    league_key=cfg.key,
                )
                league_summary = {"leagueKey": cfg.key, **result}
                if bdvm_enabled:
                    try:
                        from src.api import bdvm_signal_alerts as _bdvm_alerts  # noqa: PLC0415

                        values, assets_by_owner = _bdvm_league_data(cfg.key, contract)
                        entries = _bdvm_alerts.roster_bdvm_entries(
                            values, assets_by_owner.get(league_owner_id) or ()
                        )
                        league_summary["bdvm"] = _bdvm_alerts.process_user_bdvm_alerts(
                            username,
                            entries=entries,
                            display_name=username,
                            email=email,
                            delivery=_deliver_email_smtp,
                            league_key=cfg.key,
                        )
                    except Exception as exc:  # noqa: BLE001
                        log.warning(
                            "bdvm signal alerts failed for %s / %s: %s",
                            username,
                            cfg.key,
                            exc,
                        )
                        league_summary["bdvm"] = {"error": str(exc)}
                user_summary.append(league_summary)
            if user_summary:
                summary.append({"username": username, "byLeague": user_summary})
        return {
            "total_users_checked": len(db),
            "processed": len(summary),
            "results": summary,
            "bdvmEnabled": bdvm_enabled,
            "bdvmNewsEvents": bdvm_news_events_summary,
        }

    result = await run_in_threadpool(_run_sweep)

    # Operator alerts piggyback on the signal-alert cron so we don't
    # need another timer.  Checks: scrape success rate, circuit
    # breakers, contract health, data freshness.  Never raises.
    try:
        from src.api import ops_alerts as _ops
        from src.utils import circuit_breaker as _cb

        status_payload = _scrape_status_payload()
        # Board age, not process age (audit F-19).  This one fed the ALERTER,
        # so the wrong number here did not merely mislead a dashboard.
        data_age_hours = _board_age_hours()
        ops_summary = _ops.check_and_alert(
            status_payload=status_payload,
            circuit_snapshots=_cb.snapshot_all(),
            contract_health=contract_health,
            data_age_hours=data_age_hours,
            scrape_interval_hours=float(SCRAPE_INTERVAL_HOURS),
            delivery=_deliver_email_smtp if ALERT_TO else None,
            to_email=ALERT_TO or None,
        )
        result["opsAlerts"] = ops_summary
    except Exception as exc:  # noqa: BLE001
        log.warning("ops_alerts check failed: %s", exc)
        result["opsAlerts"] = {"error": str(exc)}

    # Per-source staleness alerts (G1 from docs/automation-audit.md).
    # Catches the "IDP Show prod cookie expired and the source went
    # silently stale for two weeks" failure mode by per-source SLA.
    # Thresholds live in ``config/source_staleness.json``; cooldown
    # state persists in ``user_kv`` under ``_system_source_health``
    # so we don't re-fire on every sweep.
    try:
        from src.api import source_health_alerts as _sha

        source_health = _build_source_health_snapshot(latest_data or latest_contract_data)
        if source_health:
            stale_summary = _sha.check_and_alert(
                source_health,
                delivery=_deliver_email_smtp if ALERT_TO else None,
                to_email=ALERT_TO or None,
            )
            result["sourceStalenessAlerts"] = stale_summary
    except Exception as exc:  # noqa: BLE001
        log.warning("source_health_alerts check failed: %s", exc)
        result["sourceStalenessAlerts"] = {"error": str(exc)}

    return JSONResponse(content=result, headers={"Cache-Control": "no-store"})


# ── LEAGUE NARRATIVE ARTICLES (preview / recap) ─────────────────────
# Public read endpoints + admin-only generation trigger.  Articles are
# persisted to ``exports/narratives/<season>/week-<NN>/<mode>-<id>.json``
# by the cron generator (see ``scripts/generate_weekly_narratives.py``
# and ``.github/workflows/weekly-narratives.yml``).  These endpoints
# serve them — they do NOT generate on read so a slow Anthropic round
# trip never blocks a page load.


@app.get("/api/league/articles")
async def get_league_articles(request: Request):
    """List narrative articles, optionally filtered by season/week.

    Query params:
      * ``season`` (optional) — restrict to one season label.
      * ``week`` (optional) — restrict to one week within season.

    Returns a flat array of {season, week, mode, matchupId, title,
    generatedAt, home, away, kicker} so the frontend list page can
    render a slate without N follow-up requests.  Heavy fields
    (body, lede) are omitted from the index response — clients fetch
    them via the single-article endpoint.
    """
    from src.public_league import matchup_narrative as _mn

    season = (request.query_params.get("season") or "").strip() or None
    week_raw = (request.query_params.get("week") or "").strip()
    try:
        week_filter = int(week_raw) if week_raw else None
    except ValueError:
        return JSONResponse(
            status_code=400,
            content={"error": "bad_request", "message": "week must be an integer"},
        )

    items = _mn.list_articles(season=season)
    if week_filter is not None:
        items = [r for r in items if int(r.get("week") or -1) == week_filter]

    # Hydrate index entries with the small subset of article fields
    # needed for a slate list (title, kicker, home/away identity).
    enriched = []
    for entry in items:
        full = _mn.load_article(
            entry["season"],
            entry["week"],
            entry["matchupId"],
            entry["mode"],
        )
        if not full:
            continue
        enriched.append(
            {
                "season": entry["season"],
                "week": entry["week"],
                "mode": entry["mode"],
                "matchupId": entry["matchupId"],
                "title": full.get("title"),
                "kicker": full.get("kicker"),
                "angleUsed": full.get("angleUsed"),
                "isChampionship": full.get("isChampionship", False),
                "roundLabel": full.get("roundLabel", ""),
                "home": full.get("home", {}),
                "away": full.get("away", {}),
                "generatedAt": full.get("generatedAt"),
                "wordCount": full.get("wordCount", 0),
            }
        )

    return JSONResponse(
        content={
            "articles": enriched,
            "total": len(enriched),
            "season": season,
            "week": week_filter,
        },
        headers={"Cache-Control": "public, max-age=120"},
    )


@app.get("/api/league/articles/{season}/{week}/{matchup_id}/{mode}")
async def get_league_article(season: str, week: int, matchup_id: int, mode: str):
    """Single article — full body, ready to render."""
    from src.public_league import matchup_narrative as _mn

    if mode not in {"preview", "recap"}:
        return JSONResponse(
            status_code=400,
            content={"error": "bad_request", "message": "mode must be preview|recap"},
        )
    article = _mn.load_article(season, week, matchup_id, mode)
    if article is None:
        return JSONResponse(
            status_code=404,
            content={
                "error": "not_found",
                "message": (f"No {mode} article on disk for {season} W{week} matchup {matchup_id}"),
            },
        )
    return JSONResponse(
        content=article,
        # Cache for 5 minutes; new articles propagate via cron not on-demand.
        headers={"Cache-Control": "public, max-age=300"},
    )


@app.post("/api/league/articles/generate")
async def post_generate_league_article(request: Request):
    """Admin trigger: generate a single article on demand.

    Body shape:
        {
          "season": "2025",         // optional, defaults to current season
          "week": 17,                // optional, detector picks live week
          "matchupId": 1,            // required for single-article runs
          "mode": "preview" | "recap",
          "force": true|false        // optional, default false
        }

    Returns the generated article on success, 404 if the matchup
    can't be found in the snapshot, 503 if the Anthropic SDK isn't
    configured.

    NOTE: this is the synchronous, single-matchup path — full slate
    generation is the cron's job.  Hold a session at the wheel; this
    will block for 10-30 seconds while Claude generates.
    """
    session_or_err = _require_admin_session(request)
    if isinstance(session_or_err, JSONResponse):
        return session_or_err

    try:
        body = await request.json()
    except Exception:  # noqa: BLE001
        body = {}
    if not isinstance(body, dict):
        body = {}

    mode = str(body.get("mode") or "").strip().lower()
    if mode not in {"preview", "recap"}:
        return JSONResponse(
            status_code=400,
            content={"error": "bad_request", "message": "mode must be preview|recap"},
        )
    matchup_id = body.get("matchupId")
    if matchup_id is None:
        return JSONResponse(
            status_code=400,
            content={"error": "bad_request", "message": "matchupId required"},
        )
    try:
        matchup_id = int(matchup_id)
    except (TypeError, ValueError):
        return JSONResponse(
            status_code=400,
            content={"error": "bad_request", "message": "matchupId must be an integer"},
        )
    explicit_week = body.get("week")
    if explicit_week is not None:
        try:
            explicit_week = int(explicit_week)
        except (TypeError, ValueError):
            return JSONResponse(
                status_code=400,
                content={"error": "bad_request", "message": "week must be an integer"},
            )
    explicit_season = body.get("season")
    explicit_season = str(explicit_season) if explicit_season else None
    force = bool(body.get("force"))

    # Resolve league via the standard resolver (so admins can override
    # via ?leagueKey=).
    try:
        league_cfg = _resolve_league_for_request(request, body=body)
    except LeagueResolutionError as err:
        return err.json_response()

    if anthropic is None:
        return JSONResponse(
            status_code=503,
            content={"error": "anthropic_unavailable", "message": "anthropic SDK not installed"},
        )
    api_key = os.getenv("ANTHROPIC_API_KEY", "").strip()
    if not api_key:
        return JSONResponse(
            status_code=503,
            content={
                "error": "anthropic_unavailable",
                "message": "ANTHROPIC_API_KEY not configured",
            },
        )

    # Build the snapshot off the event loop — it does ~85 HTTP GETs
    # against Sleeper.
    from src.public_league import matchup_narrative as _mn
    from src.public_league.snapshot import build_public_snapshot

    snapshot = await run_in_threadpool(
        build_public_snapshot,
        league_cfg.sleeper_league_id,
    )
    current = snapshot.current_season
    if current is None:
        return JSONResponse(
            status_code=503,
            content={"error": "snapshot_unavailable", "message": "Sleeper snapshot empty"},
        )

    season = explicit_season or current.season
    if explicit_week is not None:
        week = explicit_week
    else:
        from src.public_league import matchup_preview as _mp

        detected_week, detected_mode = _mp._detect_current_week(current)  # noqa: SLF001
        if detected_week == 0:
            return JSONResponse(
                status_code=404,
                content={"error": "week_not_found", "message": "no live week detected"},
            )
        if mode == "recap" and detected_mode == "preview":
            week = max(1, detected_week - 1)
        else:
            week = detected_week

    if not force:
        existing = _mn.load_article(season, week, matchup_id, mode)
        if existing is not None:
            return JSONResponse(
                status_code=200,
                content={"article": existing, "regenerated": False},
            )

    brief = _mn.build_brief(
        snapshot,
        season=season,
        week=week,
        matchup_id=matchup_id,
        mode=mode,
    )
    if brief is None:
        return JSONResponse(
            status_code=404,
            content={
                "error": "matchup_not_found",
                "message": f"no matchup_id={matchup_id} in {season} W{week}",
            },
        )
    prior = _mn.collect_prior_articles(season, n=6)
    client = anthropic.AsyncAnthropic(api_key=api_key)
    try:
        article = await _mn.generate_article(
            client=client,
            brief=brief,
            prior_articles=prior,
        )
    except Exception as exc:  # noqa: BLE001 — collapse SDK / network / parse to one structured error
        # generate_article raises RuntimeError for malformed JSON, but
        # SDK layer can raise APIStatusError / RateLimitError /
        # APIConnectionError / asyncio.TimeoutError before parsing.
        # All of these are operational failures the admin caller wants
        # to retry against, not unstructured 500s.
        log.warning(
            "league_article_generation_failed",
            extra={
                "season": season,
                "week": week,
                "matchupId": matchup_id,
                "mode": mode,
                "error_type": type(exc).__name__,
            },
        )
        return JSONResponse(
            status_code=502,
            content={
                "error": "generation_failed",
                "errorType": type(exc).__name__,
                "message": str(exc),
            },
        )
    _mn.save_article(article)
    return JSONResponse(
        status_code=200,
        content={"article": article, "regenerated": True},
    )


# Static file mount for backend-generated assets (CSS, images if any).
if STATIC_DIR.exists():
    app.mount("/static", StaticFiles(directory=str(STATIC_DIR)), name="static")


# ── SLEEPER INTEL — "Sharp Tracker" (Phase 5) ───────────────────────────
#
# PFK-style league-mate market intelligence: crawls the active
# league's members across ALL of their other Sleeper leagues (public
# API, no auth) and aggregates per-asset buy/sell activity.  All
# pipeline logic lives in ``src/intel/``; this section is only the
# HTTP surface.
#
# Intel is roster-scoped, therefore LEAGUE-scoped (CLAUDE.md): every
# read resolves the requested league via _resolve_league_for_request
# and serves that league's snapshot partition
# (data/intel/snapshot_<leagueKey>.json).  A league with no snapshot
# yet gets the standard 503 data_not_ready.
#
# Endpoints (all PRIVATE — no public cache headers; raw Sleeper
# league IDs are never exposed in responses):
#   GET  /api/intel/summary          — asset board sorted by signalStrength
#   GET  /api/intel/player           — per-asset drill-down (?playerId= / ?name=)
#   GET  /api/intel/member/{ownerId} — one member's cross-league profile
#   POST /api/intel/refresh          — 202 + daemon-thread crawl (409 when running)
#   GET  /api/intel/refresh/status   — crawl status + snapshot staleness
#
# The refresh + status routes accept EITHER a session cookie or a
# bearer token (``INTEL_REFRESH_TOKEN``, falling back to
# ``SIGNAL_ALERT_CRON_TOKEN``) so the daily GitHub Actions cron
# (.github/workflows/intel-refresh.yml) can drive them.  They are
# added to the self-authed allowlist below so the session-cookie
# middleware defers to the endpoints' own auth check — same model as
# /api/signal-alerts/run.

from src.intel import service as _intel_service  # noqa: E402

_SELF_AUTHED_API_EXACT = _SELF_AUTHED_API_EXACT | {
    "/api/intel/refresh",
    "/api/intel/refresh/status",
}

INTEL_REFRESH_TOKEN = os.getenv("INTEL_REFRESH_TOKEN", "").strip() or SIGNAL_ALERT_CRON_TOKEN

_INTEL_PRIVATE_CACHE_HEADERS = {"Cache-Control": "private, max-age=60, stale-while-revalidate=300"}

# Allow-listed so a query string cannot reach an arbitrary window name
# (signals.window_bounds raises on unknown ones) or an unsupported sort.
_INTEL_DEFAULT_WINDOW = "30d"
_INTEL_ALLOWED_WINDOWS = ("7d", "30d", "90d", "all")
_INTEL_ALLOWED_SORTS = ("net", "volume", "buys", "sells", "strength", "velocity")

# ── Manual-refresh cooldown (D13) ────────────────────────────────────
#
# A refresh spends hundreds of budgeted Sleeper calls over several
# minutes.  The process lock stops CONCURRENT crawls (409), but nothing
# stopped a signed-in user from re-triggering the moment each run
# finished — an unbounded serial drain on someone else's API, from any
# account with a session.
#
# So manual triggers get a per-user cooldown.  The CRON IS EXEMPT: it
# authenticates by bearer token and is the intended scheduled driver,
# and throttling it would defeat the schedule it exists to keep.
_INTEL_MANUAL_REFRESH_COOLDOWN_SEC = float(os.getenv("INTEL_MANUAL_REFRESH_COOLDOWN_SEC", "600"))
_intel_manual_refresh_at: dict[str, float] = {}
_intel_manual_refresh_lock = threading.Lock()


def _intel_refresh_cooldown_remaining(user_key: str) -> int:
    """Seconds left on this user's cooldown, 0 when clear."""
    if _INTEL_MANUAL_REFRESH_COOLDOWN_SEC <= 0:
        return 0
    now = time.monotonic()
    with _intel_manual_refresh_lock:
        last = _intel_manual_refresh_at.get(user_key)
    if last is None:
        return 0
    elapsed = now - last
    if elapsed >= _INTEL_MANUAL_REFRESH_COOLDOWN_SEC:
        return 0
    return int(_INTEL_MANUAL_REFRESH_COOLDOWN_SEC - elapsed) + 1


def _intel_refresh_mark_triggered(user_key: str) -> None:
    """Stamp the cooldown.  Called only AFTER a crawl actually starts,
    so a 409 (someone else already crawling) does not burn the caller's
    window for work they never got."""
    with _intel_manual_refresh_lock:
        _intel_manual_refresh_at[user_key] = time.monotonic()
        # Bound the dict — this is a long-lived process and the key set
        # grows with distinct users.
        if len(_intel_manual_refresh_at) > 512:
            cutoff = time.monotonic() - _INTEL_MANUAL_REFRESH_COOLDOWN_SEC
            for key in [k for k, v in _intel_manual_refresh_at.items() if v < cutoff]:
                _intel_manual_refresh_at.pop(key, None)


def _intel_refresh_reset_for_tests() -> None:
    with _intel_manual_refresh_lock:
        _intel_manual_refresh_at.clear()


# Rate limit for the bearer-rejection warnings below: they fire on
# UNAUTHENTICATED requests, so an unthrottled log line would be a
# journal-spam vector from the public host.  At most one line per
# interval, tracked with a cheap monotonic timestamp guard.
_INTEL_AUTH_LOG_INTERVAL_SEC = 60.0
_intel_auth_log_last_monotonic = 0.0


def _intel_auth_log(message: str, *args) -> None:
    global _intel_auth_log_last_monotonic
    now = time.monotonic()
    if now - _intel_auth_log_last_monotonic < _INTEL_AUTH_LOG_INTERVAL_SEC:
        return
    _intel_auth_log_last_monotonic = now
    log.warning(message, *args)


def _intel_bearer_auth_ok(request: Request) -> bool:
    """True when the request carries the intel cron bearer token.

    Logs WHICH branch rejected a presented bearer (never the token
    itself, and no metadata about the configured secret beyond a
    length-EQUALITY boolean — the journal is quoted in ops issues
    per the workflow's 401 checklist, so even the secret's length
    must not leak) so a cron 401 is diagnosable from the journal:
    "no token configured" means the .env var never reached this
    process (check the dynasty unit's EnvironmentFile + that the
    line is plain ``KEY=value`` — systemd ignores ``export`` lines —
    and that the service restarted after the edit); "mismatch"
    means the workflow secret and the server value differ.
    """
    header = (request.headers.get("authorization") or "").strip()
    has_bearer = header.lower().startswith("bearer ")
    if not INTEL_REFRESH_TOKEN:
        if has_bearer:
            _intel_auth_log(
                "intel bearer auth rejected: token presented but neither "
                "INTEL_REFRESH_TOKEN nor SIGNAL_ALERT_CRON_TOKEN is configured "
                "server-side (env not loaded?)"
            )
        return False
    if not has_bearer:
        return False
    presented = header.split(None, 1)[1].strip()
    # Compare as BYTES: Starlette decodes header values as latin-1,
    # and ``hmac.compare_digest`` raises TypeError on non-ASCII str
    # input — which would turn a garbage bearer into an unhandled
    # 500 instead of a 401.  surrogateescape keeps any str encodable.
    presented_b = presented.encode("utf-8", "surrogateescape")
    configured_b = INTEL_REFRESH_TOKEN.encode("utf-8", "surrogateescape")
    if hmac.compare_digest(presented_b, configured_b):
        return True
    _intel_auth_log(
        "intel bearer auth rejected: presented token does not match the "
        "configured INTEL_REFRESH_TOKEN (lengths match: %s)",
        len(presented) == len(INTEL_REFRESH_TOKEN),
    )
    return False


def _intel_id_to_player() -> dict:
    """NFL-wide Sleeper-ID → display-name map from the loaded
    contract (empty when no contract is loaded yet — asset rows then
    fall back to raw-id labels, never an error)."""
    try:
        sleeper_block = (latest_contract_data or {}).get("sleeper") or {}
        id_map = sleeper_block.get("idToPlayer")
        return id_map if isinstance(id_map, dict) else {}
    except Exception:  # noqa: BLE001
        return {}


def _intel_name_to_player_id(name: str) -> str | None:
    """Reverse lookup: display name → Sleeper player id (first match,
    case-insensitive) via the loaded contract's id map."""
    target = str(name or "").strip().lower()
    if not target:
        return None
    for pid, label in _intel_id_to_player().items():
        if str(label or "").strip().lower() == target:
            return str(pid)
    return None


def _intel_not_ready_response(league_key: str) -> JSONResponse:
    """503 ``data_not_ready`` — the requested league has no intel
    snapshot yet (no refresh has completed for it).  Same convention
    as the other league-scoped endpoints."""
    return JSONResponse(
        status_code=503,
        content={
            "error": "data_not_ready",
            "message": (
                f"No intel snapshot for league {league_key!r} yet — "
                "trigger POST /api/intel/refresh or wait for the daily cron."
            ),
            "leagueKey": league_key,
        },
        headers={"Cache-Control": "no-store"},
    )


@app.get("/api/intel/summary")
async def get_intel_summary(request: Request):
    """Insider Trading board: per-asset trade buy/sell/net/volume over an
    explicit window, served from the normalized ledger.

    League-scoped (intel is roster-scoped → league-scoped): resolves the
    requested league, and every ledger query is filtered to that
    league's member pool — the ledger itself is global.

    ``window`` selects ONE lens (7d/30d/90d/all, default 30d) and
    ``sort`` the ordering.  Counts are TRADES ONLY; waiver and
    free-agent activity is served separately by
    ``/api/intel/waiver-interest`` so a claim can never render as a buy.
    """
    try:
        league_cfg = _resolve_league_for_request(request)
    except LeagueResolutionError as err:
        return err.json_response()
    if not await run_in_threadpool(_intel_service.snapshot_ready, league_cfg.key):
        return _intel_not_ready_response(league_cfg.key)
    try:
        limit = int(request.query_params.get("limit") or 100)
    except (TypeError, ValueError):
        limit = 100
    limit = max(1, min(500, limit))
    window = (request.query_params.get("window") or "").strip() or _INTEL_DEFAULT_WINDOW
    if window not in _INTEL_ALLOWED_WINDOWS:
        window = _INTEL_DEFAULT_WINDOW
    sort = (request.query_params.get("sort") or "").strip() or "net"
    if sort not in _INTEL_ALLOWED_SORTS:
        sort = "net"
    payload = await run_in_threadpool(
        _intel_service.build_summary_payload,
        league_cfg.key,
        limit=limit,
        id_to_player=_intel_id_to_player(),
        window=window,
        sort=sort,
    )
    return JSONResponse(content=payload, headers=_INTEL_PRIVATE_CACHE_HEADERS)


def _intel_contract_context() -> dict:
    """Position, value and team maps from the loaded contract.

    Every piece is optional: a missing contract yields empty maps, which
    makes the partner-fit and value-match terms abstain rather than
    taking the whole lead list down with them.
    """
    try:
        contract = latest_contract_data or {}
        sleeper_block = contract.get("sleeper") or {}
        teams = sleeper_block.get("teams") or []
        positions = {}
        values = {}
        for row in contract.get("playersArray") or []:
            if not isinstance(row, dict):
                continue
            pid = str(row.get("sleeperId") or row.get("playerId") or "").strip()
            if not pid:
                continue
            pos = str(row.get("position") or "").strip().upper()
            if pos:
                positions[pid] = pos
            val = row.get("rankDerivedValue")
            if isinstance(val, (int, float)) and val > 0:
                values[pid] = float(val)
        return {"teams": teams, "positions": positions, "values": values}
    except Exception:  # noqa: BLE001 — leads must survive a missing contract
        log.exception("intel: contract context unavailable")
        return {"teams": [], "positions": {}, "values": {}}


def _build_intel_leads(league_cfg, asset_id: str, mode: str) -> dict:
    """Assemble one lead list.  Runs in a threadpool — pure sync."""
    from src.intel import lead_service, roster_shape

    ctx = _intel_contract_context()
    teams = ctx["teams"]
    positions = ctx["positions"]
    values = ctx["values"]

    signals_by_owner = roster_shape.team_signals(
        teams, positions, league_cfg.roster_settings, value_by_player=values
    )
    owner = roster_shape.owner_of_player(teams, asset_id)
    matchable = roster_shape.matchable_values(teams, values)

    # SELL mode asks who WANTS what I hold, so it scores their BUYING.
    # BUY mode asks who would PART with it, so it scores their SELLING.
    direction = "buy" if mode == "sell" else "sell"

    # In sell mode the caller owns the asset, so they are not a lead for
    # their own player; in buy mode the current owner IS the lead.
    exclude = []
    our_signal = None
    if mode == "sell" and owner:
        exclude = [owner]
        our_signal = signals_by_owner.get(owner)

    return lead_service.build_leads(
        league_key=league_cfg.key,
        asset_id=asset_id,
        position=positions.get(str(asset_id)),
        direction=direction,
        roster_signals=signals_by_owner,
        our_roster=our_signal,
        target_value=values.get(str(asset_id)),
        matchable_values_by_owner=matchable,
        owner_of_asset=owner,
        home_league_ids=[str(league_cfg.sleeper_league_id or "")],
        exclude_owner_ids=exclude,
    )


@app.post("/api/intel/leads")
async def post_intel_leads(request: Request):
    """Insider Trading leads for ONE asset.

    Body: ``{"assetId": "...", "mode": "sell"|"buy", "leagueKey": "..."}``

    SELL mode ("I want to move X") ranks league-mates by how much they
    look like they want X.  BUY mode ("I want X") surfaces the current
    owner and what they look likely to want back.

    Both read the SAME cross-league observations from opposite sides.
    The score is a RANKING of who to approach, never a probability that
    anyone accepts — ``limitations`` in the payload says so, and
    declined offers are not recorded by Sleeper at all.
    """
    try:
        body = await request.json()
    except Exception:  # noqa: BLE001
        body = {}
    if not isinstance(body, dict):
        body = {}

    # Body parsed BEFORE the resolver so a body leagueKey reaches it —
    # the POST convention used by the other league-scoped endpoints.
    try:
        league_cfg = _resolve_league_for_request(request, body=body)
    except LeagueResolutionError as err:
        return err.json_response()

    asset_id = str(body.get("assetId") or body.get("playerId") or "").strip()
    if not asset_id:
        name = str(body.get("name") or "").strip()
        if name:
            asset_id = _intel_name_to_player_id(name) or ""
    if not asset_id:
        return JSONResponse(
            status_code=400,
            content={
                "error": "missing_asset",
                "message": "Provide assetId (Sleeper id or pick:<season>:<round>) or name.",
            },
            headers={"Cache-Control": "no-store"},
        )

    mode = str(body.get("mode") or "sell").strip().lower()
    if mode not in ("sell", "buy"):
        mode = "sell"

    if not await run_in_threadpool(_intel_service.snapshot_ready, league_cfg.key):
        return _intel_not_ready_response(league_cfg.key)

    payload = await run_in_threadpool(_build_intel_leads, league_cfg, asset_id, mode)
    return JSONResponse(content=payload, headers=_INTEL_PRIVATE_CACHE_HEADERS)


@app.get("/api/intel/waiver-interest")
async def get_intel_waiver_interest(request: Request):
    """Waiver and free-agent activity, under its OWN label.

    A SEPARATE endpoint rather than a flag on the board, because the
    defect this whole change undoes was waiver churn rendering as trade
    "buys".  Fields here are ``adds``/``drops``, never ``buys``/``sells``.
    """
    try:
        league_cfg = _resolve_league_for_request(request)
    except LeagueResolutionError as err:
        return err.json_response()
    if not await run_in_threadpool(_intel_service.snapshot_ready, league_cfg.key):
        return _intel_not_ready_response(league_cfg.key)
    try:
        limit = int(request.query_params.get("limit") or 100)
    except (TypeError, ValueError):
        limit = 100
    limit = max(1, min(500, limit))
    window = (request.query_params.get("window") or "").strip() or _INTEL_DEFAULT_WINDOW
    if window not in _INTEL_ALLOWED_WINDOWS:
        window = _INTEL_DEFAULT_WINDOW
    payload = await run_in_threadpool(
        _intel_service.build_waiver_interest_payload,
        league_cfg.key,
        limit=limit,
        id_to_player=_intel_id_to_player(),
        window=window,
    )
    return JSONResponse(content=payload, headers=_INTEL_PRIVATE_CACHE_HEADERS)


@app.get("/api/intel/player")
async def get_intel_player(request: Request):
    """Per-asset intel drill-down.  ``?playerId=`` (Sleeper id or
    ``pick:<season>:<round>``) preferred; ``?name=`` resolves through
    the loaded contract's id map.  League-scoped like the summary.

    ``?window=`` takes the same allow-list as the board and MUST be
    passed the window the row was rendered from."""
    try:
        league_cfg = _resolve_league_for_request(request)
    except LeagueResolutionError as err:
        return err.json_response()
    if not await run_in_threadpool(_intel_service.snapshot_ready, league_cfg.key):
        return _intel_not_ready_response(league_cfg.key)
    asset_id = str(request.query_params.get("playerId") or "").strip()
    if not asset_id:
        name = str(request.query_params.get("name") or "").strip()
        if not name:
            return JSONResponse(
                status_code=400,
                content={"error": "missing_param", "message": "playerId or name required"},
                headers={"Cache-Control": "no-store"},
            )
        asset_id = _intel_name_to_player_id(name) or ""
        if not asset_id:
            return JSONResponse(
                status_code=404,
                content={"error": "unknown_player", "message": f"No Sleeper id for {name!r}"},
                headers={"Cache-Control": "no-store"},
            )
    # Same allow-list as the board.  The drill-down MUST be scoped to
    # the window the row was rendered from, or a 90d-only asset expands
    # to "no league-mate holds or traded this asset" under a row that
    # just reported a buy.
    window = (request.query_params.get("window") or "").strip() or _INTEL_DEFAULT_WINDOW
    if window not in _INTEL_ALLOWED_WINDOWS:
        window = _INTEL_DEFAULT_WINDOW
    payload = await run_in_threadpool(
        _intel_service.build_player_payload,
        league_cfg.key,
        asset_id,
        id_to_player=_intel_id_to_player(),
        window=window,
    )
    if payload is None:
        return JSONResponse(
            status_code=404,
            content={
                "error": "no_intel",
                "message": "No intel recorded for this asset",
                "leagueKey": league_cfg.key,
            },
            headers={"Cache-Control": "no-store"},
        )
    return JSONResponse(content=payload, headers=_INTEL_PRIVATE_CACHE_HEADERS)


@app.get("/api/intel/member/{owner_id}")
async def get_intel_member(owner_id: str, request: Request):
    """One league-mate's cross-league profile: league count/names,
    truncation flag, and their recent per-asset activity.
    League-scoped like the summary."""
    try:
        league_cfg = _resolve_league_for_request(request)
    except LeagueResolutionError as err:
        return err.json_response()
    if not await run_in_threadpool(_intel_service.snapshot_ready, league_cfg.key):
        return _intel_not_ready_response(league_cfg.key)
    payload = await run_in_threadpool(
        _intel_service.build_member_payload,
        league_cfg.key,
        owner_id,
        id_to_player=_intel_id_to_player(),
    )
    if payload is None:
        return JSONResponse(
            status_code=404,
            content={
                "error": "unknown_member",
                "message": f"No intel for owner {owner_id!r}",
                "leagueKey": league_cfg.key,
            },
            headers={"Cache-Control": "no-store"},
        )
    return JSONResponse(content=payload, headers=_INTEL_PRIVATE_CACHE_HEADERS)


@app.post("/api/intel/refresh")
async def post_intel_refresh(request: Request):
    """Kick off a crawl on a daemon thread and return 202 immediately
    — the crawl takes minutes (budgeted Sleeper calls with an
    inter-call sleep) and must never run inline.  409 when a run is
    already active.  Session OR bearer auth (see section header).

    ``?leagueKey=all`` refreshes every ACTIVE registry league
    sequentially (the cron's mode); any other key resolves through
    the standard league resolver."""
    is_cron = _intel_bearer_auth_ok(request)
    session = None if is_cron else _get_auth_session(request)
    if not is_cron and not session:
        return JSONResponse(
            status_code=401,
            content={"error": "auth_required", "message": "Sign-in or bearer token required."},
            headers={"Cache-Control": "no-store"},
        )

    # D13: per-user cooldown on MANUAL triggers.  The cron (bearer) is
    # exempt — it is the intended scheduled driver.  A crawl is minutes
    # of budgeted Sleeper calls, and the process lock only prevented
    # concurrent runs, not a user re-triggering after each one finished.
    cooldown_key = ""
    if not is_cron:
        cooldown_key = (
            str((session or {}).get("username") if isinstance(session, dict) else session)
            or "anonymous"
        )
        # ORDER MATTERS: "a crawl is already running" is both the more
        # informative answer and the more actionable one, so 409 wins
        # when both apply.  Checking the cooldown first turned every
        # mid-crawl retry into an opaque 429.  The gap between this
        # check and the start below is racy, but the process lock still
        # guarantees correctness — the worst case is a 429 where a 409
        # would have read better.
        already_running = bool(_intel_service.refresh_status().get("isRunning"))
        remaining = 0 if already_running else _intel_refresh_cooldown_remaining(cooldown_key)
        if remaining > 0:
            return JSONResponse(
                status_code=429,
                content={
                    "error": "refresh_cooldown",
                    "message": (
                        f"A manual refresh was triggered recently. Try again in "
                        f"{remaining}s, or wait for the daily crawl."
                    ),
                    "retryAfterSeconds": remaining,
                },
                headers={"Cache-Control": "no-store", "Retry-After": str(remaining)},
            )

    # ``leagueKey=all`` — refresh EVERY active league sequentially
    # under the single crawl lock.  This is what the daily cron
    # sends: a bearer request has no user session, so the normal
    # resolver would silently fall back to the default league and
    # non-default leagues would stay data_not_ready forever.
    requested_key = str(request.query_params.get("leagueKey") or "").strip().lower()
    if requested_key == "all":
        active = _league_registry.active_leagues()
        if not active:
            return JSONResponse(
                status_code=404,
                content={
                    "error": "no_leagues_configured",
                    "message": "No active leagues configured on this server",
                },
                headers={"Cache-Control": "no-store"},
            )
        try:
            status = _intel_service.start_refresh_async(
                leagues=[
                    {"leagueKey": cfg.key, "sleeperLeagueId": cfg.sleeper_league_id}
                    for cfg in active
                ],
            )
        except _intel_service.RefreshAlreadyRunning:
            return JSONResponse(
                status_code=409,
                content={
                    "error": "already_running",
                    "alreadyRunning": True,
                    "status": _intel_service.refresh_status(),
                },
                headers={"Cache-Control": "no-store"},
            )
        if cooldown_key:
            _intel_refresh_mark_triggered(cooldown_key)
        return JSONResponse(
            status_code=202,
            content={
                "message": "Intel refresh started for all active leagues",
                "leagueKey": "all",
                "leagueKeys": [cfg.key for cfg in active],
                "status": status,
            },
            headers={"Cache-Control": "no-store"},
        )

    try:
        league_cfg = _resolve_league_for_request(request)
    except LeagueResolutionError as err:
        return err.json_response()
    try:
        status = _intel_service.start_refresh_async(
            league_key=league_cfg.key,
            sleeper_league_id=league_cfg.sleeper_league_id,
        )
    except _intel_service.RefreshAlreadyRunning:
        return JSONResponse(
            status_code=409,
            content={
                "error": "already_running",
                "alreadyRunning": True,
                "status": _intel_service.refresh_status(),
            },
            headers={"Cache-Control": "no-store"},
        )
    if cooldown_key:
        _intel_refresh_mark_triggered(cooldown_key)
    return JSONResponse(
        status_code=202,
        content={
            "message": "Intel refresh started in background",
            "leagueKey": league_cfg.key,
            "status": status,
        },
        headers={"Cache-Control": "no-store"},
    )


@app.get("/api/intel/refresh/status")
async def get_intel_refresh_status(request: Request):
    """Crawl status (process-global — one run at a time) + the
    resolved league's snapshot staleness.  Session OR bearer auth so
    the cron workflow can poll it."""
    if not _intel_bearer_auth_ok(request) and not _get_auth_session(request):
        return JSONResponse(
            status_code=401,
            content={"error": "auth_required", "message": "Sign-in or bearer token required."},
            headers={"Cache-Control": "no-store"},
        )
    # League resolution here only scopes the staleness stamp — the
    # run status itself is global, so a resolution failure (e.g. a
    # bare-bones dev box with no registry) degrades to status-only.
    league_key = None
    try:
        league_key = _resolve_league_for_request(request).key
    except LeagueResolutionError:
        league_key = None
    return JSONResponse(
        content=_intel_service.refresh_status(league_key),
        headers={"Cache-Control": "no-store"},
    )


# ── SHARP TRACKER ──────────────────────────────────────────────────────
# A SEPARATE product from Insider Trading above.  Its cohort is
# skill-qualified and drawn from every league we observe, so it is
# deliberately NOT league-scoped — no ``_resolve_league_for_request``
# here, and that omission is the point.  Filtering this board to the
# caller's league would silently rebuild the merged feature this work
# exists to undo.

from src.sharp import service as _sharp_service  # noqa: E402

# Register the sharp market routes EXPLICITLY rather than relying on the
# import-time side effect inside ``_sharp_service``.
#
# That side effect only finds this app when ``server`` is the first thing
# to import the module. Anything that imports ``src.sharp.service``
# earlier — a test module, a script, another package — runs the
# registrar against a not-yet-existing app, and Python's module cache
# then means importing it here re-runs nothing. The routes silently
# never attach and every ``/api/sharp/market`` request 404s.
#
# ``_register_http_routes`` is idempotent (it returns early when the
# path is already present), so calling it here is safe alongside the
# module-level call and the self-heal in ``cohort_status``.
_sharp_service.register_http_routes()

# ``curated_service`` was previously imported by nothing but its own test,
# which left /api/sharp/people, /api/sharp/people/{id}, /api/sharp/review,
# /api/sharp/review/{id}, /api/sharp/curated/summary and
# /api/sharp/curated/refresh returning 404, and with them the whole
# /market/sharp-people and /admin/sharp-identities surface.  Registered
# explicitly here for the same reason as the tracker above: an import-time
# side effect does not re-run if the module is already in ``sys.modules``.
from src.sharp import curated_service as _sharp_curated_service  # noqa: E402

_sharp_curated_service._register_http_routes()


@app.get("/api/sharp/cohort")
async def get_sharp_cohort(request: Request):
    """Cohort status: observable / evaluable / qualified / uncertain.

    Always 200 with an explicit ``status`` rather than a 503 — "the
    cohort is still being built" is a real, expected state on a growing
    network, and the page renders it as an explanation.  The four tiers
    are always reported separately so "not built yet" can never be
    mistaken for "nobody qualified".
    """
    payload = await run_in_threadpool(_sharp_service.cohort_status)
    return JSONResponse(
        content=payload,
        headers={"Cache-Control": "private, max-age=300, stale-while-revalidate=900"},
    )


@app.get("/api/sharp/roster-percentage")
async def get_sharp_roster_percentage(request: Request):
    """Which players the sharp cohort rosters most, and how reliably.

    Shares the Buy/Sell Tracker's cohort exactly — both boards resolve
    their pool through ``src/sharp/cohort.py::cohort_members`` — so it is
    global for the same reason the cohort endpoint is, and takes no
    ``leagueKey``.

    Declared HERE rather than registered from ``src/sharp/service.py``.
    That module's self-registration only finds the app when it is
    imported by ``server.py`` first; when a test imports it earlier the
    routes never attach, which is the failure
    ``tests/sharp/test_public_api_allowlist.py`` hits under a full-suite
    run.  A plain decorator has no such ordering dependency.

    Always 200 with an explicit ``status``: an empty roster store is a
    real state on a cohort whose rosters have not been collected yet,
    and the page explains it rather than showing an error.
    """
    query = request.query_params

    def _bool(name: str) -> bool:
        return str(query.get(name) or "").strip().lower() in ("1", "true", "yes")

    try:
        payload = await run_in_threadpool(
            _sharp_service.roster_percentage_payload,
            contract=latest_contract_data,
            qualification=str(query.get("qualification") or "all"),
            position=str(query.get("position") or "all"),
            platform=str(query.get("platform") or "all"),
            league_format=str(query.get("format") or "all"),
            contention=str(query.get("contention") or "all"),
            experience=str(query.get("experience") or "all"),
            sort=str(query.get("sort") or "rostered"),
            limit=int(query.get("limit") or 50),
            include_picks=_bool("includePicks"),
        )
    except (ValueError, TypeError) as exc:
        return JSONResponse(
            status_code=400,
            content={"error": "bad_request", "message": str(exc)},
            headers={"Cache-Control": "no-store"},
        )
    except Exception as exc:  # noqa: BLE001
        log.exception("sharp roster percentage failed")
        return JSONResponse(
            status_code=503,
            content={"error": "sharp_roster_percentage_unavailable", "message": str(exc)},
            headers={"Cache-Control": "no-store"},
        )
    return JSONResponse(
        content=payload,
        headers={"Cache-Control": "private, max-age=300, stale-while-revalidate=900"},
    )


@app.get("/api/sharp/roster-percentage/audit")
async def get_sharp_roster_percentage_audit(request: Request):
    """Every roster behind one player's percentage, listed individually.

    The manual-verification surface required before this feature could
    be called done: a count is checkable against the underlying rosters
    without database access, and the excluded rosters are reported with
    their reasons alongside.
    """
    asset_id = str(request.query_params.get("assetId") or "").strip()
    if not asset_id:
        return JSONResponse(
            status_code=400,
            content={"error": "missing_param", "message": "assetId required"},
            headers={"Cache-Control": "no-store"},
        )
    try:
        payload = await run_in_threadpool(
            _sharp_service.roster_percentage_audit_payload,
            asset_id,
            qualification=str(request.query_params.get("qualification") or "all"),
        )
    except (ValueError, TypeError) as exc:
        return JSONResponse(
            status_code=400,
            content={"error": "bad_request", "message": str(exc)},
            headers={"Cache-Control": "no-store"},
        )
    except Exception as exc:  # noqa: BLE001
        log.exception("sharp roster percentage audit failed")
        return JSONResponse(
            status_code=503,
            content={"error": "sharp_roster_percentage_unavailable", "message": str(exc)},
            headers={"Cache-Control": "no-store"},
        )
    return JSONResponse(content=payload, headers={"Cache-Control": "private, max-age=60"})


# ── PLAYER CONTEXT (contracts / snap share / depth chart) ───────────────
# R2 player-profile surface.  Serves per-player blocks from the
# ``src/playerctx`` snapshot (``data/playerctx/snapshot.json``,
# produced by ``scripts/refresh_playerctx.py``).  This is GLOBAL player
# metadata — public NFL data joined to the Sleeper pool — so it follows
# the scoring-profile side of the CLAUDE.md split: no league resolution,
# no per-league branching.  "No snapshot yet" is a normal state (the UI
# degrades silently), so missing data is a clean 404, never a 5xx.
_playerctx_cache: dict[str, object] = {"snapshot": None, "mtime": None}


def _playerctx_snapshot() -> dict | None:
    """Mtime-cached read of the playerctx snapshot.  Re-reads only when
    the file's mtime changes (weekly refresh cadence), so the per-player
    endpoint never pays the ~1 MB JSON parse per request."""
    from src.playerctx.store import SNAPSHOT_PATH, load_snapshot

    try:
        mtime = SNAPSHOT_PATH.stat().st_mtime
    except OSError:
        _playerctx_cache["snapshot"] = None
        _playerctx_cache["mtime"] = None
        return None
    if _playerctx_cache["mtime"] != mtime:
        _playerctx_cache["snapshot"] = load_snapshot()
        _playerctx_cache["mtime"] = mtime
    return _playerctx_cache["snapshot"]  # type: ignore[return-value]


def _playerctx_lookup(player_id: str) -> dict | None:
    """Resolve a Sleeper player id to its context record, or None.
    Record keys are gsis ids when known, else ``sleeper:<id>`` — always
    go through ``sleeperIndex`` (docs/playerctx.md consumer rules)."""
    snapshot = _playerctx_snapshot()
    if not snapshot:
        return None
    record_key = (snapshot.get("sleeperIndex") or {}).get(str(player_id))
    if not record_key:
        return None
    record = (snapshot.get("players") or {}).get(record_key)
    if not isinstance(record, dict):
        return None
    return {"player": record, "generatedAt": snapshot.get("generatedAt")}


@app.get("/api/playerctx/player")
async def get_playerctx_player(request: Request):
    """Per-player context blocks (contract / snaps / depth) for the
    profile card.  ``?playerId=`` is the Sleeper id (the only stable
    join key the frontend holds).  404 with a machine-readable error
    when the snapshot is missing or doesn't cover the player — both are
    normal, silent-degrade states for the UI."""
    player_id = str(request.query_params.get("playerId") or "").strip()
    if not player_id:
        return JSONResponse(
            status_code=400,
            content={"error": "missing_param", "message": "playerId required"},
            headers={"Cache-Control": "no-store"},
        )
    payload = await run_in_threadpool(_playerctx_lookup, player_id)
    if payload is None:
        return JSONResponse(
            status_code=404,
            content={
                "error": "no_context",
                "message": "No player-context data for this player",
            },
            headers={"Cache-Control": "no-store"},
        )
    # Weekly-cadence data: cacheable for an hour client-side.
    return JSONResponse(
        content=payload,
        headers={"Cache-Control": "private, max-age=3600"},
    )


# ── MAIN ────────────────────────────────────────────────────────────────
if __name__ == "__main__":
    import uvicorn

    print()
    print("  ╔═══════════════════════════════════════════╗")
    print("  ║   Dynasty Trade Calculator — Server       ║")
    print(f"  ║   Dashboard: http://localhost:{PORT:<13}║")
    print(f"  ║   Scrape interval: {SCRAPE_INTERVAL_HOURS}h{' ' * 21}║")
    print(f"  ║   Alerts: {'ON → ' + ALERT_TO[:20] if ALERT_ENABLED else 'OFF':<30}║")
    print("  ╚═══════════════════════════════════════════╝")
    print()

    uvicorn.run(
        "server:app",
        host=HOST,
        port=PORT,
        log_level="info",
        reload=False,
    )
